#!/usr/bin/env python3
"""
Excel Allocation System - Web Application
Admin can upload allocation and data files, Agent can upload status files
"""

from flask import (
    Flask,
    render_template_string,
    request,
    jsonify,
    send_file,
    redirect,
    session,
    url_for,
    flash,
)
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
import pandas as pd
import os
import re
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import hashlib
import tempfile
import io
import uuid
import json
from functools import wraps
from urllib.parse import quote
from dotenv import load_dotenv
import base64

# Google OAuth imports
from google.auth.transport import requests
from google.oauth2 import id_token
import requests as req

# Scheduler for reminder emails
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger
from apscheduler.triggers.cron import CronTrigger
import pytz
import resend

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB max file size
app.config["SECRET_KEY"] = os.environ.get(
    "SECRET_KEY", "your-secret-key-change-in-production"
)

# Configure Flask to trust proxy headers (required for Railway/Heroku HTTPS detection)
# This allows Flask to detect HTTPS from X-Forwarded-Proto header
if os.environ.get("DATABASE_URL") or os.environ.get("RAILWAY_ENVIRONMENT"):
    # Production environment - trust proxy headers
    from werkzeug.middleware.proxy_fix import ProxyFix

    app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

# Database configuration
DATABASE_URL = os.environ.get("DATABASE_URL")
if DATABASE_URL:
    # For Railway/Heroku deployment
    if DATABASE_URL.startswith("postgres://"):
        DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
else:
    # For local development
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///excel_allocation.db"

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# Initialize database
db = SQLAlchemy(app)
migrate = Migrate(app, db)

# Resend email configuration
resend.api_key = os.environ.get("RESEND_API_KEY")

# Global variable to store agent allocations data for reminders
agent_allocations_for_reminders = None

# Google OAuth Configuration
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET")
GOOGLE_DISCOVERY_URL = "https://accounts.google.com/.well-known/openid_configuration"


# Database Models
class User(db.Model):
    """User model for authentication and employee management"""

    __tablename__ = "users"

    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(
        db.String(80), unique=True, nullable=True
    )  # Made nullable for OAuth users
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(
        db.String(255), nullable=True
    )  # Made nullable for OAuth users
    role = db.Column(db.String(20), nullable=False, default="agent")  # admin, agent
    name = db.Column(db.String(100), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)
    # Google OAuth fields
    google_id = db.Column(db.String(100), unique=True, nullable=True)
    auth_provider = db.Column(db.String(20), default="local")  # local, google

    # Relationships
    sessions = db.relationship(
        "UserSession", backref="user", lazy=True, cascade="all, delete-orphan"
    )
    allocations = db.relationship("Allocation", backref="user", lazy=True)

    def set_password(self, password):
        """Hash and set password"""
        # Use pbkdf2:sha256 method to avoid scrypt compatibility issues
        self.password_hash = generate_password_hash(password, method="pbkdf2:sha256")

    def check_password(self, password):
        """Check if provided password matches hash"""
        # Check if hash uses scrypt (which isn't available)
        if self.password_hash and self.password_hash.startswith("scrypt:"):
            # Can't verify scrypt hash - return False to trigger rehashing in login route
            return False

        try:
            return check_password_hash(self.password_hash, password)
        except (AttributeError, ValueError) as e:
            # Handle case where hashlib.scrypt is not available or other hash issues
            if "scrypt" in str(e).lower() or (
                self.password_hash and "scrypt" in self.password_hash.lower()
            ):
                # If hash uses scrypt, we can't verify it - return False to trigger rehashing
                return False
            # For other errors, re-raise
            raise

    def to_dict(self):
        """Convert user to dictionary"""
        return {
            "id": self.id,
            "username": self.username,
            "email": self.email,
            "role": self.role,
            "name": self.name,
            "is_active": self.is_active,
            "created_at": self.created_at.isoformat() if self.created_at else None,
            "last_login": self.last_login.isoformat() if self.last_login else None,
        }


class UserSession(db.Model):
    """User session model for database-based session management"""

    __tablename__ = "user_sessions"

    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    session_data = db.Column(db.Text)  # JSON string
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    expires_at = db.Column(db.DateTime, nullable=False)
    is_active = db.Column(db.Boolean, default=True)

    def set_data(self, data):
        """Set session data as JSON string"""
        self.session_data = json.dumps(data)

    def get_data(self):
        """Get session data from JSON string"""
        if self.session_data:
            return json.loads(self.session_data)
        return {}

    def is_expired(self):
        """Check if session is expired"""
        return datetime.utcnow() > self.expires_at


class Allocation(db.Model):
    """Allocation model for storing file processing data"""

    __tablename__ = "allocations"

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    allocation_filename = db.Column(db.String(255))
    data_filename = db.Column(db.String(255))
    allocation_data = db.Column(db.Text)  # JSON string
    data_file_data = db.Column(db.Text)  # JSON string
    processing_result = db.Column(db.Text)
    agent_allocations_data = db.Column(db.Text)  # JSON string
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(
        db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow
    )

    def set_allocation_data(self, data):
        """Set allocation data as JSON string"""
        if data is not None:
            # Convert pandas DataFrames to JSON-serializable format
            if isinstance(data, dict):
                serializable_data = {}
                for key, value in data.items():
                    if isinstance(value, pd.DataFrame):
                        # Convert DataFrame to records and handle Timestamps
                        df_records = value.to_dict("records")
                        # Convert any Timestamp objects to strings
                        for record in df_records:
                            for k, v in record.items():
                                if hasattr(v, "isoformat"):  # Check if it's a Timestamp
                                    record[k] = v.isoformat()
                        serializable_data[key] = df_records
                    else:
                        serializable_data[key] = value
                self.allocation_data = json.dumps(serializable_data)
            elif isinstance(data, pd.DataFrame):
                # Convert DataFrame to records and handle Timestamps
                df_records = data.to_dict("records")
                # Convert any Timestamp objects to strings
                for record in df_records:
                    for k, v in record.items():
                        if hasattr(v, "isoformat"):  # Check if it's a Timestamp
                            record[k] = v.isoformat()
                self.allocation_data = json.dumps(df_records)
            else:
                self.allocation_data = json.dumps(data)
        else:
            self.allocation_data = None

    def get_allocation_data(self):
        """Get allocation data from JSON string"""
        if self.allocation_data:
            data = json.loads(self.allocation_data)
            # Convert back to pandas DataFrames if needed
            if isinstance(data, dict):
                converted_data = {}
                for key, value in data.items():
                    if (
                        isinstance(value, list)
                        and len(value) > 0
                        and isinstance(value[0], dict)
                    ):
                        # This looks like a DataFrame converted to records
                        converted_data[key] = pd.DataFrame(value)
                    else:
                        converted_data[key] = value
                return converted_data
            elif isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
                # This looks like a single DataFrame converted to records
                return pd.DataFrame(data)
            else:
                return data
        return None

    def set_data_file_data(self, data):
        """Set data file data as JSON string"""
        if data is not None:
            # Convert pandas DataFrames to JSON-serializable format
            if isinstance(data, dict):
                serializable_data = {}
                for key, value in data.items():
                    if isinstance(value, pd.DataFrame):
                        # Convert DataFrame to records and handle Timestamps
                        df_records = value.to_dict("records")
                        # Convert any Timestamp objects to strings
                        for record in df_records:
                            for k, v in record.items():
                                if hasattr(v, "isoformat"):  # Check if it's a Timestamp
                                    record[k] = v.isoformat()
                        serializable_data[key] = df_records
                    else:
                        serializable_data[key] = value
                self.data_file_data = json.dumps(serializable_data)
            elif isinstance(data, pd.DataFrame):
                # Convert DataFrame to records and handle Timestamps
                df_records = data.to_dict("records")
                # Convert any Timestamp objects to strings
                for record in df_records:
                    for k, v in record.items():
                        if hasattr(v, "isoformat"):  # Check if it's a Timestamp
                            record[k] = v.isoformat()
                self.data_file_data = json.dumps(df_records)
            else:
                self.data_file_data = json.dumps(data)
        else:
            self.data_file_data = None

    def get_data_file_data(self):
        """Get data file data from JSON string"""
        if self.data_file_data:
            data = json.loads(self.data_file_data)
            # Convert back to pandas DataFrames if needed
            if isinstance(data, dict):
                converted_data = {}
                for key, value in data.items():
                    if (
                        isinstance(value, list)
                        and len(value) > 0
                        and isinstance(value[0], dict)
                    ):
                        # This looks like a DataFrame converted to records
                        converted_data[key] = pd.DataFrame(value)
                    else:
                        converted_data[key] = value
                return converted_data
            elif isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
                # This looks like a single DataFrame converted to records
                return pd.DataFrame(data)
            else:
                return data
        return None

    def set_agent_allocations_data(self, data):
        """Set agent allocations data as JSON string"""
        if data is not None:
            # Convert pandas DataFrames to JSON-serializable format
            if isinstance(data, dict):
                serializable_data = {}
                for key, value in data.items():
                    if isinstance(value, pd.DataFrame):
                        # Convert DataFrame to records and handle Timestamps
                        df_records = value.to_dict("records")
                        # Convert any Timestamp objects to strings
                        for record in df_records:
                            for k, v in record.items():
                                if hasattr(v, "isoformat"):  # Check if it's a Timestamp
                                    record[k] = v.isoformat()
                        serializable_data[key] = df_records
                    else:
                        serializable_data[key] = value
                self.agent_allocations_data = json.dumps(serializable_data)
            elif isinstance(data, pd.DataFrame):
                # Convert DataFrame to records and handle Timestamps
                df_records = data.to_dict("records")
                # Convert any Timestamp objects to strings
                for record in df_records:
                    for k, v in record.items():
                        if hasattr(v, "isoformat"):  # Check if it's a Timestamp
                            record[k] = v.isoformat()
                self.agent_allocations_data = json.dumps(df_records)
            else:
                self.agent_allocations_data = json.dumps(data)
        else:
            self.agent_allocations_data = None

    def get_agent_allocations_data(self):
        """Get agent allocations data from JSON string"""
        if self.agent_allocations_data:
            data = json.loads(self.agent_allocations_data)
            # Convert back to pandas DataFrames if needed
            if isinstance(data, dict):
                converted_data = {}
                for key, value in data.items():
                    if (
                        isinstance(value, list)
                        and len(value) > 0
                        and isinstance(value[0], dict)
                    ):
                        # This looks like a DataFrame converted to records
                        converted_data[key] = pd.DataFrame(value)
                    else:
                        converted_data[key] = value
                return converted_data
            elif isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
                # This looks like a single DataFrame converted to records
                return pd.DataFrame(data)
            else:
                return data
        return None


class AgentWorkFile(db.Model):
    """Agent work file model for storing agent uploads"""

    __tablename__ = "agent_work_files"

    id = db.Column(db.Integer, primary_key=True)
    agent_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    file_data = db.Column(db.Text)  # JSON string of processed data
    upload_date = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(
        db.String(50), default="uploaded"
    )  # uploaded, processed, consolidated
    notes = db.Column(db.Text)  # Optional notes from agent

    # Relationships
    agent = db.relationship("User", backref="work_files")

    def set_file_data(self, data):
        """Set file data as JSON string"""
        if data is not None:
            # Convert pandas DataFrames to JSON-serializable format
            if isinstance(data, dict):
                serializable_data = {}
                for key, value in data.items():
                    if isinstance(value, pd.DataFrame):
                        # Convert DataFrame to records and handle Timestamps
                        df_records = value.to_dict("records")
                        # Convert any Timestamp objects to strings
                        for record in df_records:
                            for k, v in record.items():
                                if hasattr(v, "isoformat"):  # Check if it's a Timestamp
                                    record[k] = v.isoformat()
                        serializable_data[key] = df_records
                    else:
                        serializable_data[key] = value
                self.file_data = json.dumps(serializable_data)
            elif isinstance(data, pd.DataFrame):
                # Convert DataFrame to records and handle Timestamps
                df_records = data.to_dict("records")
                # Convert any Timestamp objects to strings
                for record in df_records:
                    for k, v in record.items():
                        if hasattr(v, "isoformat"):  # Check if it's a Timestamp
                            record[k] = v.isoformat()
                self.file_data = json.dumps(df_records)
            else:
                self.file_data = json.dumps(data)
        else:
            self.file_data = None

    def get_file_data(self):
        """Get file data from JSON string"""
        if self.file_data:
            data = json.loads(self.file_data)
            # Convert back to pandas DataFrames if needed
            if isinstance(data, dict):
                converted_data = {}
                for key, value in data.items():
                    if (
                        isinstance(value, list)
                        and len(value) > 0
                        and isinstance(value[0], dict)
                    ):
                        # This looks like a DataFrame converted to records
                        converted_data[key] = pd.DataFrame(value)
                    else:
                        converted_data[key] = value
                return converted_data
            elif isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
                # This looks like a single DataFrame converted to records
                return pd.DataFrame(data)
            else:
                return data
        return None


# Global variables to store session data (fallback for backward compatibility)
allocation_data = None
data_file_data = None
allocation_filename = None
data_filename = None
processing_result = None

# Agent processing result
agent_processing_result = None
agent_allocations_data = None
agent_insurance_agent_names = (
    None  # Store agent names for Agent Insurance sheet formatting
)


# Database helper functions
def init_database():
    """Initialize database and create default users"""
    with app.app_context():
        db.create_all()

        # Create default admin user if it doesn't exist, or reset password if it exists
        admin_user = User.query.filter_by(username="admin").first()
        if not admin_user:
            admin_user = User(
                username="admin",
                email="admin@example.com",
                role="admin",
                name="Administrator",
            )
            admin_user.set_password("admin123")
            db.session.add(admin_user)
        else:
            # Reset password to admin123 if user already exists
            admin_user.set_password("admin123")
            admin_user.role = "admin"  # Ensure role is admin
            admin_user.is_active = True  # Ensure user is active

        # Note: Agent users will be created automatically via Google OAuth
        # No need to create static agent accounts

        db.session.commit()


def get_user_by_username(username):
    """Get user by username"""
    return User.query.filter_by(username=username, is_active=True).first()


def create_user_session(user_id, session_data=None, expires_hours=24):
    """Create a new user session"""
    expires_at = datetime.utcnow() + timedelta(hours=expires_hours)
    session = UserSession(user_id=user_id, expires_at=expires_at)
    if session_data:
        session.set_data(session_data)
    db.session.add(session)
    db.session.commit()
    return session


def get_user_session(session_id):
    """Get user session by ID"""
    return UserSession.query.filter_by(id=session_id, is_active=True).first()


def delete_user_session(session_id):
    """Delete user session"""
    session = UserSession.query.filter_by(id=session_id).first()
    if session:
        session.is_active = False
        db.session.commit()


def cleanup_expired_sessions():
    """Clean up expired sessions"""
    expired_sessions = UserSession.query.filter(
        UserSession.expires_at < datetime.utcnow()
    ).all()
    for session in expired_sessions:
        session.is_active = False
    db.session.commit()


def get_or_create_allocation(user_id):
    """Get or create allocation record for user"""
    allocation = Allocation.query.filter_by(user_id=user_id).first()
    if not allocation:
        allocation = Allocation(user_id=user_id)
        db.session.add(allocation)
        db.session.commit()
    return allocation


def save_allocation_data(
    user_id,
    allocation_data=None,
    allocation_filename=None,
    data_file_data=None,
    data_filename=None,
    processing_result=None,
    agent_allocations_data=None,
):
    """Save allocation data to database"""
    allocation = get_or_create_allocation(user_id)

    if allocation_data is not None:
        allocation.set_allocation_data(allocation_data)
    if allocation_filename is not None:
        allocation.allocation_filename = allocation_filename
    if data_file_data is not None:
        allocation.set_data_file_data(data_file_data)
    if data_filename is not None:
        allocation.data_filename = data_filename
    if processing_result is not None:
        allocation.processing_result = processing_result
    if agent_allocations_data is not None:
        allocation.set_agent_allocations_data(agent_allocations_data)

    allocation.updated_at = datetime.utcnow()
    db.session.commit()
    return allocation


def get_allocation_data(user_id):
    """Get allocation data from database"""
    allocation = Allocation.query.filter_by(user_id=user_id).first()
    if allocation:
        return {
            "allocation_data": allocation.get_allocation_data(),
            "allocation_filename": allocation.allocation_filename,
            "data_file_data": allocation.get_data_file_data(),
            "data_filename": allocation.data_filename,
            "processing_result": allocation.processing_result,
            "agent_allocations_data": allocation.get_agent_allocations_data(),
        }
    return None


def save_agent_work_file(agent_id, filename, file_data, notes=None):
    """Save agent work file to database"""
    work_file = AgentWorkFile(agent_id=agent_id, filename=filename, notes=notes)
    work_file.set_file_data(file_data)
    db.session.add(work_file)
    db.session.commit()
    return work_file


def get_agent_work_files(agent_id=None):
    """Get agent work files, optionally filtered by agent"""
    if agent_id:
        return (
            AgentWorkFile.query.filter_by(agent_id=agent_id)
            .order_by(AgentWorkFile.upload_date.desc())
            .all()
        )
    return AgentWorkFile.query.order_by(AgentWorkFile.upload_date.desc()).all()


def get_all_agent_work_files():
    """Get all agent work files for consolidation (admin view)"""
    # Return all files regardless of status so admin can see all uploaded files
    return AgentWorkFile.query.order_by(AgentWorkFile.upload_date.desc()).all()


# Template filter to convert datetime to IST
@app.template_filter("to_ist")
def to_ist_filter(dt):
    """Convert datetime to IST (Asia/Kolkata) timezone"""
    if dt is None:
        return None

    # If datetime is naive (no timezone info), assume it's UTC
    if dt.tzinfo is None:
        dt = pytz.UTC.localize(dt)

    # Convert to IST
    ist_timezone = pytz.timezone("Asia/Kolkata")
    ist_time = dt.astimezone(ist_timezone)

    return ist_time


# Google OAuth helper functions
def get_google_provider_cfg():
    """Get Google OAuth provider configuration"""
    # Use hardcoded Google OAuth endpoints instead of discovery
    return {
        "authorization_endpoint": "https://accounts.google.com/o/oauth2/v2/auth",
        "token_endpoint": "https://oauth2.googleapis.com/token",
        "userinfo_endpoint": "https://www.googleapis.com/oauth2/v3/userinfo",
    }


def verify_google_token(token):
    """Verify Google OAuth token and return user info"""
    try:
        # Verify the token
        idinfo = id_token.verify_oauth2_token(
            token, requests.Request(), GOOGLE_CLIENT_ID
        )

        # Verify the issuer
        if idinfo["iss"] not in ["accounts.google.com", "https://accounts.google.com"]:
            raise ValueError("Wrong issuer.")

        return {
            "google_id": idinfo["sub"],
            "email": idinfo["email"],
            "name": idinfo["name"],
            "picture": idinfo.get("picture", ""),
        }
    except ValueError as e:
        return None


def get_or_create_google_user(google_user_info):
    """Get existing user or create new user from Google OAuth info"""
    # First try to find by Google ID
    user = User.query.filter_by(google_id=google_user_info["google_id"]).first()

    if user:
        return user

    # If not found by Google ID, try to find by email
    user = User.query.filter_by(email=google_user_info["email"]).first()

    if user:
        # Update existing user with Google ID
        user.google_id = google_user_info["google_id"]
        user.auth_provider = "google"
        user.name = google_user_info["name"]
        db.session.commit()
        return user

    # Create new user
    user = User(
        email=google_user_info["email"],
        name=google_user_info["name"],
        google_id=google_user_info["google_id"],
        auth_provider="google",
        role="agent",  # Default role for OAuth users
        is_active=True,
    )
    db.session.add(user)
    db.session.commit()
    return user


# Authentication decorators
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Check for database session first
        db_session_id = session.get("db_session_id")
        if db_session_id:
            db_session = get_user_session(db_session_id)
            if db_session and not db_session.is_expired():
                # Update session data in Flask session
                session_data = db_session.get_data()
                session.update(session_data)
                return f(*args, **kwargs)
            else:
                # Session expired, clean up
                if db_session:
                    delete_user_session(db_session_id)
                session.clear()

        # Fallback to Flask session
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Check for database session first
        db_session_id = session.get("db_session_id")
        if db_session_id:
            db_session = get_user_session(db_session_id)
            if db_session and not db_session.is_expired():
                session_data = db_session.get_data()
                if session_data.get("user_role") != "admin":
                    flash("Access denied. Admin privileges required.", "error")
                    return redirect(url_for("dashboard"))
                session.update(session_data)
                return f(*args, **kwargs)
            else:
                if db_session:
                    delete_user_session(db_session_id)
                session.clear()

        # Fallback to Flask session
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("user_role") != "admin":
            flash("Access denied. Admin privileges required.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated_function


def agent_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Check for database session first
        db_session_id = session.get("db_session_id")
        if db_session_id:
            db_session = get_user_session(db_session_id)
            if db_session and not db_session.is_expired():
                session_data = db_session.get_data()
                if session_data.get("user_role") != "agent":
                    flash("Access denied. Agent privileges required.", "error")
                    return redirect(url_for("dashboard"))
                session.update(session_data)
                return f(*args, **kwargs)
            else:
                if db_session:
                    delete_user_session(db_session_id)
                session.clear()

        # Fallback to Flask session
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("user_role") != "agent":
            flash("Access denied. Agent privileges required.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated_function


# Email helper function using Resend
def send_email_with_resend(
    to_email,
    subject,
    html_content,
    text_content=None,
    attachment_data=None,
    attachment_filename=None,
):
    """
    Send email using Resend API

    Args:
        to_email: Recipient email address
        subject: Email subject
        html_content: HTML email content
        text_content: Plain text email content (optional)
        attachment_data: BytesIO or bytes object for attachment (optional)
        attachment_filename: Filename for attachment (optional)

    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        email_data = {
            "from": os.environ.get("RESEND_FROM_EMAIL", "onboarding@excellabs.shop"),
            "to": to_email,
            "subject": subject,
            "html": html_content,
        }

        # Add text content if provided
        if text_content:
            email_data["text"] = text_content

        # Add attachment if provided
        if attachment_data and attachment_filename:
            # Convert BytesIO to bytes if needed
            if isinstance(attachment_data, io.BytesIO):
                attachment_bytes = attachment_data.getvalue()
            else:
                attachment_bytes = attachment_data

            # Encode to base64
            attachment_base64 = base64.b64encode(attachment_bytes).decode("utf-8")

            email_data["attachments"] = [
                {"filename": attachment_filename, "content": attachment_base64}
            ]

        r = resend.Emails.send(email_data)

        # Check if email was sent successfully
        # Resend API returns an object with 'id' field on success
        if isinstance(r, dict) and "id" in r:
            return True, f"Email sent successfully to {to_email}"
        elif hasattr(r, "id"):
            return True, f"Email sent successfully to {to_email}"
        else:
            return False, f"Failed to send email: {str(r)}"

    except Exception as e:
        return False, f"Error sending email: {str(e)}"


# Login Template
LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Login - Excel Allocation System</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { 
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }
        .login-container { 
            background: white; 
            border-radius: 15px; 
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            padding: 40px;
            width: 100%;
            max-width: 400px;
        }
        .login-header {
            text-align: center;
            margin-bottom: 30px;
        }
        .login-header h1 {
            color: #333;
            font-size: 2em;
            margin-bottom: 10px;
        }
        .login-header p {
            color: #666;
            font-size: 1.1em;
        }
        .form-group {
            margin-bottom: 20px;
        }
        .form-group label {
            display: block;
            margin-bottom: 8px;
            font-weight: 600;
            color: #555;
        }
        .form-group input {
            width: 100%;
            padding: 12px;
            border: 2px solid #ddd;
            border-radius: 8px;
            font-size: 16px;
            transition: border-color 0.3s;
        }
        .form-group input:focus {
            outline: none;
            border-color: #667eea;
        }
        .login-btn {
            width: 100%;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 12px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-size: 16px;
            font-weight: 600;
            transition: transform 0.2s;
        }
        .login-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(0,0,0,0.2);
        }
        .alert {
            padding: 12px;
            border-radius: 8px;
            margin-bottom: 20px;
            font-weight: 500;
        }
        .alert-error {
            background: #f8d7da;
            color: #721c24;
            border: 1px solid #f5c6cb;
        }
        .alert-success {
            background: #d4edda;
            color: #155724;
            border: 1px solid #c3e6cb;
        }
        .demo-credentials {
            background: #f8f9fa;
            padding: 15px;
            border-radius: 8px;
            margin-top: 20px;
            font-size: 14px;
        }
        .demo-credentials h4 {
            color: #333;
            margin-bottom: 10px;
        }
        .demo-credentials p {
            color: #666;
            margin: 5px 0;
        }
        .demo-credentials strong {
            color: #333;
        }
        .google-login-btn {
            width: 100%;
            background: #4285f4;
            color: white;
            padding: 12px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-size: 16px;
            font-weight: 600;
            transition: transform 0.2s;
            margin-top: 15px;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 10px;
        }
        .google-login-btn:hover {
            background: #3367d6;
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(66, 133, 244, 0.3);
        }
        .divider {
            text-align: center;
            margin: 20px 0;
            position: relative;
            color: #666;
        }
        .divider::before {
            content: '';
            position: absolute;
            top: 50%;
            left: 0;
            right: 0;
            height: 1px;
            background: #ddd;
        }
        .divider span {
            background: white;
            padding: 0 15px;
            position: relative;
        }
    </style>
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
</head>
<body>
    <div class="login-container">
        <div class="login-header">
            <h1><i class="fas fa-file-excel"></i> Excel Allocation System</h1>
            <p>Agents: Use Google Login | Admins: Use username/password</p>
        </div>
        
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, message in messages %}
                    <div class="alert alert-{{ category }}">{{ message }}</div>
                {% endfor %}
            {% endif %}
        {% endwith %}
        
        <form method="POST" action="{{ url_for('login') }}">
            <div class="form-group">
                <label for="username"><i class="fas fa-user"></i> Username</label>
                <input type="text" id="username" name="username" required>
            </div>
            
            <div class="form-group">
                <label for="password"><i class="fas fa-lock"></i> Password</label>
                <input type="password" id="password" name="password" required>
            </div>
            
            <button type="submit" class="login-btn">
                <i class="fas fa-sign-in-alt"></i> Login
            </button>
        </form>
        
        <div class="divider">
            <span>OR</span>
        </div>
        
        {% if GOOGLE_CLIENT_ID %}
        <a href="{{ url_for('google_login') }}" class="google-login-btn">
            <i class="fab fa-google"></i> Login with Google
        </a>
        {% else %}
        <div style="background: #fff3cd; border: 1px solid #ffeaa7; color: #856404; padding: 15px; border-radius: 8px; margin-top: 15px; text-align: center;">
            <i class="fas fa-exclamation-triangle"></i>
            <strong>Google OAuth not configured</strong><br>
            <small>Contact administrator to set up Google OAuth for agent login</small>
        </div>
            {% endif %}
    </div>
</body>
</html>
"""

# HTML Template for Excel Allocation System
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel Allocation System</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { 
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        .container { 
            max-width: 1400px; 
            margin: 0 auto; 
            background: white; 
            border-radius: 15px; 
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            overflow: hidden;
        }
        .header { 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white; 
            padding: 30px; 
            text-align: center; 
        }
        .header h1 { font-size: 2.5em; margin-bottom: 10px; }
        .header p { font-size: 1.2em; opacity: 0.9; }
        
        .role-selector {
            display: flex;
            justify-content: center;
            gap: 20px;
            margin-top: 20px;
        }
        .role-btn {
            padding: 12px 24px;
            border: none;
            border-radius: 25px;
            background: rgba(255, 255, 255, 0.2);
            color: white;
            cursor: pointer;
            transition: all 0.3s ease;
            font-weight: 600;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        .role-btn:hover {
            background: rgba(255, 255, 255, 0.3);
            transform: translateY(-2px);
        }
        .role-btn.active {
            background: rgba(255, 255, 255, 0.9);
            color: #667eea;
            box-shadow: 0 4px 15px rgba(255, 255, 255, 0.3);
        }
        
        .admin-tab-content {
            display: none;
        }
        
        .admin-tab-content.active {
            display: block;
        }
        
        .admin-tab-btn.active {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
            color: white !important;
            border-bottom-color: #667eea !important;
        }
        
        /* Toast Notifications */
        .toast-container {
            position: fixed;
            top: 20px;
            right: 20px;
            z-index: 9999;
        }
        
        .toast {
            background: white;
            border-radius: 8px;
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
            padding: 16px 20px;
            margin-bottom: 10px;
            min-width: 300px;
            max-width: 400px;
            display: flex;
            align-items: center;
            gap: 12px;
            transform: translateX(100%);
            transition: transform 0.3s ease;
            border-left: 4px solid #28a745;
        }
        
        .toast.show {
            transform: translateX(0);
        }
        
        .toast.success {
            border-left-color: #28a745;
            background: linear-gradient(135deg, #d4edda 0%, #f8f9fa 100%);
            border: 1px solid #c3e6cb;
        }
        
        .toast.error {
            border-left-color: #dc3545;
            background: linear-gradient(135deg, #f8d7da 0%, #f8f9fa 100%);
            border: 1px solid #f5c6cb;
        }
        
        .toast.warning {
            border-left-color: #ffc107;
        }
        
        .toast.info {
            border-left-color: #17a2b8;
        }
        
        .toast-icon {
            font-size: 20px;
            flex-shrink: 0;
        }
        
        .toast.success .toast-icon {
            color: #28a745;
        }
        
        .toast.error .toast-icon {
            color: #dc3545;
        }
        
        .toast.warning .toast-icon {
            color: #ffc107;
        }
        
        .toast.info .toast-icon {
            color: #17a2b8;
        }
        
        .toast-content {
            flex: 1;
        }
        
        .toast-title {
            font-weight: 600;
            margin: 0 0 4px 0;
            color: #333;
        }
        
        /* Loader/Progress Bar */
        .loader-overlay {
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: rgba(0, 0, 0, 0.5);
            display: flex;
            justify-content: center;
            align-items: center;
            z-index: 10000;
            opacity: 0;
            visibility: hidden;
            transition: opacity 0.3s ease, visibility 0.3s ease;
        }
        
        .loader-overlay.show {
            opacity: 1;
            visibility: visible;
        }
        
        .loader-container {
            background: white;
            border-radius: 15px;
            padding: 40px 50px;
            box-shadow: 0 10px 40px rgba(0, 0, 0, 0.3);
            text-align: center;
            max-width: 400px;
        }
        
        .loader-spinner {
            border: 4px solid #f3f3f3;
            border-top: 4px solid #667eea;
            border-radius: 50%;
            width: 50px;
            height: 50px;
            animation: spin 1s linear infinite;
            margin: 0 auto 20px;
        }
        
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        
        .loader-text {
            font-size: 18px;
            color: #667eea;
            font-weight: 600;
            margin: 0;
        }
        
        .loader-subtitle {
            font-size: 14px;
            color: #666;
            margin-top: 8px;
        }
        
        .progress-bar-container {
            width: 100%;
            height: 6px;
            background: #e0e0e0;
            border-radius: 3px;
            margin-top: 20px;
            overflow: hidden;
        }
        
        .progress-bar {
            height: 100%;
            background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
            border-radius: 3px;
            width: 0%;
            transition: width 0.3s ease;
            animation: progress 1.5s ease-in-out infinite;
        }
        
        @keyframes progress {
            0% { width: 0%; }
            50% { width: 70%; }
            100% { width: 100%; }
        }
        
        .toast.success .toast-title {
            color: #155724;
        }
        
        .toast.error .toast-title {
            color: #721c24;
        }
        
        .toast-message {
            margin: 0;
            color: #666;
            font-size: 14px;
        }
        
        .toast.success .toast-message {
            color: #155724;
        }
        
        .toast.error .toast-message {
            color: #721c24;
        }
        
        .toast-close {
            background: none;
            border: none;
            font-size: 18px;
            color: #999;
            cursor: pointer;
            padding: 0;
            width: 20px;
            height: 20px;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        
        .toast-close:hover {
            color: #333;
        }
        
        .content { padding: 30px; }
        .panel { display: none; }
        .panel.active { display: block; }
        
        .section { 
            margin: 25px 0; 
            padding: 25px; 
            border: 1px solid #e0e0e0; 
            border-radius: 10px; 
            background: #fafafa;
        }
        .section h3 { 
            color: #333; 
            margin-bottom: 20px; 
            font-size: 1.4em;
            border-bottom: 2px solid #667eea;
            padding-bottom: 10px;
        }
        
        .upload-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(400px, 1fr));
            gap: 30px;
            margin-bottom: 30px;
        }
        
        .upload-card {
            background: white;
            border-radius: 15px;
            padding: 25px;
            border: 2px dashed #dee2e6;
            transition: all 0.3s ease;
            text-align: center;
        }
        .upload-card:hover {
            border-color: #667eea;
            transform: translateY(-5px);
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.1);
        }
        
        .upload-header {
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 15px;
            margin-bottom: 20px;
        }
        .upload-header i {
            font-size: 1.5rem;
            color: #27ae60;
        }
        .upload-header h4 {
            color: #2c3e50;
            font-size: 1.3rem;
        }
        
        .form-group { margin: 15px 0; }
        label { 
            display: block; 
            margin-bottom: 8px; 
            font-weight: 600; 
            color: #555;
        }
        input[type="file"] { 
            width: 100%; 
            padding: 12px; 
            border: 2px solid #ddd; 
            border-radius: 8px; 
            font-size: 16px;
            transition: border-color 0.3s;
        }
        input[type="file"]:focus { 
            outline: none; 
            border-color: #667eea; 
        }
        
        button { 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white; 
            padding: 12px 25px; 
            border: none; 
            border-radius: 8px; 
            cursor: pointer; 
            margin: 5px; 
            font-size: 16px;
            font-weight: 600;
            transition: transform 0.2s;
        }
        button:hover { 
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(0,0,0,0.2);
        }
        button:disabled {
            background: #bdc3c7;
            cursor: not-allowed;
            transform: none;
            box-shadow: none;
        }
        
        .process-btn {
            background: linear-gradient(135deg, #27ae60, #2ecc71);
            font-size: 18px;
            padding: 15px 40px;
            border-radius: 25px;
            display: flex;
            align-items: center;
            gap: 10px;
            margin: 20px auto;
        }
        
        .file-status { 
            background: #f8f9fa; 
            padding: 15px; 
            border-radius: 8px; 
            margin: 15px 0; 
            border-left: 4px solid #667eea;
        }
        .status-success { 
            background: #d4edda; 
            color: #155724; 
            border-color: #c3e6cb; 
        }
        .status-info { 
            background: #d1ecf1; 
            color: #0c5460; 
            border-color: #bee5eb; 
        }
        
        .status-message {
            background: #f3e5f5;
            border: 2px solid #9c27b0;
            color: #4a148c;
            padding: 20px;
            border-radius: 10px;
            margin: 15px 0;
            font-size: 16px;
            line-height: 1.6;
            white-space: pre-line;
            box-shadow: 0 2px 8px rgba(156, 39, 176, 0.2);
        }
        
        .processing-status {
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0, 0, 0, 0.8);
            display: none;
            justify-content: center;
            align-items: center;
            z-index: 9999;
            text-align: center;
            color: white;
        }
        
        .processing-content {
            background: white;
            color: #333;
            padding: 40px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.3);
            max-width: 500px;
            width: 90%;
        }
        
        .spinner {
            width: 50px;
            height: 50px;
            border: 4px solid #e9ecef;
            border-top: 4px solid #667eea;
            border-radius: 50%;
            animation: spin 1s linear infinite;
            margin: 0 auto 20px;
        }
        
        .progress-container {
            width: 100%;
            background-color: #e0e0e0;
            border-radius: 10px;
            margin: 20px 0;
            overflow: hidden;
        }
        
        .progress-bar {
            height: 30px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            border-radius: 10px;
            width: 0%;
            transition: width 0.3s ease;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: bold;
        }
        
        .progress-text {
            margin-top: 10px;
            font-size: 16px;
            color: #666;
        }
        
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        
        .results-section {
            background: #f8f9fa;
            border-radius: 15px;
            padding: 25px;
            border-left: 5px solid #27ae60;
            margin: 20px 0;
        }
        
        .results-section h3 {
            color: #27ae60;
            margin-bottom: 20px;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        
        .results-content {
            background: white;
            border-radius: 10px;
            padding: 20px;
            border: 1px solid #e9ecef;
            white-space: pre-wrap;
            font-family: 'Courier New', monospace;
            max-height: 400px;
            overflow-y: auto;
        }
        
        .coming-soon {
            text-align: center;
            padding: 60px 20px;
            color: #7f8c8d;
        }
        .coming-soon i {
            font-size: 4rem;
            margin-bottom: 20px;
            color: #bdc3c7;
        }
        .coming-soon h3 {
            font-size: 1.5rem;
            margin-bottom: 10px;
        }
        
        .reset-btn {
            background: linear-gradient(135deg, #ff6b6b, #ee5a52);
            color: white;
            border: none;
            padding: 12px 24px;
            border-radius: 8px;
            cursor: pointer;
            font-size: 16px;
            font-weight: 600;
            transition: all 0.3s ease;
            box-shadow: 0 4px 15px rgba(255, 107, 107, 0.3);
        }
        .reset-btn:hover {
            background: linear-gradient(135deg, #ff5252, #d32f2f);
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(255, 107, 107, 0.4);
        }
        
        .priority-panel {
            background: white;
            border-radius: 10px;
            padding: 20px;
            margin: 10px 0;
            border: 2px solid #e9ecef;
        }
        
        .priority-panel.active {
            border-color: #667eea;
            box-shadow: 0 4px 15px rgba(102, 126, 234, 0.1);
        }
        
        .tab-button {
            transition: all 0.3s ease;
            opacity: 0.7;
        }
        
        .tab-button:hover {
            opacity: 1;
            transform: translateY(-2px);
        }
        
        .tab-button.active {
            opacity: 1;
            transform: translateY(-2px);
            box-shadow: 0 4px 8px rgba(0,0,0,0.2);
        }
        
        @media (max-width: 768px) {
            .upload-grid {
                grid-template-columns: 1fr;
            }
            .role-selector {
                flex-direction: column;
                align-items: center;
            }
            .header h1 {
                font-size: 2rem;
            }
        }
        
        /* Table styling */
        .agent-table tbody tr:hover {
            background-color: #f8f9fa;
        }
        
        .agent-table .process-btn:hover {
            transform: scale(1.05);
        }
        
        /* Modal styling */
        .modal {
            animation: fadeIn 0.3s ease;
        }
        
        .modal-content {
            animation: slideIn 0.3s ease;
        }
        
        .close:hover {
            opacity: 0.7;
        }
        
        @keyframes fadeIn {
            from { opacity: 0; }
            to { opacity: 1; }
        }
        
        @keyframes slideIn {
            from { transform: translateY(-50px); opacity: 0; }
            to { transform: translateY(0); opacity: 1; }
        }
        
        .modal-table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
        }
        
        .modal-table th,
        .modal-table td {
            padding: 8px 12px;
            text-align: left;
            border-bottom: 1px solid #e9ecef;
        }
        
        .modal-table th {
            background-color: #f8f9fa;
            font-weight: 600;
            color: #333;
        }
        
        .modal-table tr:hover {
            background-color: #f8f9fa;
        }
        
        /* Style for serial number column */
        .modal-table th:first-child,
        .modal-table td:first-child {
            text-align: center;
            width: 60px;
            font-weight: 600;
            color: #667eea;
            background-color: #f0f2ff;
        }
        
        .modal-table th:first-child {
            background-color: #e8ecff;
        }
        
        /* Show all agent rows */
        .agent-row {
            display: table-row;
        }
    </style>
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
</head>
<body>
    <div class="container">
        <div class="header">
            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px;">
                <div>
            <h1><i class="fas fa-file-excel"></i> Excel Allocation System</h1>
            <p>Upload and process Excel files for allocation management</p>
                </div>
                <div style="display: flex; align-items: center; gap: 15px;">
                    <div style="color: white; text-align: right;">
                        <div style="font-size: 1.1em; font-weight: 600;">Welcome, {{ session.user_name }}</div>
                        <div style="font-size: 0.9em; opacity: 0.8;">{{ session.user_role.title() }} User</div>
                    </div>
                    <a href="{{ url_for('logout') }}" style="
                        background: rgba(255, 255, 255, 0.2);
                        color: white;
                        padding: 8px 16px;
                        border-radius: 20px;
                        text-decoration: none;
                        font-weight: 600;
                        transition: all 0.3s ease;
                        display: flex;
                        align-items: center;
                        gap: 8px;
                    " onmouseover="this.style.background='rgba(255, 255, 255, 0.3)'" onmouseout="this.style.background='rgba(255, 255, 255, 0.2)'">
                        <i class="fas fa-sign-out-alt"></i> Logout
                    </a>
                </div>
            </div>
            
            {% if session.user_role == 'admin' %}
            {% endif %}
        </div>

        <div class="content">
            <!-- Admin Panel -->
            <div id="admin-panel" class="panel active">
                <!-- Admin Tab Navigation -->
                <div class="admin-tabs" style="margin-bottom: 30px;">
                    <div class="tab-nav" style="display: flex; border-bottom: 2px solid #e9ecef; margin-bottom: 20px;">
                        <button class="admin-tab-btn active" onclick="switchAdminTab('file-management')" style="padding: 12px 24px; border: none; background: #f8f9fa; color: #666; cursor: pointer; border-bottom: 3px solid transparent; transition: all 0.3s;">
                            <i class="fas fa-upload"></i> File Management
                        </button>
                        <button class="admin-tab-btn" onclick="switchAdminTab('agent-allocation')" style="padding: 12px 24px; border: none; background: #f8f9fa; color: #666; cursor: pointer; border-bottom: 3px solid transparent; transition: all 0.3s;">
                            <i class="fas fa-users"></i> Agent Allocation
                        </button>
                        <button class="admin-tab-btn" onclick="switchAdminTab('agent-consolidation')" style="padding: 12px 24px; border: none; background: #f8f9fa; color: #666; cursor: pointer; border-bottom: 3px solid transparent; transition: all 0.3s;">
                            <i class="fas fa-compress-arrows-alt"></i> Agent Consolidation
                        </button>
                        <button class="admin-tab-btn" onclick="switchAdminTab('system-settings')" style="padding: 12px 24px; border: none; background: #f8f9fa; color: #666; cursor: pointer; border-bottom: 3px solid transparent; transition: all 0.3s;">
                            <i class="fas fa-cog"></i> System Settings
                        </button>
                    </div>
                </div>
                
                <!-- File Management Tab -->
                <div id="file-management-tab" class="admin-tab-content active">
                <div class="upload-grid">
                    <div class="upload-card">
                        <form action="/upload_allocation" method="post" enctype="multipart/form-data" id="allocation-form">
                            <div class="form-group">
                                <input type="file" id="allocation_file" name="file" accept=".xlsx,.xls" required>
                            </div>
                            <button type="submit" id="allocation-btn">📤 Upload Staff Details</button>
                        </form>
                    </div>

                    <div class="upload-card">
                        <form action="/upload_data" method="post" enctype="multipart/form-data" id="data-form">
                            <div class="form-group">
                                <input type="file" id="data_file" name="file" accept=".xlsx,.xls" required>
                            </div>
                            <button type="submit" id="data-btn">📤 Upload Insurance Details</button>
                        </form>
                    </div>
                </div>

                <!-- File Status -->
                <div class="section">
                    <h3>📊 File Status</h3>
                    <div class="file-status">
                        {% if allocation_filename %}
                            <div class="status-success">
                                ✅ Allocation File: {{ allocation_filename }}<br>
                                📋 Sheets: {{ allocation_data.keys() | list | length if allocation_data else 0 }}
                            </div>
                        {% else %}
                            <div class="status-info">
                                ℹ️ No agent allocation details file uploaded yet.
                            </div>
                        {% endif %}
                        
                        {% if data_filename %}
                            <div class="status-success">
                                ✅ Data File: {{ data_filename }}<br>
                                📋 Sheets: {{ data_file_data.keys() | list | length if data_file_data else 0 }}
                            </div>
                        {% else %}
                            <div class="status-info">
                                ℹ️ No insurance details file uploaded yet.
                            </div>
                        {% endif %}
                    </div>
                </div>


                <!-- Processing Section -->
                {% if data_file_data %}
                <div class="section">
                    <h3>🔄 Process Data File</h3>
                    
                    <!-- Priority Date Selection -->
                    <div class="section" style="background: #f8f9fa; margin-bottom: 20px;">
                        
                        <!-- Priority Tabs -->
                        <div class="tab-container" style="margin-bottom: 20px;">
                            <div class="tab-buttons" style="display: flex; border-bottom: 2px solid #ddd;">
                                <div class="tab-button active" id="first-priority-tab" onclick="switchPriorityTab('first')" style="padding: 12px 24px; cursor: pointer; background: #27ae60; color: white; border-radius: 8px 8px 0 0; margin-right: 2px; font-weight: bold;">First Priority</div>
                                <div class="tab-button" id="second-priority-tab" onclick="switchPriorityTab('second')" style="padding: 12px 24px; cursor: pointer; background: #f39c12; color: white; border-radius: 8px 8px 0 0; margin-right: 2px; font-weight: bold;">Second Priority</div>
                                <div class="tab-button" id="third-priority-tab" onclick="switchPriorityTab('third')" style="padding: 12px 24px; cursor: pointer; background: #e74c3c; color: white; border-radius: 8px 8px 0 0; font-weight: bold;">Third Priority</div>
                            </div>
                        </div>
                        
                        <!-- First Priority Panel -->
                        <div id="first-priority-panel" class="priority-panel" style="display: block;">
                            
                            
                            <!-- Calendar Input for First Priority Dates -->
                            <div class="form-group">
                                <div id="calendar_container" style="border: 1px solid #ddd; padding: 15px; background: white; border-radius: 8px; margin: 10px 0;"></div>
                                <div id="selected_dates_info" style="background: #f8f9fa; padding: 10px; border-radius: 5px; border: 1px solid #e9ecef;">
                                    <strong>Selected First Priority Dates:</strong> <span id="selected_count">0</span> <span id="selected_text">dates selected</span>
                                    <div id="selected_dates_list" style="margin-top: 5px; font-size: 12px; color: #666;"></div>
                                </div>
                            </div>
                            
                            <!-- Receive Date Panel (appears when appointment dates are selected) -->
                            <div id="receive-date-panel" style="display: none; margin-top: 20px; padding: 15px; background: #f0f8ff; border: 1px solid #b3d9ff; border-radius: 8px;">
                                <h4 style="margin: 0 0 15px 0; color: #2c5aa0; font-size: 16px;">
                                    <i class="fas fa-calendar-check"></i> Second Level Priority - Receive Dates
                                </h4>
                                <p style="margin: 0 0 15px 0; color: #666; font-size: 14px;">
                                    Select receive dates for each appointment date. Each appointment date has its own independent receive date selections.
                                </p>
                                <div id="appointment_receive_dates_container">
                                    <!-- Individual receive date panels for each appointment date will be populated here -->
                                </div>
                            </div>
                        </div>
                        
                        <!-- Second Priority Panel -->
                        <div id="second-priority-panel" class="priority-panel" style="display: none;">
                            
                            
                            <!-- Calendar Input for Second Priority Dates -->
                            <div class="form-group">
                                <div id="calendar_container_second" style="border: 1px solid #ddd; padding: 15px; background: white; border-radius: 8px; margin: 10px 0;"></div>
                                <div id="selected_dates_info_second" style="background: #f8f9fa; padding: 10px; border-radius: 5px; border: 1px solid #e9ecef;">
                                    <strong>Selected Second Priority Dates:</strong> <span id="selected_count_second">0</span> <span id="selected_text_second">dates selected</span>
                                    <div id="selected_dates_list_second" style="margin-top: 5px; font-size: 12px; color: #666;">No dates selected</div>
                                </div>
                            </div>
                        </div>
                        
                        <!-- Third Priority Panel -->
                        <div id="third-priority-panel" class="priority-panel" style="display: none;">
                            <p>All remaining dates will be automatically assigned "Third Priority":</p>
                            
                            <!-- Info about Third Priority -->
                            
                            <!-- Show remaining dates that will be Third Priority -->
                            <div class="form-group">
                                <div id="third_priority_dates_info" style="background: #f8f9fa; padding: 10px; border-radius: 5px; border: 1px solid #e9ecef;">
                                    <strong>Remaining Dates:</strong> <span id="third_priority_count">0</span> dates will be Third Priority
                                    <div id="third_priority_dates_list" style="margin-top: 5px; font-size: 12px; color: #666;"></div>
                                </div>
                            </div>
                        </div>
                        
                        <form action="/process_files" method="post" id="process-form">
                            <button type="submit" class="process-btn" id="process-btn">
                                <i class="fas fa-cogs"></i> Process Data File
                            </button>
                        </form>
                    </div>
                    
                    <div class="processing-status" id="processing-status">
                        <div class="processing-content">
                            <div class="spinner"></div>
                            <h3>Processing Data File...</h3>
                            <div class="progress-container">
                                <div class="progress-bar" id="progress-bar">0%</div>
                            </div>
                            <div class="progress-text" id="progress-text">Initializing...</div>
                        </div>
                    </div>
                </div>
                {% endif %}

                <!-- Status Messages -->
                {% if processing_result %}
                <div class="section">
                    <h3>📢 Processing Results</h3>
                    <div class="status-message">
                        {{ processing_result | safe }}
                    </div>
                </div>
                {% endif %}

                <!-- Download Section -->
                {% if processing_result and 'Priority processing completed successfully' in processing_result %}
                <div class="section">
                    <h3>💾 Download your Excel file with updated Priority Status assignments.</h3>
                    <form action="/download_result" method="post">
                        <button type="submit" class="process-btn" style="background: linear-gradient(135deg, #3498db, #2980b9);">
                            <i class="fas fa-download"></i> Download Processed File
                        </button>
                    </form>
                </div>
                {% endif %}

                </div>
                
                <!-- Agent Allocation Tab -->
                <div id="agent-allocation-tab" class="admin-tab-content">
                    <!-- Individual Agent Downloads -->
                    {% if agent_allocations_data %}
                    <div class="section">
                        <h3>👥 Agent Allocation Overview</h3>
                        <p>View and manage agent allocations. Each agent has been assigned specific rows based on their capacity and the allocation rules.</p>
                        
                        <div style="margin-bottom: 15px; text-align: right; display: flex; gap: 10px; justify-content: flex-end;">
                            <button type="button" class="process-btn" style="background: linear-gradient(135deg, #3498db, #2980b9); font-size: 14px; padding: 10px 20px; border: none; border-radius: 6px; color: white; cursor: pointer; transition: transform 0.2s; font-weight: 600;" onclick="viewShiftTimes()" id="view-shift-btn">
                                <i class="fas fa-clock"></i> View Shift Times
                            </button>
                            <button type="button" class="process-btn" style="background: linear-gradient(135deg, #27ae60, #2ecc71); font-size: 14px; padding: 10px 20px; border: none; border-radius: 6px; color: white; cursor: pointer; transition: transform 0.2s; font-weight: 600;" onclick="approveAllAllocations()" id="approve-all-btn">
                                <i class="fas fa-check-double"></i> Approve All Allocations
                            </button>
                        </div>
                        
                        <div style="overflow-x: auto; margin-top: 15px;">
                            <table class="agent-table" style="width: 100%; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                                <thead>
                                    <tr style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white;">
                                        <th style="padding: 15px; text-align: center; font-weight: 600; border: none; width: 60px;">Sr No</th>
                                        <th style="padding: 15px; text-align: left; font-weight: 600; border: none;">Agent Name</th>
                                        <th style="padding: 15px; text-align: center; font-weight: 600; border: none;">Allocated</th>
                                        <th style="padding: 15px; text-align: center; font-weight: 600; border: none;">Capacity</th>
                                        <th style="padding: 15px; text-align: center; font-weight: 600; border: none;">Actions</th>
                                    </tr>
                                </thead>
                                <tbody id="agentTableBody">
                                    {% for agent in agent_allocations_data %}
                                    <tr class="agent-row" style="border-bottom: 1px solid #e9ecef; transition: background-color 0.2s;" data-index="{{ loop.index0 }}">
                                        <td style="padding: 15px; text-align: center; font-weight: 600; color: #667eea;">{{ loop.index }}</td>
                                        <td style="padding: 15px; font-weight: 500; color: #333;">{{ agent.name }}</td>
                                        <td style="padding: 15px; text-align: center; color: #27ae60; font-weight: 600;">{{ agent.allocated }}</td>
                                        <td style="padding: 15px; text-align: center; color: #666;">{{ agent.capacity }}</td>
                                        <td style="padding: 15px; text-align: center;">
                                            <div style="display: flex; gap: 8px; justify-content: center;">
                                                <button type="button" class="process-btn view-btn" style="background: linear-gradient(135deg, #f39c12, #e67e22); font-size: 12px; padding: 6px 12px; border: none; border-radius: 4px; color: white; cursor: pointer; transition: transform 0.2s;" onclick="viewAgentAllocation('{{ agent.name }}')">
                                                    <i class="fas fa-eye"></i> View
                                                </button>
                                                <button type="button" class="process-btn approve-btn" style="background: linear-gradient(135deg, #3498db, #2980b9); font-size: 12px; padding: 6px 12px; border: none; border-radius: 4px; color: white; cursor: pointer; transition: transform 0.2s;" onclick="approveAllocation('{{ agent.name }}')">
                                                    <i class="fas fa-check"></i> Approve
                                                </button>
                                            </div>
                                        </td>
                                    </tr>
                                    {% endfor %}
                                </tbody>
                            </table>
                        </div>
                    </div>
                    {% else %}
                    <div class="section">
                        <h3>👥 Agent Allocation Overview</h3>
                        <div style="background: #f8f9fa; padding: 20px; border-radius: 10px;">
                            <p style="color: #666;">No agent allocation data available. Please upload allocation and data files, then process them to see agent allocations.</p>
                        </div>
                    </div>
                    {% endif %}
                    
                    <!-- Shift Times Modal -->
                    <div id="shiftTimesModal" class="modal" style="display: none; position: fixed; z-index: 1000; left: 0; top: 0; width: 100%; height: 100%; background-color: rgba(0,0,0,0.5);">
                        <div class="modal-content" style="background-color: #fefefe; margin: 5% auto; padding: 0; border: none; border-radius: 10px; width: 90%; max-width: 1400px; max-height: 85vh; overflow: hidden; box-shadow: 0 10px 40px rgba(0,0,0,0.3);">
                            <div style="background: linear-gradient(135deg, #3498db, #2980b9); color: white; padding: 20px; border-radius: 10px 10px 0 0; display: flex; justify-content: space-between; align-items: center;">
                                <h2 style="margin: 0; font-size: 1.5em;"><i class="fas fa-clock"></i> Agent Shift Times</h2>
                                <span class="close" onclick="document.getElementById('shiftTimesModal').style.display='none'" style="color: white; font-size: 28px; font-weight: bold; cursor: pointer; line-height: 1;">&times;</span>
                            </div>
                            <div style="padding: 20px; max-height: calc(85vh - 80px); overflow-y: auto;" id="shiftTimesContent">
                                <!-- Content will be loaded here -->
                            </div>
                        </div>
                    </div>
                    
                    <!-- Agent Allocation Modal -->
                    <div id="agentModal" class="modal" style="display: none; position: fixed; z-index: 1000; left: 0; top: 0; width: 100%; height: 100%; background-color: rgba(0,0,0,0.5);">
                        <div class="modal-content" style="background-color: #fefefe; margin: 5% auto; padding: 0; border: none; border-radius: 10px; width: 90%; max-width: 1200px; max-height: 80vh; overflow: hidden; box-shadow: 0 10px 30px rgba(0,0,0,0.3);">
                            <div class="modal-header" style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; display: flex; justify-content: space-between; align-items: center;">
                                <h2 style="margin: 0; font-size: 1.5em;" id="modalAgentName">Agent Allocation</h2>
                                <span class="close" style="color: white; font-size: 28px; font-weight: bold; cursor: pointer; transition: opacity 0.3s;">&times;</span>
                            </div>
                            <div class="modal-body" style="padding: 20px; max-height: 60vh; overflow-y: auto;">
                                <div id="modalContent">
                                    <div style="text-align: center; padding: 40px;">
                                        <i class="fas fa-spinner fa-spin" style="font-size: 2em; color: #667eea;"></i>
                                        <p style="margin-top: 15px; color: #666;">Loading agent allocation data...</p>
                                    </div>
                                </div>
                            </div>
                            <div class="modal-footer" style="background: #f8f9fa; padding: 15px 20px; border-top: 1px solid #e9ecef; display: flex; justify-content: space-between; align-items: center;">
                                <div id="modalStats" style="color: #666; font-size: 14px;"></div>
                                <div style="display: flex; gap: 10px;">
                                    <button id="downloadBtn" class="process-btn" style="background: linear-gradient(135deg, #27ae60, #2ecc71); padding: 8px 16px; border: none; border-radius: 5px; color: white; cursor: pointer; font-size: 14px;">
                                        <i class="fas fa-download"></i> Download Excel
                                    </button>
                                    <button class="close-btn process-btn" style="background: linear-gradient(135deg, #95a5a6, #7f8c8d); padding: 8px 16px; border: none; border-radius: 5px; color: white; cursor: pointer; font-size: 14px;">
                                        Close
                                    </button>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
                
                <!-- Agent Consolidation Tab -->
                <div id="agent-consolidation-tab" class="admin-tab-content">
                    <!-- Agent Files Consolidation -->
                    {% if all_agent_work_files %}
                    <div class="section">
                        <h3>📋 Agent Work Files Consolidation</h3>
                        <div style="background: #f8f9fa; padding: 20px; border-radius: 10px; margin-bottom: 20px;">
                            <h4>Available Agent Files:</h4>
                            {% for work_file in all_agent_work_files %}
                            <div style="border-bottom: {% if loop.last %}none{% else %}1px solid #dee2e6{% endif %}; padding: 10px 0; display: flex; justify-content: space-between; align-items: center;">
                                <div style="flex: 1;">
                                    <strong>{{ work_file.agent.name }}</strong> - {{ work_file.filename }}
                                    <br>
                                    <small style="color: #666;">
                                        Uploaded: {{ (work_file.upload_date | to_ist).strftime('%Y-%m-%d %I:%M %p') }} IST
                                        | Status: <span style="color: {% if work_file.status == 'uploaded' %}#28a745{% elif work_file.status == 'consolidated' %}#007bff{% else %}#6c757d{% endif %}">{{ work_file.status.title() }}</span>
                                    </small>
                                    {% if work_file.notes %}
                                    <br>
                                    <small style="color: #666;"><em>{{ work_file.notes }}</em></small>
                                    {% endif %}
                                </div>
                                <div style="margin-left: 15px;">
                                    <a href="/download_agent_work_file/{{ work_file.id }}" class="process-btn" style="padding: 8px 16px; text-decoration: none; display: inline-block; background: linear-gradient(135deg, #007bff, #0056b3); color: white; border-radius: 5px; font-size: 14px;">
                                        <i class="fas fa-download"></i> Download
                                    </a>
                                </div>
                            </div>
                            {% endfor %}
                        </div>
                        <div style="display: flex; gap: 10px; flex-wrap: wrap;">
                            <form action="/consolidate_agent_files" method="post" style="margin: 0;">
                                <button type="submit" class="process-btn" style="background: linear-gradient(135deg, #28a745, #20c997);">
                                    <i class="fas fa-compress-arrows-alt"></i> Consolidate All Agent Files
                                </button>
                            </form>
                            <form action="/clear_all_agent_files" method="post" style="margin: 0;" onsubmit="return confirm('Are you sure you want to delete all files? This action cannot be undone.');">
                                <button type="submit" class="process-btn" style="background: linear-gradient(135deg, #dc3545, #c82333);">
                                    <i class="fas fa-trash-alt"></i> Clear all files
                                </button>
                            </form>
                        </div>
                    </div>
                    {% else %}
                    <div class="section">
                        <h3>📋 Agent Work Files</h3>
                        <div style="background: #f8f9fa; padding: 20px; border-radius: 10px;">
                            <p style="color: #666;">No agent work files uploaded yet.</p>
                        </div>
                    </div>
                    {% endif %}
                </div>
                
                <!-- System Settings Tab -->
                <div id="system-settings-tab" class="admin-tab-content">
                    <!-- Reset Section -->
                    <div class="section">
                        <h3>🔄 Reset Application</h3>
                        <p>Clear all uploaded files and reset the application to start fresh. All agent work files will be preserved.</p>
                        <form action="/reset_app" method="post" onsubmit="return confirm('Are you sure you want to reset the application? This will clear all uploaded files and data. All agent work files will be preserved.')">
                            <button type="submit" class="reset-btn">🗑️ Reset Application</button>
                        </form>
                    </div>
                </div>
            </div>

            <!-- Agent Panel -->
            <div id="agent-panel" class="panel">
           
          
                
                <div class="section">
                    <h3><i class="fas fa-upload"></i> Upload Work File</h3>
                    <div class="upload-card">
                        <form id="agentUploadForm" enctype="multipart/form-data">
                            <div class="form-group">
                            
                                <input type="file" name="file" id="agentFile" accept=".xlsx,.xls" required>
                            </div>
                        
                            <button type="submit" class="process-btn" id="agentUploadBtn">
                                <i class="fas fa-upload"></i> Upload Work File
                            </button>
                        </form>
                    </div>
                </div>
                
                {% if agent_work_files %}
                <div class="section">
                    <h3><i class="fas fa-history"></i> My Uploaded Files</h3>
                    <div style="background: #f8f9fa; padding: 20px; border-radius: 10px;">
                        {% for work_file in agent_work_files %}
                        <div style="border-bottom: 1px solid #dee2e6; padding: 10px 0; {% if loop.last %}border-bottom: none;{% endif %}">
                            <div style="display: flex; justify-content: space-between; align-items: center;">
                                <div>
                                    <strong>{{ work_file.filename }}</strong>
                                    <br>
                                    <small style="color: #666;">
                                        Uploaded: {{ (work_file.upload_date | to_ist).strftime('%Y-%m-%d %I:%M %p') }} IST
                                        | Status: <span style="color: {% if work_file.status == 'uploaded' %}#28a745{% elif work_file.status == 'consolidated' %}#007bff{% else %}#6c757d{% endif %}">{{ work_file.status.title() }}</span>
                                    </small>
                                    {% if work_file.notes %}
                                    <br>
                                    <small style="color: #666;"><em>{{ work_file.notes }}</em></small>
                                    {% endif %}
                                </div>
                            </div>
                        </div>
                        {% endfor %}
                    </div>
                </div>
                {% endif %}
                
            
            </div>
        </div>
    </div>

    <!-- Toast Container -->
    <div class="toast-container" id="toastContainer"></div>

    <script>
        function switchRole(role) {
            // Update button states
            document.querySelectorAll('.role-btn').forEach(btn => btn.classList.remove('active'));
            event.target.classList.add('active');
            
            // Show/hide panels
            document.querySelectorAll('.panel').forEach(panel => panel.classList.remove('active'));
            document.getElementById(role + '-panel').classList.add('active');
        }
        
        function switchAdminTab(tabName) {
            // Update tab button states
            document.querySelectorAll('.admin-tab-btn').forEach(btn => btn.classList.remove('active'));
            event.target.classList.add('active');
            
            // Show/hide tab content
            document.querySelectorAll('.admin-tab-content').forEach(tab => tab.classList.remove('active'));
            document.getElementById(tabName + '-tab').classList.add('active');
        }
        
        // Toast Notification Functions
        function showToast(type, title, message, duration = 5000) {
            const container = document.getElementById('toastContainer');
            const toast = document.createElement('div');
            toast.className = `toast ${type}`;
            
            const icons = {
                success: 'fas fa-check-circle',
                error: 'fas fa-exclamation-circle',
                warning: 'fas fa-exclamation-triangle',
                info: 'fas fa-info-circle'
            };
            
            toast.innerHTML = `
                <i class="toast-icon ${icons[type]}"></i>
                <div class="toast-content">
                    <div class="toast-title">${title}</div>
                    <div class="toast-message">${message}</div>
                </div>
                <button class="toast-close" onclick="closeToast(this)">&times;</button>
            `;
            
            container.appendChild(toast);
            
            // Trigger animation
            setTimeout(() => toast.classList.add('show'), 100);
            
            // Auto remove
            setTimeout(() => {
                if (toast.parentNode) {
                    closeToast(toast.querySelector('.toast-close'));
                }
            }, duration);
        }
        
        function closeToast(button) {
            const toast = button.closest('.toast');
            toast.classList.remove('show');
            setTimeout(() => {
                if (toast.parentNode) {
                    toast.parentNode.removeChild(toast);
                }
            }, 300);
        }
        
        // Global toast functions for easy access
        window.showSuccessToast = (title, message) => showToast('success', title, message);
        window.showErrorToast = (title, message) => showToast('error', title, message);
        window.showWarningToast = (title, message) => showToast('warning', title, message);
        window.showInfoToast = (title, message) => showToast('info', title, message);
        
        // Auto-switch to appropriate panel based on user role
        document.addEventListener('DOMContentLoaded', function() {
            const userRole = '{{ session.user_role }}';
            if (userRole === 'agent') {
                // For agents, show agent panel and hide role selector
                document.querySelectorAll('.panel').forEach(panel => panel.classList.remove('active'));
                document.getElementById('agent-panel').classList.add('active');
                const roleSelector = document.querySelector('.role-selector');
                if (roleSelector) {
                    roleSelector.style.display = 'none';
                }
            }
            
            // Note: Flash messages are handled by the server-side template rendering
            // Toast notifications are only used for file uploads via JavaScript
        });
        
        function switchPriorityTab(priority) {
            // Update tab button states
            document.querySelectorAll('.tab-button').forEach(btn => btn.classList.remove('active'));
            document.getElementById(priority + '-priority-tab').classList.add('active');
            
            // Show/hide panels
            document.querySelectorAll('.priority-panel').forEach(panel => {
                panel.style.display = 'none';
            });
            
            const targetPanel = document.getElementById(priority + '-priority-panel');
            if (targetPanel) {
                targetPanel.style.display = 'block';
            }
            
            // Load dates for the selected priority panel and refresh displays
            if (priority === 'first') {
                loadAppointmentDates(); // Refresh First Priority display
            } else if (priority === 'second') {
                loadAppointmentDatesSecond(); // Refresh Second Priority display
            } else if (priority === 'third') {
                updateThirdPriorityInfo();
                loadReceiveDateCheckboxes(); // Load receive date checkboxes
            }
        }

        // Form submission with loading states and toast notifications
        const allocationForm = document.getElementById('allocation-form');
        if (allocationForm) {
            allocationForm.addEventListener('submit', function(e) {
                e.preventDefault(); // Prevent default form submission
                
                const btn = document.getElementById('allocation-btn');
                const fileInput = document.getElementById('allocation_file');
                
                if (!fileInput.files[0]) {
                    showErrorToast('Upload Error', 'Please select a file to upload');
                    return;
                }
                
                if (btn) {
                    btn.disabled = true;
                    btn.innerHTML = '<i class="fas fa-spinner fa-spin"></i> Uploading...';
                }
                
                const formData = new FormData();
                formData.append('file', fileInput.files[0]);
                
                fetch('/upload_allocation', {
                    method: 'POST',
                    body: formData
                })
                .then(response => {
                    if (!response.ok) {
                        throw new Error(`HTTP error! status: ${response.status}`);
                    }
                    return response.text();
                })
                .then(() => {
                    showSuccessToast('Upload Successful', 'Allocation file uploaded successfully!');
                    // Reset form
                    allocationForm.reset();
                    // Don't show loader yet - wait until both files are uploaded
                    // Reload page to show updated status
                    setTimeout(() => {
                        window.location.reload();
                    }, 1500);
                })
                    .catch(error => {
                        showErrorToast('Upload Failed', 'Error uploading allocation file. Please try again.');
                    })
                .finally(() => {
                    if (btn) {
                        btn.disabled = false;
                        btn.innerHTML = '<i class="fas fa-upload"></i> Upload Allocation File';
                    }
                });
            });
        }

        const dataForm = document.getElementById('data-form');
        if (dataForm) {
            dataForm.addEventListener('submit', function(e) {
                e.preventDefault(); // Prevent default form submission
                
                const btn = document.getElementById('data-btn');
                const fileInput = document.getElementById('data_file');
                
                if (!fileInput.files[0]) {
                    showErrorToast('Upload Error', 'Please select a file to upload');
                    return;
                }
                
                if (btn) {
                    btn.disabled = true;
                    btn.innerHTML = '<i class="fas fa-spinner fa-spin"></i> Uploading...';
                }
                
                const formData = new FormData();
                formData.append('file', fileInput.files[0]);
                
                fetch('/upload_data', {
                    method: 'POST',
                    body: formData
                })
                .then(response => {
                    if (!response.ok) {
                        throw new Error(`HTTP error! status: ${response.status}`);
                    }
                    return response.text();
                })
                .then(() => {
                    showSuccessToast('Upload Successful', 'Data file uploaded successfully!');
                    // Reset form
                    dataForm.reset();
                    // Mark that files were just uploaded (both files should be present now)
                    sessionStorage.setItem('filesJustUploaded', 'true');
                    // Show loader before reloading
                    showLoader();
                    // Reload page to show updated status
                    setTimeout(() => {
                        window.location.reload();
                    }, 1500);
                })
                    .catch(error => {
                        showErrorToast('Upload Failed', 'Error uploading data file. Please try again.');
                    })
                .finally(() => {
                    if (btn) {
                        btn.disabled = false;
                        btn.innerHTML = '<i class="fas fa-upload"></i> Upload Data File';
                    }
                });
            });
        }

        const processForm = document.getElementById('process-form');
        if (processForm) {
            processForm.addEventListener('submit', function(e) {
                e.preventDefault();
                processFiles();
            });
        }
        
        // Agent upload form handler
        const agentUploadForm = document.getElementById('agentUploadForm');
        if (agentUploadForm) {
            agentUploadForm.addEventListener('submit', function(e) {
                e.preventDefault();
                uploadAgentWorkFile();
            });
        }
        
        // Populate date fields when page loads
        document.addEventListener('DOMContentLoaded', function() {
            // Only load appointment dates if files have been uploaded
            // Check if data file container exists and has content before showing loader
            const calendarContainer = document.getElementById('calendar_container');
            if (calendarContainer) {
                // Try to load appointment dates, but don't show loader on initial page load
                // Loader will only show if files were just uploaded (handled in upload handlers)
                loadAppointmentDatesWithoutLoader();
            }
        });
        
        // Version of loadAppointmentDates that checks if loader should be shown (for initial page load after file upload)
        function loadAppointmentDatesWithoutLoader() {
            const calendarContainer = document.getElementById('calendar_container');
            if (!calendarContainer) return;
            
            // Check if loader was shown before page reload (e.g., after data file upload)
            // If loader overlay is visible or was just uploaded, we'll show it when dates are successfully loaded
            const wasJustUploaded = sessionStorage.getItem('filesJustUploaded') === 'true';
            sessionStorage.removeItem('filesJustUploaded');
            
            // Show loading message in the container
            calendarContainer.innerHTML = '<p style="color: #666; font-style: italic; text-align: center; padding: 20px;">Checking for uploaded files...</p>';
            
            // Fetch appointment dates from server
            fetch('/get_appointment_dates')
                .then(response => {
                    return response.json();
                })
                .then(data => {
                    if (data.error) {
                        calendarContainer.innerHTML = `<p style="color: #666; font-style: italic; text-align: center; padding: 20px;">${data.error}</p>`;
                        // Hide loader if there's an error
                        if (isLoaderVisible()) {
                            hideLoader();
                        }
                        return;
                    }
                    
                    const dates = data.appointment_dates;
                    const datesWithCounts = data.appointment_dates_with_counts;
                    const columnName = data.column_name;
                    
                    if (!dates || dates.length === 0) {
                        calendarContainer.innerHTML = '<p style="color: #666; font-style: italic; text-align: center; padding: 20px;">No appointment dates found in the file.</p>';
                        // Hide loader if no dates found
                        if (isLoaderVisible()) {
                            hideLoader();
                        }
                        return;
                    }
                    
                    // If both files exist and we have dates, show loader (if files were just uploaded)
                    if (wasJustUploaded && dates && dates.length > 0) {
                        showLoader();
                    }
                    
                    // Store appointment dates
                    appointmentDates = new Set(dates);
                    // Directly show checkbox list (no calendar view)
                    // Loader will be hidden in showFallbackDateList after dates are displayed
                    showFallbackDateList(datesWithCounts, columnName);
                    updateSelectedDatesInfo();
                })
                .catch(error => {
                    calendarContainer.innerHTML = '<p style="color: #666; font-style: italic; text-align: center; padding: 20px;">No data file uploaded yet.</p>';
                    // Hide loader on error
                    if (isLoaderVisible()) {
                        hideLoader();
                    }
                });
        }
        
        
        // Global variables for calendar
        let currentDate = new Date();
        let appointmentDates = new Set();
        let selectedDates = new Set();
        let selectedSecondDates = new Set();
        
        // Store receive date selections per appointment date
        let receiveDateSelections = new Map(); // appointmentDate -> Set of selected receive dates
        
        // Loader functions
        function showLoader() {
            const loader = document.getElementById('loader-overlay');
            if (loader) {
                loader.classList.add('show');
            }
        }
        
        function hideLoader() {
            const loader = document.getElementById('loader-overlay');
            if (loader) {
                loader.classList.remove('show');
            }
        }
        
        function isLoaderVisible() {
            const loader = document.getElementById('loader-overlay');
            return loader && loader.classList.contains('show');
        }
        
        function loadAppointmentDates() {
            const calendarContainer = document.getElementById('calendar_container');
            if (!calendarContainer) return;
            
            // Show loader when starting to load (assuming both files exist when this is called manually)
            showLoader();
            
            // Always try to load appointment dates (file might be uploaded via form submission)
            calendarContainer.innerHTML = '<p style="color: #666; font-style: italic; text-align: center; padding: 20px;">Loading appointment dates...</p>';
            
            // Fetch appointment dates from server
            fetch('/get_appointment_dates')
                .then(response => {
                    return response.json();
                })
                .then(data => {
                    if (data.error) {
                        calendarContainer.innerHTML = `<p style="color: #e74c3c; text-align: center; padding: 20px;">Error: ${data.error}</p>`;
                        // Hide loader if there's an error (likely files not uploaded yet)
                        hideLoader();
                        return;
                    }
                    
                    const dates = data.appointment_dates;
                    const datesWithCounts = data.appointment_dates_with_counts;
                    const columnName = data.column_name;
                    
                    if (!dates || dates.length === 0) {
                        calendarContainer.innerHTML = '<p style="color: #666; font-style: italic; text-align: center; padding: 20px;">No appointment dates found in the file.</p>';
                        // Hide loader if no dates found
                        hideLoader();
                        return;
                    }
                    
                    // Store appointment dates
                    appointmentDates = new Set(dates);
                    // Directly show checkbox list (no calendar view)
                    // Loader will be hidden in showFallbackDateList after dates are displayed
                    showFallbackDateList(datesWithCounts, columnName);
                    updateSelectedDatesInfo();
                })
                .catch(error => {
                    // Hide loader on error
                    hideLoader();
                    calendarContainer.innerHTML = `<p style="color: #e74c3c; text-align: center; padding: 20px;">Error loading appointment dates: ${error.message}</p>`;
                });
        }
        
        function loadAppointmentDatesSecond() {
            const calendarContainer = document.getElementById('calendar_container_second');
            if (!calendarContainer) return;
            
            // Show loader when starting to load appointment dates
            showLoader();
            
            // Always try to load appointment dates (file might be uploaded via form submission)
            calendarContainer.innerHTML = '<p style="color: #666; font-style: italic; text-align: center; padding: 20px;">Loading appointment dates...</p>';
            
            // Fetch appointment dates from server
            fetch('/get_appointment_dates')
                .then(response => {
                    return response.json();
                })
                .then(data => {
                    if (data.error) {
                        calendarContainer.innerHTML = `<p style="color: #e74c3c; text-align: center; padding: 20px;">Error: ${data.error}</p>`;
                        hideLoader();
                        return;
                    }
                    
                    const dates = data.appointment_dates;
                    const datesWithCounts = data.appointment_dates_with_counts;
                    const columnName = data.column_name;
                    
                    if (!dates || dates.length === 0) {
                        calendarContainer.innerHTML = '<p style="color: #666; font-style: italic; text-align: center; padding: 20px;">No appointment dates found in the file.</p>';
                        hideLoader();
                        return;
                    }
                    
                    
                    // Store appointment dates
                    appointmentDates = new Set(dates);
                    // Directly show checkbox list (no calendar view)
                    // Loader will be hidden in showFallbackDateListSecond after dates are displayed
                    showFallbackDateListSecond(datesWithCounts, columnName);
                    updateSelectedDatesInfoSecond();
                })
                .catch(error => {
                    // Hide loader on error
                    hideLoader();
                    calendarContainer.innerHTML = `<p style="color: #e74c3c; text-align: center; padding: 20px;">Error loading appointment dates: ${error.message}</p>`;
                });
        }
        
        function loadReceiveDateCheckboxes() {
            const appointmentReceiveDatesContainer = document.getElementById('appointment_receive_dates_container');
            if (!appointmentReceiveDatesContainer) return;
            
            // Get selected appointment dates
            const selectedAppointmentDates = getSelectedAppointmentDates();
            
            if (selectedAppointmentDates.length === 0) {
                appointmentReceiveDatesContainer.innerHTML = '<p style="color: #666; font-style: italic; text-align: center; padding: 20px;">No appointment dates selected.</p>';
                return;
            }
            
            // Create individual panels for each appointment date
            let html = '';
            selectedAppointmentDates.forEach((appointmentDate, appointmentIndex) => {
                const appointmentDateObj = new Date(appointmentDate);
                const appointmentFormatted = appointmentDateObj.toLocaleDateString('en-US', { 
                    year: 'numeric', 
                    month: 'short', 
                    day: 'numeric',
                    weekday: 'short'
                });
                
                html += `
                    <div class="appointment-receive-panel" style="margin-bottom: 20px; padding: 15px; background: white; border: 1px solid #e9ecef; border-radius: 8px;">
                        <h5 style="margin: 0 0 10px 0; color: #2c5aa0; font-size: 14px;">
                            <i class="fas fa-calendar"></i> ${appointmentFormatted}
                        </h5>
                        <div id="receive_dates_${appointmentDate.replace(/-/g, '_')}" style="display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 10px;">
                            <p style="color: #666; font-style: italic; text-align: center; padding: 10px; width: 100%;">Loading receive dates...</p>
                        </div>
                        <div id="receive_info_${appointmentDate.replace(/-/g, '_')}" style="background: #f8f9fa; padding: 8px; border-radius: 5px; border: 1px solid #e9ecef; font-size: 12px;">
                            <strong>Selected:</strong> <span id="receive_count_${appointmentDate.replace(/-/g, '_')}">0</span> receive dates
                        </div>
                    </div>
                `;
            });
            
            appointmentReceiveDatesContainer.innerHTML = html;
            
            // Load receive dates for each appointment date
            selectedAppointmentDates.forEach((appointmentDate, appointmentIndex) => {
                loadReceiveDatesForAppointment(appointmentDate);
            });
        }
        
        function loadReceiveDatesForAppointment(appointmentDate) {
            const containerId = `receive_dates_${appointmentDate.replace(/-/g, '_')}`;
            const container = document.getElementById(containerId);
            if (!container) return;
            
            // Build query parameters for this specific appointment date
            const url = `/get_receive_dates?appointment_dates=${appointmentDate}`;
            
            // Fetch receive dates for this specific appointment date
            fetch(url)
                .then(response => response.json())
                .then(data => {
                    if (data.error) {
                        container.innerHTML = `<p style="color: #e74c3c; text-align: center; padding: 10px;">Error: ${data.error}</p>`;
                        return;
                    }
                    
                    const dates = data.receive_dates;
                    
                    if (!dates || dates.length === 0) {
                        container.innerHTML = '<p style="color: #666; font-style: italic; text-align: center; padding: 10px;">No receive dates found for this appointment date.</p>';
                        return;
                    }
                    
                    // Get saved selections for this appointment date
                    const savedSelections = receiveDateSelections.get(appointmentDate) || new Set();
                    
                    // Display receive dates as checkboxes with saved selections
                    let html = '';
                    dates.forEach((date, index) => {
                        const dateObj = new Date(date);
                        const dayName = dateObj.toLocaleDateString('en-US', { weekday: 'short' });
                        const formattedDate = dateObj.toLocaleDateString('en-US', { 
                            year: 'numeric', 
                            month: 'short', 
                            day: 'numeric' 
                        });
                        
                        // Check if this receive date should be selected based on saved selections
                        // If no saved selections exist, default to all selected
                        const isSelected = savedSelections.size === 0 ? true : savedSelections.has(date);
                        const uniqueId = `receive_${appointmentDate.replace(/-/g, '_')}_${index}`;
                        
                        html += `
                            <label style="display: flex; align-items: center; padding: 6px 10px; border: 1px solid #ddd; border-radius: 4px; background: white; cursor: pointer; transition: all 0.3s; min-width: 120px; font-size: 12px;" 
                                   onclick="toggleReceiveDateForAppointment('${appointmentDate}', '${date}', ${index})">
                                <input type="checkbox" id="${uniqueId}" data-date="${date}" ${isSelected ? 'checked' : ''} style="margin-right: 6px; transform: scale(0.9);">
                                <div>
                                    <div style="font-weight: bold; font-size: 12px;">${formattedDate}</div>
                                    <div style="color: #666; font-size: 10px;">${dayName}</div>
                                </div>
                            </label>
                        `;
                    });
                    
                    container.innerHTML = html;
                    
                    // Update receive date info for this appointment date
                    updateReceiveDateInfoForAppointment(appointmentDate);
                })
                .catch(error => {
                    container.innerHTML = `<p style="color: #e74c3c; text-align: center; padding: 10px;">Error loading receive dates: ${error.message}</p>`;
                });
        }
        
        function toggleReceiveDateForAppointment(appointmentDate, dateStr, index) {
            const uniqueId = `receive_${appointmentDate.replace(/-/g, '_')}_${index}`;
            const checkbox = document.getElementById(uniqueId);
            if (!checkbox) return;
            
            checkbox.checked = !checkbox.checked;
            
            // Update the visual state
            const container = checkbox.closest('label');
            if (checkbox.checked) {
                container.style.background = '#e3f2fd';
                container.style.borderColor = '#2196f3';
            } else {
                container.style.background = 'white';
                container.style.borderColor = '#ddd';
            }
            
            // Save the receive date selection for this specific appointment date
            saveReceiveDateSelectionForAppointment(appointmentDate);
            
            // Update receive date info for this appointment date
            updateReceiveDateInfoForAppointment(appointmentDate);
        }
        
        function saveReceiveDateSelectionForAppointment(appointmentDate) {
            // Get current receive date selections for this specific appointment date
            const containerId = `receive_dates_${appointmentDate.replace(/-/g, '_')}`;
            const container = document.getElementById(containerId);
            if (!container) return;
            
            const currentSelections = new Set();
            const checkboxes = container.querySelectorAll('input[type="checkbox"]');
            checkboxes.forEach(checkbox => {
                if (checkbox.checked) {
                    currentSelections.add(checkbox.dataset.date);
                }
            });
            
            // Save the selections for this appointment date
            receiveDateSelections.set(appointmentDate, currentSelections);
        }
        
        function updateReceiveDateInfoForAppointment(appointmentDate) {
            const countElement = document.getElementById(`receive_count_${appointmentDate.replace(/-/g, '_')}`);
            if (!countElement) return;
            
            const containerId = `receive_dates_${appointmentDate.replace(/-/g, '_')}`;
            const container = document.getElementById(containerId);
            if (!container) return;
            
            const checkboxes = container.querySelectorAll('input[type="checkbox"]');
            const selectedCount = Array.from(checkboxes).filter(cb => cb.checked).length;
            
            countElement.textContent = selectedCount;
        }
        
        function updateReceiveDateInfo() {
            const selectedCount = document.getElementById('receive_selected_count');
            const selectedText = document.getElementById('receive_selected_text');
            const selectedList = document.getElementById('receive_selected_list');
            
            if (!selectedCount || !selectedText || !selectedList) return;
            
            // Get all checked receive date checkboxes
            const checkboxes = document.querySelectorAll('input[id^="receive_checkbox_"]:checked');
            const selectedDates = Array.from(checkboxes).map(cb => cb.dataset.date);
            
            selectedCount.textContent = selectedDates.length;
            selectedText.textContent = selectedDates.length === 1 ? 'date selected' : 'dates selected';
            
            if (selectedDates.length > 0) {
                const formattedDates = selectedDates.map(date => {
                    const dateObj = new Date(date);
                    return dateObj.toLocaleDateString('en-US', { 
                        month: 'short', 
                        day: 'numeric' 
                    });
                });
                selectedList.textContent = formattedDates.join(', ');
            } else {
                selectedList.textContent = 'No receive dates selected';
            }
        }
        
        function initializeCalendar() {
            renderCalendar();
        }
        
        function renderCalendar() {
            const year = currentDate.getFullYear();
            const month = currentDate.getMonth();
            
            // Update header
            const monthYearElement = document.getElementById('current_month_year');
            if (monthYearElement) {
                monthYearElement.textContent = currentDate.toLocaleDateString('en-US', { 
                    year: 'numeric', 
                    month: 'long' 
                });
            }
            
            // Get first day of month and number of days
            const firstDay = new Date(year, month, 1);
            const lastDay = new Date(year, month + 1, 0);
            const daysInMonth = lastDay.getDate();
            const startingDayOfWeek = firstDay.getDay();
            
            // Day headers
            const dayHeaders = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat'];
            let calendarHTML = '';
            
            // Add day headers
            dayHeaders.forEach(day => {
                calendarHTML += `<div style="text-align: center; font-weight: bold; padding: 8px; background: #f8f9fa; border: 1px solid #dee2e6;">${day}</div>`;
            });
            
            // Add empty cells for days before month starts
            for (let i = 0; i < startingDayOfWeek; i++) {
                calendarHTML += `<div style="height: 40px; border: 1px solid #dee2e6; background: #f8f9fa;"></div>`;
            }
            
            // Add days of the month
            for (let day = 1; day <= daysInMonth; day++) {
                const dateStr = `${year}-${String(month + 1).padStart(2, '0')}-${String(day).padStart(2, '0')}`;
                const isAppointmentDate = appointmentDates.has(dateStr);
                const isSelected = selectedDates.has(dateStr);
                const isSelectedSecond = selectedSecondDates.has(dateStr);
                const isToday = isTodayDate(year, month, day);
                
                let cellClass = 'calendar-day';
                let cellStyle = 'height: 40px; border: 1px solid #dee2e6; display: flex; align-items: center; justify-content: center; cursor: pointer; position: relative;';
                
                if (isToday) {
                    cellStyle += ' background: #e3f2fd; font-weight: bold;';
                } else if (isAppointmentDate) {
                    cellStyle += ' background: #fff3e0;';
                } else {
                    cellStyle += ' background: #f8f9fa; color: #6c757d;';
                }
                
                if (isSelected) {
                    cellStyle += ' background: #4caf50; color: white; font-weight: bold;';
                } else if (isSelectedSecond) {
                    cellStyle += ' background: #f39c12; color: white; font-weight: bold;';
                }
                
                if (!isAppointmentDate) {
                    cellStyle += ' cursor: not-allowed; opacity: 0.5;';
                }
                
                calendarHTML += `
                    <div class="${cellClass}" 
                         data-date="${dateStr}" 
                         style="${cellStyle}"
                         onclick="${isAppointmentDate ? `toggleDate('${dateStr}')` : ''}">
                        ${day}
                        ${isAppointmentDate ? '<div style="position: absolute; top: 2px; right: 2px; width: 6px; height: 6px; background: #ff9800; border-radius: 50%;"></div>' : ''}
                    </div>
                `;
            }
            
            // Update calendar grid
            const calendarGrid = document.getElementById('calendar_grid');
            if (calendarGrid) {
                calendarGrid.innerHTML = calendarHTML;
            }
            
            // Update selected dates info
            updateSelectedDatesInfo();
        }
        
        function isTodayDate(year, month, day) {
            const today = new Date();
            return today.getFullYear() === year && 
                   today.getMonth() === month && 
                   today.getDate() === day;
        }
        
        function getSelectedAppointmentDates() {
            return Array.from(selectedDates);
        }
        
        
        function toggleDate(dateStr) {
            if (!appointmentDates.has(dateStr)) return;
            
            if (selectedDates.has(dateStr)) {
                selectedDates.delete(dateStr);
            } else {
                // Remove from Second Priority if it was selected there
                if (selectedSecondDates.has(dateStr)) {
                    selectedSecondDates.delete(dateStr);
                    updateSelectedDatesInfoSecond();
                    syncFallbackCheckboxesSecond();
                }
                selectedDates.add(dateStr);
            }
            
            renderCalendar();
            syncFallbackCheckboxes();
            updateThirdPriorityInfo(); // Update Third Priority info when First Priority changes
            
            // Show/hide receive date panel based on whether any appointment dates are selected
            const receiveDatePanel = document.getElementById('receive-date-panel');
            if (receiveDatePanel) {
                if (selectedDates.size > 0) {
                    receiveDatePanel.style.display = 'block';
                    // Always reload receive dates when appointment dates change
                    loadReceiveDateCheckboxes();
                } else {
                    receiveDatePanel.style.display = 'none';
                }
            }
        }
        
        function previousMonth() {
            currentDate.setMonth(currentDate.getMonth() - 1);
            renderCalendar();
        }
        
        function nextMonth() {
            currentDate.setMonth(currentDate.getMonth() + 1);
            renderCalendar();
        }
        
        function updateSelectedDatesInfo() {
            const selectedCount = document.getElementById('selected_count');
            const selectedText = document.getElementById('selected_text');
            const selectedDatesList = document.getElementById('selected_dates_list');
            
            if (selectedCount) {
                selectedCount.textContent = selectedDates.size;
            }
            
            if (selectedText) {
                selectedText.textContent = selectedDates.size === 1 ? 'date selected' : 'dates selected';
            }
            
            if (selectedDatesList) {
                if (selectedDates.size === 0) {
                    selectedDatesList.textContent = 'No dates selected';
                } else {
                    const sortedDates = Array.from(selectedDates).sort();
                    const formattedDates = sortedDates.map(date => {
                        const dateObj = new Date(date);
                        return dateObj.toLocaleDateString('en-US', { 
                            month: 'short', 
                            day: 'numeric' 
                        });
                    });
                    selectedDatesList.textContent = formattedDates.join(', ');
                }
            }
            // Keep toggle button label in sync
            const btn = document.getElementById('toggle-select-btn');
            if (btn) {
                const total = appointmentDates ? appointmentDates.size : 0;
                if (selectedDates.size === total && total > 0) {
                    btn.textContent = 'Deselect All Dates';
                    btn.style.background = '#e74c3c';
                } else {
                    btn.textContent = 'Select All Dates';
                    btn.style.background = '#27ae60';
                }
            }
        }
        
        function showFallbackDateList(datesWithCounts, columnName) {
            const calendarContainer = document.getElementById('calendar_container');
            if (!calendarContainer) {
                // Hide loader even if container doesn't exist
                hideLoader();
                return;
            }
            
            try {
                // Ensure variables are initialized
                if (typeof selectedDates === 'undefined') {
                    selectedDates = new Set();
                }
                if (typeof selectedSecondDates === 'undefined') {
                    selectedSecondDates = new Set();
                }
            
            let html = `
                <div style="text-align: center; margin-bottom: 20px;">
                    <p>Click on dates to select them for First Priority:</p>
                </div>
                <div style="display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 10px; max-height: 300px; overflow-y: auto;">
            `;
            
                datesWithCounts.forEach((dateData, index) => {
                    const date = dateData.date;
                    const rowCount = dateData.row_count || 0;
                const dateObj = new Date(date);
                const dayName = dateObj.toLocaleDateString('en-US', { weekday: 'long' });
                const formattedDate = dateObj.toLocaleDateString('en-US', { 
                    year: 'numeric', 
                    month: 'short', 
                    day: 'numeric' 
                });
                
                const isSelectedInFirst = selectedDates.has(date);
                const isSelectedInSecond = selectedSecondDates.has(date);
                const isDisabled = isSelectedInSecond;
                
                let itemStyle = 'display: flex; align-items: center; padding: 10px; border: 2px solid #e0e0e0; border-radius: 8px; background: #f9f9f9; cursor: pointer; transition: all 0.3s;';
                let textStyle = 'font-weight: bold; font-size: 16px;';
                let dayStyle = 'color: #666; font-size: 14px;';
                let countStyle = 'color: #666; font-size: 12px; margin-top: 2px;';
                
                if (isSelectedInFirst) {
                    itemStyle = 'display: flex; align-items: center; padding: 10px; border: 2px solid #4caf50; border-radius: 8px; background: #4caf50; color: white; cursor: pointer; transition: all 0.3s;';
                    textStyle = 'font-weight: bold; font-size: 16px; color: white;';
                    dayStyle = 'color: rgba(255,255,255,0.8); font-size: 14px;';
                    countStyle = 'color: rgba(255,255,255,0.9); font-size: 12px; margin-top: 2px;';
                } else if (isDisabled) {
                    itemStyle = 'display: flex; align-items: center; padding: 10px; border: 2px solid #f39c12; border-radius: 8px; background: #f39c12; color: white; cursor: not-allowed; opacity: 0.7; transition: all 0.3s;';
                    textStyle = 'font-weight: bold; font-size: 16px; color: white;';
                    dayStyle = 'color: rgba(255,255,255,0.8); font-size: 14px;';
                    countStyle = 'color: rgba(255,255,255,0.9); font-size: 12px; margin-top: 2px;';
                }
                
                html += `
                    <div style="${itemStyle}"
                         onclick="${isDisabled ? '' : `toggleDate('${date}')`}" 
                         id="date_${index}">
                        <input type="checkbox" id="checkbox_${index}" data-date="${date}" style="margin-right: 10px; transform: scale(1.2);" ${isDisabled ? 'disabled' : ''}>
                        <div>
                            <div style="${textStyle}">${formattedDate}${isDisabled ? ' (Second Priority)' : ''}</div>
                            <div style="${dayStyle}">${dayName}</div>
                                <div style="${countStyle}">${rowCount} rows</div>
                        </div>
                    </div>
                `;
            });
            
            html += '</div>';
            calendarContainer.innerHTML = html;
            // Sync checkboxes to current selection
            syncFallbackCheckboxes();
            } catch (error) {
                console.error('Error displaying dates (First Priority):', error);
                console.error('Error details:', error.message, error.stack);
                calendarContainer.innerHTML = '<p style="color: #e74c3c; text-align: center; padding: 20px;">Error displaying dates: ' + (error.message || 'Unknown error') + '. Please try again.</p>';
            } finally {
                // Always hide loader after attempting to display dates
                // Use setTimeout to ensure DOM updates complete
                setTimeout(function() {
                    hideLoader();
                }, 100);
            }
        }
        
        function showFallbackDateListSecond(datesWithCounts, columnName) {
            const calendarContainer = document.getElementById('calendar_container_second');
            if (!calendarContainer) {
                // Hide loader even if container doesn't exist
                hideLoader();
                return;
            }
            
            try {
                // Ensure variables are initialized
                if (typeof selectedDates === 'undefined') {
                    selectedDates = new Set();
                }
                if (typeof selectedSecondDates === 'undefined') {
                    selectedSecondDates = new Set();
                }
            
            let html = `
                <div style="text-align: center; margin-bottom: 20px;">
                    <p>Click on dates to select them for Second Priority:</p>
                </div>
                <div style="display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 10px; max-height: 300px; overflow-y: auto;">
            `;
            
                datesWithCounts.forEach((dateData, index) => {
                    const date = dateData.date;
                    const rowCount = dateData.row_count || 0;
                const dateObj = new Date(date);
                const dayName = dateObj.toLocaleDateString('en-US', { weekday: 'long' });
                const formattedDate = dateObj.toLocaleDateString('en-US', { 
                    year: 'numeric', 
                    month: 'short', 
                    day: 'numeric' 
                });
                
                const isSelectedInFirst = selectedDates.has(date);
                const isSelectedInSecond = selectedSecondDates.has(date);
                const isDisabled = isSelectedInFirst;
                
                let itemStyle = 'display: flex; align-items: center; padding: 10px; border: 2px solid #e0e0e0; border-radius: 8px; background: #f9f9f9; cursor: pointer; transition: all 0.3s;';
                let textStyle = 'font-weight: bold; font-size: 16px;';
                let dayStyle = 'color: #666; font-size: 14px;';
                let countStyle = 'color: #666; font-size: 12px; margin-top: 2px;';
                
                if (isSelectedInSecond) {
                    itemStyle = 'display: flex; align-items: center; padding: 10px; border: 2px solid #f39c12; border-radius: 8px; background: #f39c12; color: white; cursor: pointer; transition: all 0.3s;';
                    textStyle = 'font-weight: bold; font-size: 16px; color: white;';
                    dayStyle = 'color: rgba(255,255,255,0.8); font-size: 14px;';
                    countStyle = 'color: rgba(255,255,255,0.9); font-size: 12px; margin-top: 2px;';
                } else if (isDisabled) {
                    itemStyle = 'display: flex; align-items: center; padding: 10px; border: 2px solid #4caf50; border-radius: 8px; background: #4caf50; color: white; cursor: not-allowed; opacity: 0.7; transition: all 0.3s;';
                    textStyle = 'font-weight: bold; font-size: 16px; color: white;';
                    dayStyle = 'color: rgba(255,255,255,0.8); font-size: 14px;';
                    countStyle = 'color: rgba(255,255,255,0.9); font-size: 12px; margin-top: 2px;';
                }
                
                html += `
                    <div style="${itemStyle}"
                         onclick="${isDisabled ? '' : `toggleDateSecond('${date}')`}" 
                         id="date_second_${index}">
                        <input type="checkbox" id="checkbox_second_${index}" data-date="${date}" style="margin-right: 10px; transform: scale(1.2);" ${isDisabled ? 'disabled' : ''}>
                        <div>
                            <div style="${textStyle}">${formattedDate}${isDisabled ? ' (First Priority)' : ''}</div>
                            <div style="${dayStyle}">${dayName}</div>
                            <div style="${countStyle}">${rowCount} rows</div>
                        </div>
                    </div>
                `;
            });
            
            html += '</div>';
            calendarContainer.innerHTML = html;
            // Sync checkboxes to current selection
            syncFallbackCheckboxesSecond();
            } catch (error) {
                console.error('Error displaying dates (Second Priority):', error);
                console.error('Error details:', error.message, error.stack);
                calendarContainer.innerHTML = '<p style="color: #e74c3c; text-align: center; padding: 20px;">Error displaying dates: ' + (error.message || 'Unknown error') + '. Please try again.</p>';
            } finally {
                // Always hide loader after attempting to display dates
                // Use setTimeout to ensure DOM updates complete
                setTimeout(function() {
                    hideLoader();
                }, 100);
            }
        }
        
        function toggleSelectAllDates() {
            const btn = document.getElementById('toggle-select-btn');
            const total = appointmentDates ? appointmentDates.size : 0;
            const selected = selectedDates ? selectedDates.size : 0;
            const shouldSelectAll = selected < total;
            if (shouldSelectAll) {
                // Select all
                selectedDates = new Set();
                appointmentDates.forEach(d => selectedDates.add(d));
            } else {
                // Deselect all
                selectedDates.clear();
            }
            renderCalendar();
            updateSelectedDatesInfo();
            syncFallbackCheckboxes();
            // Update button label and style
            if (btn) {
                if (selectedDates.size === total && total > 0) {
                    btn.textContent = 'Deselect All Dates';
                    btn.style.background = '#e74c3c';
                } else {
                    btn.textContent = 'Select All Dates';
                    btn.style.background = '#27ae60';
                }
            }
        }

        function syncFallbackCheckboxes() {
            const checkboxes = document.querySelectorAll('#calendar_container input[type="checkbox"][data-date]');
            if (!checkboxes || checkboxes.length === 0) return;
            checkboxes.forEach(cb => {
                const d = cb.getAttribute('data-date');
                cb.checked = selectedDates.has(d);
            });
        }
        
        function toggleSelectAllSecondDates() {
            const btn = document.getElementById('toggle-select-second-btn');
            const total = appointmentDates ? appointmentDates.size : 0;
            const selected = selectedSecondDates ? selectedSecondDates.size : 0;
            const shouldSelectAll = selected < total;
            if (shouldSelectAll) {
                // Select all
                selectedSecondDates = new Set();
                appointmentDates.forEach(d => selectedSecondDates.add(d));
            } else {
                // Deselect all
                selectedSecondDates.clear();
            }
            updateSelectedDatesInfoSecond();
            syncFallbackCheckboxesSecond();
            // Update button label and style
            if (btn) {
                if (selectedSecondDates.size === total && total > 0) {
                    btn.textContent = 'Deselect All Dates';
                    btn.style.background = '#e74c3c';
                } else {
                    btn.textContent = 'Select All Dates';
                    btn.style.background = '#f39c12';
                }
            }
        }
        
        function toggleDateSecond(dateStr) {
            if (!appointmentDates.has(dateStr)) return;
            
            if (selectedSecondDates.has(dateStr)) {
                selectedSecondDates.delete(dateStr);
            } else {
                // Remove from First Priority if it was selected there
                if (selectedDates.has(dateStr)) {
                    selectedDates.delete(dateStr);
                    renderCalendar();
                    syncFallbackCheckboxes();
                }
                selectedSecondDates.add(dateStr);
            }
            
            updateSelectedDatesInfoSecond();
            syncFallbackCheckboxesSecond();
            updateThirdPriorityInfo(); // Update Third Priority info when Second Priority changes
        }
        
        function updateSelectedDatesInfoSecond() {
            const selectedCount = document.getElementById('selected_count_second');
            const selectedText = document.getElementById('selected_text_second');
            const selectedDatesList = document.getElementById('selected_dates_list_second');
            
            if (selectedCount) {
                selectedCount.textContent = selectedSecondDates.size;
            }
            
            if (selectedText) {
                selectedText.textContent = selectedSecondDates.size === 1 ? 'date selected' : 'dates selected';
            }
            
            if (selectedDatesList) {
                if (selectedSecondDates.size === 0) {
                    selectedDatesList.textContent = 'No dates selected';
                } else {
                    const sortedDates = Array.from(selectedSecondDates).sort();
                    const formattedDates = sortedDates.map(date => {
                        const dateObj = new Date(date);
                        return dateObj.toLocaleDateString('en-US', { 
                            month: 'short', 
                            day: 'numeric' 
                        });
                    });
                    selectedDatesList.textContent = formattedDates.join(', ');
                }
            }
            // Keep toggle button label in sync
            const btn = document.getElementById('toggle-select-second-btn');
            if (btn) {
                const total = appointmentDates ? appointmentDates.size : 0;
                if (selectedSecondDates.size === total && total > 0) {
                    btn.textContent = 'Deselect All Dates';
                    btn.style.background = '#e74c3c';
                } else {
                    btn.textContent = 'Select All Dates';
                    btn.style.background = '#f39c12';
                }
            }
        }
        
        function syncFallbackCheckboxesSecond() {
            const checkboxes = document.querySelectorAll('#calendar_container_second input[type="checkbox"][data-date]');
            if (!checkboxes || checkboxes.length === 0) return;
            checkboxes.forEach(cb => {
                const d = cb.getAttribute('data-date');
                cb.checked = selectedSecondDates.has(d);
            });
        }
        
        function updateThirdPriorityInfo() {
            // Calculate remaining dates that will be Third Priority
            const allDates = new Set(appointmentDates);
            const firstPriorityDates = new Set(selectedDates);
            const secondPriorityDates = new Set(selectedSecondDates);
            
            // Find dates that are not in First or Second Priority
            const thirdPriorityDates = new Set();
            allDates.forEach(date => {
                if (!firstPriorityDates.has(date) && !secondPriorityDates.has(date)) {
                    thirdPriorityDates.add(date);
                }
            });
            
            // Update the display
            const thirdPriorityCount = document.getElementById('third_priority_count');
            const thirdPriorityDatesList = document.getElementById('third_priority_dates_list');
            
            if (thirdPriorityCount) {
                thirdPriorityCount.textContent = thirdPriorityDates.size;
            }
            
            if (thirdPriorityDatesList) {
                if (thirdPriorityDates.size === 0) {
                    thirdPriorityDatesList.textContent = 'No remaining dates (all dates are assigned to First or Second Priority)';
                } else {
                    const sortedDates = Array.from(thirdPriorityDates).sort();
                    const formattedDates = sortedDates.map(date => {
                        const dateObj = new Date(date);
                        return dateObj.toLocaleDateString('en-US', { 
                            month: 'short', 
                            day: 'numeric' 
                        });
                    });
                    thirdPriorityDatesList.textContent = formattedDates.join(', ');
                }
            }
        }
        
        function selectBusinessDays() {
            // Clear all first
            clearAllDates();
            // This function is now simplified since we removed the business day checkboxes
            // Users can select dates directly from the calendar
        }
        
        
        function getNextBusinessDay(startDate, n) {
            let currentDate = new Date(startDate);
            let businessDaysCount = 0;
            
            while (businessDaysCount < n) {
                currentDate.setDate(currentDate.getDate() + 1);
                // Check if it's a weekday (Monday=1, Sunday=0)
                if (currentDate.getDay() >= 1 && currentDate.getDay() <= 5) {
                    businessDaysCount++;
                }
            }
            
            return currentDate.toISOString().split('T')[0];
        }
        
        function processFiles() {
            // Add selected calendar dates to form
            const form = document.getElementById('process-form');
            if (form) {
                // Remove existing hidden inputs for appointment dates
                const existingFirstInputs = form.querySelectorAll('input[name="appointment_dates"]');
                existingFirstInputs.forEach(input => input.remove());
                
                const existingSecondInputs = form.querySelectorAll('input[name="appointment_dates_second"]');
                existingSecondInputs.forEach(input => input.remove());
                
                // Add First Priority selected dates as hidden inputs
                selectedDates.forEach(date => {
                    const input = document.createElement('input');
                    input.type = 'hidden';
                    input.name = 'appointment_dates';
                    input.value = date;
                    form.appendChild(input);
                });
                
                // Add Second Priority selected dates as hidden inputs
                selectedSecondDates.forEach(date => {
                    const input = document.createElement('input');
                    input.type = 'hidden';
                    input.name = 'appointment_dates_second';
                    input.value = date;
                    form.appendChild(input);
                });
                
                // Add receive dates as hidden inputs from all appointment dates
                selectedDates.forEach(appointmentDate => {
                    const containerId = `receive_dates_${appointmentDate.replace(/-/g, '_')}`;
                    const container = document.getElementById(containerId);
                    if (container) {
                        const receiveCheckboxes = container.querySelectorAll('input[type="checkbox"]:checked');
                        receiveCheckboxes.forEach(checkbox => {
                            const input = document.createElement('input');
                            input.type = 'hidden';
                            input.name = 'receive_dates';
                            input.value = checkbox.dataset.date;
                            form.appendChild(input);
                        });
                    }
                });
                
                // If no dates selected for First Priority, add all appointment dates as fallback
                if (selectedDates.size === 0) {
                    appointmentDates.forEach(date => {
                        const input = document.createElement('input');
                        input.type = 'hidden';
                        input.name = 'appointment_dates';
                        input.value = date;
                        form.appendChild(input);
                    });
                }
                
                // Add debug inputs to see what's being sent
                const debugFirstInput = document.createElement('input');
                debugFirstInput.type = 'hidden';
                debugFirstInput.name = 'debug_selected_count';
                debugFirstInput.value = selectedDates.size;
                form.appendChild(debugFirstInput);
                
                const debugSecondInput = document.createElement('input');
                debugSecondInput.type = 'hidden';
                debugSecondInput.name = 'debug_selected_count_second';
                debugSecondInput.value = selectedSecondDates.size;
                form.appendChild(debugSecondInput);
            }
            
            const processingStatus = document.getElementById('processing-status');
            const processBtn = document.getElementById('process-btn');
            
            if (processingStatus) {
                processingStatus.style.display = 'flex';
            }
            if (processBtn) {
                processBtn.disabled = true;
                processBtn.textContent = 'Processing...';
            }
            
            // Simulate progress updates
            let progress = 0;
            const progressBar = document.getElementById('progress-bar');
            const progressText = document.getElementById('progress-text');
            
            if (!progressBar || !progressText) {
                return;
            }
            
            const progressInterval = setInterval(() => {
                progress += Math.random() * 15;
                if (progress > 90) progress = 90;
                
                progressBar.style.width = progress + '%';
                progressBar.textContent = Math.round(progress) + '%';
                
                if (progress < 30) {
                    progressText.textContent = 'Reading files...';
                } else if (progress < 60) {
                    progressText.textContent = 'Analyzing appointment dates...';
                } else if (progress < 90) {
                    progressText.textContent = 'Assigning priorities...';
                } else {
                    progressText.textContent = 'Finalizing results...';
                }
            }, 200);
            
            // Make AJAX request with form body
            const formData = new FormData(form);
            fetch('/process_files', {
                method: 'POST',
                body: new URLSearchParams(formData)
            })
            .then(response => response.text())
            .then(html => {
                clearInterval(progressInterval);
                if (progressBar) {
                    progressBar.style.width = '100%';
                    progressBar.textContent = '100%';
                }
                if (progressText) {
                    progressText.textContent = 'Processing complete!';
                }
                
                setTimeout(() => {
                    document.body.innerHTML = html;
                }, 1000);
            })
            .catch(error => {
                clearInterval(progressInterval);
                if (progressText) {
                    progressText.textContent = 'Error: ' + error.message;
                }
            });
        }
        
        function uploadAgentWorkFile() {
            const form = document.getElementById('agentUploadForm');
            const fileInput = document.getElementById('agentFile');
            const notesInput = document.getElementById('agentNotes');
            const uploadBtn = document.getElementById('agentUploadBtn');
            
            if (!fileInput.files[0]) {
                showErrorToast('Upload Error', 'Please select a file to upload');
                return;
            }
            
            // Show loading state
            if (uploadBtn) {
                uploadBtn.disabled = true;
                uploadBtn.innerHTML = '<i class="fas fa-spinner fa-spin"></i> Uploading...';
            }
            
            const formData = new FormData();
            formData.append('file', fileInput.files[0]);
            formData.append('notes', notesInput ? notesInput.value || '' : '');
            
            fetch('/upload_work_file', {
                method: 'POST',
                body: formData
            })
            .then(response => {
                return response.json().then(data => {
                    if (!response.ok) {
                        throw {data: data, status: response.status};
                    }
                    return data;
                });
            })
            .then(data => {
                if (data.success) {
                    showSuccessToast('Upload Successful', data.message);
                    // Reset form
                    if (form) form.reset();
                    // Reload page to show updated file list after a short delay
                    setTimeout(() => {
                        window.location.reload();
                    }, 2000);
                } else {
                    showErrorToast('Upload Failed', data.message || 'Upload failed');
                }
            })
            .catch(error => {
                // Handle validation errors with missing columns
                if (error.data && error.data.missing_columns) {
                    let errorMessage = error.data.message || 'File validation failed.';
                    if (error.data.missing_columns.length > 0) {
                        errorMessage += '<br><br><strong>Missing Columns:</strong><ul style="margin: 10px 0; padding-left: 20px;">';
                        error.data.missing_columns.forEach(col => {
                            errorMessage += `<li>${col}</li>`;
                        });
                        errorMessage += '</ul>';
                    }
                    showErrorToast('File Validation Error', errorMessage);
                } else if (error.data && error.data.message) {
                    showErrorToast('Upload Failed', error.data.message);
                } else {
                    showErrorToast('Upload Error', 'Error uploading file. Please try again.');
                }
            })
            .finally(() => {
                // Reset button state
                if (uploadBtn) {
                    uploadBtn.disabled = false;
                    uploadBtn.innerHTML = '<i class="fas fa-upload"></i> Upload Work File';
                }
            });
        }
        
        // Initialize agent table when page loads
        document.addEventListener('DOMContentLoaded', function() {
            // Update serial numbers for all agent rows
            const agentRows = document.querySelectorAll('.agent-row');
            agentRows.forEach((row, index) => {
                const srNoCell = row.querySelector('td:first-child');
                if (srNoCell) {
                    srNoCell.textContent = index + 1;
                }
            });
        });
        
        function approveAllocation(agentName) {
            if (confirm(`Are you sure you want to approve the allocation for ${agentName}? This will send an email with the allocated data.`)) {
                // Add visual feedback
                const button = event.target;
                const originalText = button.innerHTML;
                button.innerHTML = '<i class="fas fa-spinner fa-spin"></i> Sending Email...';
                button.disabled = true;
                
                // Send approval email
                fetch('/send_approval_email', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        agent_name: agentName
                    })
                })
                .then(response => response.json())
                .then(data => {
                    if (data.success) {
                        button.innerHTML = '<i class="fas fa-check"></i> Email Sent';
                        button.style.background = 'linear-gradient(135deg, #27ae60, #2ecc71)';
                        showSuccessToast('Email Sent', data.message);
                    } else {
                        button.innerHTML = originalText;
                        button.disabled = false;
                        showErrorToast('Email Failed', data.message);
                    }
                })
                .catch(error => {
                    button.innerHTML = originalText;
                    button.disabled = false;
                    showErrorToast('Email Error', `Error sending email: ${error.message}`);
                });
            }
        }
        
        function approveAllAllocations() {
            if (confirm('Are you sure you want to approve ALL allocations? This will send emails to all agents with their allocated data.')) {
                const button = document.getElementById('approve-all-btn');
                const originalText = button.innerHTML;
                button.innerHTML = '<i class="fas fa-spinner fa-spin"></i> Processing...';
                button.disabled = true;
                
                // Send approval for all allocations
                fetch('/approve_all_allocations', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    }
                })
                .then(response => response.json())
                .then(data => {
                    if (data.success) {
                        button.innerHTML = '<i class="fas fa-check-double"></i> All Approved';
                        button.style.background = 'linear-gradient(135deg, #27ae60, #2ecc71)';
                        showSuccessToast('All Allocations Approved', data.message);
                        
                        // Update individual approve buttons to show as sent
                        const approveButtons = document.querySelectorAll('.approve-btn');
                        approveButtons.forEach(btn => {
                            if (!btn.disabled) {
                                btn.innerHTML = '<i class="fas fa-check"></i> Email Sent';
                                btn.style.background = 'linear-gradient(135deg, #27ae60, #2ecc71)';
                                btn.disabled = true;
                            }
                        });
                    } else {
                        button.innerHTML = originalText;
                        button.disabled = false;
                        showErrorToast('Approval Failed', data.message);
                    }
                })
                .catch(error => {
                    button.innerHTML = originalText;
                    button.disabled = false;
                    showErrorToast('Error', `Error approving allocations: ${error.message}`);
                });
            }
        }
        
        function viewShiftTimes() {
            // Show modal
            const modal = document.getElementById('shiftTimesModal');
            if (!modal) {
                showErrorToast('Error', 'Shift times modal not found');
                return;
            }
            modal.style.display = 'block';
            
            // Show loading state
            const modalContent = document.getElementById('shiftTimesContent');
            modalContent.innerHTML = `
                <div style="text-align: center; padding: 40px;">
                    <i class="fas fa-spinner fa-spin" style="font-size: 2em; color: #667eea;"></i>
                    <p style="margin-top: 15px; color: #666;">Loading shift information...</p>
                </div>
            `;
            
            // Fetch shift times
            fetch('/view_shift_times', {
                method: 'GET',
                headers: {
                    'Content-Type': 'application/json',
                }
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    let html = `<h3 style="margin-top: 0; color: #333;">⏰ Shift Times Overview (${data.total_agents} agents)</h3>`;
                    html += `<div style="overflow-x: auto; margin-top: 15px;">`;
                    html += `<table class="modal-table" style="width: 100%; border-collapse: collapse;">`;
                    html += `<thead><tr style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white;">`;
                    html += `<th style="padding: 12px; text-align: center;">Sr No</th>`;
                    html += `<th style="padding: 12px; text-align: left;">Agent Name</th>`;
                    html += `<th style="padding: 12px; text-align: left;">Email</th>`;
                    html += `<th style="padding: 12px; text-align: center;">Original Shift Time</th>`;
                    html += `<th style="padding: 12px; text-align: center;">Parsed Start Time</th>`;
                    html += `<th style="padding: 12px; text-align: center;">Shift Group</th>`;
                    html += `<th style="padding: 12px; text-align: center;">Capacity</th>`;
                    html += `<th style="padding: 12px; text-align: center;">Allocated</th>`;
                    html += `</tr></thead><tbody>`;
                    
                    data.agents.forEach((agent, index) => {
                        const rowColor = agent.shift_start_time_parsed === 'Not parsed' ? '#fff3cd' : 'white';
                        html += `<tr style="background-color: ${rowColor}; border-bottom: 1px solid #e9ecef;">`;
                        html += `<td style="padding: 10px; text-align: center; font-weight: 600; color: #667eea;">${index + 1}</td>`;
                        html += `<td style="padding: 10px;"><strong>${agent.agent_name || 'N/A'}</strong><br><small style="color: #666;">ID: ${agent.agent_id || 'N/A'}</small></td>`;
                        html += `<td style="padding: 10px;">${agent.email || 'Not set'}</td>`;
                        html += `<td style="padding: 10px; text-align: center;"><code>${agent.shift_time_original || 'Not set'}</code></td>`;
                        html += `<td style="padding: 10px; text-align: center;"><strong style="color: ${agent.shift_start_time_parsed === 'Not parsed' ? '#dc3545' : '#28a745'};">${agent.shift_start_time_display || 'Not parsed'}</strong></td>`;
                        html += `<td style="padding: 10px; text-align: center;"><span style="background: ${agent.shift_group === 1 ? '#e3f2fd' : agent.shift_group === 2 ? '#fff3cd' : '#f3e5f5'}; padding: 5px 10px; border-radius: 4px; font-size: 12px;">${agent.shift_group_name || 'Not set'}</span></td>`;
                        html += `<td style="padding: 10px; text-align: center;">${agent.capacity || 0}</td>`;
                        html += `<td style="padding: 10px; text-align: center;">${agent.allocated || 0}</td>`;
                        html += `</tr>`;
                    });
                    
                    html += `</tbody></table></div>`;
                    
                    modalContent.innerHTML = html;
                } else {
                    modalContent.innerHTML = `<div style="padding: 20px; color: #dc3545;"><strong>Error:</strong> ${data.error || 'Failed to load shift times'}</div>`;
                }
            })
            .catch(error => {
                modalContent.innerHTML = `<div style="padding: 20px; color: #dc3545;"><strong>Error:</strong> ${error.message}</div>`;
            });
        }
        
        function viewAgentAllocation(agentName) {
            const modal = document.getElementById('agentModal');
            const modalAgentName = document.getElementById('modalAgentName');
            const modalContent = document.getElementById('modalContent');
            const modalStats = document.getElementById('modalStats');
            const downloadBtn = document.getElementById('downloadBtn');
            
            // Show modal and set agent name
            modal.style.display = 'block';
            modalAgentName.textContent = `${agentName} - Allocation Details`;
            
            // Show loading state
            modalContent.innerHTML = `
                <div style="text-align: center; padding: 40px;">
                    <i class="fas fa-spinner fa-spin" style="font-size: 2em; color: #667eea;"></i>
                    <p style="margin-top: 15px; color: #666;">Loading allocation data for ${agentName}...</p>
                </div>
            `;
            
            // Fetch agent allocation data
            fetch('/get_agent_allocation', {
                method: 'POST',
                headers: {
                    'Content-Type': 'application/json',
                },
                body: JSON.stringify({ agent_name: agentName })
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    // Display the data table
                    modalContent.innerHTML = data.html_table;
                    
                    // Update statistics
                    const stats = data.stats;
                    modalStats.innerHTML = `
                        <strong>Allocation Summary:</strong> 
                        ${stats.total_rows} rows allocated | 
                        Capacity: ${stats.capacity} | 
                        First Priority: ${stats.first_priority} | 
                        Second Priority: ${stats.second_priority} | 
                        Third Priority: ${stats.third_priority}
                    `;
                    
                    // Set up download button
                    downloadBtn.onclick = function() {
                        // Create a form and submit it to download the file
                        const form = document.createElement('form');
                        form.method = 'POST';
                        form.action = '/download_agent_file';
                        
                        const input = document.createElement('input');
                        input.type = 'hidden';
                        input.name = 'agent_name';
                        input.value = agentName;
                        
                        form.appendChild(input);
                        document.body.appendChild(form);
                        form.submit();
                        document.body.removeChild(form);
                    };
                } else {
                    modalContent.innerHTML = `
                        <div style="text-align: center; padding: 40px; color: #e74c3c;">
                            <i class="fas fa-exclamation-triangle" style="font-size: 2em;"></i>
                            <p style="margin-top: 15px;">Error loading allocation data: ${data.error}</p>
                        </div>
                    `;
                }
            })
            .catch(error => {
                modalContent.innerHTML = `
                    <div style="text-align: center; padding: 40px; color: #e74c3c;">
                        <i class="fas fa-exclamation-triangle" style="font-size: 2em;"></i>
                        <p style="margin-top: 15px;">Error loading allocation data: ${error.message}</p>
                    </div>
                `;
            });
        }
        
        // Modal close functionality
        function closeModal() {
            const modal = document.getElementById('agentModal');
            if (modal) {
                modal.style.display = 'none';
            }
        }
        
        // Set up modal close event listeners
        document.addEventListener('DOMContentLoaded', function() {
            // Close modal when clicking outside of it
            document.addEventListener('click', function(event) {
                const modal = document.getElementById('agentModal');
                if (modal && event.target === modal) {
                    closeModal();
                }
            });
            
            // Close modal when clicking X button
            document.addEventListener('click', function(event) {
                if (event.target.classList.contains('close')) {
                    closeModal();
                }
            });
            
            // Close modal when clicking close button in footer
            document.addEventListener('click', function(event) {
                if (event.target.classList.contains('close-btn')) {
                    closeModal();
                }
            });
            
            // Close modal when pressing Escape key
            document.addEventListener('keydown', function(event) {
                if (event.key === 'Escape') {
                    closeModal();
                }
            });
        });
    </script>
    
    <!-- Toast Notification Container -->
    <div id="toastContainer" style="position: fixed; top: 20px; right: 20px; z-index: 10000; display: flex; flex-direction: column; gap: 10px;"></div>
    
    <script>
    // Toast notification system - using CSS classes for proper styling
    function showToast(type, title, message, duration = 5000) {
        const container = document.getElementById('toastContainer');
        if (!container) {
            alert(message); // Fallback to alert
            return;
        }
        
        const toast = document.createElement('div');
        toast.className = `toast ${type}`;
        
        const icons = {
            success: 'fas fa-check-circle',
            error: 'fas fa-exclamation-circle',
            warning: 'fas fa-exclamation-triangle',
            info: 'fas fa-info-circle'
        };
        
        toast.innerHTML = `
            <i class="toast-icon ${icons[type]}"></i>
            <div class="toast-content">
                <div class="toast-title">${title}</div>
                <div class="toast-message">${message}</div>
            </div>
            <button class="toast-close" onclick="closeToast(this)">&times;</button>
        `;
        
        container.appendChild(toast);
        
        // Trigger animation
        setTimeout(() => toast.classList.add('show'), 100);
        
        // Auto remove
        setTimeout(() => {
            if (toast.parentNode) {
                closeToast(toast.querySelector('.toast-close'));
            }
        }, duration);
    }
    
    function closeToast(button) {
        const toast = button.closest('.toast');
        toast.classList.remove('show');
        setTimeout(() => {
            if (toast.parentNode) {
                toast.parentNode.removeChild(toast);
            }
        }, 300);
    }
    
    // Global toast functions for easy access
    window.showSuccessToast = (title, message) => showToast('success', title, message);
    window.showErrorToast = (title, message) => showToast('error', title, message);
    window.showWarningToast = (title, message) => showToast('warning', title, message);
    window.showInfoToast = (title, message) => showToast('info', title, message);
    </script>
    
    <!-- Loader/Progress Bar -->
    <div id="loader-overlay" class="loader-overlay">
        <div class="loader-container">
            <div class="loader-spinner"></div>
            <p class="loader-text">Loading Appointment Dates</p>
            <p class="loader-subtitle">Please wait while we process your files...</p>
            <div class="progress-bar-container">
                <div class="progress-bar"></div>
            </div>
        </div>
    </div>
</body>
</html>
"""


def get_business_days_until_date(start_date, target_date):
    """Calculate business days between start_date and target_date (excluding weekends)"""
    from datetime import timedelta

    if target_date < start_date:
        return -1  # Past date

    current_date = start_date
    business_days = 0

    while current_date < target_date:
        current_date += timedelta(days=1)
        # Check if it's a weekday (Monday=0, Sunday=6)
        if current_date.weekday() < 5:  # Monday to Friday
            business_days += 1

    return business_days


# Global variable to cache insurance name mapping
_insurance_name_mapping = None
_insurance_name_mapping_loaded = False
_formatted_insurance_names = set()  # Track formatted insurance names
_formatted_insurance_details = []  # Track original -> formatted mappings

# Insurance corrected list mappings (stored directly in code since file will be deleted)
CORRECTED_LIST_MAPPINGS = {
    "Always Care": "Always Care",
    "Always Care Dental Benefits": "Always Care",
    "BCBS Arizona": "BCBS Arizona",
    "BCBS Arizona FEP": "BCBS Arizona",
    "BCBS California Dental Plan": "BCBS California",
    "BCBS California FEP": "BCBS California FEP",
    "BCBS FEP BLUEDENTAL": "BCBS FEP",
    "BCBS FEP Dental": "BCBS FEP",
    "BCBS FEP Program": "BCBS FEP",
    "BCBS FEPOREGON": "BCBS FEP",
    "BCBS Federal Dental": "BCBS Federal",
    "BCBS Federal Gov`t": "BCBS Federal",
    "BCBS IDAHO": "BCBS IDAHO",
    "BCBS Idaho": "BCBS IDAHO",
    "BCBS Illinois  Federal": "BCBS Illinois",
    "BCBS Oregon FEP Program": "BCBS Oregon FEP Program",
    "BCBS Tennessee Federal Gov`t": "BCBS Tennessee Federal",
    "Beam Insurance Administrators": "Beam",
    "Benefit & Risk Management (BRMS  CA)": "Benefit & Risk Management",
    "Best Life": "Best Life",
    "Best Life & Health Insurance Co.": "Best Life",
    "BlueCross BlueShield AZ": "BCBS Arizona",
    "BlueShield AZ": "BCBS Arizona",
    "CCPOA": "CCPOA",
    "CENTRAL STATES": "CENTRAL STATES",
    "CONVERSION DEFAULT  Do NOT Delete! Change Pt Ins!": "CONVERSION DEFAULT  Do NOT Delete! Change Pt Ins!",
    "CarePlus": "CarePlus",
    "Careington Benefit Solutions": "Careington Benefit Solutions",
    "Central States Health & Life Co. Of Omaha": "Central States Health & Life Co. Of Omaha",
    "Cigna": "Cigna",
    "Community Dental Associates": "Community Dental Associates",
    "Core Five Solutions": "Core Five Solutions",
    "Cypress Ancillary Benefits": "Cypress Ancillary Benefits",
    "DD $2000 MAX": "DD $2000 MAX",
    "DD California Federal Plan": "DD California Federal Plan",
    "DD California Federal Services": "DD California Federal Plan",
    "DD DI": "DD DI",
    "DD Dental Choice": "DD Dental Choice",
    "DD FE": "DD FEP",
    "DD Fed Govt": "DD FEP",
    "DD Federal Employee Dental Pro": "DD FEP",
    "DD Federal Government Programs": "DD FEP",
    "DD GE": "DD GE",
    "DD GeorgiaBasic": "DD GeorgiaBasic",
    "DD IO": "DD IO",
    "DD Idaho": "DD Idaho",
    "DD Individual": "DD Individual",
    "DD Individual Plan": "DD Individual",
    "DD Indv": "DD Individual",
    "DD Ins Company": "DD Ins Company",
    "DD Insurance Colorado": "DD Colorado",
    "DD Iowa": "DD Iowa",
    "DD KA": "DD KA",
    "DD KE": "DD KE",
    "DD M": "DD M",
    "DD Mass": "DD Mass",
    "DD NO": "DD NO",
    "DD PE": "DD PE",
    "DD PL": "DD PL",
    "DD PLAN OF Wisconsin.": "DD Wisconsin.",
    "DD PP": "DD",
    "DD PPO": "DD",
    "DD Plan": "DD",
    "DD Plan Of Arizona": "DD Arizona",
    "DD Plan of Arizona": "DD Arizona",
    "DD RH": "DD Rhode Island",
    "DD Rhode Island": "DD Rhode Island",
    "DD SO": "DD SO",
    "DD TE": "DD Tennesse",
    "DD VI": "DD VI",
    "DD Wisconsin INDV": "DD Wisconsin INDV",
    "DD plan": "DD plan",
    "DDIC": "DDIC",
    "DELTA": "DD",
    "DENCAP Dental Plans": "DENCAP Dental Plans",
    "Delt Dental of CA": "DD California",
    "Delta": "DD",
    "Delta Deltal premier": "DD",
    "Delta Denta": "DD",
    "Delta MN": "DD Minnesota",
    "Delta WI": "DD Wisconsin",
    "Delta Wi": "DD Wisconsin",
    "Delta of WA": "DD of Washington",
    "DeltaCare USA": "DeltaCare USA",
    "Dental Claims": "Dental Claims",
    "Dental Claims Administrator": "Dental Claims",
    "FEP Blue Dental": "FEP Blue Dental",
    "FEP BlueDental": "FEP BlueDental",
    "Fiedler Dentistry Membership Plan": "Fiedler Dentistry Membership Plan",
    "LINE CONSTRUCTION  LINECO": "LIneco",
    "LIneco": "LIneco",
    "Liberty Dental Plan": "Liberty Dental",
    "Lincoln Financial Group": "Lincoln Financial Group",
    "Lincoln Financial Group (Lincoln Nationa": "Lincoln Financial Group",
    "Line Construction Benefit Fund": "LIneco",
    "Manhattan Life": "Manhattan Life",
    "Medical Mutual": "Medical Mutual",
    "Medico Insurance Company": "Medico Insurance Company",
    "Meritain": "Meritain",
    "Meritain Health": "Meritain",
    "Met": "Metlife",
    "Metlife": "Metlife",
    "Metropolitan": "Metropolitan",
    "Moonlight Graham": "Moonlight Graham",
    "Mutual Omaha": "Mutual Omaha",
    "NECAIBEW Welfare Trust Fund": "NECAIBEW Welfare Trust Fund",
    "NHW": "NHW",
    "NTCA": "NTCA",
    "NTCA Benefits": "NTCA",
    "National Elevator Industry Health Benefit Plan": "National Elevator Industry Plan",
    "National Elevator Industry Plan": "National Elevator Industry Plan",
    "Network Health Wisconsin": "Network Health Wisconsin",
    "Nippon Life Insurance": "Nippon Life Insurance",
    "Novartis Corporation": "Novartis Corporation",
    "OSF MedAdvantage": "OSF MedAdvantage",
    "Oakland County Discount Plan": "Oakland County Discount Plan",
    "Operating Engineers Local #49": "Operating Engineers Local #49",
    "PACIFIC SOURCE": "PACIFIC SOURCE",
    "PacificSource Health Plans": "PacificSource Health Plans",
    "Paramount Dental": "Paramount Dental",
    "Perio Membership Plan August": "Perio Membership Plan August",
    "Physician's Mutual": "Physicians Mutual",
    "Physicians Mutual": "Physicians Mutual",
    "Plan for Health": "Plan for Health",
    "Prairie States": "Prairie States",
    "Principal": "Principal",
    "Principlal": "Principal",
    "Professional Benefits Administr": "Professional Benefits Administr",
    "REGENCE BCBS": "REGENCE BCBS",
    "Regarding Dentistry  Membership": "Regarding Dentistry  Membership",
    "Reliance Standard": "Reliance Standard",
    "Renaissance": "Renaissance",
    "Renaissance Life and Health": "Renaissance",
    "Renaissance, Dental": "Renaissance",
    "SIHO": "SIHO",
    "Secure Care Dental": "Secure Care Dental",
    "Security Life Ins of America": "Security Life Ins of America",
    "Simple Dental": "Simple Dental",
    "Standard Life Insurance": "Standard Life Insurance",
    "Strong Family Health": "Strong Family Health",
    "Sunlife": "Sunlife",
    "Superior Dental Care": "Superior Dental Care",
    "THE UNITED FURNITURE WORKERS INSURANCE F": "THE UNITED FURNITURE WORKERS INSURANCE F",
    "Team Care": "Teamcare",
    "Teamcare": "Teamcare",
    "Texas International Life Ins Co": "Texas International Life Ins Co",
    "The Benefit Group": "The Benefit Group",
    "Tricare": "Tricare",
    "TruAssure Insurance Company": "TruAssure",
    "UHC": "UHC",
    "UMR": "UMR",
    "US Health Group": "US Health Group",
    "United Concordia": "UCCI",
    "Unum": "Unum",
    "WilsonMcShane Corporation": "WilsonMcShane Corporation",
}


def clean_insurance_name(name):
    """Remove spaces and special characters from the beginning and end of insurance name"""
    if not name or pd.isna(name):
        return name

    name_str = str(name)
    # Remove spaces and special characters from start and end
    # Special characters: - . , ; : | / \ _ ( ) [ ] { } * # @ $ % ^ & + = ~ ` ' " < > ?
    name_str = re.sub(r'^[\s\-.,;:|/\\_()\[\]{}*#@$%^&+=\~`\'"<>?]+', "", name_str)
    name_str = re.sub(r'[\s\-.,;:|/\\_()\[\]{}*#@$%^&+=\~`\'"<>?]+$', "", name_str)
    return name_str.strip()


def load_insurance_name_mapping():
    """Load insurance name mapping from Insurance Uniform Name.xlsx file and CORRECTED_LIST_MAPPINGS dictionary"""
    global _insurance_name_mapping, _insurance_name_mapping_loaded

    if _insurance_name_mapping_loaded:
        return _insurance_name_mapping

    _insurance_name_mapping = {}
    total_mappings = 0

    # Load from Insurance Uniform Name.xlsx
    mapping_file1 = "Insurance Uniform Name.xlsx"
    try:
        if os.path.exists(mapping_file1):
            df = pd.read_excel(mapping_file1)
            count = 0
            # Create mapping dictionary: original name -> uniform name
            # Handle case-insensitive matching by storing both original and lowercase keys
            for _, row in df.iterrows():
                original = (
                    str(row["Insurance"]).strip() if pd.notna(row["Insurance"]) else ""
                )
                uniform = (
                    str(row["Insurance New"]).strip()
                    if pd.notna(row["Insurance New"])
                    else ""
                )

                if original and uniform:
                    # Clean both original and formatted names
                    original = clean_insurance_name(original)
                    uniform = clean_insurance_name(uniform)

                    if original and uniform:
                        # Store with original case
                        _insurance_name_mapping[original] = uniform
                        # Also store with lowercase for case-insensitive lookup
                        _insurance_name_mapping[original.lower()] = uniform
                        count += 1

            total_mappings += count
        else:
            pass
    except Exception as e:
        pass

    # Load from CORRECTED_LIST_MAPPINGS dictionary (stored directly in code)
    try:
        count = 0
        for original, formatted in CORRECTED_LIST_MAPPINGS.items():
            # Clean both original and formatted names
            original_clean = clean_insurance_name(original)
            formatted_clean = clean_insurance_name(formatted)

            if original_clean and formatted_clean:
                # Store with original case
                _insurance_name_mapping[original_clean] = formatted_clean
                # Also store with lowercase for case-insensitive lookup
                _insurance_name_mapping[original_clean.lower()] = formatted_clean
                count += 1

        total_mappings += count
    except Exception as e:
        pass

    if total_mappings > 0:
        pass
    else:
        pass

    _insurance_name_mapping_loaded = True
    return _insurance_name_mapping


def format_insurance_company_name(insurance_text):
    """Format insurance company name for better allocation matching - uses Insurance Uniform Name.xlsx mapping first"""
    global _formatted_insurance_names, _formatted_insurance_details

    if pd.isna(insurance_text):
        return insurance_text

    insurance_str = clean_insurance_name(insurance_text)
    if not insurance_str:
        return insurance_text  # Return original if cleaning results in empty

    original_name = insurance_str  # Keep original for tracking
    matched_from_mapping = False  # Track if matched from mapping file

    # Ensure insurance_str is a string
    if not isinstance(insurance_str, str):
        insurance_str = str(insurance_str) if insurance_str is not None else ""

    # Handle special cases first (after cleaning)
    if insurance_str.upper() == "NO INSURANCE":
        formatted = clean_insurance_name("No Insurance")
    elif insurance_str.upper() == "PATIENT NOT FOUND":
        formatted = clean_insurance_name("PATIENT NOT FOUND")
    elif insurance_str.upper() == "DUPLICATE":
        formatted = clean_insurance_name("DUPLICATE")
    elif insurance_str.upper() == "UNKNOWN":
        formatted = clean_insurance_name("Unknown")
    else:
        # Load insurance name mapping (will only load once)
        mapping = load_insurance_name_mapping()
        formatted = None

        # Try exact match first (with original text)
        if insurance_str in mapping:
            formatted = clean_insurance_name(str(mapping[insurance_str]))
            matched_from_mapping = True
        else:
            # Extract company name before "Ph#" or phone numbers for matching
            if "Ph#" in insurance_str:
                company_name = insurance_str.split("Ph#")[0]
            elif re.search(r"Ph#:?-?\s*\(?\d{3}", insurance_str):
                # Handle various phone number patterns
                company_name = re.split(r"Ph#:?-?\s*\(?\d", insurance_str)[0]
            else:
                company_name = insurance_str

            # Clean the company name
            company_name = clean_insurance_name(company_name)

            # Try matching with cleaned company name
            if company_name and company_name in mapping:
                formatted = clean_insurance_name(str(mapping[company_name]))
                matched_from_mapping = True
            elif company_name:
                company_name_str = (
                    str(company_name)
                    if not isinstance(company_name, str)
                    else company_name
                )
                if company_name_str.lower() in mapping:
                    formatted = clean_insurance_name(
                        str(mapping[company_name_str.lower()])
                    )
                    matched_from_mapping = True

            # Try matching original string (lowercase) as fallback
            if not formatted:
                insurance_str_safe = (
                    str(insurance_str)
                    if not isinstance(insurance_str, str)
                    else insurance_str
                )
                if insurance_str_safe.lower() in mapping:
                    formatted = clean_insurance_name(
                        str(mapping[insurance_str_safe.lower()])
                    )
                    matched_from_mapping = True

            # Remove "Primary" and "Secondary" text
            if not formatted and company_name:
                company_name = re.sub(
                    r"\s*\(Primary\)", "", company_name, flags=re.IGNORECASE
                )
                company_name = re.sub(
                    r"\s*\(Secondary\)", "", company_name, flags=re.IGNORECASE
                )
                company_name = re.sub(
                    r"\s*Primary", "", company_name, flags=re.IGNORECASE
                )
                company_name = re.sub(
                    r"\s*Secondary", "", company_name, flags=re.IGNORECASE
                )
                company_name = clean_insurance_name(company_name)

                # Try matching again after removing Primary/Secondary
                if company_name and company_name in mapping:
                    formatted = clean_insurance_name(str(mapping[company_name]))
                    matched_from_mapping = True
                elif company_name:
                    company_name_str = (
                        str(company_name)
                        if not isinstance(company_name, str)
                        else company_name
                    )
                    if company_name_str.lower() in mapping:
                        formatted = clean_insurance_name(
                            str(mapping[company_name_str.lower()])
                        )
                        matched_from_mapping = True

        # If no match found in mapping, use fallback logic (existing code continues below)
        if not formatted:
            formatted = None  # Will be set by fallback logic

    # Continue with fallback logic if no mapping match found
    if not formatted:
        # Use existing fallback formatting logic
        # State abbreviations mapping
        STATE_ABBREVIATIONS = {
            "AL": "Alabama",
            "AK": "Alaska",
            "AR": "Arkansas",
            "AZ": "Arizona",
            "CA": "California",
            "CL": "California",
            "CO": "Colorado",
            "CT": "Connecticut",
            "DE": "Delaware",
            "DC": "District of Columbia",
            "FL": "Florida",
            "GA": "Georgia",
            "HI": "Hawaii",
            "ID": "Idaho",
            "IL": "Illinois",
            "IN": "Indiana",
            "IA": "Iowa",
            "KS": "Kansas",
            "KY": "Kentucky",
            "LA": "Louisiana",
            "ME": "Maine",
            "MD": "Maryland",
            "MA": "Massachusetts",
            "MI": "Michigan",
            "MN": "Minnesota",
            "MS": "Mississippi",
            "MO": "Missouri",
            "MT": "Montana",
            "NE": "Nebraska",
            "NV": "Nevada",
            "NH": "New Hampshire",
            "NJ": "New Jersey",
            "NM": "New Mexico",
            "NY": "New York",
            "NC": "North Carolina",
            "ND": "North Dakota",
            "OH": "Ohio",
            "OK": "Oklahoma",
            "OR": "Oregon",
            "PA": "Pennsylvania",
            "RI": "Rhode Island",
            "SC": "South Carolina",
            "SD": "South Dakota",
            "TN": "Tennessee",
            "TX": "Texas",
            "UT": "Utah",
            "VT": "Vermont",
            "VA": "Virginia",
            "WA": "Washington",
            "WV": "West Virginia",
            "WI": "Wisconsin",
            "WY": "Wyoming",
        }

        # Common state name typos and variations
        STATE_TYPO_CORRECTIONS = {
            "californi": "California",  # Missing last letter
            "californa": "California",  # Common typo
            "californai": "California",  # Letter order typo
            "colarado": "Colorado",  # Missing 'o' typo
            "minnesotta": "Minnesota",  # Extra 't' typo
        }

        def expand_state_abbreviations(text):
            """Expand state abbreviations to full state names"""
            for abbr, full_name in STATE_ABBREVIATIONS.items():
                pattern = r"\b" + re.escape(abbr) + r"\b"
                text = re.sub(pattern, full_name, text, flags=re.IGNORECASE)
            return text

        def correct_state_typos(text):
            """Correct common state name typos"""
            if not text:
                return text

            # Replace any occurrence of typo words with correct spelling (case-insensitive)
            for typo, correct in STATE_TYPO_CORRECTIONS.items():
                pattern = r"\b" + re.escape(typo) + r"\b"
                text = re.sub(pattern, correct, text, flags=re.IGNORECASE)

            return text

        def format_state_name(state_text):
            """Format state name: first letter capital, rest lowercase (handles all caps, mixed case, etc.)"""
            if not state_text:
                return state_text

            # Handle multi-word state names (e.g., "NEW YORK" -> "New York", "NORTH CAROLINA" -> "North Carolina")
            words = state_text.split()
            formatted_words = []
            for word in words:
                if word.isupper() and len(word) > 1:
                    # If word is all caps, convert to title case (first letter capital, rest lowercase)
                    formatted_words.append(word.capitalize())
                elif word.isupper() and len(word) == 1:
                    # Single letter stays as is (for abbreviations like "I" in "Rhode Island")
                    formatted_words.append(word)
                elif word.islower():
                    # All lowercase, capitalize first letter
                    formatted_words.append(word.capitalize())
                elif word.istitle():
                    # Already in title case (e.g., "New"), keep as is
                    formatted_words.append(word)
                else:
                    # Mixed case or other, convert to title case
                    formatted_words.append(word.capitalize())

            return " ".join(formatted_words)

        # Handle Delta Dental variations - normalize to "DD {state}" format
        if re.search(r"\bDD\b", company_name, re.IGNORECASE):
            # Handle existing "DD" patterns like "DD California", "DD of California", "DD CA", "DD PLAN OF Wisconsin"
            dd_match = re.search(
                r"\bDD\b\s+(?:plan\s+of\s+|of\s+)?(.+)", company_name, re.IGNORECASE
            )
            if dd_match:
                state = clean_insurance_name(dd_match.group(1))
                # Remove "PLAN OF" if it appears in the state text
                state = re.sub(r"\bplan\s+of\s+", "", state, flags=re.IGNORECASE)
                # Remove common suffixes
                state = re.sub(r"\s*\(.*?\)", "", state)
                # Remove trailing periods, pipes, and other special characters
                state = re.sub(r"[|.]+\s*$", "", state)
                state = correct_state_typos(state)
                state = expand_state_abbreviations(state)
                state = format_state_name(state)
                formatted = clean_insurance_name(f"DD {state}")
            else:
                formatted = clean_insurance_name("DD")
        elif re.search(r"delta\s+dental", company_name, re.IGNORECASE):
            # Handle "Delta Dental" variations
            delta_match = re.search(
                r"delta\s+dental\s+(?:of\s+)?(.+)", company_name, re.IGNORECASE
            )
            if delta_match:
                state = clean_insurance_name(delta_match.group(1))
                # Remove "PLAN OF" if it appears in the state text
                state = re.sub(r"\bplan\s+of\s+", "", state, flags=re.IGNORECASE)
                # Remove common suffixes
                state = re.sub(r"\s*\(.*?\)", "", state)
                # Remove trailing periods, pipes, and other special characters
                state = re.sub(r"[|.]+\s*$", "", state)
                state = correct_state_typos(state)
                state = expand_state_abbreviations(state)
                state = format_state_name(state)
                formatted = clean_insurance_name(f"DD {state}")
            else:
                formatted = clean_insurance_name("DD")

        # Handle Anthem variations FIRST (before BCBS to avoid conflicts)
        elif re.search(
            r"anthem|blue\s+cross.*anthem|anthem.*blue\s+cross",
            company_name,
            re.IGNORECASE,
        ):
            formatted = clean_insurance_name("Anthem")

        # Handle BCBS variations (including BC/BS with slash)
        elif re.search(
            r"bc\s*/\s*bs|bcbs|bc\s+of|blue\s+cross|blue\s+shield|bcbbs",
            company_name,
            re.IGNORECASE,
        ):
            # Check for "BCBS / BLUE SHEILD", "BCBS Blue Shiel", "BCBS Blue Shield" -> just "BCBS"
            # Handles: "shiel" (without 'd'), "shield" (correct spelling), "sheild" (misspelling)
            if re.search(
                r"bcbs\s*/\s*blue\s+(shiel|shield|sheild)", company_name, re.IGNORECASE
            ) or re.search(
                r"bcbs\s+blue\s+(shiel|shield|sheild)", company_name, re.IGNORECASE
            ):
                formatted = clean_insurance_name("BCBS")
            # Handle BCBBS typo
            elif re.search(r"bcbbs", company_name, re.IGNORECASE):
                formatted = clean_insurance_name("BCBS")
            # Check for full "Blue Cross Blue Shield" pattern first
            elif re.search(
                r"blue\s+cross\s+blue\s+shield", company_name, re.IGNORECASE
            ):
                bcbs_match = re.search(
                    r"blue\s+cross\s+blue\s+shield\s+(?:of\s+)?(.+)",
                    company_name,
                    re.IGNORECASE,
                )
                if bcbs_match:
                    state = bcbs_match.group(1)
                    # Remove trailing dashes and extra text like "- federal", "- Federal", etc.
                    state = re.sub(
                        r"\s*-\s*(federal|Federal|FEDERAL).*$",
                        "",
                        state,
                        flags=re.IGNORECASE,
                    )
                    # Remove common suffixes in parentheses
                    state = re.sub(r"\s*\(.*?\)", "", state)
                    state = clean_insurance_name(state)
                    state = re.sub(r"[|.]+\s*$", "", state)
                    state = correct_state_typos(state)
                    state = expand_state_abbreviations(state)
                    state = format_state_name(state)
                    formatted = (
                        clean_insurance_name(f"BCBS {state}")
                        if state
                        else clean_insurance_name("BCBS")
                    )
                else:
                    formatted = clean_insurance_name("BCBS")
            # Handle BC/BS patterns
            elif re.search(r"bc/bs", company_name, re.IGNORECASE):
                bcbs_match = re.search(
                    r"bc/bs\s+(?:of\s+)?(.+)", company_name, re.IGNORECASE
                )
                if bcbs_match:
                    state = bcbs_match.group(1)
                    state = re.sub(
                        r"\s*-\s*(federal|Federal|FEDERAL).*$",
                        "",
                        state,
                        flags=re.IGNORECASE,
                    )
                    state = re.sub(r"\s*\(.*?\)", "", state)
                    state = clean_insurance_name(state)
                    state = re.sub(r"[|.]+\s*$", "", state)
                    state = correct_state_typos(state)
                    state = expand_state_abbreviations(state)
                    state = format_state_name(state)
                    formatted = (
                        clean_insurance_name(f"BCBS {state}")
                        if state
                        else clean_insurance_name("BCBS")
                    )
                else:
                    formatted = clean_insurance_name("BCBS")
            # Handle BC Of patterns
            elif re.search(r"bc\s+of", company_name, re.IGNORECASE):
                bcbs_match = re.search(r"bc\s+of\s+(.+)", company_name, re.IGNORECASE)
                if bcbs_match:
                    state = bcbs_match.group(1)
                    state = re.sub(
                        r"\s*-\s*(federal|Federal|FEDERAL).*$",
                        "",
                        state,
                        flags=re.IGNORECASE,
                    )
                    state = re.sub(r"\s*\(.*?\)", "", state)
                    state = clean_insurance_name(state)
                    state = re.sub(r"[|.]+\s*$", "", state)
                    state = correct_state_typos(state)
                    state = expand_state_abbreviations(state)
                    state = format_state_name(state)
                    formatted = (
                        clean_insurance_name(f"BCBS {state}")
                        if state
                        else clean_insurance_name("BCBS")
                    )
                else:
                    formatted = clean_insurance_name("BCBS")
            # Handle other BCBS patterns
            else:
                bcbs_match = re.search(
                    r"(?:bcbs|blue\s+cross|blue\s+shield)\s+(?:of\s+)?(.+)",
                    company_name,
                    re.IGNORECASE,
                )
                if bcbs_match:
                    state = bcbs_match.group(1)
                    # Remove trailing dashes and extra text like "- federal", "- Federal", etc.
                    state = re.sub(
                        r"\s*-\s*(federal|Federal|FEDERAL).*$",
                        "",
                        state,
                        flags=re.IGNORECASE,
                    )
                    # Remove common suffixes in parentheses
                    state = re.sub(r"\s*\(.*?\)", "", state)
                    # Remove trailing dashes and special characters
                    state = clean_insurance_name(state)
                    # Remove trailing periods, pipes, and other special characters
                    state = re.sub(r"[|.]+\s*$", "", state)
                    state = correct_state_typos(state)
                    state = expand_state_abbreviations(state)
                    state = format_state_name(state)
                    if state:
                        formatted = clean_insurance_name(f"BCBS {state}")
                    else:
                        formatted = clean_insurance_name("BCBS")
                else:
                    formatted = clean_insurance_name("BCBS")

        # Handle other specific companies
        elif re.search(r"metlife|met\s+life", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Metlife")
        elif re.search(r"cigna", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Cigna")
        elif re.search(r"aarp", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("AARP")
        elif re.search(
            r"uhc|united\s*healthcare|united\s*health\s*care",
            company_name,
            re.IGNORECASE,
        ):
            formatted = clean_insurance_name("UHC")
        elif re.search(r"teamcare", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Teamcare")
        elif re.search(r"humana", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Humana")
        elif re.search(r"aetna", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Aetna")
        elif re.search(r"guardian", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Guardian")
        elif re.search(r"g\s*e\s*h\s*a", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("GEHA")
        elif re.search(r"principal", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Principal")
        elif re.search(r"ameritas", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Ameritas")
        elif re.search(r"physicians\s+mutual", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Physicians Mutual")
        elif re.search(r"mutual\s+of\s+omaha", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Mutual Omaha")
        elif re.search(r"sunlife|sun\s+life", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Sunlife")
        elif re.search(r"careington", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Careington Benefit Solutions")
        elif re.search(r"automated\s+benefit", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Automated Benefit Services Inc")
        elif re.search(r"regence", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("REGENCE BCBS")
        elif re.search(r"united\s+concordia", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("United Concordia")
        elif re.search(r"medical\s+mutual", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Medical Mutual")
        elif re.search(r"unum", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Unum")
        elif re.search(r"wilson\s+mcshane", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Wilson McShane- Delta Dental")
        elif re.search(r"dentaquest", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Dentaquest")
        elif re.search(r"umr", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("UMR")
        elif re.search(r"adn\s+administrators", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("ADN Administrators")
        elif re.search(r"beam", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Beam")
        elif re.search(r"liberty(?:\s+dental)?", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Liberty Dental Plan")
        elif re.search(r"ucci", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("UCCI")
        elif re.search(
            r"ccpoa|cc\s*poa|c\s+c\s+p\s+o\s+a", company_name, re.IGNORECASE
        ):
            formatted = clean_insurance_name("CCPOA")
        elif re.search(r"kansas\s+city", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Kansas City")
        elif re.search(r"the\s+guardian", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("The Guardian")
        elif re.search(r"community\s+dental", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Community Dental Associates")
        elif re.search(r"northeast\s+delta\s+dental", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Northeast Delta Dental")
        elif re.search(r"equitable", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Equitable")
        elif re.search(r"manhattan\s+life", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Manhattan Life")
        elif re.search(
            r"standard\s+(?:life\s+)?insurance", company_name, re.IGNORECASE
        ):
            formatted = clean_insurance_name("Standard Life Insurance")
        elif re.search(r"keenan", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Keenan")
        elif re.search(r"plan\s+for\s+health", company_name, re.IGNORECASE):
            formatted = clean_insurance_name("Plan for Health")
        elif re.search(r"conversion\s+default", company_name, re.IGNORECASE):
            formatted = clean_insurance_name(
                "CONVERSION DEFAULT - Do NOT Delete! Change Pt Ins!"
            )
        elif re.search(r"health\s*partners", company_name, re.IGNORECASE):
            # Check if it has "of [State]" pattern
            hp_match = re.search(
                r"health\s*partners\s+of\s+(.+)", company_name, re.IGNORECASE
            )
            if hp_match:
                state = hp_match.group(1).strip()
                state = clean_insurance_name(state)
                state = format_state_name(state)
                formatted = clean_insurance_name(f"Health Partners {state}")
            else:
                formatted = clean_insurance_name("Health Partners")
        elif re.search(r"network\s+health", company_name, re.IGNORECASE):
            # Check if it has "Wisconsin" in the name
            if re.search(r"wisconsin", company_name, re.IGNORECASE):
                formatted = clean_insurance_name("Network Health Wisconsin")
            else:
                formatted = clean_insurance_name("Network Health Go")
        else:
            # If no specific pattern matches, return the cleaned company name
            formatted = (
                clean_insurance_name(company_name) if company_name else company_name
            )

    # Track formatted names (only if different from original)
    if formatted and formatted != original_name:
        if original_name not in _formatted_insurance_names:
            _formatted_insurance_names.add(original_name)
            _formatted_insurance_details.append(
                {
                    "original": original_name,
                    "formatted": formatted,
                    "from_mapping": matched_from_mapping,
                }
            )

    # Ensure final output is cleaned (remove any spaces/special chars before/after)
    return clean_insurance_name(formatted) if formatted else formatted


def print_formatted_insurance_companies():
    """Print list of all formatted insurance companies to console"""
    global _formatted_insurance_details

    if not _formatted_insurance_details:
        return

    # Group by source (mapping vs fallback)
    from_mapping = [d for d in _formatted_insurance_details if d["from_mapping"]]
    from_fallback = [d for d in _formatted_insurance_details if not d["from_mapping"]]

    if from_mapping:
        for i, detail in enumerate(from_mapping, 1):
            pass

    if from_fallback:
        for i, detail in enumerate(from_fallback, 1):
            pass


# DD INS group mapping - these companies should be treated as part of "DD INS" or "INS" group
DD_INS_GROUP = [
    "DD California",
    "DD Florida",
    "DD Texas",
    "DD Pennsylvania",
    "DD New York",
    "DD Alabama",
    "DD Georgia",
    "DD Delaware",
]

# DD Toolkit group mapping - these companies should be treated as part of "DD Toolkit", "DD Toolkits", or "DD" group
DD_TOOLKIT_GROUP = [
    "DD New Mexico",
    "DD Ohio",
    "DD Indiana",
    "DD Michigan",  # Note: user mentioned "Michigen" but correct spelling is "Michigan"
    "DD Minnesota",
    "DD Tennessee",
    "DD Arizona",
    "DD North Carolina",
    "DD California Federal",
]


def expand_insurance_groups(insurance_list_str):
    """
    Expand insurance group names to include all companies in those groups.
    Handles:
    - "DD INS" or "INS" -> expands to DD_INS_GROUP companies
    - "DD Toolkit", "DD Toolkits", or "DD" (when used as group) -> expands to DD_TOOLKIT_GROUP companies

    Args:
        insurance_list_str: String containing insurance companies separated by ; , or |

    Returns:
        String with group names expanded to individual companies
    """
    if pd.isna(insurance_list_str) or not insurance_list_str:
        return insurance_list_str

    # Ensure input is a string
    if not isinstance(insurance_list_str, str):
        insurance_list_str = (
            str(insurance_list_str) if insurance_list_str is not None else ""
        )

    value_str = str(insurance_list_str)
    # Split by common delimiters and ensure each component is a string
    companies = [
        str(comp).strip()
        for comp in re.split(r"[;,\|]", value_str)
        if comp and str(comp).strip()
    ]

    expanded_companies = []
    has_dd_ins = False
    has_ins = False
    has_dd_toolkit = False
    has_dd_toolkits = False
    has_dd_group = False

    for comp in companies:
        comp_str = str(comp) if comp is not None else ""
        comp_lower = comp_str.lower().strip()

        # Check for "DD INS" or "INS" (case-insensitive)
        if comp_lower == "dd ins" or comp_lower == "ins":
            if "dd" in comp_lower:
                has_dd_ins = True
            else:
                has_ins = True
            # Don't add the group name itself, we'll add the group companies
        # Check for "DD Toolkit", "DD Toolkits", or "DD" (as group name)
        elif comp_lower == "dd toolkit":
            has_dd_toolkit = True
        elif comp_lower == "dd toolkits":
            has_dd_toolkits = True
        elif comp_lower == "dd":
            # "DD" can be a group name OR a standalone company name
            # We'll treat it as a group name if it appears alone or with other groups
            # To be safe, we'll check if there are other group keywords nearby
            has_dd_group = True
        else:
            # Keep other companies as-is
            expanded_companies.append(comp)

    # Add all DD INS group companies if DD INS or INS was found
    if has_dd_ins or has_ins:
        for dd_ins_company in DD_INS_GROUP:
            # Check if company is already in the list (case-insensitive)
            if not any(
                str(existing).lower() == str(dd_ins_company).lower()
                for existing in expanded_companies
            ):
                expanded_companies.append(dd_ins_company)
        expansion_type = "DD INS" if has_dd_ins else "INS"

    # Add all DD Toolkit group companies if DD Toolkit/Toolkits/DD was found
    if has_dd_toolkit or has_dd_toolkits or has_dd_group:
        for dd_toolkit_company in DD_TOOLKIT_GROUP:
            # Check if company is already in the list (case-insensitive)
            if not any(
                str(existing).lower() == str(dd_toolkit_company).lower()
                for existing in expanded_companies
            ):
                expanded_companies.append(dd_toolkit_company)
        expansion_type = (
            "DD Toolkit"
            if has_dd_toolkit
            else ("DD Toolkits" if has_dd_toolkits else "DD")
        )

    # Join back with semicolon - ensure all items are strings
    expanded_companies_str = [
        str(comp).strip()
        for comp in expanded_companies
        if comp is not None and not pd.isna(comp)
    ]
    return (
        "; ".join(expanded_companies_str)
        if expanded_companies_str
        else insurance_list_str
    )


def should_skip_row_for_allocation(idx, processed_df, remark_col):
    """
    Check if a row should be skipped for allocation (e.g., "Not to work" remark).
    Returns True if row should be skipped, False otherwise.
    """
    if remark_col and remark_col in processed_df.columns:
        if pd.notna(processed_df.at[idx, remark_col]):
            remark_val = str(processed_df.at[idx, remark_col]).strip().upper()
            # Skip rows with "Not to work" remark (check multiple variations)
            remark_val_clean = remark_val.replace("-", " ").replace("_", " ").strip()
            if (
                "NOT TO WORK" in remark_val
                or remark_val == "NOT TO WORK"
                or "NOTTOWORK" in remark_val.replace(" ", "")
                or remark_val_clean == "NOT TO WORK"
            ):
                return True
    return False


# Maximum secondary insurance rows that can be allocated to "Sec + XX" agents
MAX_SECONDARY_ROWS_PER_AGENT = 10

# Maximum rows per appointment date that can be allocated to any agent
MAX_ROWS_PER_APPOINTMENT_DATE = 20


def get_available_secondary_slots(agent):
    """
    Get the number of secondary insurance rows that can still be allocated to a "Sec + XX" agent (max 10 total).
    Returns the number of available slots for secondary insurance rows.
    """
    secondary_count = agent.get("secondary_insurance_count", 0)
    return max(0, MAX_SECONDARY_ROWS_PER_AGENT - secondary_count)


def can_allocate_row_by_appointment_date(
    agent, row_idx, processed_df, appointment_date_col
):
    """
    Check if a row can be allocated to an agent based on appointment date limit (max 20 per date).
    Returns True if row can be allocated, False otherwise.
    Also updates the agent's appointment_date_counts if allocation is allowed.
    """
    if not appointment_date_col or appointment_date_col not in processed_df.columns:
        return True  # No appointment date column - allow allocation

    if row_idx >= len(processed_df):
        return False

    # Initialize appointment date tracking for agent if not exists
    if "appointment_date_counts" not in agent:
        agent["appointment_date_counts"] = {}

    # Get appointment date for this row
    appointment_date = None
    if pd.notna(processed_df.at[row_idx, appointment_date_col]):
        appointment_date_val = processed_df.at[row_idx, appointment_date_col]
        # Convert to date if it's datetime
        if hasattr(appointment_date_val, "date"):
            appointment_date = appointment_date_val.date()
        elif isinstance(appointment_date_val, str):
            try:
                # Try to parse date string
                from datetime import datetime

                appointment_date = datetime.strptime(
                    appointment_date_val.split()[0], "%Y-%m-%d"
                ).date()
            except:
                appointment_date = str(appointment_date_val).strip()
        else:
            appointment_date = str(appointment_date_val).strip()

    if appointment_date:
        # Convert to string key for dictionary
        date_key = str(appointment_date)
        # Check if agent already has 20 rows for this appointment date
        current_count = agent["appointment_date_counts"].get(date_key, 0)
        if current_count < MAX_ROWS_PER_APPOINTMENT_DATE:
            # Update count and allow allocation
            agent["appointment_date_counts"][date_key] = current_count + 1
            return True
        else:
            # Agent already has 20 rows for this appointment date
            return False
    else:
        # If no appointment date, allow allocation (fallback)
        return True


def safe_extend_row_indices(
    agent,
    row_indices_list,
    processed_df,
    remark_col,
    agent_name,
    appointment_date_col=None,
):
    """
    Safely extend agent's row_indices, filtering out "Not to work" rows.
    Also enforces limit of 20 rows per appointment date per agent.
    Returns the number of rows actually allocated.
    """
    filtered_indices = [
        idx
        for idx in row_indices_list
        if not should_skip_row_for_allocation(idx, processed_df, remark_col)
    ]

    # Filter by appointment date limit (max 20 rows per appointment date per agent)
    if appointment_date_col and appointment_date_col in processed_df.columns:
        final_filtered_indices = []
        for idx in filtered_indices:
            if idx < len(processed_df):
                if can_allocate_row_by_appointment_date(
                    agent,
                    idx,
                    processed_df,
                    appointment_date_col,
                ):
                    final_filtered_indices.append(idx)
        filtered_indices = final_filtered_indices

    if filtered_indices:
        agent["row_indices"].extend(filtered_indices)
        agent["allocated"] += len(filtered_indices)
        for idx in filtered_indices:
            processed_df.at[idx, "Agent Name"] = agent_name

    return len(filtered_indices)


# Insurance companies that "Afreen Ansari" can work with (in addition to her regular list)
AFREEN_ANSARI_ADDITIONAL_INSURANCE = [
    "Beam",
    "Best Life",
    "Careington Benefit Solutions",
    "Cypress Ancillary Benefits",
    "Delta Denta WI",
    "Delta MN CDMN1 07000",
    "Employer Driven Insurance Services (E.D.I.S",
    "Meritain",
    "MODA HEALTH",
    "NESIKA HEALTH LLC",
    "NTCA",
    "ODS/Moda",
    "Pacific Life Dental",
    "PACIFIC SOURCE",
    "Physicians Mutual",
    "Purchase & Referred Care",
    "Reliance Standard",
    "Southwest Service Administrators",
    "Superior Dental Care",
    "TRUST DENTAL CLAIMS",
    "UHC",
    "UMR",
    "Unum",
    "Westfield Dental Membership",
    "Zenith American Solutions",
]


def check_insurance_match(
    row_insurance, agent_insurance_list, is_senior=False, agent_name=None
):
    """
    Check if a row's insurance company matches any of the agent's insurance capabilities.
    Handles formatting variations, especially for DD Toolkit/Toolkits/DD INS variations.

    Args:
        row_insurance: Insurance company name from the data row
        agent_insurance_list: List of insurance companies the agent can work with
        is_senior: Whether the agent is senior (can work with any insurance)
        agent_name: Name of the agent (for special cases like "Afreen Ansari")

    Returns:
        True if the row insurance matches agent capabilities, False otherwise
    """
    if is_senior:
        return True

    # Special case: "Afreen Ansari" can work with additional insurance companies
    if agent_name == "Afreen Ansari":
        # Format row insurance for comparison
        formatted_row_insurance = format_insurance_company_name(row_insurance)
        if pd.isna(formatted_row_insurance) or not formatted_row_insurance:
            formatted_row_insurance = (
                str(row_insurance).strip() if row_insurance is not None else ""
            )
        else:
            formatted_row_insurance = str(formatted_row_insurance).strip()

        formatted_row_lower = (
            formatted_row_insurance.lower().strip() if formatted_row_insurance else ""
        )

        # Check against Afreen Ansari's additional insurance list
        for allowed_insurance in AFREEN_ANSARI_ADDITIONAL_INSURANCE:
            allowed_lower = str(allowed_insurance).strip().lower()
            # Exact match or substring match
            if (
                formatted_row_lower == allowed_lower
                or allowed_lower in formatted_row_lower
                or formatted_row_lower in allowed_lower
            ):
                return True

    if not agent_insurance_list:
        return True

    # Format the row insurance company name
    formatted_row_insurance = format_insurance_company_name(row_insurance)
    # Handle NaN/None/float values
    if pd.isna(formatted_row_insurance) or not formatted_row_insurance:
        if pd.isna(row_insurance):
            formatted_row_insurance = ""
        else:
            formatted_row_insurance = (
                str(row_insurance).strip() if row_insurance is not None else ""
            )
    else:
        formatted_row_insurance = str(formatted_row_insurance).strip()

    formatted_row_lower = (
        formatted_row_insurance.lower().strip() if formatted_row_insurance else ""
    )

    # Check against each agent capability
    for comp in agent_insurance_list:
        # Skip if comp is NaN/None
        if pd.isna(comp) or comp is None:
            continue

        # Format the agent's insurance capability
        formatted_comp = format_insurance_company_name(comp)
        # Handle NaN/None/float values
        if pd.isna(formatted_comp) or not formatted_comp:
            formatted_comp = str(comp).strip() if comp is not None else ""
        else:
            formatted_comp = str(formatted_comp).strip()

        comp_lower = formatted_comp.lower().strip() if formatted_comp else ""

        # Exact match
        if formatted_row_lower == comp_lower:
            return True

        # Substring match (handles partial matches)
        if formatted_row_lower in comp_lower or comp_lower in formatted_row_lower:
            return True

        # Handle DD Toolkit/Toolkits variations
        # Check if row insurance is in DD_TOOLKIT_GROUP and agent has DD Toolkit/Toolkits/DD capability
        if formatted_row_insurance in DD_TOOLKIT_GROUP:
            if (
                "dd toolkit" in comp_lower
                or "dd toolkits" in comp_lower
                or comp_lower == "dd"
            ):
                return True

        # Check if row insurance name contains "DD Toolkit" and agent has DD Toolkit/Toolkits/DD capability
        if "dd toolkit" in formatted_row_lower or "dd toolkits" in formatted_row_lower:
            if (
                "dd toolkit" in comp_lower
                or "dd toolkits" in comp_lower
                or comp_lower == "dd"
            ):
                return True

        # Handle DD INS variations
        # Check if row insurance is in DD_INS_GROUP and agent has DD INS/INS capability
        if formatted_row_insurance in DD_INS_GROUP:
            if "dd ins" in comp_lower or comp_lower == "ins":
                return True

        # Check if row insurance name contains "DD INS" and agent has DD INS/INS capability
        if "dd ins" in formatted_row_lower or formatted_row_lower == "ins":
            if "dd ins" in comp_lower or comp_lower == "ins":
                return True

        # Check if row insurance starts with "DD" and agent has "DD" capability (for standalone DD)
        if formatted_row_lower.startswith("dd ") and comp_lower == "dd":
            return True

        # Check if row insurance is a DD company (starts with "DD") and agent has DD Toolkit/Toolkits/DD INS/INS capability
        if formatted_row_lower.startswith("dd "):
            # Check if it's a DD Toolkit company
            if formatted_row_insurance in DD_TOOLKIT_GROUP:
                if (
                    "dd toolkit" in comp_lower
                    or "dd toolkits" in comp_lower
                    or comp_lower == "dd"
                ):
                    return True
            # Check if it's a DD INS company
            if formatted_row_insurance in DD_INS_GROUP:
                if "dd ins" in comp_lower or comp_lower == "ins":
                    return True

    return False


def format_insurance_column_in_dataframe(df, column_name):
    """Format insurance company names in a DataFrame column"""
    if column_name not in df.columns:
        return df

    original_count = len(df[column_name].dropna())

    # Apply formatting
    df[column_name] = df[column_name].apply(format_insurance_company_name)

    formatted_count = len(df[column_name].dropna())

    return df


def detect_and_assign_new_insurance_companies(
    data_df,
    agent_data,
    insurance_carrier_col,
    insurance_working_col,
    agent_name_col=None,
):
    """Detect new insurance companies in data file and automatically assign them to senior agents"""
    try:
        if not insurance_carrier_col or not insurance_working_col:
            return agent_data, []

        # Get all insurance companies from data file
        data_insurance_companies = set()
        for _, row in data_df.iterrows():
            if pd.notna(row[insurance_carrier_col]):
                company = str(row[insurance_carrier_col]).strip()
                if company and company.lower() != "unknown":
                    data_insurance_companies.add(company)

        # Get all insurance companies currently assigned to agents
        agent_insurance_companies = set()
        for _, row in agent_data.iterrows():
            if pd.notna(row[insurance_working_col]):
                companies_str = str(row[insurance_working_col])
                companies = [
                    comp.strip()
                    for comp in companies_str.replace(",", ";")
                    .replace("|", ";")
                    .split(";")
                    if comp.strip()
                ]
                for comp in companies:
                    if comp.lower() != "senior":
                        agent_insurance_companies.add(comp)

        # Find new insurance companies
        new_insurance_companies = data_insurance_companies - agent_insurance_companies

        if not new_insurance_companies:
            return agent_data, []

        # Find senior agents (those with 'senior' in their Insurance List column)
        senior_agents = []
        for idx, row in agent_data.iterrows():
            if pd.notna(row[insurance_working_col]):
                companies_str = str(row[insurance_working_col])
                if "senior" in companies_str.lower():
                    senior_agents.append(idx)

        # Console log senior agents found
        if senior_agents:
            for idx in senior_agents:
                if agent_name_col and agent_name_col in agent_data.columns:
                    agent_name = agent_data.iloc[idx][agent_name_col]
                else:
                    agent_name = f"Agent {idx}"
        else:
            pass

        # Assign new insurance companies to senior agents
        updated_agents = []
        for idx, row in agent_data.iterrows():
            if idx in senior_agents:
                # Add new insurance companies to senior agents
                current_companies = (
                    str(row[insurance_working_col])
                    if pd.notna(row[insurance_working_col])
                    else ""
                )
                new_companies_str = "; ".join(
                    [
                        str(comp).strip()
                        for comp in new_insurance_companies
                        if comp is not None and not pd.isna(comp)
                    ]
                )

                if current_companies:
                    updated_companies = f"{current_companies}; {new_companies_str}"
                else:
                    updated_companies = new_companies_str

                # Update the row
                row_copy = row.copy()
                row_copy[insurance_working_col] = updated_companies
                updated_agents.append(row_copy)
            else:
                updated_agents.append(row)

        # Convert back to DataFrame
        updated_agent_data = pd.DataFrame(updated_agents)

        return updated_agent_data, list(new_insurance_companies)

    except Exception as e:
        return agent_data, []


def parse_excel_date(value):
    """
    Robustly parse dates from Excel that works consistently across Windows and Mac.
    Handles:
    - Excel serial numbers (days since 1899-12-30)
    - String dates
    - Already parsed datetime objects
    - Filters out invalid dates (like Excel's 1900 leap year bug artifacts)
    """
    from datetime import datetime, timedelta, date
    import pandas as pd

    if pd.isna(value):
        return None

    # If it's already a date object (not datetime), return it
    if isinstance(value, date) and not isinstance(value, datetime):
        # Validate the date is in a reasonable range
        if 2000 <= value.year <= 2100:
            return value
        return None

    # If it's already a datetime object, extract date
    if isinstance(value, datetime):
        if 2000 <= value.year <= 2100:
            return value.date()
        return None

    # If it has a date() method (like pandas Timestamp), use it
    if hasattr(value, "date") and callable(value.date):
        try:
            date_obj = value.date()
            if isinstance(date_obj, date) and 2000 <= date_obj.year <= 2100:
                return date_obj
        except (AttributeError, ValueError):
            pass

    # Try to parse as Excel serial number (common on Windows)
    # Excel stores dates as days since 1899-12-30 (with a bug treating 1900 as leap year)
    try:
        if isinstance(value, (int, float)):
            # Excel serial number: days since 1899-12-30
            excel_epoch = datetime(1899, 12, 30)
            days = float(value)

            # Only process if it looks like a reasonable Excel date serial number
            # Excel dates typically range from ~1 (1900-01-01) to ~73050 (2099-12-31)
            # But we only want dates from 2000 onwards to avoid false positives
            # 2000-01-01 is approximately serial number 36526
            if 36526 <= days <= 73050:
                parsed_date = excel_epoch + timedelta(days=int(days))
                # Additional validation: check if year is reasonable (2000-2100)
                if 2000 <= parsed_date.year <= 2100:
                    return parsed_date.date()
    except (ValueError, TypeError, OverflowError, OSError):
        pass

    # Try to parse string dates with explicit MM/DD/YYYY format first
    # This handles formats like "11/17/2025 12:00:00 AM" or "11/17/2025"
    if isinstance(value, str):
        value_str = str(value).strip()

        # Try MM/DD/YYYY format (US format)
        # Handle formats like:
        # - "11/17/2025"
        # - "11/17/2025 12:00:00 AM"
        # - "11/17/2025 12:00:00"
        try:
            # Remove time component if present
            date_part = value_str.split()[0] if " " in value_str else value_str

            # Try MM/DD/YYYY format (US format)
            if "/" in date_part:
                parts = date_part.split("/")
                if len(parts) == 3:
                    first, second, year = parts
                    first_int = int(first)
                    second_int = int(second)
                    year_int = int(year)

                    # Determine if it's MM/DD/YYYY or DD/MM/YYYY
                    # If first > 12, it must be DD/MM/YYYY (month can't be > 12 in MM/DD)
                    # If second > 12, it must be MM/DD/YYYY (day can't be > 12 in DD/MM)
                    # Otherwise, prefer MM/DD/YYYY as user specified
                    if first_int > 12:
                        # Must be DD/MM/YYYY (first=day, second=month)
                        try:
                            parsed_date = datetime(year_int, second_int, first_int)
                            if 2000 <= parsed_date.year <= 2100:
                                return parsed_date.date()
                        except ValueError:
                            pass
                    elif second_int > 12:
                        # Must be MM/DD/YYYY (first=month, second=day)
                        try:
                            parsed_date = datetime(year_int, first_int, second_int)
                            if 2000 <= parsed_date.year <= 2100:
                                return parsed_date.date()
                        except ValueError:
                            pass
                    else:
                        # Ambiguous (both <= 12) - try MM/DD/YYYY first (as user specified)
                        # Try as MM/DD/YYYY: first=month, second=day
                        try:
                            parsed_date = datetime(year_int, first_int, second_int)
                            if 2000 <= parsed_date.year <= 2100:
                                return parsed_date.date()
                        except ValueError:
                            # If MM/DD fails, try DD/MM as fallback
                            try:
                                parsed_date = datetime(year_int, second_int, first_int)
                                if 2000 <= parsed_date.year <= 2100:
                                    return parsed_date.date()
                            except ValueError:
                                pass
        except (ValueError, TypeError, IndexError):
            pass

    # Try pandas to_datetime with dayfirst=False to prefer MM/DD/YYYY format
    try:
        # Use dayfirst=False (default) to prefer MM/DD/YYYY over DD/MM/YYYY
        parsed = pd.to_datetime(value, errors="coerce", dayfirst=False)
        if pd.notna(parsed):
            parsed_date = parsed.to_pydatetime()
            # Filter out invalid dates (before 2000 or after 2100)
            if 2000 <= parsed_date.year <= 2100:
                return parsed_date.date()
    except (ValueError, TypeError, OverflowError, OSError):
        pass

    # If all parsing fails, return None
    return None


def get_nth_business_day(start_date, n):
    """Get the nth business day from start_date"""
    from datetime import timedelta

    current_date = start_date
    business_days_count = 0

    while business_days_count < n:
        current_date += timedelta(days=1)
        # Check if it's a weekday (Monday=0, Sunday=6)
        if current_date.weekday() < 5:  # Monday to Friday
            business_days_count += 1

    return current_date


def process_allocation_files(allocation_df, data_df):
    """Process data file with priority assignment based on business days calendar"""
    try:
        from datetime import datetime, timedelta
        import pandas as pd

        # Use data_df as the main file to process (ignore allocation_df for now)
        processed_df = data_df.copy()

        # Find the appointment date column (case-insensitive search)
        appointment_date_col = None
        for col in processed_df.columns:
            if "appointment" in col.lower() and "date" in col.lower():
                appointment_date_col = col
                break

        if appointment_date_col is None:
            return (
                f"❌ Error: 'Appointment Date' column not found in data file.\nAvailable columns: {list(processed_df.columns)}",
                None,
            )

        # Convert appointment date column to datetime and remove time component
        # Use robust date parsing that works consistently across Windows and Mac
        try:
            processed_df[appointment_date_col] = processed_df[
                appointment_date_col
            ].apply(parse_excel_date)
        except Exception as e:
            return f"❌ Error converting appointment dates: {str(e)}", None

        # Get today's date
        today = datetime.now().date()

        # Check if Priority Status column exists, if not create it
        if "Priority Status" not in processed_df.columns:
            processed_df["Priority Status"] = ""

        # Convert Priority Status column to object type to avoid dtype warnings
        processed_df["Priority Status"] = processed_df["Priority Status"].astype(
            "object"
        )

        # Calculate business day targets
        first_business_day = get_nth_business_day(today, 1)
        second_business_day = get_nth_business_day(today, 2)
        seventh_business_day = get_nth_business_day(today, 7)

        # Count statistics
        total_rows = len(processed_df)
        first_priority_count = 0
        invalid_dates = 0

        # Process each row
        for idx, row in processed_df.iterrows():
            appointment_date = row[appointment_date_col]

            # Skip rows with invalid dates
            if pd.isna(appointment_date):
                processed_df.at[idx, "Priority Status"] = "Invalid Date"
                invalid_dates += 1
                continue

            # Convert to date if it's datetime
            if hasattr(appointment_date, "date"):
                appointment_date = appointment_date.date()

            # Check if appointment date matches First Priority criteria
            if (
                appointment_date == today
                or appointment_date == first_business_day
                or appointment_date == second_business_day
                or appointment_date == seventh_business_day
            ):
                processed_df.at[idx, "Priority Status"] = "First Priority"
                first_priority_count += 1
            else:
                # Keep blank for now as requested
                processed_df.at[idx, "Priority Status"] = ""

        # Generate result message
        result_message = f"""✅ Priority processing completed successfully!

📊 Processing Statistics:
- Total rows processed: {total_rows}
- First Priority: {first_priority_count} rows
- Other rows: {total_rows - first_priority_count - invalid_dates} rows (kept blank for now)
- Invalid dates: {invalid_dates} rows

📅 Business Day Calendar Logic Applied:
1. First Priority: Same day, 1st business day, 2nd business day, and 7th business day from today
2. Second Priority: (To be implemented later)
3. Third Priority: (To be implemented later)

📅 Business Day Targets:
- Today: {today.strftime('%Y-%m-%d (%A)')}
- 1st Business Day: {first_business_day.strftime('%Y-%m-%d (%A)')}
- 2nd Business Day: {second_business_day.strftime('%Y-%m-%d (%A)')}
- 7th Business Day: {seventh_business_day.strftime('%Y-%m-%d (%A)')}

📋 Updated column: 'Priority Status'
📅 Based on column: '{appointment_date_col}'

🔍 Sample of processed data:
{processed_df[['Priority Status', appointment_date_col]].head(10).to_string()}

💾 Ready to download the processed result file!"""

        return result_message, processed_df

    except Exception as e:
        return f"❌ Error during processing: {str(e)}", None


def process_allocation_files_with_dates(
    allocation_df,
    data_df,
    selected_dates,
    custom_dates,
    appointment_dates,
    appointment_dates_second=None,
    receive_dates=None,
):
    """Process data file with priority assignment and generate agent allocation summary"""
    global agent_allocations_data
    try:
        from datetime import datetime, timedelta
        import pandas as pd

        # Use data_df as the main file to process
        processed_df = data_df.copy()

        # Find the appointment date column, receive date column, insurance carrier column, remark column, and secondary insurance column
        appointment_date_col = None
        receive_date_col = None
        insurance_carrier_col = None
        remark_col = None
        secondary_insurance_col = None
        for col in processed_df.columns:
            if "appointment" in col.lower() and "date" in col.lower():
                appointment_date_col = col
            elif "receive" in col.lower() and "date" in col.lower():
                receive_date_col = col
            elif (
                "dental" in col.lower()
                and "primary" in col.lower()
                and "ins" in col.lower()
                and "carr" in col.lower()
            ):
                insurance_carrier_col = col
            elif col.lower() in ["remark", "remarks"]:
                remark_col = col
            elif (
                "dental" in col.lower()
                and "secondary" in col.lower()
                and "ins" in col.lower()
                and "carr" in col.lower()
            ):
                secondary_insurance_col = col

        if appointment_date_col is None:
            return (
                f"❌ Error: 'Appointment Date' column not found in data file.\nAvailable columns: {list(processed_df.columns)}",
                None,
            )

        # Convert appointment date column to datetime and remove time component
        # Use robust date parsing that works consistently across Windows and Mac
        try:
            processed_df[appointment_date_col] = processed_df[
                appointment_date_col
            ].apply(parse_excel_date)
        except Exception as e:
            return f"❌ Error converting appointment dates: {str(e)}", None

        # Parse receive date column if it exists (using robust date parsing)
        if receive_date_col and receive_date_col in processed_df.columns:
            try:
                processed_df[receive_date_col] = processed_df[receive_date_col].apply(
                    parse_excel_date
                )
            except Exception as e:
                # If receive date parsing fails, log but don't fail the whole process
                print(f"Warning: Error parsing receive dates: {str(e)}")

        # Check if Priority Status column exists, if not create it
        if "Priority Status" not in processed_df.columns:
            processed_df["Priority Status"] = ""

        # Convert Priority Status column to object type
        processed_df["Priority Status"] = processed_df["Priority Status"].astype(
            "object"
        )

        # Build list of priority dates from selection (as strings)
        first_priority_dates = set(appointment_dates) if appointment_dates else set()
        second_priority_dates = (
            set(appointment_dates_second) if appointment_dates_second else set()
        )

        # Count statistics
        total_rows = len(processed_df)
        first_priority_count = 0
        second_priority_count = 0
        third_priority_count = 0
        invalid_dates = 0

        # Collect Third Priority dates
        third_priority_dates_set = set()

        # Process each row
        for idx, row in processed_df.iterrows():
            appointment_date = row[appointment_date_col]

            # Skip rows with invalid dates
            if pd.isna(appointment_date):
                processed_df.at[idx, "Priority Status"] = "Invalid Date"
                invalid_dates += 1
                continue

            # Convert appointment date to string and handle different formats
            appointment_date_str = str(appointment_date)

            # If it's a datetime string like '2025-11-03 00:00:00', extract just the date part
            if " " in appointment_date_str:
                appointment_date_str = appointment_date_str.split(" ")[0]

            # Convert calendar dates (YYYY-MM-DD) to YYYY-MM-DD format for comparison
            def convert_calendar_to_original_format(calendar_date):
                try:
                    from datetime import datetime

                    # Parse YYYY-MM-DD format
                    dt = datetime.strptime(calendar_date, "%Y-%m-%d")
                    # Return in YYYY-MM-DD format for comparison
                    return dt.strftime("%Y-%m-%d")
                except:
                    return calendar_date

            # Convert priority dates to YYYY-MM-DD format for comparison
            first_priority_dates_yyyy_mm_dd = set()
            for calendar_date in first_priority_dates:
                converted_date = convert_calendar_to_original_format(calendar_date)
                first_priority_dates_yyyy_mm_dd.add(converted_date)

            second_priority_dates_yyyy_mm_dd = set()
            for calendar_date in second_priority_dates:
                converted_date = convert_calendar_to_original_format(calendar_date)
                second_priority_dates_yyyy_mm_dd.add(converted_date)

            # Check if appointment date is in First Priority dates
            if appointment_date_str in first_priority_dates_yyyy_mm_dd:
                # Additional filtering: check receive dates if provided
                should_include = True
                if (
                    receive_dates
                    and receive_date_col
                    and receive_date_col in processed_df.columns
                ):
                    receive_date = row[receive_date_col]
                    if not pd.isna(receive_date) and receive_date is not None:
                        # Convert receive date to string format for comparison
                        receive_date_str = (
                            receive_date.strftime("%Y-%m-%d")
                            if hasattr(receive_date, "strftime")
                            else str(receive_date)
                        )

                        # Convert receive dates to YYYY-MM-DD format for comparison
                        receive_dates_yyyy_mm_dd = set()
                        for calendar_date in receive_dates:
                            converted_date = convert_calendar_to_original_format(
                                calendar_date
                            )
                            receive_dates_yyyy_mm_dd.add(converted_date)

                        # Only include if receive date is in selected receive dates
                        if receive_date_str not in receive_dates_yyyy_mm_dd:
                            should_include = False

                if should_include:
                    processed_df.at[idx, "Priority Status"] = "First Priority"
                    first_priority_count += 1
                else:
                    # If receive date is not selected, assign to Second Priority
                    processed_df.at[idx, "Priority Status"] = "Second Priority"
                    second_priority_count += 1
            # Check if appointment date is in Second Priority dates
            elif appointment_date_str in second_priority_dates_yyyy_mm_dd:
                processed_df.at[idx, "Priority Status"] = "Second Priority"
                second_priority_count += 1
            else:
                # All remaining dates get Third Priority
                processed_df.at[idx, "Priority Status"] = "Third Priority"
                third_priority_count += 1
                # Add to Third Priority dates set (convert back to calendar format for display)
                try:
                    from datetime import datetime

                    dt = datetime.strptime(appointment_date_str, "%Y-%m-%d")
                    calendar_date = dt.strftime("%Y-%m-%d")
                    third_priority_dates_set.add(calendar_date)
                except:
                    # If conversion fails, use the original string
                    third_priority_dates_set.add(appointment_date_str)

        # Generate agent allocation summary if allocation_df is provided
        agent_summary = ""
        if allocation_df is not None:
            try:
                # Get the "main" sheet from allocation data, fallback to first sheet if "main" doesn't exist
                agent_df = None
                if "main" in allocation_df:
                    agent_df = allocation_df["main"]
                elif len(allocation_df) > 0:
                    agent_df = list(allocation_df.values())[0]

                if agent_df is None:
                    agent_summary = "\n⚠️ No sheets found in allocation file."
                    return processed_df, agent_summary

                # Find agent name, ID, counts, insurance list, exceptions, email, role, shift time, shift group, domain, allocation preference, and ins do not allocate columns
                agent_name_col = None
                agent_id_col = None
                cc_col = None  # Current Capacity column (prioritized)
                counts_col = None  # Fallback to TFD/Capacity/Count
                insurance_working_col = None
                insurance_needs_training_col = None
                insurance_do_not_allocate_col = None
                email_col = None
                role_col = None
                category_col = (
                    None  # Category column for roles (Senior, Auditor, Junior, Trainee)
                )
                shift_time_col = None
                shift_group_col = None
                domain_col = None
                allocation_preference_col = None
                status_col = None
                for col in agent_df.columns:
                    col_lower = col.lower()
                    if "agent" in col_lower and "name" in col_lower:
                        agent_name_col = col
                    elif col_lower == "id":
                        agent_id_col = col
                    elif (
                        col_lower == "cc"
                        or col_lower == "current capacity"
                        or (col_lower.startswith("cc") and "capacity" in col_lower)
                    ):
                        cc_col = col  # Current Capacity column (prioritized)
                    elif (
                        col_lower == "tfd"
                        or col_lower == "capacity"
                        or col_lower == "count"
                    ):
                        if not cc_col:  # Only use TFD/capacity/count if CC not found
                            counts_col = (
                                col  # Fallback to TFD/capacity/count if CC not found
                            )
                    elif "insurance" in col_lower and "list" in col_lower:
                        insurance_working_col = col
                    elif "exception" in col_lower:
                        insurance_needs_training_col = col
                    elif ("ins" in col_lower or "insurance" in col_lower) and (
                        "do" in col_lower
                        and "not" in col_lower
                        and "allocate" in col_lower
                    ):
                        insurance_do_not_allocate_col = col
                    elif "email" in col_lower and "id" in col_lower:
                        email_col = col
                    elif col_lower == "category":
                        category_col = (
                            col  # Category column (Senior, Auditor, Junior, Trainee)
                        )
                    elif (
                        col_lower == "role"
                        or col_lower == "job role"
                        or col_lower == "position"
                        or ("role" in col_lower and "type" in col_lower)
                    ):
                        role_col = col
                    elif "shift" in col_lower and "time" in col_lower:
                        shift_time_col = col
                    elif "shift" in col_lower and "group" in col_lower:
                        shift_group_col = col
                    elif col_lower == "domain":
                        domain_col = col
                    elif "allocation" in col_lower and "preference" in col_lower:
                        allocation_preference_col = col
                    elif col_lower == "status":
                        status_col = col

                # Use CC column if available, otherwise fallback to counts_col
                capacity_col = cc_col if cc_col else counts_col

                if agent_name_col and capacity_col:
                    # Get agent data with their capacities and insurance capabilities
                    columns_to_select = [agent_name_col, capacity_col]
                    if agent_id_col:
                        columns_to_select.append(agent_id_col)
                    if insurance_working_col:
                        columns_to_select.append(insurance_working_col)
                    if insurance_needs_training_col:
                        columns_to_select.append(insurance_needs_training_col)
                    if insurance_do_not_allocate_col:
                        columns_to_select.append(insurance_do_not_allocate_col)
                    if email_col:
                        columns_to_select.append(email_col)
                    if role_col:
                        columns_to_select.append(role_col)
                    if category_col:
                        columns_to_select.append(category_col)
                    if shift_time_col:
                        columns_to_select.append(shift_time_col)
                    if shift_group_col:
                        columns_to_select.append(shift_group_col)
                    if domain_col:
                        columns_to_select.append(domain_col)
                    if allocation_preference_col:
                        columns_to_select.append(allocation_preference_col)
                    if status_col:
                        columns_to_select.append(status_col)

                    agent_data = agent_df[columns_to_select].dropna(
                        subset=[agent_name_col, capacity_col]
                    )

                    # Filter out agents with "Status" = "No Allocation"
                    if status_col:
                        agent_data = agent_data[
                            ~(
                                agent_data[status_col]
                                .astype(str)
                                .str.strip()
                                .str.upper()
                                .isin(["NO ALLOCATION", "NOALLOCATION"])
                            )
                        ]

                    # Filter out "Auditor" role based on Category column (priority) or role column (fallback)
                    if category_col:
                        # Filter based on Category column (case-insensitive)
                        agent_data = agent_data[
                            ~agent_data[category_col]
                            .astype(str)
                            .str.lower()
                            .str.strip()
                            .isin(["auditor"])
                        ]
                    elif role_col:
                        # Fallback: Filter based on role column (case-insensitive)
                        agent_data = agent_data[
                            ~agent_data[role_col]
                            .astype(str)
                            .str.lower()
                            .str.strip()
                            .isin(["auditor", "caller"])
                        ]
                    else:
                        # If no category or role column found, check if agent name column contains these roles (case-insensitive)
                        agent_data = agent_data[
                            ~agent_data[agent_name_col]
                            .astype(str)
                            .str.lower()
                            .str.strip()
                            .isin(["auditor", "caller"])
                        ]

                    # Add empty columns if not found
                    if not insurance_working_col:
                        agent_data["Insurance List"] = ""
                        insurance_working_col = "Insurance List"

                    # Detect and assign new insurance companies to senior agents
                    if insurance_carrier_col and insurance_working_col:
                        agent_data, new_insurance_companies = (
                            detect_and_assign_new_insurance_companies(
                                processed_df,
                                agent_data,
                                insurance_carrier_col,
                                insurance_working_col,
                                agent_name_col,
                            )
                        )
                        if new_insurance_companies:
                            agent_summary += f"\n🆕 New insurance companies detected and assigned to senior agents: {', '.join([str(comp).strip() for comp in new_insurance_companies if comp is not None and not pd.isna(comp)])}"
                    if not insurance_needs_training_col:
                        agent_data["Exceptions"] = ""
                        insurance_needs_training_col = "Exceptions"

                    total_agents = len(agent_data)

                    # Calculate total capacity with proper type conversion
                    total_capacity = 0
                    for _, row in agent_data.iterrows():
                        try:
                            if pd.notna(row[capacity_col]):
                                capacity = int(
                                    float(str(row[capacity_col]).replace(",", ""))
                                )
                                total_capacity += capacity
                        except (ValueError, TypeError):
                            continue

                    # Create capability-based allocation
                    agent_allocations = []

                    # First, prepare agent data with their capabilities
                    for _, row in agent_data.iterrows():
                        agent_name = (
                            str(row[agent_name_col]).strip()
                            if pd.notna(row[agent_name_col])
                            else "Unknown"
                        )

                        # Create unique agent_id: Use ID if available, otherwise use name + index as fallback
                        if agent_id_col and pd.notna(row[agent_id_col]):
                            agent_id = str(row[agent_id_col]).strip()
                        else:
                            # Fallback: Use name + row index to ensure uniqueness
                            agent_id = f"{agent_name}_{row.name}"

                        # Handle different data types in capacity column (CC or counts_col)
                        try:
                            if pd.notna(row[capacity_col]):
                                capacity = int(
                                    float(str(row[capacity_col]).replace(",", ""))
                                )
                            else:
                                capacity = 0
                        except (ValueError, TypeError):
                            capacity = 0

                        # Get insurance companies this agent can work with and check if senior
                        insurance_companies = []
                        is_senior = False

                        # Check if agent is senior based on Category column (priority) or Insurance List
                        if category_col and pd.notna(row[category_col]):
                            agent_category = str(row[category_col]).strip().lower()
                            if "senior" in agent_category:
                                is_senior = True

                        if insurance_working_col and pd.notna(
                            row[insurance_working_col]
                        ):
                            # Split by common delimiters and clean up
                            companies_str = str(row[insurance_working_col])
                            companies = [
                                comp.strip()
                                for comp in companies_str.replace(",", ";")
                                .replace("|", ";")
                                .split(";")
                                if comp.strip()
                            ]

                            # Check if agent is senior (if not already determined by Category)
                            if not is_senior and any(
                                "senior" in comp.lower() for comp in companies
                            ):
                                is_senior = True

                            if is_senior:
                                # For senior agents, they can work with any insurance company
                                insurance_companies = ["ALL_COMPANIES"]
                            else:
                                insurance_companies = companies

                        # Get insurance companies this agent needs training for
                        insurance_needs_training = []
                        if insurance_needs_training_col and pd.notna(
                            row[insurance_needs_training_col]
                        ):
                            # Split by common delimiters and clean up
                            training_str = str(row[insurance_needs_training_col])
                            training_companies = [
                                comp.strip()
                                for comp in training_str.replace(",", ";")
                                .replace("|", ";")
                                .split(";")
                                if comp.strip()
                            ]
                            insurance_needs_training = training_companies

                        # Get insurance companies this agent should NOT be allocated
                        insurance_do_not_allocate = []
                        if insurance_do_not_allocate_col and pd.notna(
                            row[insurance_do_not_allocate_col]
                        ):
                            # Split by common delimiters and clean up
                            do_not_allocate_str = str(
                                row[insurance_do_not_allocate_col]
                            )
                            do_not_allocate_companies = [
                                comp.strip()
                                for comp in do_not_allocate_str.replace(",", ";")
                                .replace("|", ";")
                                .split(";")
                                if comp.strip()
                            ]
                            # Format insurance company names for matching
                            for comp in do_not_allocate_companies:
                                formatted_comp = format_insurance_company_name(comp)
                                insurance_do_not_allocate.append(formatted_comp)

                        # Get agent email
                        agent_email = ""
                        if email_col and pd.notna(row[email_col]):
                            agent_email = str(row[email_col]).strip()

                        # Get domain value
                        agent_domain = None
                        if domain_col and pd.notna(row[domain_col]):
                            agent_domain = str(row[domain_col]).strip().upper()

                        # Get allocation preference value and check if it contains PB, NTC, or Single
                        allocation_preference = None
                        has_pb_preference = False
                        has_ntc_preference = False
                        has_single_preference = False
                        if allocation_preference_col and pd.notna(
                            row[allocation_preference_col]
                        ):
                            allocation_preference_raw = row[allocation_preference_col]
                            allocation_preference = (
                                str(allocation_preference_raw).strip().upper()
                            )
                            # Check if allocation preference is exactly "PB" or contains "PB"
                            # This handles values like "PB", "PB+NTC", "MIX+PB", etc.
                            has_pb_preference = (
                                allocation_preference == "PB"
                                or "PB" in allocation_preference
                            )
                            # Check if allocation preference contains "NTC"
                            # Valid values: "Sec+NTC", "Sec+Mix+NTC", "Mix+NTC", "NTC"
                            has_ntc_preference = "NTC" in allocation_preference
                            # Check if allocation preference is exactly "Single"
                            # Agents with "Single" should only get rows from one insurance company
                            has_single_preference = allocation_preference == "SINGLE"
                            # Check if allocation preference contains "Mix"
                            # Agents with "Mix" should get multiple insurance company rows
                            has_mix_preference = "MIX" in allocation_preference
                            # Check if allocation preference contains "Sec" (e.g., "Sec + Single", "Sec + NTC", "Sec + Mix")
                            # Agents with "Sec + X" should first get rows with secondary insurance, then apply X logic
                            has_sec_preference = "SEC" in allocation_preference
                            # Check if allocation preference is "Sec + Single" or "SEC+SINGLE"
                            # This is a specific case of Sec preference
                            has_sec_single_preference = (
                                "SEC" in allocation_preference
                                and "SINGLE" in allocation_preference
                            )
                        # Debug: Store raw allocation preference for troubleshooting
                        allocation_preference_raw_value = (
                            str(row[allocation_preference_col]).strip()
                            if allocation_preference_col
                            and pd.notna(row[allocation_preference_col])
                            else None
                        )

                        # Get shift group (1=day, 2=afternoon, 3=night) to help parse ambiguous times
                        shift_group = None
                        if shift_group_col and pd.notna(row[shift_group_col]):
                            try:
                                shift_group = int(
                                    float(str(row[shift_group_col]).strip())
                                )
                            except (ValueError, TypeError):
                                shift_group = None

                        # Parse shift time (could be a range like "10-7pm", "1-10pm", "7-5 am")
                        shift_start_time = None
                        if shift_time_col and pd.notna(row[shift_time_col]):
                            shift_time_str = str(row[shift_time_col]).strip()
                            try:
                                from datetime import time as dt_time
                                import re

                                # Try parsing various time formats
                                if isinstance(row[shift_time_col], pd.Timestamp):
                                    shift_start_time = row[shift_time_col].time()
                                elif "-" in shift_time_str:
                                    # Parse time range (e.g., "10-7pm", "1-10pm", "7-5 am", "10am-7pm")
                                    # Extract start time (first part before the dash)
                                    parts = shift_time_str.split("-")
                                    if len(parts) >= 2:
                                        start_time_str = parts[0].strip()
                                        end_time_str = parts[1].strip()

                                        # Check if end time has AM/PM indicator
                                        has_end_am = "am" in end_time_str.lower()
                                        has_end_pm = "pm" in end_time_str.lower()
                                        has_start_am = "am" in start_time_str.lower()
                                        has_start_pm = "pm" in start_time_str.lower()

                                        # Extract start hour (could be just a number like "10" or "7")
                                        start_match = re.search(
                                            r"(\d{1,2})", start_time_str
                                        )
                                        if start_match:
                                            hour = int(start_match.group(1))
                                            minute = 0  # Default to 0 minutes if not specified

                                            # Check for explicit AM/PM in start time
                                            if has_start_am:
                                                if hour == 12:
                                                    hour = 0  # 12 AM = 0
                                            elif has_start_pm:
                                                if hour != 12:
                                                    hour += (
                                                        12  # Convert to 24-hour format
                                                    )
                                            else:
                                                # No AM/PM in start time - infer from context using Shift Group if available
                                                # Extract end hour for comparison
                                                end_match = re.search(
                                                    r"(\d{1,2})", end_time_str
                                                )
                                                if end_match:
                                                    end_hour_12 = int(
                                                        end_match.group(1)
                                                    )  # 12-hour format

                                                    # Use Shift Group to help determine AM/PM (1=day, 2=afternoon, 3=night)
                                                    if shift_group == 1:
                                                        # Day shift: typically starts in AM (morning, e.g., 8-5pm, 10-7pm)
                                                        # If hour >= end_hour_12, it's likely AM (day shift starts morning)
                                                        if hour >= end_hour_12:
                                                            pass  # Keep as AM
                                                        else:
                                                            # If start < end, could still be AM for day shift
                                                            pass  # Keep as AM
                                                    elif shift_group == 2:
                                                        # Afternoon shift: typically starts in PM (afternoon, e.g., 1-10pm, 3-6pm)
                                                        # But can also start in late AM and extend into evening (e.g., 11-8pm = 11 AM to 8 PM)
                                                        if hour >= end_hour_12:
                                                            # Start >= end, likely AM (e.g., "11-8pm" = 11 AM to 8 PM)
                                                            pass  # Keep as AM
                                                        else:
                                                            # Start < end, likely PM (e.g., "1-10pm" = 1 PM to 10 PM, "6-10pm" = 6 PM to 10 PM)
                                                            if hour != 12:
                                                                hour += (
                                                                    12  # Convert to PM
                                                                )
                                                    elif shift_group == 3:
                                                        # Night shift: can start in PM and go into AM (e.g., 7-5 am, 10-7am)
                                                        if has_end_am:
                                                            # End is AM - if start >= end, it's overnight (start is PM)
                                                            if hour >= end_hour_12:
                                                                if hour != 12:
                                                                    hour += 12  # Overnight shift - start is PM
                                                            else:
                                                                # Start < end AM, likely same day AM
                                                                if hour == 12:
                                                                    hour = 0
                                                        elif has_end_pm:
                                                            # End is PM - night shift might start late PM
                                                            if hour < end_hour_12:
                                                                if hour != 12:
                                                                    hour += 12  # Late PM start
                                                            else:
                                                                # Hour >= end, might be early AM (unusual but possible)
                                                                pass  # Keep as AM

                                                    # If no shift group, use original logic
                                                    if shift_group is None:
                                                        # Infer start time AM/PM using original heuristics
                                                        if has_end_am:
                                                            # End is AM (e.g., "7-5 am", "10-6 am")
                                                            # If start hour >= end hour, it's likely an overnight shift (start is PM)
                                                            # If start hour < end hour, it's likely same day (start is AM)
                                                            if hour >= end_hour_12:
                                                                # Overnight shift - start is PM (e.g., "7-5 am" = 7 PM to 5 AM)
                                                                if hour != 12:
                                                                    hour += 12  # Convert to PM (24-hour format)
                                                                # If hour is 12, it's 12 PM (noon)
                                                            else:
                                                                # Same day shift - start is AM (e.g., "2-5 am" = 2 AM to 5 AM)
                                                                if hour == 12:
                                                                    hour = 0  # 12 AM
                                                                # Otherwise keep hour as is (AM)
                                                        elif has_end_pm:
                                                            # End is PM
                                                            # If start hour < end hour (both in 12-hour format), likely both PM (e.g., "1-10pm")
                                                            # If start hour >= end hour, likely start AM and end PM (e.g., "10-7pm", "9-5pm")
                                                            if (
                                                                hour >= end_hour_12
                                                                and hour < 12
                                                            ):
                                                                # Start hour is >= end hour, likely AM (day shift)
                                                                pass  # Keep as is (AM)
                                                            elif hour < end_hour_12:
                                                                # Start hour < end hour, likely PM (afternoon/evening shift)
                                                                if hour != 12:
                                                                    hour += 12
                                                            elif hour == 12:
                                                                # Start is 12, check if end is also 12 or less
                                                                if end_hour_12 == 12:
                                                                    hour = 12  # Noon to noon (unlikely but handle it)
                                                                else:
                                                                    hour = 12  # 12 PM (noon)

                                            # Check for minutes in start time (e.g., "10:30-7pm")
                                            minute_match = re.search(
                                                r":(\d{2})", start_time_str
                                            )
                                            if minute_match:
                                                minute = int(minute_match.group(1))

                                            shift_start_time = dt_time(
                                                hour % 24, minute
                                            )

                                elif ":" in shift_time_str:
                                    # Parse single time string (e.g., "09:00", "9:00 AM", "09:00:00")
                                    time_match = re.search(
                                        r"(\d{1,2}):(\d{2})", shift_time_str
                                    )
                                    if time_match:
                                        hour = int(time_match.group(1))
                                        minute = int(time_match.group(2))
                                        # Check for AM/PM
                                        if (
                                            "pm" in shift_time_str.lower()
                                            and hour != 12
                                        ):
                                            hour += 12
                                        elif (
                                            "am" in shift_time_str.lower()
                                            and hour == 12
                                        ):
                                            hour = 0
                                        shift_start_time = dt_time(hour, minute)
                            except Exception as e:
                                shift_start_time = None

                        # Store original shift time for admin review
                        original_shift_time = None
                        if shift_time_col and pd.notna(row[shift_time_col]):
                            original_shift_time = str(row[shift_time_col]).strip()

                        agent_allocations.append(
                            {
                                "id": agent_id,  # Unique identifier (ID column or name + index)
                                "name": agent_name,  # Display name
                                "capacity": capacity,
                                "allocated": 0,
                                "ntc_allocated": 0,  # Track number of NTC rows allocated to this agent (max 15)
                                "email": agent_email,
                                "insurance_companies": insurance_companies,
                                "insurance_needs_training": insurance_needs_training,
                                "insurance_do_not_allocate": insurance_do_not_allocate,  # Insurance companies this agent should NOT be allocated
                                "is_senior": is_senior,
                                "shift_start_time": (
                                    shift_start_time.strftime("%H:%M")
                                    if shift_start_time
                                    else None
                                ),  # Store as HH:MM string
                                "shift_time_original": original_shift_time,  # Original shift time value from Excel
                                "shift_group": shift_group,  # Shift group (1=day, 2=afternoon, 3=night)
                                "domain": agent_domain,  # Domain value (e.g., 'PB')
                                "has_pb_preference": has_pb_preference,  # Whether allocation preference contains "PB"
                                "has_ntc_preference": has_ntc_preference,  # Whether allocation preference contains "NTC"
                                "has_single_preference": has_single_preference,  # Whether allocation preference is "Single"
                                "has_mix_preference": has_mix_preference,  # Whether allocation preference contains "Mix"
                                "has_sec_preference": has_sec_preference,  # Whether allocation preference contains "Sec" (e.g., "Sec + Single", "Sec + NTC", "Sec + Mix")
                                "has_sec_single_preference": has_sec_single_preference,  # Whether allocation preference is "Sec + Single"
                                "allocation_preference_raw": (
                                    allocation_preference_raw_value
                                    if "allocation_preference_raw_value" in locals()
                                    else None
                                ),  # Raw allocation preference value for debugging
                                "row_indices": [],
                                # New field to enforce single-insurance allocation rule
                                "assigned_insurance": None,
                            }
                        )

                    # Now allocate rows based on insurance company matching and priority
                    unmatched_insurance_companies = (
                        set()
                    )  # Initialize for use in summary
                    if insurance_carrier_col:
                        # Step 1: Identify all insurance companies in the data and all agent insurance companies
                        all_data_insurance_companies = set()
                        all_agent_insurance_companies = set()

                        for idx, row in processed_df.iterrows():
                            insurance_carrier = (
                                str(row[insurance_carrier_col]).strip()
                                if pd.notna(row[insurance_carrier_col])
                                else "Unknown"
                            )
                            if (
                                insurance_carrier
                                and insurance_carrier.lower() != "unknown"
                            ):
                                all_data_insurance_companies.add(insurance_carrier)

                        # Collect all insurance companies from non-senior agents (normalize to lowercase for comparison)
                        # Exclude "Afreen Ansari" from this check so unmatched insurance companies can be allocated to her
                        agent_insurance_lower = set()
                        for agent in agent_allocations:
                            if (
                                not agent["is_senior"]
                                and agent["name"] != "Afreen Ansari"
                                and agent["insurance_companies"]
                            ):
                                for comp in agent["insurance_companies"]:
                                    if comp != "ALL_COMPANIES":
                                        agent_insurance_lower.add(comp.strip().lower())

                        # Identify unmatched insurance companies (not in any non-senior agent's list)
                        # Compare case-insensitively
                        unmatched_insurance_companies = set()
                        for data_comp in all_data_insurance_companies:
                            data_comp_lower = data_comp.lower()
                            # Check if this insurance company matches any agent's insurance companies
                            is_matched = False
                            for agent_comp_lower in agent_insurance_lower:
                                if (
                                    data_comp_lower in agent_comp_lower
                                    or agent_comp_lower in data_comp_lower
                                ):
                                    is_matched = True
                                    break
                            if not is_matched:
                                unmatched_insurance_companies.add(data_comp)

                        # Get all senior agents
                        senior_agents = [a for a in agent_allocations if a["is_senior"]]

                        # Step 2: Group data by insurance carrier and priority
                        data_by_insurance_priority = {}
                        unmatched_data_by_priority = {}
                        matched_data_by_insurance_priority = {}

                        for idx, row in processed_df.iterrows():
                            insurance_carrier = (
                                str(row[insurance_carrier_col]).strip()
                                if pd.notna(row[insurance_carrier_col])
                                else "Unknown"
                            )
                            priority = row.get("Priority Status", "Unknown")

                            if (
                                insurance_carrier.lower() == "unknown"
                                or not insurance_carrier
                            ):
                                insurance_carrier = "Unknown"

                            # Separate unmatched and matched insurance companies
                            # Unknown insurance is always unmatched (senior only)
                            # First Priority is always senior only (for both matched and unmatched)
                            is_unmatched = (
                                insurance_carrier in unmatched_insurance_companies
                                or insurance_carrier == "Unknown"
                            )

                            if is_unmatched:
                                # Store unmatched insurance companies separately (highest priority)
                                if insurance_carrier not in unmatched_data_by_priority:
                                    unmatched_data_by_priority[insurance_carrier] = {}
                                if (
                                    priority
                                    not in unmatched_data_by_priority[insurance_carrier]
                                ):
                                    unmatched_data_by_priority[insurance_carrier][
                                        priority
                                    ] = []
                                unmatched_data_by_priority[insurance_carrier][
                                    priority
                                ].append(idx)
                            else:
                                # Store matched insurance companies normally
                                if (
                                    insurance_carrier
                                    not in matched_data_by_insurance_priority
                                ):
                                    matched_data_by_insurance_priority[
                                        insurance_carrier
                                    ] = {}
                                if (
                                    priority
                                    not in matched_data_by_insurance_priority[
                                        insurance_carrier
                                    ]
                                ):
                                    matched_data_by_insurance_priority[
                                        insurance_carrier
                                    ][priority] = []
                                matched_data_by_insurance_priority[insurance_carrier][
                                    priority
                                ].append(idx)

                            # Also keep full data structure for reference
                            if insurance_carrier not in data_by_insurance_priority:
                                data_by_insurance_priority[insurance_carrier] = {}
                            if (
                                priority
                                not in data_by_insurance_priority[insurance_carrier]
                            ):
                                data_by_insurance_priority[insurance_carrier][
                                    priority
                                ] = []
                            data_by_insurance_priority[insurance_carrier][
                                priority
                            ].append(idx)

                        # Initialize INS and Toolkit group tracking (used across all allocation steps)
                        ins_group_allocations = {}  # {agent_name: count}
                        toolkit_group_allocations = {}  # {agent_name: count}
                        ins_group_companies = set(DD_INS_GROUP)
                        toolkit_group_companies = set(DD_TOOLKIT_GROUP)

                        # Identify agents with INS or Toolkit groups
                        # After expansion, DD INS/INS becomes the actual companies, so we check if agent has any DD_INS_GROUP companies
                        agents_with_ins = []
                        agents_with_toolkit = []

                        # Convert to sets for faster lookup
                        ins_group_set = set([c.upper() for c in DD_INS_GROUP])
                        toolkit_group_set = set([c.upper() for c in DD_TOOLKIT_GROUP])

                        for agent in agent_allocations:
                            insurance_list = agent.get("insurance_companies", [])
                            agent_id = agent.get("id", agent.get("name", "Unknown"))
                            agent_name = agent.get("name", "Unknown")

                            if insurance_list:
                                # Convert to uppercase set for comparison
                                agent_insurance_set = set(
                                    [
                                        c.upper().strip()
                                        for c in insurance_list
                                        if c and c != "ALL_COMPANIES"
                                    ]
                                )

                                # Debug: Show first few agents' insurance companies
                                if len(agents_with_ins) + len(agents_with_toolkit) < 5:
                                    pass

                                # Check if agent has any DD_INS_GROUP companies
                                has_ins_group = bool(
                                    agent_insurance_set.intersection(ins_group_set)
                                )
                                if has_ins_group:
                                    agents_with_ins.append(agent_name)
                                    ins_group_allocations[agent_id] = 0

                                # Check if agent has any DD_TOOLKIT_GROUP companies
                                has_toolkit_group = bool(
                                    agent_insurance_set.intersection(toolkit_group_set)
                                )
                                if has_toolkit_group:
                                    agents_with_toolkit.append(agent_name)
                                    toolkit_group_allocations[agent_id] = 0
                            else:
                                if len(agents_with_ins) + len(agents_with_toolkit) < 5:
                                    pass

                        # Initialize Agent Name column if it doesn't exist
                        if "Agent Name" not in processed_df.columns:
                            processed_df["Agent Name"] = ""

                        # Step 2.5: Global NTBP Allocation - Allocate all NTBP remark rows globally
                        # Allocate NTBP rows to agents with PB in Allocation Preference column
                        # Use CC column for current capacity
                        all_ntbp_rows = []
                        if remark_col and remark_col in processed_df.columns:
                            for idx in processed_df.index:
                                # Skip already allocated rows
                                if idx in [
                                    i
                                    for ag in agent_allocations
                                    for i in ag["row_indices"]
                                ]:
                                    continue
                                if pd.notna(processed_df.at[idx, remark_col]):
                                    row_remark = (
                                        str(processed_df.at[idx, remark_col])
                                        .strip()
                                        .upper()
                                    )
                                    # Skip rows with "Not to work" remark - they should never be allocated
                                    if (
                                        "NOT TO WORK" in row_remark
                                        or row_remark == "NOT TO WORK"
                                    ):
                                        continue
                                    if row_remark == "NTBP":
                                        all_ntbp_rows.append(idx)

                        # Find agents with PB in Allocation Preference column (NOT domain)
                        # Only agents with "PB" in their Allocation Preference column should get NTBP work
                        agents_with_pb_preference = []
                        pb_agent_names = []
                        for a in agent_allocations:
                            # Check has_pb_preference flag (set when "PB" is in Allocation Preference column)
                            # This flag is set during agent data preparation based on Allocation Preference column
                            if a.get("has_pb_preference", False):
                                agents_with_pb_preference.append(a)
                                pb_agent_names.append(a.get("name", "Unknown"))

                        # Debug output for troubleshooting
                        if all_ntbp_rows:
                            if not agents_with_pb_preference:
                                # No agents with PB preference found - NTBP rows will remain unallocated
                                # Check if allocation preference column was detected
                                if allocation_preference_col:
                                    agent_summary += f"\n⚠️ Warning: Found {len(all_ntbp_rows)} NTBP rows but no agents with 'PB' in Allocation Preference column '{allocation_preference_col}'. NTBP rows will remain unallocated."
                                else:
                                    agent_summary += f"\n⚠️ Warning: Found {len(all_ntbp_rows)} NTBP rows but 'Allocation Preference' column not found. NTBP rows will remain unallocated."
                            else:
                                pb_names_safe = [
                                    str(name).strip()
                                    for name in pb_agent_names[:5]
                                    if name is not None and not pd.isna(name)
                                ]
                                agent_summary += f"\n✅ Found {len(agents_with_pb_preference)} agent(s) with PB preference ({', '.join(pb_names_safe)}{'...' if len(pb_agent_names) > 5 else ''}) for {len(all_ntbp_rows)} NTBP rows"

                        # Only allocate NTBP rows if we have PB preference agents
                        # If no PB preference agents exist, NTBP rows will remain unallocated
                        # (they should NOT be allocated to other agents in later steps)
                        if agents_with_pb_preference and all_ntbp_rows:
                            # Calculate total capacity of PB preference agents (from CC column)
                            total_pb_capacity = sum(
                                a["capacity"] - a["allocated"]
                                for a in agents_with_pb_preference
                            )

                            if len(all_ntbp_rows) >= total_pb_capacity:
                                # Distribute equally when NTBP count >= total capacity
                                # Round-robin distribution: assign rows one by one to each agent in turn
                                # Filter to only agents with available capacity
                                available_pb_agents = [
                                    a
                                    for a in agents_with_pb_preference
                                    if a["capacity"] > a["allocated"]
                                ]

                                if available_pb_agents:
                                    agent_idx = 0
                                    for ntbp_row_idx in all_ntbp_rows:
                                        # Find next available agent with capacity (round-robin)
                                        # Keep trying until we find an agent with capacity or exhaust all options
                                        assigned = False
                                        max_attempts = (
                                            len(available_pb_agents) * 10
                                        )  # Increased attempts
                                        attempts = 0

                                        while attempts < max_attempts and not assigned:
                                            # Refresh available agents list in case some filled up
                                            available_pb_agents = [
                                                a
                                                for a in agents_with_pb_preference
                                                if a["capacity"] > a["allocated"]
                                            ]

                                            if not available_pb_agents:
                                                # No more capacity available - stop allocation
                                                break

                                            agent = available_pb_agents[
                                                agent_idx % len(available_pb_agents)
                                            ]

                                            # Check capacity and appointment date limit (max 20 per date)
                                            if agent["capacity"] > agent["allocated"]:
                                                # Check appointment date limit before allocating
                                                if can_allocate_row_by_appointment_date(
                                                    agent,
                                                    ntbp_row_idx,
                                                    processed_df,
                                                    appointment_date_col,
                                                ):
                                                    agent["row_indices"].append(
                                                        ntbp_row_idx
                                                    )
                                                    agent["allocated"] += 1
                                                    processed_df.at[
                                                        ntbp_row_idx, "Agent Name"
                                                    ] = agent["name"]
                                                    assigned = True
                                                    agent_idx += 1  # Move to next agent for next row
                                                    break
                                                else:
                                                    # Can't allocate due to appointment date limit, try next agent
                                                    agent_idx += 1
                                            else:
                                                # This agent is full, try next one
                                                agent_idx += 1

                                            attempts += 1

                                        # If we couldn't assign this row, check if any agents still have capacity
                                        if not assigned:
                                            # Final check - refresh available agents
                                            available_pb_agents = [
                                                a
                                                for a in agents_with_pb_preference
                                                if a["capacity"] > a["allocated"]
                                            ]
                                            if not available_pb_agents:
                                                # No more capacity - stop allocation
                                                break
                            else:
                                # If NTBP rows are fewer than total capacity, allocate ALL to a single agent
                                available_pb_agents = [
                                    a
                                    for a in agents_with_pb_preference
                                    if a["capacity"] > a["allocated"]
                                ]
                                if available_pb_agents:
                                    # Sort by remaining capacity (highest first) to pick best agent
                                    available_pb_agents.sort(
                                        key=lambda x: x["capacity"] - x["allocated"],
                                        reverse=True,
                                    )
                                    # Allocate ALL NTBP rows, starting with the agent with highest capacity
                                    # If that agent fills up, continue with next agent
                                    remaining_ntbp_rows = all_ntbp_rows.copy()
                                    for agent in available_pb_agents:
                                        if not remaining_ntbp_rows:
                                            break
                                        available = (
                                            agent["capacity"] - agent["allocated"]
                                        )
                                        if available > 0:
                                            rows_to_assign = min(
                                                available, len(remaining_ntbp_rows)
                                            )
                                            if rows_to_assign > 0:
                                                assigned = remaining_ntbp_rows[
                                                    :rows_to_assign
                                                ]
                                                # Safety check: Filter out any "Not to work" rows and check appointment date limit
                                                filtered_assigned = []
                                                for idx in assigned:
                                                    if remark_col and pd.notna(
                                                        processed_df.at[idx, remark_col]
                                                    ):
                                                        remark_check = (
                                                            str(
                                                                processed_df.at[
                                                                    idx, remark_col
                                                                ]
                                                            )
                                                            .strip()
                                                            .upper()
                                                        )
                                                        if (
                                                            "NOT TO WORK"
                                                            in remark_check
                                                            or remark_check
                                                            == "NOT TO WORK"
                                                        ):
                                                            continue  # Skip this row
                                                    # Check appointment date limit (max 20 per date)
                                                    if can_allocate_row_by_appointment_date(
                                                        agent,
                                                        idx,
                                                        processed_df,
                                                        appointment_date_col,
                                                    ):
                                                        filtered_assigned.append(idx)

                                                if filtered_assigned:
                                                    agent["row_indices"].extend(
                                                        filtered_assigned
                                                    )
                                                    agent["allocated"] += len(
                                                        filtered_assigned
                                                    )
                                                    for idx in filtered_assigned:
                                                        processed_df.at[
                                                            idx, "Agent Name"
                                                        ] = agent["name"]
                                                remaining_ntbp_rows = (
                                                    remaining_ntbp_rows[rows_to_assign:]
                                                )

                        # Step 3.5: Global NTC Allocation - Allocate all NTC remark rows globally
                        # Allocate NTC rows to agents with NTC in Allocation Preference column
                        # Valid Allocation Preference values: "Sec+NTC", "Sec+Mix+NTC", "Mix+NTC", "NTC"
                        # Use CC column for current capacity
                        all_ntc_rows = []
                        if remark_col and remark_col in processed_df.columns:
                            for idx in processed_df.index:
                                # Skip already allocated rows
                                if idx in [
                                    i
                                    for ag in agent_allocations
                                    for i in ag["row_indices"]
                                ]:
                                    continue
                                if pd.notna(processed_df.at[idx, remark_col]):
                                    row_remark = (
                                        str(processed_df.at[idx, remark_col])
                                        .strip()
                                        .upper()
                                    )
                                    # Skip rows with "Not to work" remark - they should never be allocated
                                    if (
                                        "NOT TO WORK" in row_remark
                                        or row_remark == "NOT TO WORK"
                                    ):
                                        continue
                                    if row_remark == "NTC":
                                        all_ntc_rows.append(idx)

                        # Find agents with NTC in Allocation Preference column
                        # Valid values: "Sec+NTC", "Sec+Mix+NTC", "Mix+NTC", "NTC"
                        agents_with_ntc_preference = []
                        ntc_agent_names = []
                        for a in agent_allocations:
                            # Check has_ntc_preference flag (set when "NTC" is in Allocation Preference column)
                            if a.get("has_ntc_preference", False):
                                agents_with_ntc_preference.append(a)
                                ntc_agent_names.append(a.get("name", "Unknown"))

                        # Debug output for troubleshooting
                        if all_ntc_rows:
                            if not agents_with_ntc_preference:
                                # No agents with NTC preference found - NTC rows will remain unallocated
                                if allocation_preference_col:
                                    agent_summary += f"\n⚠️ Warning: Found {len(all_ntc_rows)} NTC rows but no agents with 'NTC' in Allocation Preference column '{allocation_preference_col}'. NTC rows will remain unallocated."
                                else:
                                    agent_summary += f"\n⚠️ Warning: Found {len(all_ntc_rows)} NTC rows but 'Allocation Preference' column not found. NTC rows will remain unallocated."
                            else:
                                ntc_names_safe = [
                                    str(name).strip()
                                    for name in ntc_agent_names[:5]
                                    if name is not None and not pd.isna(name)
                                ]
                                agent_summary += f"\n✅ Found {len(agents_with_ntc_preference)} agent(s) with NTC preference ({', '.join(ntc_names_safe)}{'...' if len(ntc_agent_names) > 5 else ''}) for {len(all_ntc_rows)} NTC rows"

                        # Only allocate NTC rows if we have NTC preference agents
                        if agents_with_ntc_preference and all_ntc_rows:
                            # Calculate total capacity of NTC preference agents (from CC column)
                            total_ntc_capacity = sum(
                                a["capacity"] - a["allocated"]
                                for a in agents_with_ntc_preference
                            )

                            if len(all_ntc_rows) >= total_ntc_capacity:
                                # Distribute equally when NTC count >= total capacity
                                # Round-robin distribution: assign rows one by one to each agent in turn
                                # Limit: No agent can receive more than 15 NTC rows
                                available_ntc_agents = [
                                    a
                                    for a in agents_with_ntc_preference
                                    if a["capacity"] > a["allocated"]
                                    and a.get("ntc_allocated", 0) < 15
                                ]

                                if available_ntc_agents:
                                    agent_idx = 0
                                    for ntc_row_idx in all_ntc_rows:
                                        # Find next available agent with capacity (round-robin)
                                        assigned = False
                                        max_attempts = (
                                            len(available_ntc_agents) * 10
                                        )  # Increased attempts
                                        attempts = 0

                                        while attempts < max_attempts and not assigned:
                                            # Refresh available agents list in case some filled up
                                            # Limit: No agent can receive more than 15 NTC rows
                                            available_ntc_agents = [
                                                a
                                                for a in agents_with_ntc_preference
                                                if a["capacity"] > a["allocated"]
                                                and a.get("ntc_allocated", 0) < 15
                                            ]

                                            if not available_ntc_agents:
                                                # No more capacity available or all agents reached NTC limit - stop allocation
                                                break

                                            agent = available_ntc_agents[
                                                agent_idx % len(available_ntc_agents)
                                            ]

                                            if (
                                                agent["capacity"] > agent["allocated"]
                                                and agent.get("ntc_allocated", 0) < 15
                                            ):
                                                # Safety check: Verify this is not a "Not to work" row before allocating
                                                if remark_col and pd.notna(
                                                    processed_df.at[
                                                        ntc_row_idx, remark_col
                                                    ]
                                                ):
                                                    remark_check = (
                                                        str(
                                                            processed_df.at[
                                                                ntc_row_idx, remark_col
                                                            ]
                                                        )
                                                        .strip()
                                                        .upper()
                                                    )
                                                    if (
                                                        "NOT TO WORK" in remark_check
                                                        or remark_check == "NOT TO WORK"
                                                    ):
                                                        continue  # Skip this row

                                                # Check appointment date limit (max 20 per date)
                                                if can_allocate_row_by_appointment_date(
                                                    agent,
                                                    ntc_row_idx,
                                                    processed_df,
                                                    appointment_date_col,
                                                ):
                                                    agent["row_indices"].append(
                                                        ntc_row_idx
                                                    )
                                                    agent["allocated"] += 1
                                                    agent["ntc_allocated"] = (
                                                        agent.get("ntc_allocated", 0)
                                                        + 1
                                                    )
                                                    processed_df.at[
                                                        ntc_row_idx, "Agent Name"
                                                    ] = agent["name"]
                                                    assigned = True
                                                    agent_idx += 1  # Move to next agent for next row
                                                    break
                                                else:
                                                    # Can't allocate due to appointment date limit, try next agent
                                                    agent_idx += 1
                                            else:
                                                # This agent is full, try next one
                                                agent_idx += 1

                                            attempts += 1

                                        # If we couldn't assign this row, check if any agents still have capacity
                                        if not assigned:
                                            # Final check - refresh available agents
                                            # Limit: No agent can receive more than 15 NTC rows
                                            available_ntc_agents = [
                                                a
                                                for a in agents_with_ntc_preference
                                                if a["capacity"] > a["allocated"]
                                                and a.get("ntc_allocated", 0) < 15
                                            ]
                                            if not available_ntc_agents:
                                                # No more capacity or all agents reached NTC limit - stop allocation
                                                break
                            else:
                                # If NTC rows are fewer than total capacity, allocate ALL to a single agent
                                # Limit: No agent can receive more than 15 NTC rows
                                available_ntc_agents = [
                                    a
                                    for a in agents_with_ntc_preference
                                    if a["capacity"] > a["allocated"]
                                    and a.get("ntc_allocated", 0) < 15
                                ]
                                if available_ntc_agents:
                                    # Sort by remaining capacity (highest first) to pick best agent
                                    available_ntc_agents.sort(
                                        key=lambda x: x["capacity"] - x["allocated"],
                                        reverse=True,
                                    )
                                    # Allocate NTC rows, starting with the agent with highest capacity
                                    # If that agent fills up or reaches NTC limit, continue with next agent
                                    # Limit: No agent can receive more than 15 NTC rows
                                    remaining_ntc_rows = all_ntc_rows.copy()
                                    for agent in available_ntc_agents:
                                        if not remaining_ntc_rows:
                                            break
                                        available = (
                                            agent["capacity"] - agent["allocated"]
                                        )
                                        ntc_limit_remaining = 15 - agent.get(
                                            "ntc_allocated", 0
                                        )
                                        if available > 0 and ntc_limit_remaining > 0:
                                            rows_to_assign = min(
                                                available,
                                                len(remaining_ntc_rows),
                                                ntc_limit_remaining,
                                            )
                                            if rows_to_assign > 0:
                                                assigned = remaining_ntc_rows[
                                                    :rows_to_assign
                                                ]
                                                # Safety check: Filter out any "Not to work" rows and check appointment date limit
                                                filtered_assigned = []
                                                for idx in assigned:
                                                    if remark_col and pd.notna(
                                                        processed_df.at[idx, remark_col]
                                                    ):
                                                        remark_check = (
                                                            str(
                                                                processed_df.at[
                                                                    idx, remark_col
                                                                ]
                                                            )
                                                            .strip()
                                                            .upper()
                                                        )
                                                        if (
                                                            "NOT TO WORK"
                                                            in remark_check
                                                            or remark_check
                                                            == "NOT TO WORK"
                                                        ):
                                                            continue  # Skip this row
                                                    # Check appointment date limit (max 20 per date)
                                                    if can_allocate_row_by_appointment_date(
                                                        agent,
                                                        idx,
                                                        processed_df,
                                                        appointment_date_col,
                                                    ):
                                                        filtered_assigned.append(idx)

                                                if filtered_assigned:
                                                    agent["row_indices"].extend(
                                                        filtered_assigned
                                                    )
                                                    agent["allocated"] += len(
                                                        filtered_assigned
                                                    )
                                                    agent["ntc_allocated"] = agent.get(
                                                        "ntc_allocated", 0
                                                    ) + len(filtered_assigned)
                                                    for idx in filtered_assigned:
                                                        processed_df.at[
                                                            idx, "Agent Name"
                                                        ] = agent["name"]
                                                remaining_ntc_rows = remaining_ntc_rows[
                                                    rows_to_assign:
                                                ]

                        # Step 3.5.5: After NTC allocation, allocate other insurance rows to "Mix + NTC" agents if capacity remains
                        # "Mix + NTC" agents should get NTC rows first (done in Step 3.5), then other insurance rows if capacity remains
                        mix_ntc_agents = [
                            a
                            for a in agent_allocations
                            if a.get("has_mix_preference", False)
                            and a.get("has_ntc_preference", False)
                            and not a.get(
                                "has_sec_preference", False
                            )  # Exclude "Sec + Mix + NTC" (handled separately)
                            and (a["capacity"] - a["allocated"]) > 0
                        ]

                        if mix_ntc_agents and insurance_carrier_col:
                            for agent in mix_ntc_agents:
                                remaining_capacity = (
                                    agent["capacity"] - agent["allocated"]
                                )
                                if remaining_capacity <= 0:
                                    continue

                                agent_insurance_list = agent.get(
                                    "insurance_companies", []
                                )

                                # Find unallocated rows from multiple insurance companies that match agent's capabilities
                                # Exclude NTC rows (already allocated) and secondary insurance rows
                                mix_ntc_rows = []

                                for idx in processed_df.index:
                                    # Skip already allocated rows
                                    if idx in [
                                        i
                                        for ag in agent_allocations
                                        for i in ag["row_indices"]
                                    ]:
                                        continue

                                    # Skip rows with secondary insurance (those go to "Sec + X" agents)
                                    if secondary_insurance_col and pd.notna(
                                        processed_df.at[idx, secondary_insurance_col]
                                    ):
                                        secondary_val = str(
                                            processed_df.at[
                                                idx, secondary_insurance_col
                                            ]
                                        ).strip()
                                        if (
                                            secondary_val
                                            and secondary_val.lower() != "nan"
                                        ):
                                            continue

                                    # Skip NTBP rows (handled in Step 2.5)
                                    if remark_col and pd.notna(
                                        processed_df.at[idx, remark_col]
                                    ):
                                        remark_val = (
                                            str(processed_df.at[idx, remark_col])
                                            .strip()
                                            .upper()
                                        )
                                        if remark_val == "NTBP":
                                            continue
                                        # Skip NTC rows (already allocated in Step 3.5)
                                        if remark_val == "NTC":
                                            continue
                                        # Skip rows with "Not to work" remark
                                        if (
                                            "NOT TO WORK" in remark_val
                                            or remark_val == "NOT TO WORK"
                                        ):
                                            continue

                                    # Get insurance company from "Dental Primary Ins Carr" column
                                    if pd.notna(
                                        processed_df.at[idx, insurance_carrier_col]
                                    ):
                                        row_insurance_raw = str(
                                            processed_df.at[idx, insurance_carrier_col]
                                        ).strip()

                                        # Use improved matching function that handles formatting variations and "DD All"
                                        can_work = check_insurance_match(
                                            row_insurance_raw,
                                            agent_insurance_list,
                                            agent.get("is_senior", False),
                                            agent.get("name"),
                                        )

                                        if can_work:
                                            mix_ntc_rows.append(idx)

                                # Allocate other insurance rows until capacity is full
                                if mix_ntc_rows:
                                    allocated_count = 0
                                    while (
                                        remaining_capacity > 0
                                        and allocated_count < len(mix_ntc_rows)
                                    ):
                                        take = min(
                                            remaining_capacity,
                                            len(mix_ntc_rows) - allocated_count,
                                        )
                                        if take > 0:
                                            slice_rows = mix_ntc_rows[
                                                allocated_count : allocated_count + take
                                            ]
                                            # Use safe extend function to filter out "Not to work" rows
                                            actual_allocated = safe_extend_row_indices(
                                                agent,
                                                slice_rows,
                                                processed_df,
                                                remark_col,
                                                agent["name"],
                                                appointment_date_col,
                                            )
                                            allocated_count += take
                                            remaining_capacity = (
                                                agent["capacity"] - agent["allocated"]
                                            )
                                        else:
                                            break

                        # Step 3.6: Global Secondary Insurance Allocation - Allocate rows with secondary insurance to "Sec + X" agents
                        # This should happen before other allocations so "Sec + X" agents get secondary insurance rows first
                        if (
                            secondary_insurance_col
                            and secondary_insurance_col in processed_df.columns
                        ):
                            # Find all agents with "Sec + X" preference
                            sec_preference_agents = [
                                a
                                for a in agent_allocations
                                if a.get("has_sec_preference", False)
                                and (a["capacity"] - a["allocated"]) > 0
                            ]

                            # Initialize secondary insurance count for all "Sec + XX" agents
                            for agent in sec_preference_agents:
                                if "secondary_insurance_count" not in agent:
                                    agent["secondary_insurance_count"] = 0

                            if sec_preference_agents:
                                # Find all unallocated rows that have secondary insurance
                                rows_with_secondary_insurance = []
                                for idx in processed_df.index:
                                    # Skip already allocated rows
                                    if idx in [
                                        i
                                        for ag in agent_allocations
                                        for i in ag["row_indices"]
                                    ]:
                                        continue

                                    # Check if row has secondary insurance
                                    if pd.notna(
                                        processed_df.at[idx, secondary_insurance_col]
                                    ):
                                        secondary_val = str(
                                            processed_df.at[
                                                idx, secondary_insurance_col
                                            ]
                                        ).strip()
                                        if (
                                            secondary_val
                                            and secondary_val.lower() != "nan"
                                        ):
                                            # Skip "Not to work" rows before adding to list
                                            if not should_skip_row_for_allocation(
                                                idx, processed_df, remark_col
                                            ):
                                                rows_with_secondary_insurance.append(
                                                    idx
                                                )

                                if rows_with_secondary_insurance:
                                    # Allocate secondary insurance rows to "Sec + X" agents
                                    # For "Sec + Single", they can only get one insurance company
                                    # For "Sec + NTC", "Sec + Mix", etc., they can get any insurance

                                    # Separate agents by type
                                    sec_single_agents = [
                                        a
                                        for a in sec_preference_agents
                                        if a.get("has_sec_single_preference", False)
                                    ]
                                    sec_other_agents = [
                                        a
                                        for a in sec_preference_agents
                                        if not a.get("has_sec_single_preference", False)
                                    ]

                                    # First, allocate to "Sec + Single" agents (they need to stick to one insurance)
                                    if (
                                        sec_single_agents
                                        and rows_with_secondary_insurance
                                    ):
                                        # Group rows by insurance carrier for "Sec + Single" agents
                                        rows_by_carrier = {}
                                        for row_idx in rows_with_secondary_insurance:
                                            if insurance_carrier_col and pd.notna(
                                                processed_df.at[
                                                    row_idx, insurance_carrier_col
                                                ]
                                            ):
                                                carrier = str(
                                                    processed_df.at[
                                                        row_idx, insurance_carrier_col
                                                    ]
                                                ).strip()
                                                if carrier not in rows_by_carrier:
                                                    rows_by_carrier[carrier] = []
                                                # Skip "Not to work" rows before adding to list
                                                if not should_skip_row_for_allocation(
                                                    row_idx, processed_df, remark_col
                                                ):
                                                    rows_by_carrier[carrier].append(
                                                        row_idx
                                                    )

                                        # Allocate to "Sec + Single" agents, one carrier at a time
                                        for (
                                            carrier,
                                            carrier_rows,
                                        ) in rows_by_carrier.items():
                                            available_sec_single = [
                                                a
                                                for a in sec_single_agents
                                                if (a["capacity"] - a["allocated"]) > 0
                                                and a.get("assigned_insurance")
                                                in (None, carrier)
                                            ]

                                            if available_sec_single:
                                                available_sec_single.sort(
                                                    key=lambda a: a["capacity"]
                                                    - a["allocated"],
                                                    reverse=True,
                                                )

                                                # Track which agents actually received secondary insurance rows
                                                agents_with_secondary = []

                                                row_pos = 0
                                                for agent in available_sec_single:
                                                    if row_pos >= len(carrier_rows):
                                                        break
                                                    # Calculate remaining capacity
                                                    remaining_capacity = (
                                                        agent["capacity"]
                                                        - agent["allocated"]
                                                    )
                                                    # Calculate remaining secondary insurance slots (max 10)
                                                    remaining_secondary_slots = (
                                                        get_available_secondary_slots(
                                                            agent
                                                        )
                                                    )
                                                    # Take the minimum of capacity and secondary slots
                                                    remaining = min(
                                                        remaining_capacity,
                                                        remaining_secondary_slots,
                                                    )
                                                    if remaining <= 0:
                                                        continue
                                                    take = min(
                                                        remaining,
                                                        len(carrier_rows) - row_pos,
                                                    )
                                                    if take > 0:
                                                        slice_rows = carrier_rows[
                                                            row_pos : row_pos + take
                                                        ]
                                                        # Filter by appointment date limit (max 20 per date)
                                                        final_slice = []
                                                        for idx in slice_rows:
                                                            if can_allocate_row_by_appointment_date(
                                                                agent,
                                                                idx,
                                                                processed_df,
                                                                appointment_date_col,
                                                            ):
                                                                final_slice.append(idx)

                                                        if final_slice:
                                                            agent["row_indices"].extend(
                                                                final_slice
                                                            )
                                                            agent["allocated"] += len(
                                                                final_slice
                                                            )
                                                            # Track secondary insurance row count
                                                            if (
                                                                "secondary_insurance_count"
                                                                not in agent
                                                            ):
                                                                agent[
                                                                    "secondary_insurance_count"
                                                                ] = 0
                                                            agent[
                                                                "secondary_insurance_count"
                                                            ] += len(final_slice)
                                                            if (
                                                                agent.get(
                                                                    "assigned_insurance"
                                                                )
                                                                is None
                                                            ):
                                                                agent[
                                                                    "assigned_insurance"
                                                                ] = carrier
                                                            for idx in final_slice:
                                                                processed_df.at[
                                                                    idx, "Agent Name"
                                                                ] = agent["name"]
                                                            row_pos += len(final_slice)
                                                        else:
                                                            # No rows could be allocated due to appointment date limit, move to next agent
                                                            break
                                                        # Track this agent for same insurance allocation
                                                        agents_with_secondary.append(
                                                            agent
                                                        )

                                                # Remove allocated rows
                                                rows_with_secondary_insurance = [
                                                    r
                                                    for r in rows_with_secondary_insurance
                                                    if r not in carrier_rows[:row_pos]
                                                ]

                                                # After allocating secondary insurance rows to "Sec + Single" agents,
                                                # if they still have capacity, allocate same insurance company rows
                                                # BUT only if the insurance company is in the agent's Insurance List capabilities
                                                for agent in agents_with_secondary:
                                                    remaining_capacity = (
                                                        agent["capacity"]
                                                        - agent["allocated"]
                                                    )
                                                    if remaining_capacity <= 0:
                                                        continue

                                                    # Check if agent can work with this insurance company (from Insurance List)
                                                    agent_can_work = False
                                                    agent_insurance_list = agent.get(
                                                        "insurance_companies", []
                                                    )

                                                    # Senior agents can work with any insurance company
                                                    if agent.get("is_senior", False):
                                                        agent_can_work = True
                                                    elif not agent_insurance_list:
                                                        # If no specific companies listed, can work with any
                                                        agent_can_work = True
                                                    else:
                                                        # Check if the carrier is in the agent's insurance list
                                                        carrier_lower = carrier.lower()
                                                        for (
                                                            comp
                                                        ) in agent_insurance_list:
                                                            comp_lower = comp.lower()
                                                            if (
                                                                carrier_lower
                                                                in comp_lower
                                                                or comp_lower
                                                                in carrier_lower
                                                                or carrier == comp
                                                            ):
                                                                agent_can_work = True
                                                                break

                                                    # Only proceed if agent can work with this insurance company
                                                    if not agent_can_work:
                                                        continue

                                                    # Find unallocated rows with the same insurance company
                                                    # (excluding rows with secondary insurance, as those were already allocated)
                                                    # AND check that the insurance company from "Dental Primary Ins Carr" matches
                                                    same_insurance_rows = []
                                                    for idx in processed_df.index:
                                                        # Skip already allocated rows
                                                        if idx in [
                                                            i
                                                            for ag in agent_allocations
                                                            for i in ag["row_indices"]
                                                        ]:
                                                            continue
                                                        # Skip rows with secondary insurance (already allocated)
                                                        if (
                                                            secondary_insurance_col
                                                            and pd.notna(
                                                                processed_df.at[
                                                                    idx,
                                                                    secondary_insurance_col,
                                                                ]
                                                            )
                                                        ):
                                                            secondary_val = str(
                                                                processed_df.at[
                                                                    idx,
                                                                    secondary_insurance_col,
                                                                ]
                                                            ).strip()
                                                            if (
                                                                secondary_val
                                                                and secondary_val.lower()
                                                                != "nan"
                                                            ):
                                                                continue
                                                        # Check if insurance carrier from "Dental Primary Ins Carr" matches assigned carrier
                                                        # AND check if that insurance company is in agent's "Insurance List" (capabilities)
                                                        if (
                                                            insurance_carrier_col
                                                            and pd.notna(
                                                                processed_df.at[
                                                                    idx,
                                                                    insurance_carrier_col,
                                                                ]
                                                            )
                                                        ):
                                                            row_insurance = str(
                                                                processed_df.at[
                                                                    idx,
                                                                    insurance_carrier_col,
                                                                ]
                                                            ).strip()

                                                            # First check: row insurance must match assigned carrier
                                                            if row_insurance == carrier:
                                                                # Second check: row insurance must be in agent's "Insurance List" (capabilities)
                                                                can_work = False
                                                                if agent.get(
                                                                    "is_senior", False
                                                                ):
                                                                    # Senior agents can work with any insurance company
                                                                    can_work = True
                                                                elif (
                                                                    not agent_insurance_list
                                                                ):
                                                                    # If no specific companies listed, can work with any
                                                                    can_work = True
                                                                else:
                                                                    # Check if row insurance matches/is in agent's Insurance List
                                                                    row_insurance_lower = (
                                                                        row_insurance.lower()
                                                                    )
                                                                    for (
                                                                        comp
                                                                    ) in agent_insurance_list:
                                                                        comp_lower = (
                                                                            comp.lower()
                                                                        )
                                                                        if (
                                                                            row_insurance_lower
                                                                            in comp_lower
                                                                            or comp_lower
                                                                            in row_insurance_lower
                                                                            or row_insurance
                                                                            == comp
                                                                        ):
                                                                            can_work = (
                                                                                True
                                                                            )
                                                                            break

                                                                if can_work:
                                                                    # Skip "Not to work" rows before adding to list
                                                                    if not should_skip_row_for_allocation(
                                                                        idx,
                                                                        processed_df,
                                                                        remark_col,
                                                                    ):
                                                                        same_insurance_rows.append(
                                                                            idx
                                                                        )

                                                    # Allocate same insurance rows (up to remaining capacity)
                                                    # Keep allocating until capacity is full or no more rows available
                                                    allocated_count = 0
                                                    while (
                                                        remaining_capacity > 0
                                                        and allocated_count
                                                        < len(same_insurance_rows)
                                                    ):
                                                        take = min(
                                                            remaining_capacity,
                                                            len(same_insurance_rows)
                                                            - allocated_count,
                                                        )
                                                        if take > 0:
                                                            slice_rows = same_insurance_rows[
                                                                allocated_count : allocated_count
                                                                + take
                                                            ]
                                                            # Use safe extend function to filter out "Not to work" rows
                                                            actual_allocated = (
                                                                safe_extend_row_indices(
                                                                    agent,
                                                                    slice_rows,
                                                                    processed_df,
                                                                    remark_col,
                                                                    agent["name"],
                                                                )
                                                            )
                                                            allocated_count += take
                                                            remaining_capacity = (
                                                                agent["capacity"]
                                                                - agent["allocated"]
                                                            )
                                                        else:
                                                            break

                                    # Then, allocate remaining secondary insurance rows to other "Sec + X" agents
                                    if (
                                        sec_other_agents
                                        and rows_with_secondary_insurance
                                    ):
                                        sec_other_agents.sort(
                                            key=lambda a: a["capacity"]
                                            - a["allocated"],
                                            reverse=True,
                                        )

                                        row_pos = 0
                                        for agent in sec_other_agents:
                                            if row_pos >= len(
                                                rows_with_secondary_insurance
                                            ):
                                                break
                                            # Calculate remaining capacity
                                            remaining_capacity = (
                                                agent["capacity"] - agent["allocated"]
                                            )
                                            # Calculate remaining secondary insurance slots (max 10)
                                            remaining_secondary_slots = (
                                                get_available_secondary_slots(agent)
                                            )
                                            # Take the minimum of capacity and secondary slots
                                            remaining = min(
                                                remaining_capacity,
                                                remaining_secondary_slots,
                                            )
                                            if remaining <= 0:
                                                continue
                                            take = min(
                                                remaining,
                                                len(rows_with_secondary_insurance)
                                                - row_pos,
                                            )
                                            if take > 0:
                                                slice_rows = (
                                                    rows_with_secondary_insurance[
                                                        row_pos : row_pos + take
                                                    ]
                                                )
                                                # Filter out "Not to work" rows and check appointment date limit
                                                filtered_slice = []
                                                for idx in slice_rows:
                                                    if not should_skip_row_for_allocation(
                                                        idx, processed_df, remark_col
                                                    ):
                                                        # Check appointment date limit (max 20 per date)
                                                        if can_allocate_row_by_appointment_date(
                                                            agent,
                                                            idx,
                                                            processed_df,
                                                            appointment_date_col,
                                                        ):
                                                            filtered_slice.append(idx)

                                                if filtered_slice:
                                                    agent["row_indices"].extend(
                                                        filtered_slice
                                                    )
                                                    agent["allocated"] += len(
                                                        filtered_slice
                                                    )
                                                    # Track secondary insurance row count
                                                    if (
                                                        "secondary_insurance_count"
                                                        not in agent
                                                    ):
                                                        agent[
                                                            "secondary_insurance_count"
                                                        ] = 0
                                                    agent[
                                                        "secondary_insurance_count"
                                                    ] += len(filtered_slice)
                                                    for idx in filtered_slice:
                                                        processed_df.at[
                                                            idx, "Agent Name"
                                                        ] = agent["name"]
                                                actual_allocated = len(filtered_slice)
                                                # Set assigned insurance if not set (use primary insurance carrier)
                                                if (
                                                    agent.get("assigned_insurance")
                                                    is None
                                                    and insurance_carrier_col
                                                    and filtered_slice
                                                ):
                                                    if pd.notna(
                                                        processed_df.at[
                                                            filtered_slice[0],
                                                            insurance_carrier_col,
                                                        ]
                                                    ):
                                                        agent["assigned_insurance"] = (
                                                            str(
                                                                processed_df.at[
                                                                    filtered_slice[0],
                                                                    insurance_carrier_col,
                                                                ]
                                                            ).strip()
                                                        )
                                                row_pos += take

                                    # After allocating secondary insurance rows, if agents still have capacity,
                                    # allocate based on the value after "Sec +" (Single, NTC, Mix, etc.)
                                    # Check each "Sec + X" agent for remaining capacity
                                    for agent in sec_preference_agents:
                                        remaining_capacity = (
                                            agent["capacity"] - agent["allocated"]
                                        )
                                        if remaining_capacity <= 0:
                                            continue

                                        # Get the allocation preference to determine what comes after "Sec +"
                                        allocation_pref = agent.get(
                                            "allocation_preference_raw", ""
                                        )
                                        if allocation_pref:
                                            allocation_pref_upper = (
                                                str(allocation_pref).strip().upper()
                                            )

                                            # For "Sec + Single", allocate same insurance company rows
                                            if agent.get(
                                                "has_sec_single_preference", False
                                            ):
                                                # Agent already has assigned_insurance from secondary allocation
                                                assigned_ins = agent.get(
                                                    "assigned_insurance"
                                                )
                                                if assigned_ins:
                                                    # Find unallocated rows with the same insurance company
                                                    same_insurance_rows = []
                                                    for idx in processed_df.index:
                                                        # Skip already allocated rows
                                                        if idx in [
                                                            i
                                                            for ag in agent_allocations
                                                            for i in ag["row_indices"]
                                                        ]:
                                                            continue
                                                        # Skip rows with secondary insurance (already allocated)
                                                        if (
                                                            secondary_insurance_col
                                                            and pd.notna(
                                                                processed_df.at[
                                                                    idx,
                                                                    secondary_insurance_col,
                                                                ]
                                                            )
                                                        ):
                                                            secondary_val = str(
                                                                processed_df.at[
                                                                    idx,
                                                                    secondary_insurance_col,
                                                                ]
                                                            ).strip()
                                                            if (
                                                                secondary_val
                                                                and secondary_val.lower()
                                                                != "nan"
                                                            ):
                                                                continue
                                                        # Check if insurance carrier matches
                                                        if (
                                                            insurance_carrier_col
                                                            and pd.notna(
                                                                processed_df.at[
                                                                    idx,
                                                                    insurance_carrier_col,
                                                                ]
                                                            )
                                                        ):
                                                            row_carrier = str(
                                                                processed_df.at[
                                                                    idx,
                                                                    insurance_carrier_col,
                                                                ]
                                                            ).strip()
                                                            if (
                                                                row_carrier
                                                                == assigned_ins
                                                            ):
                                                                # Skip "Not to work" rows before adding to list
                                                                if not should_skip_row_for_allocation(
                                                                    idx,
                                                                    processed_df,
                                                                    remark_col,
                                                                ):
                                                                    same_insurance_rows.append(
                                                                        idx
                                                                    )

                                                    # Allocate same insurance rows
                                                    if same_insurance_rows:
                                                        take = min(
                                                            remaining_capacity,
                                                            len(same_insurance_rows),
                                                        )
                                                        if take > 0:
                                                            slice_rows = (
                                                                same_insurance_rows[
                                                                    :take
                                                                ]
                                                            )
                                                            agent["row_indices"].extend(
                                                                slice_rows
                                                            )
                                                            agent["allocated"] += take
                                                            for idx in slice_rows:
                                                                processed_df.at[
                                                                    idx, "Agent Name"
                                                                ] = agent["name"]

                                            # For "Sec + NTC", allocate NTC rows
                                            elif (
                                                "NTC" in allocation_pref_upper
                                                and "SEC" in allocation_pref_upper
                                            ):
                                                # Find unallocated NTC rows
                                                ntc_rows = []
                                                if (
                                                    remark_col
                                                    and remark_col
                                                    in processed_df.columns
                                                ):
                                                    for idx in processed_df.index:
                                                        # Skip already allocated rows
                                                        if idx in [
                                                            i
                                                            for ag in agent_allocations
                                                            for i in ag["row_indices"]
                                                        ]:
                                                            continue
                                                        # Skip rows with secondary insurance (already allocated)
                                                        if (
                                                            secondary_insurance_col
                                                            and pd.notna(
                                                                processed_df.at[
                                                                    idx,
                                                                    secondary_insurance_col,
                                                                ]
                                                            )
                                                        ):
                                                            secondary_val = str(
                                                                processed_df.at[
                                                                    idx,
                                                                    secondary_insurance_col,
                                                                ]
                                                            ).strip()
                                                            if (
                                                                secondary_val
                                                                and secondary_val.lower()
                                                                != "nan"
                                                            ):
                                                                continue
                                                        # Check if remark is NTC
                                                        if pd.notna(
                                                            processed_df.at[
                                                                idx, remark_col
                                                            ]
                                                        ):
                                                            row_remark = (
                                                                str(
                                                                    processed_df.at[
                                                                        idx, remark_col
                                                                    ]
                                                                )
                                                                .strip()
                                                                .upper()
                                                            )
                                                            if row_remark == "NTC":
                                                                ntc_rows.append(idx)

                                                # Allocate NTC rows
                                                # Limit: No agent can receive more than 15 NTC rows
                                                if ntc_rows:
                                                    ntc_limit_remaining = (
                                                        15
                                                        - agent.get("ntc_allocated", 0)
                                                    )
                                                    take = min(
                                                        remaining_capacity,
                                                        len(ntc_rows),
                                                        ntc_limit_remaining,
                                                    )
                                                    if take > 0:
                                                        slice_rows = ntc_rows[:take]
                                                        agent["row_indices"].extend(
                                                            slice_rows
                                                        )
                                                        agent["allocated"] += take
                                                        agent["ntc_allocated"] = (
                                                            agent.get(
                                                                "ntc_allocated", 0
                                                            )
                                                            + take
                                                        )
                                                        for idx in slice_rows:
                                                            processed_df.at[
                                                                idx, "Agent Name"
                                                            ] = agent["name"]

                                            # For "Sec + Mix", allocate Mix rows (multiple insurance company rows)
                                            elif (
                                                "MIX" in allocation_pref_upper
                                                and "SEC" in allocation_pref_upper
                                            ):
                                                # Find unallocated rows from multiple insurance companies that match agent's capabilities
                                                agent_insurance_list = agent.get(
                                                    "insurance_companies", []
                                                )
                                                mix_rows = []
                                                for idx in processed_df.index:
                                                    # Skip already allocated rows
                                                    if idx in [
                                                        i
                                                        for ag in agent_allocations
                                                        for i in ag["row_indices"]
                                                    ]:
                                                        continue
                                                    # Skip rows with secondary insurance (already allocated)
                                                    if (
                                                        secondary_insurance_col
                                                        and pd.notna(
                                                            processed_df.at[
                                                                idx,
                                                                secondary_insurance_col,
                                                            ]
                                                        )
                                                    ):
                                                        secondary_val = str(
                                                            processed_df.at[
                                                                idx,
                                                                secondary_insurance_col,
                                                            ]
                                                        ).strip()
                                                        if (
                                                            secondary_val
                                                            and secondary_val.lower()
                                                            != "nan"
                                                        ):
                                                            continue
                                                    # Skip rows with special remarks (NTBP, NTC)
                                                    if (
                                                        remark_col
                                                        and remark_col
                                                        in processed_df.columns
                                                    ):
                                                        if pd.notna(
                                                            processed_df.at[
                                                                idx, remark_col
                                                            ]
                                                        ):
                                                            row_remark = (
                                                                str(
                                                                    processed_df.at[
                                                                        idx, remark_col
                                                                    ]
                                                                )
                                                                .strip()
                                                                .upper()
                                                            )
                                                            if row_remark in [
                                                                "NTBP",
                                                                "NTC",
                                                            ]:
                                                                continue

                                                    # Get insurance company from "Dental Primary Ins Carr" column
                                                    if (
                                                        insurance_carrier_col
                                                        and pd.notna(
                                                            processed_df.at[
                                                                idx,
                                                                insurance_carrier_col,
                                                            ]
                                                        )
                                                    ):
                                                        row_insurance_raw = str(
                                                            processed_df.at[
                                                                idx,
                                                                insurance_carrier_col,
                                                            ]
                                                        ).strip()

                                                        # Use improved matching function that handles formatting variations and "DD All"
                                                        can_work = (
                                                            check_insurance_match(
                                                                row_insurance_raw,
                                                                agent_insurance_list,
                                                                agent.get(
                                                                    "is_senior", False
                                                                ),
                                                                agent.get("name"),
                                                            )
                                                        )

                                                        if can_work:
                                                            # Skip "Not to work" rows before adding to list
                                                            if not should_skip_row_for_allocation(
                                                                idx,
                                                                processed_df,
                                                                remark_col,
                                                            ):
                                                                mix_rows.append(idx)

                                                # Allocate Mix rows until capacity is full
                                                if mix_rows:
                                                    allocated_count = 0
                                                    while (
                                                        remaining_capacity > 0
                                                        and allocated_count
                                                        < len(mix_rows)
                                                    ):
                                                        take = min(
                                                            remaining_capacity,
                                                            len(mix_rows)
                                                            - allocated_count,
                                                        )
                                                        if take > 0:
                                                            slice_rows = mix_rows[
                                                                allocated_count : allocated_count
                                                                + take
                                                            ]
                                                            agent["row_indices"].extend(
                                                                slice_rows
                                                            )
                                                            agent["allocated"] += take
                                                            for idx in slice_rows:
                                                                processed_df.at[
                                                                    idx, "Agent Name"
                                                                ] = agent["name"]
                                                            allocated_count += take
                                                            remaining_capacity = (
                                                                agent["capacity"]
                                                                - agent["allocated"]
                                                            )
                                                        else:
                                                            break

                                            # For "Sec + Mix + NTC", allocate Mix rows first, then NTC rows
                                            elif (
                                                "MIX" in allocation_pref_upper
                                                and "NTC" in allocation_pref_upper
                                                and "SEC" in allocation_pref_upper
                                            ):
                                                # Phase 1: Allocate mixed insurance company rows (if capacity remains after secondary allocation)
                                                agent_insurance_list = agent.get(
                                                    "insurance_companies", []
                                                )
                                                mix_rows = []
                                                for idx in processed_df.index:
                                                    # Skip already allocated rows
                                                    if idx in [
                                                        i
                                                        for ag in agent_allocations
                                                        for i in ag["row_indices"]
                                                    ]:
                                                        continue
                                                    # Skip rows with secondary insurance (already allocated)
                                                    if (
                                                        secondary_insurance_col
                                                        and pd.notna(
                                                            processed_df.at[
                                                                idx,
                                                                secondary_insurance_col,
                                                            ]
                                                        )
                                                    ):
                                                        secondary_val = str(
                                                            processed_df.at[
                                                                idx,
                                                                secondary_insurance_col,
                                                            ]
                                                        ).strip()
                                                        if (
                                                            secondary_val
                                                            and secondary_val.lower()
                                                            != "nan"
                                                        ):
                                                            continue
                                                    # Skip rows with special remarks (NTBP, NTC - NTC will be handled in Phase 2)
                                                    if (
                                                        remark_col
                                                        and remark_col
                                                        in processed_df.columns
                                                    ):
                                                        if pd.notna(
                                                            processed_df.at[
                                                                idx, remark_col
                                                            ]
                                                        ):
                                                            row_remark = (
                                                                str(
                                                                    processed_df.at[
                                                                        idx, remark_col
                                                                    ]
                                                                )
                                                                .strip()
                                                                .upper()
                                                            )
                                                            if row_remark in [
                                                                "NTBP",
                                                                "NTC",
                                                            ]:
                                                                continue

                                                    # Get insurance company from "Dental Primary Ins Carr" column
                                                    if (
                                                        insurance_carrier_col
                                                        and pd.notna(
                                                            processed_df.at[
                                                                idx,
                                                                insurance_carrier_col,
                                                            ]
                                                        )
                                                    ):
                                                        row_insurance_raw = str(
                                                            processed_df.at[
                                                                idx,
                                                                insurance_carrier_col,
                                                            ]
                                                        ).strip()

                                                        # Use improved matching function that handles formatting variations and "DD All"
                                                        can_work = (
                                                            check_insurance_match(
                                                                row_insurance_raw,
                                                                agent_insurance_list,
                                                                agent.get(
                                                                    "is_senior", False
                                                                ),
                                                                agent.get("name"),
                                                            )
                                                        )

                                                        if can_work:
                                                            # Skip "Not to work" rows before adding to list
                                                            if not should_skip_row_for_allocation(
                                                                idx,
                                                                processed_df,
                                                                remark_col,
                                                            ):
                                                                mix_rows.append(idx)

                                                # Allocate Mix rows until capacity is full or no more rows
                                                if mix_rows:
                                                    allocated_count = 0
                                                    while (
                                                        remaining_capacity > 0
                                                        and allocated_count
                                                        < len(mix_rows)
                                                    ):
                                                        take = min(
                                                            remaining_capacity,
                                                            len(mix_rows)
                                                            - allocated_count,
                                                        )
                                                        if take > 0:
                                                            slice_rows = mix_rows[
                                                                allocated_count : allocated_count
                                                                + take
                                                            ]
                                                            agent["row_indices"].extend(
                                                                slice_rows
                                                            )
                                                            agent["allocated"] += take
                                                            for idx in slice_rows:
                                                                processed_df.at[
                                                                    idx, "Agent Name"
                                                                ] = agent["name"]
                                                            allocated_count += take
                                                            remaining_capacity = (
                                                                agent["capacity"]
                                                                - agent["allocated"]
                                                            )
                                                        else:
                                                            break

                                                # Phase 2: If capacity still remains, allocate NTC rows
                                                if remaining_capacity > 0:
                                                    ntc_rows = []
                                                    if (
                                                        remark_col
                                                        and remark_col
                                                        in processed_df.columns
                                                    ):
                                                        for idx in processed_df.index:
                                                            # Skip already allocated rows
                                                            if idx in [
                                                                i
                                                                for ag in agent_allocations
                                                                for i in ag[
                                                                    "row_indices"
                                                                ]
                                                            ]:
                                                                continue
                                                            # Skip rows with secondary insurance (already allocated)
                                                            if (
                                                                secondary_insurance_col
                                                                and pd.notna(
                                                                    processed_df.at[
                                                                        idx,
                                                                        secondary_insurance_col,
                                                                    ]
                                                                )
                                                            ):
                                                                secondary_val = str(
                                                                    processed_df.at[
                                                                        idx,
                                                                        secondary_insurance_col,
                                                                    ]
                                                                ).strip()
                                                                if (
                                                                    secondary_val
                                                                    and secondary_val.lower()
                                                                    != "nan"
                                                                ):
                                                                    continue
                                                            # Check if remark is NTC
                                                            if pd.notna(
                                                                processed_df.at[
                                                                    idx, remark_col
                                                                ]
                                                            ):
                                                                row_remark = (
                                                                    str(
                                                                        processed_df.at[
                                                                            idx,
                                                                            remark_col,
                                                                        ]
                                                                    )
                                                                    .strip()
                                                                    .upper()
                                                                )
                                                                if row_remark == "NTC":
                                                                    ntc_rows.append(idx)

                                                    # Allocate NTC rows until capacity is full
                                                    # Limit: No agent can receive more than 15 NTC rows
                                                    if ntc_rows:
                                                        allocated_count = 0
                                                        ntc_limit_remaining = (
                                                            15
                                                            - agent.get(
                                                                "ntc_allocated", 0
                                                            )
                                                        )
                                                        while (
                                                            remaining_capacity > 0
                                                            and allocated_count
                                                            < len(ntc_rows)
                                                            and ntc_limit_remaining > 0
                                                        ):
                                                            take = min(
                                                                remaining_capacity,
                                                                len(ntc_rows)
                                                                - allocated_count,
                                                                ntc_limit_remaining,
                                                            )
                                                            if take > 0:
                                                                slice_rows = ntc_rows[
                                                                    allocated_count : allocated_count
                                                                    + take
                                                                ]
                                                                # Use safe extend function to filter out "Not to work" rows
                                                                actual_allocated = safe_extend_row_indices(
                                                                    agent,
                                                                    slice_rows,
                                                                    processed_df,
                                                                    remark_col,
                                                                    agent["name"],
                                                                )
                                                                allocated_count += take  # Move forward in the list by take positions
                                                                agent[
                                                                    "ntc_allocated"
                                                                ] = (
                                                                    agent.get(
                                                                        "ntc_allocated",
                                                                        0,
                                                                    )
                                                                    + actual_allocated
                                                                )  # Track actual NTC rows allocated
                                                                remaining_capacity = (
                                                                    agent["capacity"]
                                                                    - agent["allocated"]
                                                                )
                                                                ntc_limit_remaining = (
                                                                    15
                                                                    - agent.get(
                                                                        "ntc_allocated",
                                                                        0,
                                                                    )
                                                                )
                                                            else:
                                                                break

                        # Also ensure "Sec + Single" agents get same insurance company rows globally after Step 3.6
                        # This handles cases where agents didn't get enough rows in the carrier-specific loop above
                        sec_single_agents_global = [
                            a
                            for a in agent_allocations
                            if a.get("has_sec_single_preference", False)
                            and (a["capacity"] - a["allocated"]) > 0
                        ]

                        if sec_single_agents_global:
                            for agent in sec_single_agents_global:
                                remaining_capacity = (
                                    agent["capacity"] - agent["allocated"]
                                )
                                if remaining_capacity <= 0:
                                    continue

                                assigned_ins = agent.get("assigned_insurance")
                                if not assigned_ins:
                                    continue  # Skip if no assigned insurance (should have been set from secondary allocation)

                                # Find unallocated rows with the same insurance company
                                same_insurance_rows = []
                                agent_insurance_list = agent.get(
                                    "insurance_companies", []
                                )

                                for idx in processed_df.index:
                                    # Skip already allocated rows
                                    if idx in [
                                        i
                                        for ag in agent_allocations
                                        for i in ag["row_indices"]
                                    ]:
                                        continue

                                    # Skip rows with secondary insurance (already allocated)
                                    if secondary_insurance_col and pd.notna(
                                        processed_df.at[idx, secondary_insurance_col]
                                    ):
                                        secondary_val = str(
                                            processed_df.at[
                                                idx, secondary_insurance_col
                                            ]
                                        ).strip()
                                        if (
                                            secondary_val
                                            and secondary_val.lower() != "nan"
                                        ):
                                            continue

                                    # Get insurance company from "Dental Primary Ins Carr" column
                                    if insurance_carrier_col and pd.notna(
                                        processed_df.at[idx, insurance_carrier_col]
                                    ):
                                        row_insurance = str(
                                            processed_df.at[idx, insurance_carrier_col]
                                        ).strip()

                                        # First check: row insurance must match assigned insurance
                                        if row_insurance == assigned_ins:
                                            # Second check: row insurance must be in agent's "Insurance List" (capabilities)
                                            can_work = False
                                            if agent.get("is_senior", False):
                                                # Senior agents can work with any insurance company
                                                can_work = True
                                            elif not agent_insurance_list:
                                                # If no specific companies listed, can work with any
                                                can_work = True
                                            else:
                                                # Check if row insurance matches/is in agent's Insurance List (capabilities)
                                                row_insurance_lower = (
                                                    row_insurance.lower()
                                                )
                                                for comp in agent_insurance_list:
                                                    comp_lower = comp.lower()
                                                    if (
                                                        row_insurance_lower
                                                        in comp_lower
                                                        or comp_lower
                                                        in row_insurance_lower
                                                        or row_insurance == comp
                                                    ):
                                                        can_work = True
                                                        break

                                            if can_work:
                                                same_insurance_rows.append(idx)

                                # Allocate same insurance rows until capacity is full
                                if same_insurance_rows:
                                    allocated_count = 0
                                    while (
                                        remaining_capacity > 0
                                        and allocated_count < len(same_insurance_rows)
                                    ):
                                        take = min(
                                            remaining_capacity,
                                            len(same_insurance_rows) - allocated_count,
                                        )
                                        if take > 0:
                                            slice_rows = same_insurance_rows[
                                                allocated_count : allocated_count + take
                                            ]
                                            # Use safe extend function to filter out "Not to work" rows
                                            actual_allocated = safe_extend_row_indices(
                                                agent,
                                                slice_rows,
                                                processed_df,
                                                remark_col,
                                                agent["name"],
                                                appointment_date_col,
                                            )
                                            allocated_count += take
                                            remaining_capacity = (
                                                agent["capacity"] - agent["allocated"]
                                            )
                                        else:
                                            break

                        # Step 3.7: Global Single Allocation - Allocate same insurance company rows to "Single" preference agents
                        # This ensures agents with "Single" preference (not "Sec + Single") get same insurance company rows to fill their capacity
                        # Exclude PB preference agents (should only get NTBP rows in Step 2.5)
                        single_preference_agents = [
                            a
                            for a in agent_allocations
                            if a.get("has_single_preference", False)
                            and not a.get(
                                "has_sec_preference", False
                            )  # Exclude "Sec + Single" (handled in Step 3.6)
                            and not a.get(
                                "has_pb_preference", False
                            )  # Exclude PB preference agents
                            and (a["capacity"] - a["allocated"]) > 0
                        ]

                        if single_preference_agents:
                            for agent in single_preference_agents:
                                remaining_capacity = (
                                    agent["capacity"] - agent["allocated"]
                                )
                                if remaining_capacity <= 0:
                                    continue

                                # Special handling for Abdul Hakim, Alisha Mulla, and Iqra Patel: Group by insurance company, then allocate from one matching company
                                agent_name = agent.get("name", "").strip()
                                agent_name_lower = agent_name.lower()
                                is_abdul_hakim = agent_name_lower == "abdul hakim"
                                is_alisha_mulla = agent_name_lower == "alisha mulla"
                                is_iqra_patel = agent_name_lower == "iqra patel"
                                needs_special_grouping = (
                                    is_abdul_hakim or is_alisha_mulla or is_iqra_patel
                                )

                                if needs_special_grouping:
                                    # Special logic for Abdul Hakim, Alisha Mulla, and Iqra Patel: Group unallocated rows by insurance company
                                    # Then find matching insurance companies from their capabilities and allocate from one
                                    # Uses formatted insurance names for better matching
                                    agent_insurance_list = agent.get(
                                        "insurance_companies", []
                                    )

                                    # Group unallocated rows by insurance company from "Dental Primary Ins" column
                                    insurance_groups = {}
                                    all_allocated_indices = [
                                        i
                                        for ag in agent_allocations
                                        for i in ag["row_indices"]
                                    ]

                                    for idx in processed_df.index:
                                        # Skip already allocated rows
                                        if idx in all_allocated_indices:
                                            continue

                                        # Skip rows with secondary insurance (those go to "Sec + X" agents)
                                        if secondary_insurance_col and pd.notna(
                                            processed_df.at[
                                                idx, secondary_insurance_col
                                            ]
                                        ):
                                            secondary_val = str(
                                                processed_df.at[
                                                    idx, secondary_insurance_col
                                                ]
                                            ).strip()
                                            if (
                                                secondary_val
                                                and secondary_val.lower() != "nan"
                                            ):
                                                continue

                                        # Skip rows with "Not to work" remark
                                        if remark_col and pd.notna(
                                            processed_df.at[idx, remark_col]
                                        ):
                                            remark_val = (
                                                str(processed_df.at[idx, remark_col])
                                                .strip()
                                                .upper()
                                            )
                                            if (
                                                "NOT TO WORK" in remark_val
                                                or remark_val == "NOT TO WORK"
                                            ):
                                                continue

                                        # Get insurance company from "Dental Primary Ins Carr" column
                                        if insurance_carrier_col and pd.notna(
                                            processed_df.at[idx, insurance_carrier_col]
                                        ):
                                            row_insurance_raw = str(
                                                processed_df.at[
                                                    idx, insurance_carrier_col
                                                ]
                                            ).strip()

                                            # Format the row insurance company name for consistent matching
                                            row_insurance = (
                                                format_insurance_company_name(
                                                    row_insurance_raw
                                                )
                                            )
                                            if not row_insurance:
                                                row_insurance = row_insurance_raw

                                            # Use improved matching function that handles formatting variations
                                            can_work = check_insurance_match(
                                                row_insurance_raw,
                                                agent_insurance_list,
                                                agent.get("is_senior", False),
                                                agent.get("name"),
                                            )

                                            if can_work:
                                                # Group by formatted insurance company name for consistency
                                                if (
                                                    row_insurance
                                                    not in insurance_groups
                                                ):
                                                    insurance_groups[row_insurance] = []
                                                insurance_groups[row_insurance].append(
                                                    idx
                                                )

                                    # Find the insurance company with the most available rows that matches capabilities
                                    # and allocate from that company until capacity is full
                                    assigned_ins = agent.get("assigned_insurance")
                                    if not assigned_ins and insurance_groups:
                                        # Choose the insurance company with the most available rows
                                        # This ensures we fill capacity efficiently
                                        assigned_ins = max(
                                            insurance_groups.keys(),
                                            key=lambda k: len(insurance_groups[k]),
                                        )
                                        agent["assigned_insurance"] = assigned_ins

                                    # If already assigned, use that insurance company
                                    if (
                                        assigned_ins
                                        and assigned_ins in insurance_groups
                                    ):
                                        same_insurance_rows = insurance_groups[
                                            assigned_ins
                                        ]
                                    else:
                                        same_insurance_rows = []

                                else:
                                    # Standard Single allocation logic for other agents
                                    # Get the assigned insurance company (should be set from previous allocations)
                                    # If not set, find first row where "Dental Primary Ins Carr" matches agent's "Insurance List" capabilities
                                    assigned_ins = agent.get("assigned_insurance")
                                if not assigned_ins:
                                    agent_insurance_list = agent.get(
                                        "insurance_companies", []
                                    )

                                    # Find first available row where insurance from "Dental Primary Ins Carr" matches agent's "Insurance List"
                                    for idx in processed_df.index:
                                        # Skip already allocated rows
                                        if idx in [
                                            i
                                            for ag in agent_allocations
                                            for i in ag["row_indices"]
                                        ]:
                                            continue

                                        # Skip rows with secondary insurance (those go to "Sec + X" agents)
                                        if secondary_insurance_col and pd.notna(
                                            processed_df.at[
                                                idx, secondary_insurance_col
                                            ]
                                        ):
                                            secondary_val = str(
                                                processed_df.at[
                                                    idx, secondary_insurance_col
                                                ]
                                            ).strip()
                                            if (
                                                secondary_val
                                                and secondary_val.lower() != "nan"
                                            ):
                                                continue

                                        # Get insurance company from "Dental Primary Ins Carr" column
                                        if insurance_carrier_col and pd.notna(
                                            processed_df.at[idx, insurance_carrier_col]
                                        ):
                                            row_insurance = str(
                                                processed_df.at[
                                                    idx, insurance_carrier_col
                                                ]
                                            ).strip()

                                            # Check if row insurance is in agent's "Insurance List" (capabilities)
                                            can_work = False
                                            if agent.get("is_senior", False):
                                                # Senior agents can work with any insurance company
                                                can_work = True
                                            elif not agent_insurance_list:
                                                # If no specific companies listed, can work with any
                                                can_work = True
                                            else:
                                                # Check if row insurance matches/is in agent's Insurance List (capabilities)
                                                row_insurance_lower = (
                                                    row_insurance.lower()
                                                )
                                                for comp in agent_insurance_list:
                                                    comp_lower = comp.lower()
                                                    if (
                                                        row_insurance_lower
                                                        in comp_lower
                                                        or comp_lower
                                                        in row_insurance_lower
                                                        or row_insurance == comp
                                                    ):
                                                        can_work = True
                                                        break

                                            if can_work:
                                                # Found matching insurance - assign this insurance company to agent
                                                assigned_ins = row_insurance
                                                agent["assigned_insurance"] = (
                                                    assigned_ins
                                                )
                                                break

                                    # For agents without special grouping logic, find unallocated rows with the same insurance company
                                    if assigned_ins and not needs_special_grouping:
                                        same_insurance_rows = []
                                    agent_insurance_list = agent.get(
                                        "insurance_companies", []
                                    )

                                    for idx in processed_df.index:
                                        # Skip already allocated rows
                                        if idx in [
                                            i
                                            for ag in agent_allocations
                                            for i in ag["row_indices"]
                                        ]:
                                            continue

                                        # Skip rows with secondary insurance (those go to "Sec + X" agents)
                                        if secondary_insurance_col and pd.notna(
                                            processed_df.at[
                                                idx, secondary_insurance_col
                                            ]
                                        ):
                                            secondary_val = str(
                                                processed_df.at[
                                                    idx, secondary_insurance_col
                                                ]
                                            ).strip()
                                            if (
                                                secondary_val
                                                and secondary_val.lower() != "nan"
                                            ):
                                                continue

                                        # Skip rows with "Not to work" remark
                                        if remark_col and pd.notna(
                                            processed_df.at[idx, remark_col]
                                        ):
                                            remark_val = (
                                                str(processed_df.at[idx, remark_col])
                                                .strip()
                                                .upper()
                                            )
                                            if (
                                                "NOT TO WORK" in remark_val
                                                or remark_val == "NOT TO WORK"
                                            ):
                                                continue

                                        # Get insurance company from "Dental Primary Ins Carr" column
                                        if insurance_carrier_col and pd.notna(
                                            processed_df.at[idx, insurance_carrier_col]
                                        ):
                                            row_insurance = str(
                                                processed_df.at[
                                                    idx, insurance_carrier_col
                                                ]
                                            ).strip()

                                            # First check: row insurance must match assigned insurance
                                            if row_insurance == assigned_ins:
                                                # Second check: row insurance must be in agent's "Insurance List" (capabilities)
                                                can_work = False
                                                if agent.get("is_senior", False):
                                                    # Senior agents can work with any insurance company
                                                    can_work = True
                                                elif not agent_insurance_list:
                                                    # If no specific companies listed, can work with any
                                                    can_work = True
                                                else:
                                                    # Check if row insurance matches/is in agent's Insurance List (capabilities)
                                                    # Also check if this is "Afreen Ansari" and insurance is in her additional list
                                                    is_afreen_ansari = (
                                                        agent.get("name")
                                                        == "Afreen Ansari"
                                                    )

                                                    # Check against Afreen Ansari's additional insurance list first
                                                    if is_afreen_ansari:
                                                        row_insurance_lower = (
                                                            row_insurance.lower()
                                                        )
                                                        for (
                                                            allowed_insurance
                                                        ) in AFREEN_ANSARI_ADDITIONAL_INSURANCE:
                                                            allowed_lower = (
                                                                str(allowed_insurance)
                                                                .strip()
                                                                .lower()
                                                            )
                                                            if (
                                                                row_insurance_lower
                                                                == allowed_lower
                                                                or allowed_lower
                                                                in row_insurance_lower
                                                                or row_insurance_lower
                                                                in allowed_lower
                                                            ):
                                                                can_work = True
                                                                break

                                                    # If not matched yet, check against agent's regular insurance list
                                                    if not can_work:
                                                        row_insurance_lower = (
                                                            row_insurance.lower()
                                                        )
                                                        for (
                                                            comp
                                                        ) in agent_insurance_list:
                                                            comp_lower = comp.lower()
                                                        if (
                                                            row_insurance_lower
                                                            in comp_lower
                                                            or comp_lower
                                                            in row_insurance_lower
                                                            or row_insurance == comp
                                                        ):
                                                            can_work = True
                                                            break

                                                if can_work:
                                                    same_insurance_rows.append(idx)
                                    else:
                                        same_insurance_rows = []

                                    # Allocate same insurance rows until capacity is full (for both Abdul Hakim and other agents)
                                    if same_insurance_rows:
                                        # Final safety check: Filter out any "Not to work" rows
                                        filtered_same_insurance_rows = [
                                            idx
                                            for idx in same_insurance_rows
                                            if not should_skip_row_for_allocation(
                                                idx, processed_df, remark_col
                                            )
                                        ]

                                        allocated_count = 0
                                        while (
                                            remaining_capacity > 0
                                            and allocated_count
                                            < len(filtered_same_insurance_rows)
                                        ):
                                            take = min(
                                                remaining_capacity,
                                                len(filtered_same_insurance_rows)
                                                - allocated_count,
                                            )
                                            if take > 0:
                                                slice_rows = filtered_same_insurance_rows[
                                                    allocated_count : allocated_count
                                                    + take
                                                ]
                                                # Use safe extend function to filter out "Not to work" rows
                                                actual_allocated = (
                                                    safe_extend_row_indices(
                                                        agent,
                                                        slice_rows,
                                                        processed_df,
                                                        remark_col,
                                                        agent["name"],
                                                    )
                                                )
                                                allocated_count += take
                                                remaining_capacity = (
                                                    agent["capacity"]
                                                    - agent["allocated"]
                                                )
                                            else:
                                                break

                        # Step 3.8: Global Mix Allocation - Allocate multiple insurance company rows to "Mix" preference agents
                        # Agents with "Mix" preference should get rows from multiple insurance companies (unlike "Single" which gets only one)
                        # Exclude "Sec + Mix" agents (handled in Step 3.6), "Mix + NTC" agents (handled in Step 3.5.5), and "PB" preference agents (should only get NTBP rows)
                        mix_preference_agents = [
                            a
                            for a in agent_allocations
                            if a.get("has_mix_preference", False)
                            and not a.get(
                                "has_sec_preference", False
                            )  # Exclude "Sec + Mix" (handled in Step 3.6)
                            and not (
                                a.get("has_ntc_preference", False)
                            )  # Exclude "Mix + NTC" (handled in Step 3.5.5)
                            and not a.get(
                                "has_pb_preference", False
                            )  # Exclude "PB" preference agents (should only get NTBP rows in Step 2.5)
                            and (a["capacity"] - a["allocated"]) > 0
                        ]

                        if mix_preference_agents and insurance_carrier_col:
                            for agent in mix_preference_agents:
                                remaining_capacity = (
                                    agent["capacity"] - agent["allocated"]
                                )
                                if remaining_capacity <= 0:
                                    continue

                                agent_insurance_list = agent.get(
                                    "insurance_companies", []
                                )

                                # Find unallocated rows from multiple insurance companies that match agent's capabilities
                                mix_rows = []

                                for idx in processed_df.index:
                                    # Skip already allocated rows
                                    if idx in [
                                        i
                                        for ag in agent_allocations
                                        for i in ag["row_indices"]
                                    ]:
                                        continue

                                    # Skip rows with secondary insurance (those go to "Sec + X" agents)
                                    if secondary_insurance_col and pd.notna(
                                        processed_df.at[idx, secondary_insurance_col]
                                    ):
                                        secondary_val = str(
                                            processed_df.at[
                                                idx, secondary_insurance_col
                                            ]
                                        ).strip()
                                        if (
                                            secondary_val
                                            and secondary_val.lower() != "nan"
                                        ):
                                            continue

                                    # Skip NTBP, NTC, and "Not to work" rows
                                    if remark_col and pd.notna(
                                        processed_df.at[idx, remark_col]
                                    ):
                                        remark_val = (
                                            str(processed_df.at[idx, remark_col])
                                            .strip()
                                            .upper()
                                        )
                                        if remark_val == "NTBP" or remark_val == "NTC":
                                            continue
                                        # Skip rows with "Not to work" remark
                                        if (
                                            "NOT TO WORK" in remark_val
                                            or remark_val == "NOT TO WORK"
                                        ):
                                            continue

                                    # Get insurance company from "Dental Primary Ins Carr" column
                                    if pd.notna(
                                        processed_df.at[idx, insurance_carrier_col]
                                    ):
                                        row_insurance_raw = str(
                                            processed_df.at[idx, insurance_carrier_col]
                                        ).strip()

                                        # Use improved matching function that handles formatting variations and "DD All"
                                        can_work = check_insurance_match(
                                            row_insurance_raw,
                                            agent_insurance_list,
                                            agent.get("is_senior", False),
                                            agent.get("name"),
                                        )

                                        if can_work:
                                            mix_rows.append(idx)

                                # Allocate mix rows until capacity is full
                                if mix_rows:
                                    allocated_count = 0
                                    while (
                                        remaining_capacity > 0
                                        and allocated_count < len(mix_rows)
                                    ):
                                        take = min(
                                            remaining_capacity,
                                            len(mix_rows) - allocated_count,
                                        )
                                        if take > 0:
                                            slice_rows = mix_rows[
                                                allocated_count : allocated_count + take
                                            ]
                                            # Use safe extend function to filter out "Not to work" rows
                                            actual_allocated = safe_extend_row_indices(
                                                agent,
                                                slice_rows,
                                                processed_df,
                                                remark_col,
                                                agent["name"],
                                                appointment_date_col,
                                            )
                                            allocated_count += take
                                            remaining_capacity = (
                                                agent["capacity"] - agent["allocated"]
                                            )
                                        else:
                                            break

                        # Step 3: FIRST PRIORITY - Allocate First Priority matched work to senior agents FIRST
                        # This takes precedence over unmatched insurance

                        # Check senior agent remaining capacity before Step 3
                        senior_capacity_before = sum(
                            a["capacity"] - a["allocated"] for a in senior_agents
                        )

                        # ONLY process First Priority matched work for senior agents
                        priority = "First Priority"
                        # Collect all unallocated First Priority matched work across all insurance carriers
                        priority_work = (
                            []
                        )  # List of (insurance_carrier, row_index) tuples
                        for (
                            insurance_carrier,
                            priority_data,
                        ) in matched_data_by_insurance_priority.items():
                            if priority in priority_data:
                                row_indices = priority_data[priority]
                                # Get unallocated indices for this insurance carrier and priority
                                unallocated_indices = [
                                    idx
                                    for idx in row_indices
                                    if idx
                                    not in [
                                        i
                                        for ag in agent_allocations
                                        for i in ag["row_indices"]
                                    ]
                                ]
                                for idx in unallocated_indices:
                                    # Skip "Not to work" rows before adding to list
                                    if not should_skip_row_for_allocation(
                                        idx, processed_df, remark_col
                                    ):
                                        priority_work.append((insurance_carrier, idx))

                        if priority_work:
                            pass

                            # Allocate all First Priority matched work to senior agents, maximizing capacity utilization
                            # But apply Domain/Remark rule: NTBP rows only to PB agents, PB agents only get NTBP rows

                            # Separate priority work into NTBP, NTC, and other rows
                            # IMPORTANT: NTBP rows should have been allocated in Step 2.5 to PB preference agents
                            # IMPORTANT: NTC rows should have been allocated in Step 3.5 to NTC preference agents
                            # Skip NTBP and NTC rows here - they should only go to their respective preference agents
                            ntbp_priority_work = []
                            ntc_priority_work = []
                            non_special_priority_work = []

                            if remark_col and remark_col in processed_df.columns:
                                for insurance_carrier, row_idx in priority_work:
                                    # Skip if already allocated (should have been allocated in Step 2.5 or 3.5)
                                    if row_idx in [
                                        i
                                        for ag in agent_allocations
                                        for i in ag["row_indices"]
                                    ]:
                                        continue

                                    row_remark = None
                                    if pd.notna(processed_df.at[row_idx, remark_col]):
                                        row_remark = (
                                            str(processed_df.at[row_idx, remark_col])
                                            .strip()
                                            .upper()
                                        )

                                    if row_remark == "NTBP":
                                        # NTBP rows should only go to PB preference agents (allocated in Step 2.5)
                                        # Skip them here - don't allocate to senior agents
                                        continue
                                    elif row_remark == "NTC":
                                        # NTC rows should only go to NTC preference agents (allocated in Step 3.5)
                                        # Skip them here - don't allocate to senior agents
                                        continue
                                    else:
                                        non_special_priority_work.append(
                                            (insurance_carrier, row_idx)
                                        )
                            else:
                                # If no remark column, all rows are non-special
                                non_special_priority_work = priority_work.copy()

                            # Separate senior agents into PB and non-PB
                            # Check both domain == 'PB' and allocation preference contains 'PB'
                            pb_senior_agents = [
                                a
                                for a in senior_agents
                                if (
                                    (
                                        a.get("domain")
                                        and str(a.get("domain")).strip().upper() == "PB"
                                    )
                                    or a.get("has_pb_preference", False)
                                )
                            ]
                            non_pb_senior_agents = [
                                a
                                for a in senior_agents
                                if not (
                                    (
                                        a.get("domain")
                                        and str(a.get("domain")).strip().upper() == "PB"
                                    )
                                    or a.get("has_pb_preference", False)
                                )
                            ]

                            # NTBP rows should have been allocated in Step 2.5 to PB preference agents only
                            # Skip NTBP allocation here - ntbp_priority_work should be empty since we skip NTBP rows above
                            # NTBP rows are NOT allocated in Step 3 - they should only go to PB preference agents (Step 2.5)

                            # Allocate non-NTBP priority work only to non-PB senior agents
                            work_idx = 0
                            while work_idx < len(non_special_priority_work):
                                available_non_pb_seniors = [
                                    a
                                    for a in non_pb_senior_agents
                                    if (a["capacity"] - a["allocated"]) > 0
                                ]
                                if not available_non_pb_seniors:
                                    break

                                available_non_pb_seniors.sort(
                                    key=lambda x: x["capacity"] - x["allocated"],
                                    reverse=True,
                                )

                                for senior_agent in available_non_pb_seniors:
                                    if work_idx >= len(non_special_priority_work):
                                        break

                                    available_capacity = (
                                        senior_agent["capacity"]
                                        - senior_agent["allocated"]
                                    )
                                    if available_capacity <= 0:
                                        continue

                                    # Collect rows that can be allocated to this agent (checking "do not allocate" list and "Single" preference)
                                    assignable_rows = []
                                    rows_processed = 0
                                    for i in range(
                                        work_idx,
                                        min(
                                            work_idx + available_capacity,
                                            len(non_special_priority_work),
                                        ),
                                    ):
                                        insurance_carrier, row_idx = (
                                            non_special_priority_work[i]
                                        )

                                        # For agents with "Single" preference, only allow rows from their assigned insurance
                                        if senior_agent.get(
                                            "has_single_preference", False
                                        ):
                                            assigned_ins = senior_agent.get(
                                                "assigned_insurance"
                                            )
                                            # If agent already has an assigned insurance, only allow that insurance
                                            if (
                                                assigned_ins is not None
                                                and assigned_ins != insurance_carrier
                                            ):
                                                rows_processed += 1
                                                continue  # Skip this row - agent with "Single" already has different insurance

                                        rows_processed += 1

                                        # Check if this insurance company is in the agent's "do not allocate" list
                                        should_not_allocate = False
                                        if senior_agent.get(
                                            "insurance_do_not_allocate"
                                        ):
                                            insurance_carrier_str = (
                                                str(insurance_carrier)
                                                if insurance_carrier
                                                else ""
                                            )
                                            for do_not_allocate_comp in senior_agent[
                                                "insurance_do_not_allocate"
                                            ]:
                                                do_not_allocate_comp_str = (
                                                    str(do_not_allocate_comp)
                                                    if do_not_allocate_comp
                                                    else ""
                                                )
                                                if (
                                                    insurance_carrier_str
                                                    and do_not_allocate_comp_str
                                                ):
                                                    if (
                                                        insurance_carrier_str.lower()
                                                        in do_not_allocate_comp_str.lower()
                                                        or do_not_allocate_comp_str.lower()
                                                        in insurance_carrier_str.lower()
                                                        or insurance_carrier_str
                                                        == do_not_allocate_comp_str
                                                    ):
                                                        should_not_allocate = True
                                                        break

                                        # Only add row if agent can be allocated this insurance company
                                        # For agents with "Single" preference, ensure they only get one insurance company
                                        if not should_not_allocate:
                                            # If agent has "Single" preference and already has assigned insurance,
                                            # only allow rows from that same insurance
                                            if senior_agent.get(
                                                "has_single_preference", False
                                            ):
                                                assigned_ins = senior_agent.get(
                                                    "assigned_insurance"
                                                )
                                                if assigned_ins is None:
                                                    # Agent doesn't have assigned insurance yet - can assign this one
                                                    assignable_rows.append(
                                                        (insurance_carrier, row_idx)
                                                    )
                                                elif assigned_ins == insurance_carrier:
                                                    # Agent already has this insurance - can assign more rows
                                                    assignable_rows.append(
                                                        (insurance_carrier, row_idx)
                                                    )
                                                # else: agent has different insurance - skip this row (already handled above)
                                            else:
                                                # Agent doesn't have "Single" preference - normal allocation
                                                assignable_rows.append(
                                                    (insurance_carrier, row_idx)
                                                )

                                    rows_to_assign = len(assignable_rows)

                                    if rows_to_assign > 0:
                                        agent_id = senior_agent.get(
                                            "id", senior_agent.get("name", "Unknown")
                                        )
                                        # Set assigned_insurance for agents with "Single" preference when they get their first allocation
                                        if (
                                            senior_agent.get(
                                                "has_single_preference", False
                                            )
                                            and senior_agent.get("assigned_insurance")
                                            is None
                                        ):
                                            # Get the insurance carrier from the first assignable row
                                            if assignable_rows:
                                                first_insurance, _ = assignable_rows[0]
                                                senior_agent["assigned_insurance"] = (
                                                    first_insurance
                                                )

                                        for (
                                            insurance_carrier,
                                            row_idx,
                                        ) in assignable_rows:
                                            # Safety check: Verify this is not a "Not to work" row before allocating
                                            if not should_skip_row_for_allocation(
                                                row_idx, processed_df, remark_col
                                            ):
                                                senior_agent["row_indices"].append(
                                                    row_idx
                                                )
                                                senior_agent["allocated"] += 1
                                                processed_df.at[
                                                    row_idx, "Agent Name"
                                                ] = senior_agent["name"]

                                            if agent_id in ins_group_allocations:
                                                insurance_carrier_upper = (
                                                    insurance_carrier.upper().strip()
                                                )
                                                if any(
                                                    insurance_carrier_upper
                                                    == ic.upper().strip()
                                                    for ic in DD_INS_GROUP
                                                ):
                                                    ins_group_allocations[agent_id] += 1
                                            if agent_id in toolkit_group_allocations:
                                                insurance_carrier_upper = (
                                                    insurance_carrier.upper().strip()
                                                )
                                                if any(
                                                    insurance_carrier_upper
                                                    == ic.upper().strip()
                                                    for ic in DD_TOOLKIT_GROUP
                                                ):
                                                    toolkit_group_allocations[
                                                        agent_id
                                                    ] += 1

                                        senior_agent["allocated"] += rows_to_assign

                                    # Increment work_idx by number of rows processed (including skipped ones)
                                    work_idx += rows_processed

                                # Log progress
                                if work_idx % 50 == 0 and work_idx < len(
                                    non_special_priority_work
                                ):
                                    pass

                        else:
                            pass

                        # Note: Second/Third Priority matched work will be allocated to non-seniors in Step 5
                        senior_capacity_after = sum(
                            a["capacity"] - a["allocated"] for a in senior_agents
                        )

                        # Step 4: Allocate unmatched insurance companies to senior agents (after First Priority matched work)
                        # Also include "Afreen Ansari" even if not marked as senior
                        # Check if we have unmatched insurance and (senior agents OR Afreen Ansari)
                        has_afreen_ansari = any(
                            a["name"] == "Afreen Ansari" for a in agent_allocations
                        )
                        if unmatched_insurance_companies and (
                            senior_agents or has_afreen_ansari
                        ):
                            for (
                                insurance_carrier,
                                priority_data,
                            ) in unmatched_data_by_priority.items():
                                # Process by priority order
                                for priority in [
                                    "First Priority",
                                    "Second Priority",
                                    "Third Priority",
                                ]:
                                    if priority in priority_data:
                                        row_indices = priority_data[priority]

                                        # Only senior agents can handle unmatched insurance
                                        # Also include "Afreen Ansari" even if not marked as senior
                                        available_senior_agents = [
                                            a
                                            for a in senior_agents
                                            if (a["capacity"] - a["allocated"]) > 0
                                        ]

                                        # Also add "Afreen Ansari" if not already in the list and has capacity
                                        # Check if she's already in available_senior_agents by name
                                        afreen_in_available = any(
                                            a["name"] == "Afreen Ansari"
                                            for a in available_senior_agents
                                        )
                                        if not afreen_in_available:
                                            for agent in agent_allocations:
                                                if agent["name"] == "Afreen Ansari":
                                                    if (
                                                        agent["capacity"]
                                                        - agent["allocated"]
                                                    ) > 0:
                                                        # Add Afreen Ansari to available agents for unmatched insurance
                                                        # We'll use the original agent object but treat her as senior for unmatched insurance
                                                        available_senior_agents.append(
                                                            agent
                                                        )
                                                        break

                                        # Filter row indices based on Domain/Remark rule: NTBP rows only go to PB agents
                                        filtered_row_indices = []
                                        if (
                                            remark_col
                                            and remark_col in processed_df.columns
                                        ):
                                            for row_idx in row_indices:
                                                row_remark = None
                                                if pd.notna(
                                                    processed_df.at[row_idx, remark_col]
                                                ):
                                                    row_remark = (
                                                        str(
                                                            processed_df.at[
                                                                row_idx, remark_col
                                                            ]
                                                        )
                                                        .strip()
                                                        .upper()
                                                    )

                                                row_is_ntbp = row_remark == "NTBP"

                                                # For unmatched insurance, we'll filter agents later, but keep all rows for now
                                                # The actual filtering will happen when assigning to specific agents
                                                filtered_row_indices.append(row_idx)
                                        else:
                                            filtered_row_indices = row_indices

                                        if (
                                            available_senior_agents
                                            and filtered_row_indices
                                        ):
                                            # Filter to only agents with available capacity (double-check before allocation)
                                            available_senior_agents = [
                                                a
                                                for a in available_senior_agents
                                                if (a["capacity"] - a["allocated"]) > 0
                                            ]

                                            if not available_senior_agents:
                                                # No senior agents with capacity - skip this insurance company
                                                continue

                                            # Distribute unmatched insurance rows among senior agents by priority
                                            # Sort by remaining capacity (highest first)
                                            available_senior_agents.sort(
                                                key=lambda x: x["capacity"]
                                                - x["allocated"],
                                                reverse=True,
                                            )

                                            # Allocate to senior agents up to their capacity
                                            row_idx = 0
                                            for senior_agent in available_senior_agents:
                                                # Check if agent still has capacity before processing
                                                if (
                                                    senior_agent["capacity"]
                                                    - senior_agent["allocated"]
                                                ) <= 0:
                                                    continue  # Skip agents that are at capacity

                                                if row_idx >= len(filtered_row_indices):
                                                    break

                                                # For unmatched insurance, treat "Afreen Ansari" as if she can work with any insurance
                                                # Check if this is Afreen Ansari
                                                is_afreen_ansari = (
                                                    senior_agent.get("name")
                                                    == "Afreen Ansari"
                                                )

                                                # Check if this insurance company is in the agent's "do not allocate" list
                                                should_not_allocate = False
                                                if senior_agent.get(
                                                    "insurance_do_not_allocate"
                                                ):
                                                    for (
                                                        do_not_allocate_comp
                                                    ) in senior_agent[
                                                        "insurance_do_not_allocate"
                                                    ]:
                                                        if (
                                                            insurance_carrier.lower()
                                                            in do_not_allocate_comp.lower()
                                                            or do_not_allocate_comp.lower()
                                                            in insurance_carrier.lower()
                                                            or insurance_carrier
                                                            == do_not_allocate_comp
                                                        ):
                                                            should_not_allocate = True
                                                            break

                                                # Skip this agent if insurance company is in their "do not allocate" list
                                                if should_not_allocate:
                                                    continue

                                                available_capacity = (
                                                    senior_agent["capacity"]
                                                    - senior_agent["allocated"]
                                                )
                                                if available_capacity > 0:
                                                    # Filter rows based on Domain/Remark rule for this specific agent
                                                    agent_domain = senior_agent.get(
                                                        "domain"
                                                    )
                                                    agent_has_pb_pref = (
                                                        senior_agent.get(
                                                            "has_pb_preference", False
                                                        )
                                                    )
                                                    # Agent is PB if domain is 'PB' OR allocation preference contains 'PB'
                                                    agent_is_pb = (
                                                        agent_domain is not None
                                                        and agent_domain.upper() == "PB"
                                                    ) or agent_has_pb_pref

                                                    # Collect rows that match this agent's domain requirement and "Single" preference
                                                    matching_rows = []
                                                    for check_idx in range(
                                                        row_idx,
                                                        min(
                                                            row_idx
                                                            + available_capacity,
                                                            len(filtered_row_indices),
                                                        ),
                                                    ):
                                                        actual_row_idx = (
                                                            filtered_row_indices[
                                                                check_idx
                                                            ]
                                                        )

                                                        # For agents with "Single" preference, only allow rows from their assigned insurance
                                                        if senior_agent.get(
                                                            "has_single_preference",
                                                            False,
                                                        ):
                                                            assigned_ins = (
                                                                senior_agent.get(
                                                                    "assigned_insurance"
                                                                )
                                                            )
                                                            # Get the insurance carrier for this row
                                                            row_insurance = None
                                                            if (
                                                                insurance_carrier_col
                                                                and pd.notna(
                                                                    processed_df.at[
                                                                        actual_row_idx,
                                                                        insurance_carrier_col,
                                                                    ]
                                                                )
                                                            ):
                                                                row_insurance = str(
                                                                    processed_df.at[
                                                                        actual_row_idx,
                                                                        insurance_carrier_col,
                                                                    ]
                                                                ).strip()
                                                            # If agent already has an assigned insurance, only allow that insurance
                                                            if (
                                                                assigned_ins is not None
                                                                and row_insurance
                                                                is not None
                                                                and assigned_ins
                                                                != row_insurance
                                                            ):
                                                                continue  # Skip this row - agent with "Single" already has different insurance

                                                        row_remark = None
                                                        if (
                                                            remark_col
                                                            and remark_col
                                                            in processed_df.columns
                                                            and pd.notna(
                                                                processed_df.at[
                                                                    actual_row_idx,
                                                                    remark_col,
                                                                ]
                                                            )
                                                        ):
                                                            row_remark = (
                                                                str(
                                                                    processed_df.at[
                                                                        actual_row_idx,
                                                                        remark_col,
                                                                    ]
                                                                )
                                                                .strip()
                                                                .upper()
                                                            )

                                                        row_is_ntbp = (
                                                            row_remark == "NTBP"
                                                        )
                                                        row_is_ntc = row_remark == "NTC"

                                                        # IMPORTANT: NTBP rows should ONLY be allocated in Step 2.5 to PB preference agents
                                                        # IMPORTANT: NTC rows should ONLY be allocated in Step 3.5 to NTC preference agents
                                                        # Skip NTBP and NTC rows here - they should not be allocated in Step 4
                                                        if row_is_ntbp:
                                                            # Skip NTBP rows - they should only go to PB preference agents (Step 2.5)
                                                            continue
                                                        elif row_is_ntc:
                                                            # Skip NTC rows - they should only go to NTC preference agents (Step 3.5)
                                                            continue
                                                        # Skip rows with "Not to work" remark
                                                        elif (
                                                            "NOT TO WORK" in row_remark
                                                            or row_remark
                                                            == "NOT TO WORK"
                                                        ):
                                                            continue
                                                        elif (
                                                            not row_is_ntbp
                                                            and not row_is_ntc
                                                            and not agent_is_pb
                                                        ):
                                                            # Non-NTBP, non-NTC row and non-PB agent - match
                                                            matching_rows.append(
                                                                actual_row_idx
                                                            )
                                                        # Otherwise: non-NTBP, non-NTC row with PB agent - no match (PB agents should only get NTBP, but NTBP is handled in Step 2.5)

                                                    rows_to_assign = len(matching_rows)
                                                    if rows_to_assign > 0:
                                                        # Set assigned_insurance for agents with "Single" preference when they get their first allocation
                                                        if (
                                                            senior_agent.get(
                                                                "has_single_preference",
                                                                False,
                                                            )
                                                            and senior_agent.get(
                                                                "assigned_insurance"
                                                            )
                                                            is None
                                                        ):
                                                            senior_agent[
                                                                "assigned_insurance"
                                                            ] = insurance_carrier

                                                        agent_id = senior_agent.get(
                                                            "id",
                                                            senior_agent.get(
                                                                "name", "Unknown"
                                                            ),
                                                        )
                                                        # Track INS and Toolkit group allocations (case-insensitive)
                                                        insurance_carrier_upper = (
                                                            insurance_carrier.upper().strip()
                                                        )
                                                        for (
                                                            assigned_row_idx
                                                        ) in matching_rows:
                                                            if (
                                                                agent_id
                                                                in ins_group_allocations
                                                            ):
                                                                if any(
                                                                    insurance_carrier_upper
                                                                    == ic.upper().strip()
                                                                    for ic in DD_INS_GROUP
                                                                ):
                                                                    ins_group_allocations[
                                                                        agent_id
                                                                    ] += 1
                                                            if (
                                                                agent_id
                                                                in toolkit_group_allocations
                                                            ):
                                                                if any(
                                                                    insurance_carrier_upper
                                                                    == ic.upper().strip()
                                                                    for ic in DD_TOOLKIT_GROUP
                                                                ):
                                                                    toolkit_group_allocations[
                                                                        agent_id
                                                                    ] += 1

                                                        # Use safe extend function to filter out "Not to work" rows
                                                        actual_allocated = (
                                                            safe_extend_row_indices(
                                                                senior_agent,
                                                                matching_rows,
                                                                processed_df,
                                                                remark_col,
                                                                senior_agent["name"],
                                                            )
                                                        )
                                                        row_idx += len(matching_rows)

                                                        # Check if agent is now at capacity - if so, break to next agent
                                                        if (
                                                            senior_agent["capacity"]
                                                            - senior_agent["allocated"]
                                                        ) <= 0:
                                                            break  # Agent is at capacity, move to next agent

                                            # If there are remaining unmatched rows that couldn't fit in senior capacity
                                            # they will be handled later or logged
                                            if row_idx < len(row_indices):
                                                pass

                        # Step 5: Allocate remaining matched insurance companies to capable agents (normal allocation)

                        for (
                            insurance_carrier,
                            priority_data,
                        ) in matched_data_by_insurance_priority.items():
                            # Process First Priority first (senior agents get priority)
                            for priority in [
                                "First Priority",
                                "Second Priority",
                                "Third Priority",
                            ]:
                                if priority in priority_data:
                                    row_indices = priority_data[priority]

                                    # Filter out already allocated rows
                                    unallocated_row_indices = [
                                        idx
                                        for idx in row_indices
                                        if idx
                                        not in [
                                            i
                                            for ag in agent_allocations
                                            for i in ag["row_indices"]
                                        ]
                                    ]

                                    # Exclude NTBP rows - they should only go to PB preference agents (allocated in Step 2.5)
                                    if (
                                        remark_col
                                        and remark_col in processed_df.columns
                                    ):
                                        filtered_indices = []
                                        for idx in unallocated_row_indices:
                                            if pd.notna(
                                                processed_df.at[idx, remark_col]
                                            ):
                                                row_remark = (
                                                    str(
                                                        processed_df.at[idx, remark_col]
                                                    )
                                                    .strip()
                                                    .upper()
                                                )
                                                if row_remark == "NTBP":
                                                    # Skip NTBP rows - they should only go to PB preference agents
                                                    continue
                                            filtered_indices.append(idx)
                                        unallocated_row_indices = filtered_indices

                                    if not unallocated_row_indices:
                                        continue

                                    # For First Priority and Unknown insurance, ONLY consider senior agents
                                    # For Second/Third Priority, EXCLUDE senior agents (they should only get First Priority)
                                    if (
                                        priority == "First Priority"
                                        or insurance_carrier == "Unknown"
                                    ):
                                        # First Priority and Unknown: ONLY senior agents
                                        agents_to_check = [
                                            a
                                            for a in agent_allocations
                                            if a["is_senior"]
                                        ]
                                        if not agents_to_check:
                                            continue
                                    else:
                                        # Second/Third Priority: EXCLUDE senior agents - they should only handle First Priority work
                                        agents_to_check = [
                                            a
                                            for a in agent_allocations
                                            if not a["is_senior"]
                                        ]
                                        if not agents_to_check:
                                            continue

                                    # Find agents who can work with this insurance company
                                    capable_agents = []
                                    for agent in agents_to_check:
                                        # Skip if agent is at capacity
                                        if agent["capacity"] - agent["allocated"] <= 0:
                                            continue

                                        # CRITICAL: Skip PB preference agents - they should ONLY get NTBP rows (allocated in Step 2.5)
                                        if agent.get("has_pb_preference", False):
                                            continue

                                        # Check if agent can work with this insurance company
                                        can_work = False

                                        # First check if this insurance company is in the agent's "do not allocate" list
                                        should_not_allocate = False
                                        if agent.get("insurance_do_not_allocate"):
                                            insurance_carrier_str = (
                                                str(insurance_carrier)
                                                if insurance_carrier
                                                else ""
                                            )
                                            for do_not_allocate_comp in agent[
                                                "insurance_do_not_allocate"
                                            ]:
                                                do_not_allocate_comp_str = (
                                                    str(do_not_allocate_comp)
                                                    if do_not_allocate_comp
                                                    else ""
                                                )
                                                if (
                                                    insurance_carrier_str
                                                    and do_not_allocate_comp_str
                                                ):
                                                    if (
                                                        insurance_carrier_str.lower()
                                                        in do_not_allocate_comp_str.lower()
                                                        or do_not_allocate_comp_str.lower()
                                                        in insurance_carrier_str.lower()
                                                        or insurance_carrier_str
                                                        == do_not_allocate_comp_str
                                                    ):
                                                        should_not_allocate = True
                                                        break

                                        # If agent should not be allocated this insurance company, skip them
                                        if should_not_allocate:
                                            continue

                                        # Senior agents can work with any insurance company (unless in do not allocate list)
                                        if agent["is_senior"]:
                                            can_work = True
                                        elif not agent[
                                            "insurance_companies"
                                        ]:  # If no specific companies listed, can work with any
                                            can_work = True
                                        else:
                                            # Check if insurance carrier matches any of the agent's working companies
                                            for comp in agent["insurance_companies"]:
                                                if (
                                                    insurance_carrier.lower()
                                                    in comp.lower()
                                                    or comp.lower()
                                                    in insurance_carrier.lower()
                                                    or insurance_carrier == comp
                                                ):
                                                    can_work = True
                                                    break

                                        # Check if agent needs training for this insurance company
                                        needs_training = False
                                        if agent["insurance_needs_training"]:
                                            for training_comp in agent[
                                                "insurance_needs_training"
                                            ]:
                                                if (
                                                    insurance_carrier.lower()
                                                    in training_comp.lower()
                                                    or training_comp.lower()
                                                    in insurance_carrier.lower()
                                                    or insurance_carrier
                                                    == training_comp
                                                ):
                                                    needs_training = True
                                                    break

                                        # Agent is capable if they can work AND don't need training
                                        # Domain/Remark filtering will happen when assigning rows
                                        if can_work and not needs_training:
                                            capable_agents.append(agent)

                                    if capable_agents:
                                        # For First Priority and Unknown, verify we only have seniors
                                        if (
                                            priority == "First Priority"
                                            or insurance_carrier == "Unknown"
                                        ):
                                            # Double-check: filter to only seniors with capacity
                                            available_senior = [
                                                a
                                                for a in capable_agents
                                                if a["is_senior"]
                                                and (a["capacity"] - a["allocated"]) > 0
                                            ]
                                            if available_senior:
                                                capable_agents = available_senior
                                            else:
                                                # No senior capacity available - skip allocation (keep unassigned)
                                                continue

                                        # IMPORTANT: NTBP rows should ONLY be allocated in Step 2.5 to PB preference agents
                                        # IMPORTANT: NTC rows should ONLY be allocated in Step 3.5 to NTC preference agents
                                        # Filter out any NTBP and NTC rows that might have slipped through - they should remain unallocated
                                        # if not allocated in Step 2.5 or Step 3.5
                                        non_special_rows = []
                                        if (
                                            remark_col
                                            and remark_col in processed_df.columns
                                        ):
                                            for r_idx in unallocated_row_indices:
                                                row_remark = None
                                                if pd.notna(
                                                    processed_df.at[r_idx, remark_col]
                                                ):
                                                    row_remark = (
                                                        str(
                                                            processed_df.at[
                                                                r_idx, remark_col
                                                            ]
                                                        )
                                                        .strip()
                                                        .upper()
                                                    )
                                                # Skip NTBP rows - they should only go to PB preference agents (Step 2.5)
                                                if row_remark == "NTBP":
                                                    continue  # Skip NTBP rows completely
                                                # Skip NTC rows - they should only go to NTC preference agents (Step 3.5)
                                                elif row_remark == "NTC":
                                                    continue  # Skip NTC rows completely
                                                # Skip rows with "Not to work" remark
                                                elif (
                                                    "NOT TO WORK" in row_remark
                                                    or row_remark == "NOT TO WORK"
                                                ):
                                                    continue  # Skip "Not to work" rows completely
                                                else:
                                                    non_special_rows.append(r_idx)
                                        else:
                                            non_special_rows = (
                                                unallocated_row_indices.copy()
                                            )

                                        # Only process non-NTBP, non-NTC rows
                                        # Partition capable agents by PB domain or allocation preference
                                        pb_agents = [
                                            a
                                            for a in capable_agents
                                            if (
                                                (
                                                    a.get("domain")
                                                    and str(a.get("domain"))
                                                    .strip()
                                                    .upper()
                                                    == "PB"
                                                )
                                                or a.get("has_pb_preference", False)
                                            )
                                        ]
                                        non_pb_agents = [
                                            a
                                            for a in capable_agents
                                            if not (
                                                (
                                                    a.get("domain")
                                                    and str(a.get("domain"))
                                                    .strip()
                                                    .upper()
                                                    == "PB"
                                                )
                                                or a.get("has_pb_preference", False)
                                            )
                                        ]

                                        def sticky_assign(rows, agents, carrier):
                                            if not rows or not agents:
                                                return

                                            # For agents with "Sec + X" preference (Sec + Single, Sec + NTC, Sec + Mix, etc.),
                                            # prioritize rows with secondary insurance first
                                            sec_agents = [
                                                a
                                                for a in agents
                                                if a.get("has_sec_preference", False)
                                                and (a["capacity"] - a["allocated"]) > 0
                                            ]

                                            # Separate rows with secondary insurance for "Sec + X" agents
                                            rows_with_secondary = []
                                            rows_without_secondary = []

                                            if secondary_insurance_col and sec_agents:
                                                for row_idx in rows:
                                                    has_secondary = False
                                                    if pd.notna(
                                                        processed_df.at[
                                                            row_idx,
                                                            secondary_insurance_col,
                                                        ]
                                                    ):
                                                        secondary_val = str(
                                                            processed_df.at[
                                                                row_idx,
                                                                secondary_insurance_col,
                                                            ]
                                                        ).strip()
                                                        if (
                                                            secondary_val
                                                            and secondary_val.lower()
                                                            != "nan"
                                                        ):
                                                            has_secondary = True

                                                    if has_secondary:
                                                        rows_with_secondary.append(
                                                            row_idx
                                                        )
                                                    else:
                                                        rows_without_secondary.append(
                                                            row_idx
                                                        )
                                            else:
                                                rows_without_secondary = rows.copy()

                                            # First, allocate rows with secondary insurance to "Sec + X" agents
                                            if rows_with_secondary and sec_agents:
                                                # Filter agents based on their specific preference after "Sec +"
                                                # For "Sec + Single", they can only take this carrier if unassigned or already assigned to this carrier
                                                # For "Sec + NTC", "Sec + Mix", etc., they can take any carrier
                                                available_sec_agents = []
                                                for a in sec_agents:
                                                    # If agent has "Sec + Single", apply Single logic
                                                    if a.get(
                                                        "has_sec_single_preference",
                                                        False,
                                                    ):
                                                        if a.get(
                                                            "assigned_insurance"
                                                        ) in (None, carrier):
                                                            available_sec_agents.append(
                                                                a
                                                            )
                                                    else:
                                                        # For "Sec + NTC", "Sec + Mix", etc., they can take any carrier
                                                        available_sec_agents.append(a)

                                                if available_sec_agents:
                                                    available_sec_agents.sort(
                                                        key=lambda a: a["capacity"]
                                                        - a["allocated"],
                                                        reverse=True,
                                                    )

                                                    row_pos = 0
                                                    for agent in available_sec_agents:
                                                        if row_pos >= len(
                                                            rows_with_secondary
                                                        ):
                                                            break
                                                        remaining = (
                                                            agent["capacity"]
                                                            - agent["allocated"]
                                                        )
                                                        if remaining <= 0:
                                                            continue
                                                        take = min(
                                                            remaining,
                                                            len(rows_with_secondary)
                                                            - row_pos,
                                                        )
                                                        if take > 0:
                                                            slice_rows = (
                                                                rows_with_secondary[
                                                                    row_pos : row_pos
                                                                    + take
                                                                ]
                                                            )
                                                            # Use safe extend function to filter out "Not to work" rows
                                                            actual_allocated = (
                                                                safe_extend_row_indices(
                                                                    agent,
                                                                    slice_rows,
                                                                    processed_df,
                                                                    remark_col,
                                                                    agent["name"],
                                                                )
                                                            )
                                                            # Set assigned insurance if not set
                                                            if (
                                                                agent.get(
                                                                    "assigned_insurance"
                                                                )
                                                                is None
                                                            ):
                                                                agent[
                                                                    "assigned_insurance"
                                                                ] = carrier
                                                            for idx in slice_rows:
                                                                processed_df.at[
                                                                    idx, "Agent Name"
                                                                ] = agent["name"]
                                                            row_pos += take

                                                    # Remove allocated rows from the list
                                                    rows_with_secondary = (
                                                        rows_with_secondary[row_pos:]
                                                    )

                                            # Combine remaining rows (secondary rows that weren't allocated + rows without secondary)
                                            remaining_rows = (
                                                rows_with_secondary
                                                + rows_without_secondary
                                            )

                                            # Phase 1: agents already on this carrier or unassigned
                                            # For agents with "Single" preference, they can only get this carrier if:
                                            # - They have no assigned insurance yet (None), OR
                                            # - They already have this carrier assigned
                                            phase1_agents = [
                                                a
                                                for a in agents
                                                if (a["capacity"] - a["allocated"]) > 0
                                                and (
                                                    # If agent has "Single" or "Sec + Single" preference, they can only get this carrier
                                                    # if they have no assigned insurance or already have this carrier
                                                    (
                                                        (
                                                            a.get(
                                                                "has_single_preference",
                                                                False,
                                                            )
                                                            or a.get(
                                                                "has_sec_single_preference",
                                                                False,
                                                            )
                                                        )
                                                        and a.get("assigned_insurance")
                                                        in (None, carrier)
                                                    )
                                                    or
                                                    # If agent has "Sec + X" (but not Single), they can take any carrier
                                                    # (they already got secondary insurance rows above, now they can get regular rows)
                                                    (
                                                        a.get(
                                                            "has_sec_preference", False
                                                        )
                                                        and not a.get(
                                                            "has_sec_single_preference",
                                                            False,
                                                        )
                                                        and a.get("assigned_insurance")
                                                        in (None, carrier)
                                                    )
                                                    or
                                                    # If agent doesn't have "Single" or "Sec" preference, use normal logic
                                                    (
                                                        not a.get(
                                                            "has_single_preference",
                                                            False,
                                                        )
                                                        and not a.get(
                                                            "has_sec_preference",
                                                            False,
                                                        )
                                                        and a.get("assigned_insurance")
                                                        in (None, carrier)
                                                    )
                                                )
                                            ]

                                            # Use remaining_rows instead of rows
                                            rows = remaining_rows
                                            # Sort by remaining capacity desc to maximize concentration
                                            phase1_agents.sort(
                                                key=lambda a: a["capacity"]
                                                - a["allocated"],
                                                reverse=True,
                                            )
                                            row_pos = 0
                                            while row_pos < len(rows) and phase1_agents:
                                                for agent in phase1_agents:
                                                    if row_pos >= len(rows):
                                                        break
                                                    remaining = (
                                                        agent["capacity"]
                                                        - agent["allocated"]
                                                    )
                                                    if remaining <= 0:
                                                        continue
                                                    take = min(
                                                        remaining, len(rows) - row_pos
                                                    )
                                                    if take <= 0:
                                                        continue
                                                    slice_rows = rows[
                                                        row_pos : row_pos + take
                                                    ]
                                                    # Use safe extend function to filter out "Not to work" rows
                                                    actual_allocated = (
                                                        safe_extend_row_indices(
                                                            agent,
                                                            slice_rows,
                                                            processed_df,
                                                            remark_col,
                                                            agent["name"],
                                                        )
                                                    )
                                                    if (
                                                        agent.get("assigned_insurance")
                                                        is None
                                                    ):
                                                        agent["assigned_insurance"] = (
                                                            carrier
                                                        )
                                                    # Track group allocations
                                                    agent_id = agent.get(
                                                        "id",
                                                        agent.get("name", "Unknown"),
                                                    )
                                                    carrier_upper = (
                                                        carrier.upper().strip()
                                                    )
                                                    if (
                                                        agent_id
                                                        in ins_group_allocations
                                                    ):
                                                        if any(
                                                            carrier_upper
                                                            == ic.upper().strip()
                                                            for ic in DD_INS_GROUP
                                                        ):
                                                            ins_group_allocations[
                                                                agent_id
                                                            ] += take
                                                    if (
                                                        agent_id
                                                        in toolkit_group_allocations
                                                    ):
                                                        if any(
                                                            carrier_upper
                                                            == ic.upper().strip()
                                                            for ic in DD_TOOLKIT_GROUP
                                                        ):
                                                            toolkit_group_allocations[
                                                                agent_id
                                                            ] += take
                                                    row_pos += take
                                            # Phase 2: remaining rows can go to agents with different carrier if their primary exhausted
                                            # Exclude "Single" and "Sec + Single" agents from Phase 2 (they should only get same insurance)
                                            if row_pos < len(rows):
                                                phase2_agents = [
                                                    a
                                                    for a in agents
                                                    if a.get("assigned_insurance")
                                                    not in (None, carrier)
                                                    and (a["capacity"] - a["allocated"])
                                                    > 0
                                                    and not a.get(
                                                        "has_single_preference", False
                                                    )
                                                    and not a.get(
                                                        "has_sec_single_preference",
                                                        False,
                                                    )
                                                ]
                                                phase2_agents.sort(
                                                    key=lambda a: a["capacity"]
                                                    - a["allocated"],
                                                    reverse=True,
                                                )
                                                while (
                                                    row_pos < len(rows)
                                                    and phase2_agents
                                                ):
                                                    for agent in phase2_agents:
                                                        if row_pos >= len(rows):
                                                            break
                                                        remaining = (
                                                            agent["capacity"]
                                                            - agent["allocated"]
                                                        )
                                                        if remaining <= 0:
                                                            continue
                                                        take = min(
                                                            remaining,
                                                            len(rows) - row_pos,
                                                        )
                                                        if take <= 0:
                                                            continue
                                                        slice_rows = rows[
                                                            row_pos : row_pos + take
                                                        ]
                                                        # Use safe extend function to filter out "Not to work" rows
                                                        actual_allocated = (
                                                            safe_extend_row_indices(
                                                                agent,
                                                                slice_rows,
                                                                processed_df,
                                                                remark_col,
                                                                agent["name"],
                                                            )
                                                        )
                                                        # Do NOT change assigned_insurance here (keep original primary)
                                                        agent_id = agent.get(
                                                            "id",
                                                            agent.get(
                                                                "name", "Unknown"
                                                            ),
                                                        )
                                                        carrier_upper = (
                                                            carrier.upper().strip()
                                                        )
                                                        if (
                                                            agent_id
                                                            in ins_group_allocations
                                                        ):
                                                            if any(
                                                                carrier_upper
                                                                == ic.upper().strip()
                                                                for ic in DD_INS_GROUP
                                                            ):
                                                                ins_group_allocations[
                                                                    agent_id
                                                                ] += take
                                                        if (
                                                            agent_id
                                                            in toolkit_group_allocations
                                                        ):
                                                            if any(
                                                                carrier_upper
                                                                == ic.upper().strip()
                                                                for ic in DD_TOOLKIT_GROUP
                                                            ):
                                                                toolkit_group_allocations[
                                                                    agent_id
                                                                ] += take
                                                        row_pos += take

                                        # Execute sticky assignment for non-NTBP, non-NTC rows only
                                        # NTBP rows should have been allocated in Step 2.5 to PB preference agents only
                                        # NTC rows should have been allocated in Step 3.5 to NTC preference agents only
                                        # If any NTBP or NTC rows remain unallocated, they will stay unallocated (not assigned to other agents)
                                        if non_special_rows:
                                            sticky_assign(
                                                non_special_rows,
                                                non_pb_agents,
                                                insurance_carrier,
                                            )
                    else:
                        # Fallback: if no insurance carrier column, use simple capacity-based allocation
                        # IMPORTANT: NTBP rows should ONLY be allocated in Step 2.5 to PB preference agents
                        # IMPORTANT: NTC rows should ONLY be allocated in Step 3.5 to NTC preference agents
                        # Skip NTBP and NTC rows here - they should remain unallocated if not allocated in Step 2.5 or 3.5
                        non_special_rows = []

                        if remark_col and remark_col in processed_df.columns:
                            for idx in range(total_rows):
                                # Skip already allocated rows (should have been allocated in Step 2.5 or 3.5)
                                if idx in [
                                    i
                                    for ag in agent_allocations
                                    for i in ag["row_indices"]
                                ]:
                                    continue

                                row_remark = None
                                if pd.notna(processed_df.at[idx, remark_col]):
                                    row_remark = (
                                        str(processed_df.at[idx, remark_col])
                                        .strip()
                                        .upper()
                                    )

                                # Skip NTBP rows - they should only go to PB preference agents (Step 2.5)
                                if row_remark == "NTBP":
                                    continue
                                # Skip NTC rows - they should only go to NTC preference agents (Step 3.5)
                                elif row_remark == "NTC":
                                    continue
                                # Skip rows with "Not to work" remark - they should never be allocated
                                elif (
                                    "NOT TO WORK" in row_remark
                                    or row_remark == "NOT TO WORK"
                                ):
                                    continue
                                else:
                                    non_special_rows.append(idx)
                        else:
                            # If no remark column, all rows are non-special
                            # But still skip already allocated rows
                            non_special_rows = [
                                idx
                                for idx in range(total_rows)
                                if idx
                                not in [
                                    i
                                    for ag in agent_allocations
                                    for i in ag["row_indices"]
                                ]
                            ]

                        # Allocate non-NTBP, non-NTC rows to all available agents (capacity-based)
                        row_idx = 0
                        for agent in agent_allocations:
                            if row_idx >= len(non_special_rows):
                                break
                            available_capacity = agent["capacity"] - agent["allocated"]
                            if available_capacity > 0:
                                actual_allocation = min(
                                    available_capacity, len(non_special_rows) - row_idx
                                )
                                if actual_allocation > 0:
                                    slice_rows = non_special_rows[
                                        row_idx : row_idx + actual_allocation
                                    ]
                                    # Use safe extend function to filter out "Not to work" rows
                                    actual_allocated = safe_extend_row_indices(
                                        agent,
                                        slice_rows,
                                        processed_df,
                                        remark_col,
                                        agent["name"],
                                    )
                                    row_idx += actual_allocation

                    # Step 6: Final Fallback - Allocate ANY remaining unallocated rows to agents with matching capabilities
                    # This ensures that unallocated work matching agent capabilities gets allocated, regardless of allocation preference
                    if insurance_carrier_col:
                        # Find all unallocated rows
                        unallocated_indices = [
                            idx
                            for idx in processed_df.index
                            if idx
                            not in [
                                i for ag in agent_allocations for i in ag["row_indices"]
                            ]
                        ]

                        if unallocated_indices:
                            # Group unallocated rows by insurance company
                            unallocated_by_insurance = {}
                            for idx in unallocated_indices:
                                # Skip rows with secondary insurance (those should go to Sec + X agents)
                                if secondary_insurance_col and pd.notna(
                                    processed_df.at[idx, secondary_insurance_col]
                                ):
                                    secondary_val = str(
                                        processed_df.at[idx, secondary_insurance_col]
                                    ).strip()
                                    if secondary_val and secondary_val.lower() != "nan":
                                        continue

                                # Skip NTBP, NTC, and "Not to work" rows
                                if remark_col and pd.notna(
                                    processed_df.at[idx, remark_col]
                                ):
                                    remark_val = (
                                        str(processed_df.at[idx, remark_col])
                                        .strip()
                                        .upper()
                                    )
                                    if remark_val == "NTBP" or remark_val == "NTC":
                                        continue
                                    # Skip rows with "Not to work" remark
                                    if (
                                        "NOT TO WORK" in remark_val
                                        or remark_val == "NOT TO WORK"
                                    ):
                                        continue

                                # Get insurance company
                                if pd.notna(
                                    processed_df.at[idx, insurance_carrier_col]
                                ):
                                    insurance_carrier = str(
                                        processed_df.at[idx, insurance_carrier_col]
                                    ).strip()
                                    if insurance_carrier:
                                        if (
                                            insurance_carrier
                                            not in unallocated_by_insurance
                                        ):
                                            unallocated_by_insurance[
                                                insurance_carrier
                                            ] = []
                                        unallocated_by_insurance[
                                            insurance_carrier
                                        ].append(idx)

                            # Allocate unallocated rows to ALL agents with capacity (not just matching capabilities)
                            # This ensures maximum utilization of agent capacity
                            for (
                                insurance_carrier,
                                row_indices_list,
                            ) in unallocated_by_insurance.items():
                                # Find ALL agents with capacity (except PB agents)
                                # Prioritize agents with matching capabilities, but include all agents with capacity
                                capable_agents = []  # Agents with matching capabilities
                                all_capable_agents = []  # All agents with capacity

                                for agent in agent_allocations:
                                    remaining_capacity = (
                                        agent["capacity"] - agent["allocated"]
                                    )
                                    if remaining_capacity <= 0:
                                        continue

                                    # CRITICAL: Skip PB preference agents - they should ONLY get NTBP rows (allocated in Step 2.5)
                                    if agent.get("has_pb_preference", False):
                                        continue

                                    # Check if this is "Afreen Ansari" - she can work with any unallocated insurance
                                    is_afreen_ansari = (
                                        agent.get("name") == "Afreen Ansari"
                                    )

                                    # Check if agent can work with this insurance company
                                    agent_insurance_list = agent.get(
                                        "insurance_companies", []
                                    )
                                    can_work = check_insurance_match(
                                        insurance_carrier,
                                        agent_insurance_list,
                                        agent.get("is_senior", False),
                                        agent.get("name"),
                                    )

                                    # Add to all_capable_agents (all agents with capacity)
                                    all_capable_agents.append(agent)

                                    # Add to capable_agents if they match capabilities or are Afreen Ansari
                                    if can_work or is_afreen_ansari:
                                        capable_agents.append(agent)

                                # Use capable_agents first (matching capabilities), then fall back to all_capable_agents
                                agents_to_use = (
                                    capable_agents
                                    if capable_agents
                                    else all_capable_agents
                                )

                                # Allocate rows to agents (round-robin)
                                # Prioritize agents with matching capabilities, but use all agents with capacity if needed
                                if agents_to_use:
                                    agent_idx = 0
                                    for row_idx in row_indices_list:
                                        # Find next agent with capacity
                                        attempts = 0
                                        max_attempts = len(agents_to_use)
                                        while attempts < max_attempts:
                                            if agent_idx >= len(agents_to_use):
                                                agent_idx = 0

                                            agent = agents_to_use[agent_idx]
                                            remaining_capacity = (
                                                agent["capacity"] - agent["allocated"]
                                            )

                                            if remaining_capacity > 0:
                                                break  # Found agent with capacity

                                            # Move to next agent
                                            agent_idx += 1
                                            attempts += 1
                                        else:
                                            # No agents with capacity left - break
                                            break

                                        if remaining_capacity > 0:
                                            # Safety check: Verify this is not a "Not to work" row before allocating
                                            if remark_col and pd.notna(
                                                processed_df.at[row_idx, remark_col]
                                            ):
                                                remark_check = (
                                                    str(
                                                        processed_df.at[
                                                            row_idx, remark_col
                                                        ]
                                                    )
                                                    .strip()
                                                    .upper()
                                                )
                                                if (
                                                    "NOT TO WORK" in remark_check
                                                    or remark_check == "NOT TO WORK"
                                                ):
                                                    continue  # Skip this row

                                            # Safety check: Verify this is not a "Not to work" row before allocating
                                            if not should_skip_row_for_allocation(
                                                row_idx, processed_df, remark_col
                                            ):
                                                agent["row_indices"].append(row_idx)
                                                agent["allocated"] += 1
                                                processed_df.at[
                                                    row_idx, "Agent Name"
                                                ] = agent["name"]
                                                agent_idx += 1
                                            else:
                                                # Skip this row if it's "Not to work", but continue to next row
                                                continue
                                        # Note: We check capacity each iteration, so agents at capacity are automatically skipped
                                        # Move to next agent for next row
                                        agent_idx += 1

                    # Soft stickiness rule: prefer keeping same insurance per agent, allow adding new carrier only after
                    # existing carrier rows are exhausted. Actual assignment handled in Step 5 logic.
                    # Here we just set assigned_insurance for agents that currently have a single carrier.
                    if insurance_carrier_col:
                        for agent in agent_allocations:
                            indices = agent.get("row_indices", [])
                            if not indices:
                                continue
                            carrier_groups = {}
                            for idx in indices:
                                if idx < len(processed_df):
                                    carrier = processed_df.at[
                                        idx, insurance_carrier_col
                                    ]
                                    carrier_groups.setdefault(carrier, 0)
                                    carrier_groups[carrier] += 1
                            if len(carrier_groups) == 1:
                                agent["assigned_insurance"] = next(
                                    iter(carrier_groups.keys())
                                )
                            elif (
                                agent.get("assigned_insurance") is None
                                and carrier_groups
                            ):
                                # Tentatively choose dominant carrier as primary
                                dominant = max(
                                    carrier_groups.items(), key=lambda kv: kv[1]
                                )[0]
                                agent["assigned_insurance"] = dominant
                    # Sort agents by name for display
                    agent_allocations.sort(key=lambda x: x["name"])

                    # CRITICAL: Deduplicate row_indices for each agent and recalculate allocated counts
                    # This prevents counting the same row multiple times
                    for agent in agent_allocations:
                        row_indices = agent.get("row_indices", [])
                        if row_indices:
                            # Remove duplicates and invalid indices
                            unique_valid_indices = [
                                idx
                                for idx in set(row_indices)
                                if idx < len(processed_df)
                            ]
                            agent["row_indices"] = unique_valid_indices
                            agent["allocated"] = len(unique_valid_indices)

                    # Calculate total allocated rows based on unique row indices across all agents
                    all_allocated_indices = set()
                    for agent in agent_allocations:
                        all_allocated_indices.update(agent.get("row_indices", []))
                    total_allocated = len(all_allocated_indices)

                    # Print INS and Toolkit group allocation summary
                    # Ensure dictionaries exist (they should be initialized earlier)
                    if "ins_group_allocations" not in locals():
                        ins_group_allocations = {}
                    if "toolkit_group_allocations" not in locals():
                        toolkit_group_allocations = {}

                    if ins_group_allocations:
                        total_ins = sum(ins_group_allocations.values())
                        # Create mapping from agent_id to agent_name for display
                        agent_id_to_name = {
                            a.get("id", a.get("name")): a.get("name")
                            for a in agent_allocations
                        }
                        for agent_id, count in sorted(ins_group_allocations.items()):
                            agent_name = agent_id_to_name.get(agent_id, agent_id)
                            pass
                        if total_ins == 0:
                            pass
                    else:
                        pass

                    if toolkit_group_allocations:
                        total_toolkit = sum(toolkit_group_allocations.values())
                        # Create mapping from agent_id to agent_name for display
                        agent_id_to_name = {
                            a.get("id", a.get("name")): a.get("name")
                            for a in agent_allocations
                        }
                        for agent_id, count in sorted(
                            toolkit_group_allocations.items()
                        ):
                            agent_name = agent_id_to_name.get(agent_id, agent_id)
                            pass
                        if total_toolkit == 0:
                            pass
                    else:
                        pass

                    # Add Agent Name column to processed_df based on allocation
                    # Initialize Agent Name column if it doesn't exist
                    if "Agent Name" not in processed_df.columns:
                        processed_df["Agent Name"] = ""

                    # Set agent name for each allocated row
                    for agent in agent_allocations:
                        agent_name = agent["name"]
                        row_indices = agent.get("row_indices", [])
                        if row_indices:
                            # Filter to only valid indices within the dataframe
                            valid_indices = [
                                idx for idx in row_indices if idx < len(processed_df)
                            ]
                            if valid_indices:
                                # Set agent name for all rows allocated to this agent
                                processed_df.loc[valid_indices, "Agent Name"] = (
                                    agent_name
                                )

                    # Store agent allocations data globally for individual downloads
                    agent_allocations_data = agent_allocations

                    # Also store for reminder system
                    global agent_allocations_for_reminders
                    agent_allocations_for_reminders = agent_allocations

                    # Calculate allocation statistics
                    # CRITICAL: Calculate total_allocated based on unique row indices to avoid duplicates
                    # Collect all unique row indices across all agents
                    # Also filter out "Not to work" rows from the count
                    all_allocated_indices = set()
                    for agent in agent_allocations:
                        row_indices = agent.get("row_indices", [])
                        # Deduplicate row_indices for this agent
                        unique_indices = list(set(row_indices))
                        # Filter out "Not to work" rows from indices
                        if remark_col and remark_col in processed_df.columns:
                            unique_indices = [
                                idx
                                for idx in unique_indices
                                if not should_skip_row_for_allocation(
                                    idx, processed_df, remark_col
                                )
                            ]
                        # Update agent's row_indices to be unique
                        agent["row_indices"] = unique_indices
                        # Update agent's allocated count to match actual unique indices
                        agent["allocated"] = len(unique_indices)
                        # Add to global set
                        all_allocated_indices.update(unique_indices)

                    # Total allocated is the count of unique row indices (excluding "Not to work" rows)
                    total_allocated = len(all_allocated_indices)
                    agents_with_work = len(
                        [a for a in agent_allocations if a["allocated"] > 0]
                    )

                    # Calculate total_rows excluding "Not to work" rows
                    # This matches what we actually allocate (we skip "Not to work" rows)
                    total_rows_excluding_not_to_work = total_rows
                    if remark_col and remark_col in processed_df.columns:
                        not_to_work_count = 0
                        for idx in processed_df.index:
                            if should_skip_row_for_allocation(
                                idx, processed_df, remark_col
                            ):
                                not_to_work_count += 1
                        total_rows_excluding_not_to_work = (
                            total_rows - not_to_work_count
                        )

                    # Get unmatched insurance companies info (if it exists from allocation process)
                    unmatched_info = ""
                    if insurance_carrier_col and unmatched_insurance_companies:
                        unmatched_list = [
                            str(comp).strip()
                            for comp in sorted(list(unmatched_insurance_companies))[:5]
                            if comp is not None and not pd.isna(comp)
                        ]
                        unmatched_info = f"\n🔴 Unmatched Insurance Companies ({len(unmatched_insurance_companies)}): {', '.join(unmatched_list)}{'...' if len(unmatched_insurance_companies) > 5 else ''}\n   ⚠️ These companies were assigned ONLY to senior agents with highest priority."

                    agent_summary = f"""
👥 Agent Allocation Summary (Capability-Based):
- Total Agents: {total_agents}
- Agents with Work: {agents_with_work}
                    - Total Rows to Allocate: {total_rows_excluding_not_to_work}
- Total Allocated: {total_allocated}
                    - Remaining Unallocated: {total_rows_excluding_not_to_work - total_allocated}
- Insurance Matching: {'Enabled' if insurance_carrier_col else 'Disabled'}
{unmatched_info}

🧷 Sticky Carrier Rule: Agents prefer to keep working on a single insurance carrier until that carrier's available rows are exhausted. Only then are additional carriers added to fill remaining capacity. Primary carrier shown below; secondary carriers appear only if needed.

📋 Agent Allocation Details:
"""
                    for i, agent in enumerate(agent_allocations):
                        insurance_info = ""
                        senior_info = " (Senior Agent)" if agent["is_senior"] else ""

                        if agent["is_senior"]:
                            insurance_info = " (Can work: Any insurance company)"
                        elif agent["insurance_companies"]:
                            insurance_list = [
                                str(comp).strip()
                                for comp in agent["insurance_companies"][:2]
                                if comp is not None and not pd.isna(comp)
                            ]
                            insurance_info = f" (Can work: {', '.join(insurance_list)}{'...' if len(agent['insurance_companies']) > 2 else ''})"

                        if agent["insurance_needs_training"]:
                            training_list = [
                                str(comp).strip()
                                for comp in agent["insurance_needs_training"][:2]
                                if comp is not None and not pd.isna(comp)
                            ]
                            training_info = f" (Needs training: {', '.join(training_list)}{'...' if len(agent['insurance_needs_training']) > 2 else ''})"
                            insurance_info += training_info

                        primary = agent.get("assigned_insurance")
                        # Derive secondary carriers (those allocated rows not matching primary)
                        secondary = []
                        if (
                            primary
                            and "assigned_insurance" in agent
                            and insurance_carrier_col
                        ):
                            allocated_carriers = set()
                            for ridx in agent.get("row_indices", []):
                                if ridx < len(processed_df):
                                    allocated_carriers.add(
                                        processed_df.at[ridx, insurance_carrier_col]
                                    )
                            secondary = [c for c in allocated_carriers if c != primary]
                        primary_info = f" | Primary: {primary}" if primary else ""
                        secondary_info = (
                            f" | Secondary: {', '.join([str(c).strip() for c in secondary[:2] if c is not None and not pd.isna(c)])}{'...' if len(secondary) > 2 else ''}"
                            if secondary
                            else ""
                        )
                        agent_summary += f"  {i+1}. {agent['name']}: {agent['allocated']}/{agent['capacity']} rows{senior_info}{insurance_info}{primary_info}{secondary_info}\n"

                    # Calculate priority distribution based on actual allocations
                    # CRITICAL: Calculate total_allocated based on unique row indices to avoid duplicates
                    all_allocated_indices = set()
                    for agent in agent_allocations:
                        row_indices = agent.get("row_indices", [])
                        # Deduplicate row_indices for this agent
                        unique_indices = list(set(row_indices))
                        # Update agent's row_indices to be unique
                        agent["row_indices"] = unique_indices
                        # Update agent's allocated count to match actual unique indices
                        agent["allocated"] = len(unique_indices)
                        # Add to global set
                        all_allocated_indices.update(unique_indices)

                    # Total allocated is the count of unique row indices
                    total_allocated = len(all_allocated_indices)
                    if total_allocated > 0:
                        agent_summary += f"""
📊 Priority Distribution (Based on Actual Allocations):
- First Priority: {first_priority_count} rows total
- Second Priority: {second_priority_count} rows total  
- Third Priority: {third_priority_count} rows total

⚠️ Note: Priority distribution will be proportional to each agent's allocated capacity.
"""
                    else:
                        agent_summary += "\n⚠️ No rows could be allocated due to capacity constraints."

                elif not agent_name_col:
                    agent_summary = (
                        "\n⚠️ Agent Name column not found in allocation file."
                    )
                elif not counts_col:
                    agent_summary = "\n⚠️ TFD column not found in allocation file."

                # Add information about insurance matching
                if insurance_carrier_col and insurance_working_col:
                    training_info = (
                        f" and '{insurance_needs_training_col}'"
                        if insurance_needs_training_col
                        else ""
                    )
                    agent_summary += f"\n✅ Insurance capability matching enabled using '{insurance_working_col}'{training_info} and '{insurance_carrier_col}' columns."
                elif insurance_carrier_col and not insurance_working_col:
                    agent_summary += f"\n⚠️ Insurance carrier column '{insurance_carrier_col}' found, but 'Insurance List' column not found in allocation file."
                elif not insurance_carrier_col and insurance_working_col:
                    agent_summary += f"\n⚠️ 'Insurance List' column found, but 'Dental Primary Ins Carr' column not found in data file."
                else:
                    agent_summary += f"\nℹ️ Insurance capability matching disabled - using simple capacity-based allocation."

                # Add information about training filtering
                if insurance_needs_training_col:
                    agent_summary += f"\n🎓 Training-based filtering enabled - agents will not be assigned work for insurance companies they need training for."

                # Add information about senior agents
                senior_count = sum(
                    1 for agent in agent_allocations if agent["is_senior"]
                )
                if senior_count > 0:
                    unmatched_note = (
                        f" Unmatched insurance companies ({len(unmatched_insurance_companies)}) are assigned ONLY to senior agents with highest priority."
                        if unmatched_insurance_companies
                        else ""
                    )
                    agent_summary += f"\n👑 Senior agents detected: {senior_count} - Senior agents can work with any insurance company and get priority for First Priority cases.{unmatched_note}"
            except Exception as e:
                agent_summary = f"\n⚠️ Error processing agent allocation: {str(e)}"

        # Generate result message
        first_priority_dates_list = sorted(list(first_priority_dates))
        second_priority_dates_list = sorted(list(second_priority_dates))
        third_priority_dates_list = sorted(list(third_priority_dates_set))
        first_priority_dates_str = (
            ", ".join(first_priority_dates_list)
            if first_priority_dates_list
            else "None"
        )
        second_priority_dates_str = (
            ", ".join(second_priority_dates_list)
            if second_priority_dates_list
            else "None"
        )
        third_priority_dates_str = (
            ", ".join(third_priority_dates_list)
            if third_priority_dates_list
            else "None"
        )

        result_message = f"""✅ Priority processing completed successfully!

📊 Processing Statistics:
- Total rows processed: {total_rows}
- First Priority: {first_priority_count} rows
- Second Priority: {second_priority_count} rows
- Third Priority: {third_priority_count} rows
- Invalid dates: {invalid_dates} rows

📅 Selected First Priority Dates: {first_priority_dates_str}
📅 Selected Second Priority Dates: {second_priority_dates_str}
📅 Third Priority Dates: {third_priority_dates_str}

📋 Updated column: 'Priority Status'
📅 Based on column: '{appointment_date_col}'{agent_summary}

💾 Ready to download the processed result file!"""

        return result_message, processed_df

    except Exception as e:
        return f"❌ Error during processing: {str(e)}", None


@app.route("/")
@login_required
def index():
    global allocation_data, data_file_data, allocation_filename, data_filename, processing_result
    global agent_processing_result, agent_allocations_data

    # Get current user
    user = get_user_by_username(session.get("user_id"))

    # Load agent work files if user is an agent
    agent_work_files = None
    if user and user.role == "agent":
        agent_work_files = get_agent_work_files(user.id)

    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Load all agent work files for admin view
    all_agent_work_files = None
    if user and user.role == "admin":
        all_agent_work_files = get_all_agent_work_files()

    return render_template_string(
        HTML_TEMPLATE,
        allocation_data=allocation_data,
        data_file_data=data_file_data,
        allocation_filename=allocation_filename,
        data_filename=data_filename,
        processing_result=processing_result,
        agent_processing_result=agent_processing_result,
        agent_allocations_data=agent_allocations_data,
        agent_work_files=agent_work_files,
        all_agent_work_files=all_agent_work_files,
        current_time=current_time,
    )


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        # Try database authentication first
        user = get_user_by_username(username)
        if user:
            password_valid = user.check_password(password)

            # If password check failed and user has scrypt hash, rehash and try again
            if (
                not password_valid
                and user.password_hash
                and (
                    "scrypt" in user.password_hash.lower()
                    or user.password_hash.startswith("scrypt:")
                )
            ):
                # Rehash the password with pbkdf2:sha256 and update user
                user.set_password(password)
                db.session.commit()
                # Verify the new hash works
                password_valid = user.check_password(password)

            if password_valid:
                # Update last login
                user.last_login = datetime.utcnow()
                db.session.commit()

                # Create database session
                session_data = {
                    "user_id": user.username,
                    "user_role": user.role,
                    "user_name": user.name,
                    "user_email": user.email,
                }
                db_session = create_user_session(user.id, session_data)

                # Set Flask session
                session["db_session_id"] = db_session.id
                session.update(session_data)

                return redirect(url_for("dashboard"))
            else:
                flash("Invalid username or password. Please try again.", "error")
        else:
            flash("Invalid username or password. Please try again.", "error")

    return render_template_string(LOGIN_TEMPLATE, GOOGLE_CLIENT_ID=GOOGLE_CLIENT_ID)


@app.route("/google-login")
def google_login():
    """Initiate Google OAuth login"""
    if not GOOGLE_CLIENT_ID:
        flash(
            "Google OAuth is not configured. Please contact administrator to set up Google OAuth for agent login.",
            "error",
        )
        return redirect(url_for("login"))

    # Get Google OAuth configuration
    google_provider_cfg = get_google_provider_cfg()
    if not google_provider_cfg:
        flash(
            "Unable to connect to Google OAuth service. Please check your internet connection and try again.",
            "error",
        )
        return redirect(url_for("login"))

    authorization_endpoint = google_provider_cfg["authorization_endpoint"]

    # Get the exact callback URL using url_for to ensure consistency
    callback_url = url_for("callback", _external=True)

    # Force HTTPS for production (Railway/Heroku)
    if os.environ.get("DATABASE_URL") or os.environ.get("RAILWAY_ENVIRONMENT"):
        # Ensure callback URL uses HTTPS in production
        if callback_url.startswith("http://"):
            callback_url = callback_url.replace("http://", "https://", 1)

    # Create request URI with properly URL-encoded redirect_uri
    redirect_uri_encoded = quote(callback_url, safe="")
    request_uri = f"{authorization_endpoint}?client_id={GOOGLE_CLIENT_ID}&redirect_uri={redirect_uri_encoded}&scope=openid email profile&response_type=code"

    return redirect(request_uri)


@app.route("/callback")
def callback():
    """Handle Google OAuth callback"""
    if not GOOGLE_CLIENT_ID or not GOOGLE_CLIENT_SECRET:
        flash("Google OAuth is not configured. Please contact administrator.", "error")
        return redirect(url_for("login"))

    # Get authorization code from the request
    code = request.args.get("code")
    if not code:
        flash("Authorization failed. Please try again.", "error")
        return redirect(url_for("login"))

    try:
        # Get Google OAuth configuration
        google_provider_cfg = get_google_provider_cfg()
        if not google_provider_cfg:
            flash(
                "Unable to connect to Google OAuth service. Please try again later.",
                "error",
            )
            return redirect(url_for("login"))

        token_endpoint = google_provider_cfg["token_endpoint"]

        # Get the exact callback URL using url_for to match what was sent initially
        callback_url = url_for("callback", _external=True)

        # Force HTTPS for production (Railway/Heroku) - must match what was sent in initial request
        if os.environ.get("DATABASE_URL") or os.environ.get("RAILWAY_ENVIRONMENT"):
            # Ensure callback URL uses HTTPS in production
            if callback_url.startswith("http://"):
                callback_url = callback_url.replace("http://", "https://", 1)

        # Exchange code for token
        token_data = {
            "code": code,
            "client_id": GOOGLE_CLIENT_ID,
            "client_secret": GOOGLE_CLIENT_SECRET,
            "redirect_uri": callback_url,
            "grant_type": "authorization_code",
        }

        token_response = req.post(
            token_endpoint,
            data=token_data,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
        )

        # Parse the tokens
        if token_response.status_code != 200:
            flash(
                "Failed to exchange authorization code for token. Please try again.",
                "error",
            )
            return redirect(url_for("login"))

        tokens = token_response.json()

        if "id_token" not in tokens:
            flash("No ID token received from Google. Please try again.", "error")
            return redirect(url_for("login"))

        # Verify the token
        google_user_info = verify_google_token(tokens["id_token"])

        if not google_user_info:
            flash("Token verification failed. Please try again.", "error")
            return redirect(url_for("login"))

        # Get or create user
        user = get_or_create_google_user(google_user_info)

        if not user.is_active:
            flash("Your account is inactive. Please contact administrator.", "error")
            return redirect(url_for("login"))

        # Update last login
        user.last_login = datetime.utcnow()
        db.session.commit()

        # Create database session
        session_data = {
            "user_id": user.email,  # Use email as user_id for OAuth users
            "user_role": user.role,
            "user_name": user.name,
            "user_email": user.email,
        }
        db_session = create_user_session(user.id, session_data)

        # Set Flask session
        session["db_session_id"] = db_session.id
        session.update(session_data)

        return redirect(url_for("dashboard"))

    except Exception as e:
        flash("Authentication failed. Please try again.", "error")
        return redirect(url_for("login"))


@app.route("/logout")
def logout():
    # Clean up database session
    db_session_id = session.get("db_session_id")
    if db_session_id:
        delete_user_session(db_session_id)

    # Clear Flask session
    session.clear()
    flash("You have been logged out successfully.", "success")
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    return redirect(url_for("index"))


@app.route("/upload_allocation", methods=["POST"])
@admin_required
def upload_allocation_file():
    global allocation_data, allocation_filename, processing_result

    if "file" not in request.files:
        flash("No file provided", "error")
        return redirect("/")

    file = request.files["file"]
    if file.filename == "":
        flash("No file selected", "error")
        return redirect("/")

    try:
        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        file.save(filename)

        # Load Excel file
        # Use parse_dates=False to prevent automatic date parsing that differs between Windows and Mac
        allocation_data = pd.read_excel(filename, sheet_name=None, parse_dates=False)

        # Focus on "main" sheet if it exists, otherwise use all sheets
        sheets_to_process = {}
        if "main" in allocation_data:
            sheets_to_process["main"] = allocation_data["main"]
        else:
            sheets_to_process = allocation_data

        # Format insurance company names in "Insurance List" column for better allocation matching
        for sheet_name, df in sheets_to_process.items():
            # Find the Insurance List column (case-insensitive)
            insurance_working_col = None
            for col in df.columns:
                col_str = str(col) if not isinstance(col, str) else col
                if "insurance" in col_str.lower() and "list" in col_str.lower():
                    insurance_working_col = col
                    break

            if insurance_working_col:
                # Format each value in Insurance List column (which may contain multiple companies separated by ; or ,)
                def format_insurance_list(value):
                    if pd.isna(value):
                        return value
                    value_str = str(value)
                    # Split by common delimiters
                    companies = [
                        comp.strip()
                        for comp in re.split(r"[;,\|]", value_str)
                        if comp.strip()
                    ]
                    # Format each company name, but preserve "senior" keyword and group names for expansion
                    formatted_companies = []
                    for comp in companies:
                        comp_str = str(comp) if comp is not None else ""
                        comp_lower = comp_str.lower()
                        if "senior" in comp_lower:
                            formatted_companies.append(comp)  # Keep senior as-is
                        elif (
                            comp_lower == "dd ins"
                            or comp_lower == "ins"
                            or comp_lower == "dd toolkit"
                            or comp_lower == "dd toolkits"
                            or comp_lower == "dd"
                        ):
                            # Keep group names as-is for later expansion
                            formatted_companies.append(comp)
                        else:
                            formatted = format_insurance_company_name(comp)
                            formatted_companies.append(formatted)
                    # Join back with semicolon
                    return "; ".join(formatted_companies)

                # First format the insurance names
                df[insurance_working_col] = df[insurance_working_col].apply(
                    format_insurance_list
                )

                # Then expand insurance groups (DD INS/INS and DD Toolkit/Toolkits/DD)
                df[insurance_working_col] = df[insurance_working_col].apply(
                    expand_insurance_groups
                )

                if "main" in allocation_data:
                    allocation_data["main"] = df
                else:
                    allocation_data[sheet_name] = df

        allocation_filename = filename

        # Update allocation_data to only include processed sheets
        if "main" in allocation_data:
            allocation_data = {"main": allocation_data["main"]}

        processing_result = f"✅ Allocation file uploaded successfully! Loaded {len(allocation_data)} sheet(s): {', '.join(list(allocation_data.keys()))}"
        flash(
            f'Allocation file uploaded successfully! Loaded {len(allocation_data)} sheet(s): {", ".join(list(allocation_data.keys()))}',
            "success",
        )

        # Clean up uploaded file
        if os.path.exists(filename):
            os.remove(filename)

        return redirect("/")

    except Exception as e:
        processing_result = f"❌ Error uploading allocation file: {str(e)}"
        flash(f"Error uploading allocation file: {str(e)}", "error")
        # Clean up uploaded file on error
        if "filename" in locals() and os.path.exists(filename):
            os.remove(filename)
        return redirect("/")


@app.route("/upload_data", methods=["POST"])
@admin_required
def upload_data_file():
    global data_file_data, data_filename, processing_result

    if "file" not in request.files:
        flash("No file provided", "error")
        return redirect("/")

    file = request.files["file"]
    if file.filename == "":
        flash("No file selected", "error")
        return redirect("/")

    try:
        # Reset tracking for new file
        global _formatted_insurance_names, _formatted_insurance_details
        _formatted_insurance_names = set()
        _formatted_insurance_details = []

        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        file.save(filename)

        # Load Excel file
        # Use parse_dates=False to prevent automatic date parsing that differs between Windows and Mac
        data_file_data = pd.read_excel(filename, sheet_name=None, parse_dates=False)

        # Format insurance company names in "Dental Primary Ins Carr" column for better allocation
        for sheet_name, df in data_file_data.items():
            # Find the insurance carrier column (case-insensitive)
            insurance_col = None
            for col in df.columns:
                if (
                    "dental" in col.lower()
                    and "primary" in col.lower()
                    and "ins" in col.lower()
                    and "carr" in col.lower()
                ):
                    insurance_col = col
                    break

            if insurance_col:
                data_file_data[sheet_name] = format_insurance_column_in_dataframe(
                    df.copy(), insurance_col
                )

        data_filename = filename

        processing_result = f"✅ Data file uploaded successfully! Loaded {len(data_file_data)} sheets: {', '.join(list(data_file_data.keys()))}"
        flash(
            f'Data file uploaded successfully! Loaded {len(data_file_data)} sheets: {", ".join(list(data_file_data.keys()))}',
            "success",
        )

        # Print formatted insurance companies list
        print_formatted_insurance_companies()

        # Clean up uploaded file
        if os.path.exists(filename):
            os.remove(filename)

        return redirect("/")

    except Exception as e:
        processing_result = f"❌ Error uploading data file: {str(e)}"
        flash(f"Error uploading data file: {str(e)}", "error")
        # Clean up uploaded file on error
        if "filename" in locals() and os.path.exists(filename):
            os.remove(filename)
        return redirect("/")


@app.route("/process_files", methods=["POST"])
@admin_required
def process_files():
    global allocation_data, data_file_data, processing_result, agent_processing_result, agent_allocations_data

    if not data_file_data:
        processing_result = "❌ Error: Please upload data file first"
        return render_template_string(
            HTML_TEMPLATE,
            allocation_data=allocation_data,
            data_file_data=data_file_data,
            allocation_filename=allocation_filename,
            data_filename=data_filename,
            processing_result=processing_result,
            agent_processing_result=agent_processing_result,
            agent_allocations_data=agent_allocations_data,
        )

    try:
        # Get the first sheet from data file
        data_df = list(data_file_data.values())[0]

        # Get selected appointment dates from calendar
        appointment_dates = request.form.getlist("appointment_dates")
        appointment_dates_second = request.form.getlist("appointment_dates_second")
        receive_dates = request.form.getlist("receive_dates")
        debug_count = request.form.get("debug_selected_count", "0")
        debug_count_second = request.form.get("debug_selected_count_second", "0")

        # Process the data file with selected dates and allocation data
        result_message, processed_df = process_allocation_files_with_dates(
            allocation_data,
            data_df,
            [],
            "",
            appointment_dates,
            appointment_dates_second,
            receive_dates,
        )

        if processed_df is not None:
            # Store the result for download
            processing_result = result_message
            # Update the data_file_data with the processed result
            data_file_data[list(data_file_data.keys())[0]] = processed_df
        else:
            processing_result = result_message

        return render_template_string(
            HTML_TEMPLATE,
            allocation_data=allocation_data,
            data_file_data=data_file_data,
            allocation_filename=allocation_filename,
            data_filename=data_filename,
            processing_result=processing_result,
            agent_processing_result=agent_processing_result,
            agent_allocations_data=agent_allocations_data,
        )

    except Exception as e:
        processing_result = f"❌ Error processing data file: {str(e)}"
        return render_template_string(
            HTML_TEMPLATE,
            allocation_data=allocation_data,
            data_file_data=data_file_data,
            allocation_filename=allocation_filename,
            data_filename=data_filename,
            processing_result=processing_result,
            agent_processing_result=agent_processing_result,
            agent_allocations_data=agent_allocations_data,
        )


@app.route("/download_result", methods=["POST"])
@admin_required
def download_result():
    global data_file_data, data_filename, agent_allocations_data, agent_insurance_agent_names

    if not data_file_data:
        return jsonify({"error": "No data to download"}), 400

    filename = request.form.get("filename", "").strip()
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"processed_data_{timestamp}.xlsx"

    try:
        # Create a temporary file
        temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx")

        try:
            with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
                # Write all existing sheets
                for sheet_name, df in data_file_data.items():
                    # Create a copy of the dataframe to avoid modifying the original
                    df_copy = df.copy()

                    # Find appointment date and received date columns and format them as MM/DD/YYYY
                    for col in df_copy.columns:
                        if ("appointment" in col.lower() and "date" in col.lower()) or (
                            "receive" in col.lower() and "date" in col.lower()
                        ):
                            # Convert to datetime and then format as MM/DD/YYYY (no time)
                            df_copy[col] = pd.to_datetime(
                                df_copy[col], errors="coerce"
                            ).dt.strftime("%m/%d/%Y")

                    df_copy.to_excel(writer, sheet_name=sheet_name, index=False)

                # Create Agent Count Summary sheet
                # Get the processed dataframe
                processed_df = (
                    list(data_file_data.values())[0] if data_file_data else None
                )

                if processed_df is not None:
                    # Check if Agent Name column exists
                    agent_name_col = None
                    for col in processed_df.columns:
                        if "agent" in col.lower() and "name" in col.lower():
                            agent_name_col = col
                            break

                    if agent_name_col:
                        # Count allocations by agent name
                        agent_counts = {}

                        # Find Remark column
                        remark_col = None
                        for col in processed_df.columns:
                            if "remark" in col.lower():
                                remark_col = col
                                break

                        # Count each row once - prioritize Remark column for NTC and Not to work
                        for idx, row in processed_df.iterrows():
                            # Check Remark column first for NTC and Not to work
                            if remark_col and remark_col in processed_df.columns:
                                remark_value = row.get(remark_col)
                                if pd.notna(remark_value):
                                    remark_str = str(remark_value).strip().upper()
                                    if remark_str == "NTC":
                                        agent_counts["NTC"] = (
                                            agent_counts.get("NTC", 0) + 1
                                        )
                                        continue  # Skip agent name counting for this row
                                    elif "NOT TO WORK" in remark_str.replace(
                                        "-", " "
                                    ).replace("_", " "):
                                        agent_counts["Not to work"] = (
                                            agent_counts.get("Not to work", 0) + 1
                                        )
                                        continue  # Skip agent name counting for this row

                            # Count by Agent Name (if not already counted as NTC or Not to work)
                            if agent_name_col in processed_df.columns:
                                agent_name_value = row.get(agent_name_col)
                                if pd.notna(agent_name_value):
                                    agent_name_str = str(agent_name_value).strip()
                                    # Check if it's NTC or Not to work in Agent Name column
                                    agent_name_upper = agent_name_str.upper()
                                    if agent_name_upper == "NTC":
                                        agent_counts["NTC"] = (
                                            agent_counts.get("NTC", 0) + 1
                                        )
                                    elif "NOT TO WORK" in agent_name_upper.replace(
                                        "-", " "
                                    ).replace("_", " "):
                                        agent_counts["Not to work"] = (
                                            agent_counts.get("Not to work", 0) + 1
                                        )
                                    else:
                                        # Regular agent name
                                        agent_counts[agent_name_str] = (
                                            agent_counts.get(agent_name_str, 0) + 1
                                        )

                        # Convert to list of tuples and sort by count (descending)
                        summary_list = [
                            (row_label, count)
                            for row_label, count in agent_counts.items()
                        ]
                        summary_list.sort(
                            key=lambda x: x[1], reverse=True
                        )  # Sort by count in descending order

                        # Calculate grand total
                        grand_total = sum(count for _, count in summary_list)

                        # Create summary dataframe
                        if summary_list:
                            summary_df = pd.DataFrame(
                                summary_list,
                                columns=["Row Labels", "Count of Agent Name"],
                            )
                            # Add grand total row
                            grand_total_row = pd.DataFrame(
                                [["Grand Total", grand_total]],
                                columns=["Row Labels", "Count of Agent Name"],
                            )
                            summary_df = pd.concat(
                                [summary_df, grand_total_row], ignore_index=True
                            )
                        else:
                            # Create empty summary with just grand total
                            summary_df = pd.DataFrame(
                                [["Grand Total", 0]],
                                columns=["Row Labels", "Count of Agent Name"],
                            )

                        # Write summary sheet
                        summary_df.to_excel(
                            writer, sheet_name="Agent Count Summary", index=False
                        )

                # Create Priority Status sheet
                if processed_df is not None:
                    # Find Priority Status column
                    priority_status_col = None
                    appointment_date_col = None

                    for col in processed_df.columns:
                        if "priority" in col.lower() and "status" in col.lower():
                            priority_status_col = col
                        if "appointment" in col.lower() and "date" in col.lower():
                            appointment_date_col = col

                    if priority_status_col and appointment_date_col:
                        # Get all unique appointment dates (store both YYYY-MM-DD for matching and MM/DD/YYYY for display)
                        appointment_dates_dict = (
                            {}
                        )  # key: YYYY-MM-DD, value: MM/DD/YYYY

                        for idx, row in processed_df.iterrows():
                            appt_date = row.get(appointment_date_col)
                            if pd.notna(appt_date):
                                # Convert to date object
                                date_obj = None
                                if hasattr(appt_date, "date"):
                                    date_obj = appt_date.date()
                                elif hasattr(appt_date, "strftime"):
                                    # Try to parse if it's a string
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        date_obj = appt_date
                                else:
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        continue

                                if date_obj:
                                    date_key = date_obj.strftime("%Y-%m-%d")
                                    date_display = date_obj.strftime("%m/%d/%Y")
                                    appointment_dates_dict[date_key] = date_display

                        # Sort dates by key (YYYY-MM-DD) and create lists
                        sorted_date_keys = sorted(appointment_dates_dict.keys())
                        appointment_dates = sorted_date_keys  # For matching
                        appointment_dates_display = [
                            appointment_dates_dict[key] for key in sorted_date_keys
                        ]  # For display

                        # Create pivot data structure
                        priority_rows = ["First Priority", "Second Priority"]
                        priority_data = {}

                        # Initialize counts for each priority and date
                        for priority in priority_rows:
                            priority_data[priority] = {}
                            for date in appointment_dates:
                                priority_data[priority][date] = 0

                        # Count rows by priority and date
                        for idx, row in processed_df.iterrows():
                            priority_status = row.get(priority_status_col)
                            appt_date = row.get(appointment_date_col)

                            if pd.notna(priority_status) and pd.notna(appt_date):
                                priority_str = str(priority_status).strip()

                                # Convert appointment date to YYYY-MM-DD format for matching
                                date_obj = None
                                if hasattr(appt_date, "date"):
                                    date_obj = appt_date.date()
                                elif hasattr(appt_date, "strftime"):
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        # Try parsing as string
                                        try:
                                            date_obj = pd.to_datetime(
                                                str(appt_date)
                                            ).date()
                                        except:
                                            continue
                                else:
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        continue

                                if date_obj:
                                    date_str = date_obj.strftime("%Y-%m-%d")

                                    if (
                                        priority_str in priority_rows
                                        and date_str in appointment_dates
                                    ):
                                        priority_data[priority_str][date_str] = (
                                            priority_data[priority_str].get(date_str, 0)
                                            + 1
                                        )

                        # Calculate grand totals for each priority
                        priority_totals = {}
                        for priority in priority_rows:
                            priority_totals[priority] = sum(
                                priority_data[priority].values()
                            )

                        # Calculate grand total for all priorities
                        overall_grand_total = sum(priority_totals.values())

                        # Calculate totals for each date column
                        date_totals = {}
                        for date in appointment_dates:
                            date_totals[date] = sum(
                                priority_data[priority][date]
                                for priority in priority_rows
                            )

                        # Build the dataframe
                        # Columns: Row Labels, Grand Total, [date columns with MM/DD/YYYY format]
                        columns = [
                            "Row Labels",
                            "Grand Total",
                        ] + appointment_dates_display
                        rows_data = []

                        # Add priority rows
                        for priority in priority_rows:
                            row_data = [priority, priority_totals[priority]]
                            for date_key in appointment_dates:
                                row_data.append(priority_data[priority][date_key])
                            rows_data.append(row_data)

                        # Add Grand Total row
                        grand_total_row = ["Grand Total", overall_grand_total]
                        for date_key in appointment_dates:
                            grand_total_row.append(date_totals[date_key])
                        rows_data.append(grand_total_row)

                        # Create dataframe
                        priority_df = pd.DataFrame(rows_data, columns=columns)

                        # Write Priority Status sheet
                        priority_df.to_excel(
                            writer, sheet_name="Priority Status", index=False
                        )

                # Create Priority Remark sheet
                if processed_df is not None:
                    # Find Priority Status, Remark, and Appointment Date columns
                    priority_status_col = None
                    remark_col = None
                    appointment_date_col = None

                    for col in processed_df.columns:
                        if "priority" in col.lower() and "status" in col.lower():
                            priority_status_col = col
                        if "remark" in col.lower():
                            remark_col = col
                        if "appointment" in col.lower() and "date" in col.lower():
                            appointment_date_col = col

                    if priority_status_col and appointment_date_col:
                        # Get all unique appointment dates (reuse the logic from Priority Status sheet)
                        appointment_dates_dict = (
                            {}
                        )  # key: YYYY-MM-DD, value: MM/DD/YYYY

                        for idx, row in processed_df.iterrows():
                            appt_date = row.get(appointment_date_col)
                            if pd.notna(appt_date):
                                # Convert to date object
                                date_obj = None
                                if hasattr(appt_date, "date"):
                                    date_obj = appt_date.date()
                                elif hasattr(appt_date, "strftime"):
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        date_obj = appt_date
                                else:
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        continue

                                if date_obj:
                                    date_key = date_obj.strftime("%Y-%m-%d")
                                    date_display = date_obj.strftime("%m/%d/%Y")
                                    appointment_dates_dict[date_key] = date_display

                        # Sort dates by key (YYYY-MM-DD) and create lists
                        sorted_date_keys = sorted(appointment_dates_dict.keys())
                        appointment_dates = sorted_date_keys  # For matching
                        appointment_dates_display = [
                            appointment_dates_dict[key] for key in sorted_date_keys
                        ]  # For display

                        # Define priority and remark categories
                        priority_rows = ["First Priority", "Second Priority"]
                        remark_types = ["NTBP", "Not to work", "Workable", "NTC"]

                        # Initialize data structure: priority -> remark -> date -> count
                        priority_remark_data = {}
                        for priority in priority_rows:
                            priority_remark_data[priority] = {}
                            for remark in remark_types:
                                priority_remark_data[priority][remark] = {}
                                for date_key in appointment_dates:
                                    priority_remark_data[priority][remark][date_key] = 0

                        # Count rows by priority, remark, and date
                        for idx, row in processed_df.iterrows():
                            priority_status = row.get(priority_status_col)
                            appt_date = row.get(appointment_date_col)
                            remark_value = row.get(remark_col) if remark_col else None

                            if pd.notna(priority_status) and pd.notna(appt_date):
                                priority_str = str(priority_status).strip()

                                # Convert appointment date to YYYY-MM-DD format
                                date_obj = None
                                if hasattr(appt_date, "date"):
                                    date_obj = appt_date.date()
                                elif hasattr(appt_date, "strftime"):
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        try:
                                            date_obj = pd.to_datetime(
                                                str(appt_date)
                                            ).date()
                                        except:
                                            continue
                                else:
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        continue

                                if date_obj and priority_str in priority_rows:
                                    date_str = date_obj.strftime("%Y-%m-%d")

                                    if date_str in appointment_dates:
                                        # Determine remark type
                                        remark_type = "Workable"  # Default

                                        if pd.notna(remark_value):
                                            remark_str = (
                                                str(remark_value).strip().upper()
                                            )

                                            if remark_str == "NTBP":
                                                remark_type = "NTBP"
                                            elif remark_str == "NTC":
                                                remark_type = "NTC"
                                            elif "NOT TO WORK" in remark_str.replace(
                                                "-", " "
                                            ).replace("_", " "):
                                                remark_type = "Not to work"
                                            # else: Workable (default)

                                        if remark_type in remark_types:
                                            priority_remark_data[priority_str][
                                                remark_type
                                            ][date_str] = (
                                                priority_remark_data[priority_str][
                                                    remark_type
                                                ].get(date_str, 0)
                                                + 1
                                            )

                        # Calculate totals
                        # Totals for each priority+remark combination
                        priority_remark_totals = {}
                        for priority in priority_rows:
                            priority_remark_totals[priority] = {}
                            for remark in remark_types:
                                priority_remark_totals[priority][remark] = sum(
                                    priority_remark_data[priority][remark].values()
                                )

                        # Totals for each date column
                        date_totals = {}
                        for date_key in appointment_dates:
                            date_totals[date_key] = sum(
                                priority_remark_data[priority][remark][date_key]
                                for priority in priority_rows
                                for remark in remark_types
                            )

                        # Overall grand total
                        overall_grand_total = sum(
                            priority_remark_totals[priority][remark]
                            for priority in priority_rows
                            for remark in remark_types
                        )

                        # Build the dataframe
                        # Columns: Row Labels, Grand Total, [date columns]
                        columns = [
                            "Row Labels",
                            "Grand Total",
                        ] + appointment_dates_display
                        rows_data = []

                        # Add rows for each priority and remark combination
                        for priority in priority_rows:
                            # Add priority header row with totals
                            priority_total = sum(
                                priority_remark_totals[priority].values()
                            )
                            priority_row_data = [priority, priority_total]
                            for date_key in appointment_dates:
                                # Sum all remarks for this priority and date
                                date_total = sum(
                                    priority_remark_data[priority][remark][date_key]
                                    for remark in remark_types
                                )
                                priority_row_data.append(date_total)
                            rows_data.append(priority_row_data)

                            # Add remark sub-rows for this priority
                            for remark in remark_types:
                                row_label = remark  # Just the remark name
                                row_total = priority_remark_totals[priority][remark]

                                row_data = [row_label, row_total]
                                for date_key in appointment_dates:
                                    row_data.append(
                                        priority_remark_data[priority][remark][date_key]
                                    )
                                rows_data.append(row_data)

                        # Add Grand Total row
                        grand_total_row = ["Grand Total", overall_grand_total]
                        for date_key in appointment_dates:
                            grand_total_row.append(date_totals[date_key])
                        rows_data.append(grand_total_row)

                        # Create dataframe
                        priority_remark_df = pd.DataFrame(rows_data, columns=columns)

                        # Write Priority Remark sheet
                        priority_remark_df.to_excel(
                            writer, sheet_name="Priority Remark", index=False
                        )

                # Create Today Allocation sheet
                if processed_df is not None:
                    # Find Agent Name, Appointment Date, and Remark columns
                    agent_name_col = None
                    appointment_date_col = None
                    remark_col = None

                    for col in processed_df.columns:
                        if "agent" in col.lower() and "name" in col.lower():
                            agent_name_col = col
                        if "appointment" in col.lower() and "date" in col.lower():
                            appointment_date_col = col
                        if "remark" in col.lower():
                            remark_col = col

                    if agent_name_col and appointment_date_col and remark_col:
                        # Get all unique appointment dates (reuse the logic from Priority Status sheet)
                        appointment_dates_dict = (
                            {}
                        )  # key: YYYY-MM-DD, value: MM/DD/YYYY

                        for idx, row in processed_df.iterrows():
                            appt_date = row.get(appointment_date_col)
                            if pd.notna(appt_date):
                                # Convert to date object
                                date_obj = None
                                if hasattr(appt_date, "date"):
                                    date_obj = appt_date.date()
                                elif hasattr(appt_date, "strftime"):
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        date_obj = appt_date
                                else:
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        continue

                                if date_obj:
                                    date_key = date_obj.strftime("%Y-%m-%d")
                                    date_display = date_obj.strftime("%m/%d/%Y")
                                    appointment_dates_dict[date_key] = date_display

                        # Sort dates by key (YYYY-MM-DD) and create lists
                        sorted_date_keys = sorted(appointment_dates_dict.keys())
                        appointment_dates = sorted_date_keys  # For matching
                        appointment_dates_display = [
                            appointment_dates_dict[key] for key in sorted_date_keys
                        ]  # For display

                        # Initialize data structure: agent_name -> date -> count
                        agent_allocation_data = {}

                        # Get all unique agent names
                        agent_names = set()
                        for idx, row in processed_df.iterrows():
                            agent_name = row.get(agent_name_col)
                            remark_value = row.get(remark_col)

                            # Only include agents with "Workable" remark
                            if pd.notna(remark_value):
                                remark_str = str(remark_value).strip().upper()
                                if remark_str == "WORKABLE":
                                    if pd.notna(agent_name) and str(agent_name).strip():
                                        agent_name_str = str(agent_name).strip()
                                        # Skip NTC and Not to work as they're not agents
                                        agent_name_upper = agent_name_str.upper()
                                        if (
                                            agent_name_upper != "NTC"
                                            and "NOT TO WORK"
                                            not in agent_name_upper.replace(
                                                "-", " "
                                            ).replace("_", " ")
                                        ):
                                            agent_names.add(agent_name_str)

                        # Initialize counts for each agent and date
                        for agent_name in agent_names:
                            agent_allocation_data[agent_name] = {}
                            for date_key in appointment_dates:
                                agent_allocation_data[agent_name][date_key] = 0

                        # Count allocations by agent name and date (only for Workable remark)
                        for idx, row in processed_df.iterrows():
                            agent_name = row.get(agent_name_col)
                            appt_date = row.get(appointment_date_col)
                            remark_value = row.get(remark_col)

                            # Only process rows with "Workable" remark
                            if pd.isna(remark_value):
                                continue

                            remark_str = str(remark_value).strip().upper()
                            if remark_str != "WORKABLE":
                                continue

                            if pd.notna(agent_name) and pd.notna(appt_date):
                                agent_name_str = str(agent_name).strip()
                                # Skip NTC and Not to work
                                agent_name_upper = agent_name_str.upper()
                                if (
                                    agent_name_upper == "NTC"
                                    or "NOT TO WORK"
                                    in agent_name_upper.replace("-", " ").replace(
                                        "_", " "
                                    )
                                ):
                                    continue

                                # Convert appointment date to YYYY-MM-DD format
                                date_obj = None
                                if hasattr(appt_date, "date"):
                                    date_obj = appt_date.date()
                                elif hasattr(appt_date, "strftime"):
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        try:
                                            date_obj = pd.to_datetime(
                                                str(appt_date)
                                            ).date()
                                        except:
                                            continue
                                else:
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        continue

                                if date_obj and agent_name_str in agent_names:
                                    date_str = date_obj.strftime("%Y-%m-%d")

                                    if date_str in appointment_dates:
                                        agent_allocation_data[agent_name_str][
                                            date_str
                                        ] = (
                                            agent_allocation_data[agent_name_str].get(
                                                date_str, 0
                                            )
                                            + 1
                                        )

                        # Calculate totals
                        # Totals for each agent
                        agent_totals = {}
                        for agent_name in agent_names:
                            agent_totals[agent_name] = sum(
                                agent_allocation_data[agent_name].values()
                            )

                        # Totals for each date column
                        date_totals = {}
                        for date_key in appointment_dates:
                            date_totals[date_key] = sum(
                                agent_allocation_data[agent_name][date_key]
                                for agent_name in agent_names
                            )

                        # Overall grand total
                        overall_grand_total = sum(agent_totals.values())

                        # Build the dataframe
                        # Columns: Row Labels, Grand Total, [date columns]
                        columns = [
                            "Row Labels",
                            "Grand Total",
                        ] + appointment_dates_display
                        rows_data = []

                        # Add Grand Total row first
                        grand_total_row = ["Grand Total", overall_grand_total]
                        for date_key in appointment_dates:
                            grand_total_row.append(date_totals[date_key])
                        rows_data.append(grand_total_row)

                        # Sort agent names for consistent ordering
                        sorted_agent_names = sorted(agent_names)

                        # Add rows for each agent
                        for agent_name in sorted_agent_names:
                            agent_total = agent_totals[agent_name]

                            row_data = [agent_name, agent_total]
                            for date_key in appointment_dates:
                                row_data.append(
                                    agent_allocation_data[agent_name][date_key]
                                )
                            rows_data.append(row_data)

                        # Create dataframe
                        today_allocation_df = pd.DataFrame(rows_data, columns=columns)

                        # Write Today Allocation sheet
                        today_allocation_df.to_excel(
                            writer, sheet_name="Today Allocation", index=False
                        )

                # Create NTBP Allocation sheet
                if processed_df is not None:
                    # Find Agent Name, Appointment Date, and Remark columns
                    agent_name_col = None
                    appointment_date_col = None
                    remark_col = None

                    for col in processed_df.columns:
                        if "agent" in col.lower() and "name" in col.lower():
                            agent_name_col = col
                        if "appointment" in col.lower() and "date" in col.lower():
                            appointment_date_col = col
                        if "remark" in col.lower():
                            remark_col = col

                    if agent_name_col and appointment_date_col and remark_col:
                        # Get all unique appointment dates (reuse the logic from Priority Status sheet)
                        appointment_dates_dict = (
                            {}
                        )  # key: YYYY-MM-DD, value: MM/DD/YYYY

                        for idx, row in processed_df.iterrows():
                            appt_date = row.get(appointment_date_col)
                            if pd.notna(appt_date):
                                # Convert to date object
                                date_obj = None
                                if hasattr(appt_date, "date"):
                                    date_obj = appt_date.date()
                                elif hasattr(appt_date, "strftime"):
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        date_obj = appt_date
                                else:
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        continue

                                if date_obj:
                                    date_key = date_obj.strftime("%Y-%m-%d")
                                    date_display = date_obj.strftime("%m/%d/%Y")
                                    appointment_dates_dict[date_key] = date_display

                        # Sort dates by key (YYYY-MM-DD) and create lists
                        sorted_date_keys = sorted(appointment_dates_dict.keys())
                        appointment_dates = sorted_date_keys  # For matching
                        appointment_dates_display = [
                            appointment_dates_dict[key] for key in sorted_date_keys
                        ]  # For display

                        # Initialize data structure: agent_name -> date -> count
                        agent_ntbp_allocation_data = {}

                        # Get all unique agent names (only for rows with NTBP remark)
                        agent_names = set()
                        for idx, row in processed_df.iterrows():
                            agent_name = row.get(agent_name_col)
                            remark_value = row.get(remark_col)

                            # Only include rows with NTBP remark
                            if pd.notna(remark_value):
                                remark_str = str(remark_value).strip().upper()
                                if remark_str == "NTBP":
                                    if pd.notna(agent_name) and str(agent_name).strip():
                                        agent_name_str = str(agent_name).strip()
                                        # Skip NTC and Not to work as they're not agents
                                        agent_name_upper = agent_name_str.upper()
                                        if (
                                            agent_name_upper != "NTC"
                                            and "NOT TO WORK"
                                            not in agent_name_upper.replace(
                                                "-", " "
                                            ).replace("_", " ")
                                        ):
                                            agent_names.add(agent_name_str)

                        # Initialize counts for each agent and date
                        for agent_name in agent_names:
                            agent_ntbp_allocation_data[agent_name] = {}
                            for date_key in appointment_dates:
                                agent_ntbp_allocation_data[agent_name][date_key] = 0

                        # Count NTBP allocations by agent name and date
                        for idx, row in processed_df.iterrows():
                            agent_name = row.get(agent_name_col)
                            appt_date = row.get(appointment_date_col)
                            remark_value = row.get(remark_col)

                            # Only process rows with NTBP remark
                            if pd.isna(remark_value):
                                continue  # Skip rows without remarks

                            remark_str = str(remark_value).strip().upper()
                            if remark_str != "NTBP":
                                continue  # Skip rows that are not NTBP

                            if pd.notna(agent_name) and pd.notna(appt_date):
                                agent_name_str = str(agent_name).strip()
                                # Skip NTC and Not to work
                                agent_name_upper = agent_name_str.upper()
                                if (
                                    agent_name_upper == "NTC"
                                    or "NOT TO WORK"
                                    in agent_name_upper.replace("-", " ").replace(
                                        "_", " "
                                    )
                                ):
                                    continue

                                # Convert appointment date to YYYY-MM-DD format
                                date_obj = None
                                if hasattr(appt_date, "date"):
                                    date_obj = appt_date.date()
                                elif hasattr(appt_date, "strftime"):
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        try:
                                            date_obj = pd.to_datetime(
                                                str(appt_date)
                                            ).date()
                                        except:
                                            continue
                                else:
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        continue

                                if date_obj and agent_name_str in agent_names:
                                    date_str = date_obj.strftime("%Y-%m-%d")

                                    if date_str in appointment_dates:
                                        agent_ntbp_allocation_data[agent_name_str][
                                            date_str
                                        ] = (
                                            agent_ntbp_allocation_data[
                                                agent_name_str
                                            ].get(date_str, 0)
                                            + 1
                                        )

                        # Calculate totals
                        # Totals for each agent
                        agent_totals = {}
                        for agent_name in agent_names:
                            agent_totals[agent_name] = sum(
                                agent_ntbp_allocation_data[agent_name].values()
                            )

                        # Totals for each date column
                        date_totals = {}
                        for date_key in appointment_dates:
                            date_totals[date_key] = sum(
                                agent_ntbp_allocation_data[agent_name][date_key]
                                for agent_name in agent_names
                            )

                        # Overall grand total
                        overall_grand_total = sum(agent_totals.values())

                        # Build the dataframe
                        # Columns: Row Labels, Grand Total, [date columns]
                        columns = [
                            "Row Labels",
                            "Grand Total",
                        ] + appointment_dates_display
                        rows_data = []

                        # Add Grand Total row first
                        grand_total_row = ["Grand Total", overall_grand_total]
                        for date_key in appointment_dates:
                            grand_total_row.append(date_totals[date_key])
                        rows_data.append(grand_total_row)

                        # Sort agent names for consistent ordering
                        sorted_agent_names = sorted(agent_names)

                        # Add rows for each agent
                        for agent_name in sorted_agent_names:
                            agent_total = agent_totals[agent_name]

                            row_data = [agent_name, agent_total]
                            for date_key in appointment_dates:
                                row_data.append(
                                    agent_ntbp_allocation_data[agent_name][date_key]
                                )
                            rows_data.append(row_data)

                        # Create dataframe
                        ntbp_allocation_df = pd.DataFrame(rows_data, columns=columns)

                        # Write NTBP Allocation sheet
                        ntbp_allocation_df.to_excel(
                            writer, sheet_name="NTBP Allocation", index=False
                        )

                # Create NTC Allocation sheet
                if processed_df is not None:
                    # Find Agent Name, Appointment Date, and Remark columns
                    agent_name_col = None
                    appointment_date_col = None
                    remark_col = None

                    for col in processed_df.columns:
                        if "agent" in col.lower() and "name" in col.lower():
                            agent_name_col = col
                        if "appointment" in col.lower() and "date" in col.lower():
                            appointment_date_col = col
                        if "remark" in col.lower():
                            remark_col = col

                    if agent_name_col and appointment_date_col and remark_col:
                        # Get all unique appointment dates (reuse the logic from Priority Status sheet)
                        appointment_dates_dict = (
                            {}
                        )  # key: YYYY-MM-DD, value: MM/DD/YYYY

                        for idx, row in processed_df.iterrows():
                            appt_date = row.get(appointment_date_col)
                            if pd.notna(appt_date):
                                # Convert to date object
                                date_obj = None
                                if hasattr(appt_date, "date"):
                                    date_obj = appt_date.date()
                                elif hasattr(appt_date, "strftime"):
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        date_obj = appt_date
                                else:
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        continue

                                if date_obj:
                                    date_key = date_obj.strftime("%Y-%m-%d")
                                    date_display = date_obj.strftime("%m/%d/%Y")
                                    appointment_dates_dict[date_key] = date_display

                        # Sort dates by key (YYYY-MM-DD) and create lists
                        sorted_date_keys = sorted(appointment_dates_dict.keys())
                        appointment_dates = sorted_date_keys  # For matching
                        appointment_dates_display = [
                            appointment_dates_dict[key] for key in sorted_date_keys
                        ]  # For display

                        # Initialize data structure: agent_name -> date -> count
                        agent_ntc_allocation_data = {}
                        ntc_row_data = {}  # For rows where Agent Name is "NTC" or empty

                        # Initialize NTC row data
                        for date_key in appointment_dates:
                            ntc_row_data[date_key] = 0

                        # Get all unique agent names (only for rows with NTC remark)
                        agent_names = set()
                        for idx, row in processed_df.iterrows():
                            agent_name = row.get(agent_name_col)
                            remark_value = row.get(remark_col)

                            # Only include rows with NTC remark
                            if pd.notna(remark_value):
                                remark_str = str(remark_value).strip().upper()
                                if remark_str == "NTC":
                                    if pd.notna(agent_name) and str(agent_name).strip():
                                        agent_name_str = str(agent_name).strip()
                                        agent_name_upper = agent_name_str.upper()
                                        # If agent name is "NTC", it goes to NTC row, not agent names
                                        if (
                                            agent_name_upper != "NTC"
                                            and "NOT TO WORK"
                                            not in agent_name_upper.replace(
                                                "-", " "
                                            ).replace("_", " ")
                                        ):
                                            agent_names.add(agent_name_str)

                        # Initialize counts for each agent and date
                        for agent_name in agent_names:
                            agent_ntc_allocation_data[agent_name] = {}
                            for date_key in appointment_dates:
                                agent_ntc_allocation_data[agent_name][date_key] = 0

                        # Count NTC allocations by agent name and date
                        for idx, row in processed_df.iterrows():
                            agent_name = row.get(agent_name_col)
                            appt_date = row.get(appointment_date_col)
                            remark_value = row.get(remark_col)

                            # Only process rows with NTC remark
                            if pd.isna(remark_value):
                                continue  # Skip rows without remarks

                            remark_str = str(remark_value).strip().upper()
                            if remark_str != "NTC":
                                continue  # Skip rows that are not NTC

                            if pd.notna(appt_date):
                                # Convert appointment date to YYYY-MM-DD format
                                date_obj = None
                                if hasattr(appt_date, "date"):
                                    date_obj = appt_date.date()
                                elif hasattr(appt_date, "strftime"):
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        try:
                                            date_obj = pd.to_datetime(
                                                str(appt_date)
                                            ).date()
                                        except:
                                            continue
                                else:
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        continue

                                if date_obj:
                                    date_str = date_obj.strftime("%Y-%m-%d")

                                    if date_str in appointment_dates:
                                        # Check if agent name is valid or if it should go to NTC row
                                        if (
                                            pd.notna(agent_name)
                                            and str(agent_name).strip()
                                        ):
                                            agent_name_str = str(agent_name).strip()
                                            agent_name_upper = agent_name_str.upper()

                                            if agent_name_upper == "NTC":
                                                # Count in NTC row
                                                ntc_row_data[date_str] = (
                                                    ntc_row_data.get(date_str, 0) + 1
                                                )
                                            elif agent_name_str in agent_names:
                                                # Count in agent row
                                                agent_ntc_allocation_data[
                                                    agent_name_str
                                                ][date_str] = (
                                                    agent_ntc_allocation_data[
                                                        agent_name_str
                                                    ].get(date_str, 0)
                                                    + 1
                                                )
                                        else:
                                            # Empty agent name - count in NTC row
                                            ntc_row_data[date_str] = (
                                                ntc_row_data.get(date_str, 0) + 1
                                            )

                        # Calculate totals
                        # Totals for each agent
                        agent_totals = {}
                        for agent_name in agent_names:
                            agent_totals[agent_name] = sum(
                                agent_ntc_allocation_data[agent_name].values()
                            )

                        # Total for NTC row
                        ntc_row_total = sum(ntc_row_data.values())

                        # Totals for each date column
                        date_totals = {}
                        for date_key in appointment_dates:
                            date_totals[date_key] = (
                                sum(
                                    agent_ntc_allocation_data[agent_name][date_key]
                                    for agent_name in agent_names
                                )
                                + ntc_row_data[date_key]
                            )

                        # Overall grand total
                        overall_grand_total = sum(agent_totals.values()) + ntc_row_total

                        # Build the dataframe
                        # Columns: Row Labels, Grand Total, [date columns]
                        columns = [
                            "Row Labels",
                            "Grand Total",
                        ] + appointment_dates_display
                        rows_data = []

                        # Add Grand Total row first
                        grand_total_row = ["Grand Total", overall_grand_total]
                        for date_key in appointment_dates:
                            grand_total_row.append(date_totals[date_key])
                        rows_data.append(grand_total_row)

                        # Sort agent names for consistent ordering
                        sorted_agent_names = sorted(agent_names)

                        # Add rows for each agent
                        for agent_name in sorted_agent_names:
                            agent_total = agent_totals[agent_name]

                            row_data = [agent_name, agent_total]
                            for date_key in appointment_dates:
                                row_data.append(
                                    agent_ntc_allocation_data[agent_name][date_key]
                                )
                            rows_data.append(row_data)

                        # Add NTC row
                        ntc_row = ["NTC", ntc_row_total]
                        for date_key in appointment_dates:
                            ntc_row.append(ntc_row_data[date_key])
                        rows_data.append(ntc_row)

                        # Create dataframe
                        ntc_allocation_df = pd.DataFrame(rows_data, columns=columns)

                        # Write NTC Allocation sheet
                        ntc_allocation_df.to_excel(
                            writer, sheet_name="NTC Allocation", index=False
                        )

                # Create NTC Insurance Name and counts sheet
                if processed_df is not None:
                    # Find Insurance Carrier and Remark columns
                    insurance_carrier_col = None
                    remark_col = None

                    for col in processed_df.columns:
                        if (
                            (
                                "dental" in col.lower()
                                and "primary" in col.lower()
                                and "ins" in col.lower()
                            )
                            or ("insurance" in col.lower() and "carrier" in col.lower())
                            or ("insurance" in col.lower() and "name" in col.lower())
                        ):
                            insurance_carrier_col = col
                        if "remark" in col.lower():
                            remark_col = col

                    if insurance_carrier_col and remark_col:
                        # Count insurance companies with NTC remark
                        insurance_ntc_counts = {}

                        for idx, row in processed_df.iterrows():
                            remark_value = row.get(remark_col)
                            insurance_value = row.get(insurance_carrier_col)

                            # Only process rows with NTC remark
                            if pd.isna(remark_value):
                                continue  # Skip rows without remarks

                            remark_str = str(remark_value).strip().upper()
                            if remark_str != "NTC":
                                continue  # Skip rows that are not NTC

                            # Count by insurance company name
                            if pd.notna(insurance_value):
                                insurance_name = str(insurance_value).strip()
                                if (
                                    insurance_name
                                ):  # Only count non-empty insurance names
                                    insurance_ntc_counts[insurance_name] = (
                                        insurance_ntc_counts.get(insurance_name, 0) + 1
                                    )

                        # Convert to list of tuples and sort by count (descending)
                        insurance_list = [
                            (insurance_name, count)
                            for insurance_name, count in insurance_ntc_counts.items()
                        ]
                        insurance_list.sort(
                            key=lambda x: x[1], reverse=True
                        )  # Sort by count in descending order

                        # Calculate grand total
                        grand_total = sum(count for _, count in insurance_list)

                        # Create summary dataframe
                        if insurance_list:
                            ntc_insurance_df = pd.DataFrame(
                                insurance_list,
                                columns=["Row Labels", "Count of Agent Name"],
                            )
                            # Add grand total row
                            grand_total_row = pd.DataFrame(
                                [["Grand Total", grand_total]],
                                columns=["Row Labels", "Count of Agent Name"],
                            )
                            ntc_insurance_df = pd.concat(
                                [ntc_insurance_df, grand_total_row], ignore_index=True
                            )
                        else:
                            # Create empty summary with just grand total
                            ntc_insurance_df = pd.DataFrame(
                                [["Grand Total", 0]],
                                columns=["Row Labels", "Count of Agent Name"],
                            )

                        # Write NTC Insurance Name and counts sheet
                        ntc_insurance_df.to_excel(
                            writer,
                            sheet_name="NTC Insurance Name and counts",
                            index=False,
                        )

                # Create Agent \ Insurance sheet
                if processed_df is not None:
                    # Find Agent Name and Insurance Carrier columns
                    agent_name_col = None
                    insurance_carrier_col = None

                    for col in processed_df.columns:
                        if "agent" in col.lower() and "name" in col.lower():
                            agent_name_col = col
                        if (
                            (
                                "dental" in col.lower()
                                and "primary" in col.lower()
                                and "ins" in col.lower()
                            )
                            or ("insurance" in col.lower() and "carrier" in col.lower())
                            or ("insurance" in col.lower() and "name" in col.lower())
                        ):
                            insurance_carrier_col = col

                    if agent_name_col and insurance_carrier_col:
                        # Initialize data structure: agent_name -> insurance_name -> count
                        agent_insurance_data = {}

                        # Get all unique agent names (excluding NTC and Not to work)
                        agent_names = set()
                        for idx, row in processed_df.iterrows():
                            agent_name = row.get(agent_name_col)
                            if pd.notna(agent_name) and str(agent_name).strip():
                                agent_name_str = str(agent_name).strip()
                                agent_name_upper = agent_name_str.upper()
                                # Skip NTC and Not to work as they're not agents
                                if (
                                    agent_name_upper != "NTC"
                                    and "NOT TO WORK"
                                    not in agent_name_upper.replace("-", " ").replace(
                                        "_", " "
                                    )
                                ):
                                    agent_names.add(agent_name_str)

                        # Initialize counts for each agent and insurance
                        for agent_name in agent_names:
                            agent_insurance_data[agent_name] = {}

                        # Count rows by agent name and insurance company
                        for idx, row in processed_df.iterrows():
                            agent_name = row.get(agent_name_col)
                            insurance_value = row.get(insurance_carrier_col)

                            if pd.notna(agent_name) and str(agent_name).strip():
                                agent_name_str = str(agent_name).strip()
                                agent_name_upper = agent_name_str.upper()

                                # Skip NTC and Not to work
                                if (
                                    agent_name_upper == "NTC"
                                    or "NOT TO WORK"
                                    in agent_name_upper.replace("-", " ").replace(
                                        "_", " "
                                    )
                                ):
                                    continue

                                if agent_name_str in agent_names:
                                    # Count by insurance company
                                    if pd.notna(insurance_value):
                                        insurance_name = str(insurance_value).strip()
                                        if (
                                            insurance_name
                                        ):  # Only count non-empty insurance names
                                            agent_insurance_data[agent_name_str][
                                                insurance_name
                                            ] = (
                                                agent_insurance_data[
                                                    agent_name_str
                                                ].get(insurance_name, 0)
                                                + 1
                                            )

                        # Calculate totals for each agent
                        agent_totals = {}
                        for agent_name in agent_names:
                            agent_totals[agent_name] = sum(
                                agent_insurance_data[agent_name].values()
                            )

                        # Calculate overall grand total
                        overall_grand_total = sum(agent_totals.values())

                        # Build the dataframe
                        # Columns: Row Labels, Count of Insurance rows
                        columns = ["Row Labels", "Count of Insurance rows"]
                        rows_data = []

                        # Sort agent names for consistent ordering
                        sorted_agent_names = sorted(agent_names)

                        # Add rows for each agent and their insurance companies
                        for agent_name in sorted_agent_names:
                            # Add agent name row with total
                            agent_total = agent_totals[agent_name]
                            rows_data.append([agent_name, agent_total])

                            # Add insurance company sub-rows for this agent
                            insurance_companies = sorted(
                                agent_insurance_data[agent_name].items(),
                                key=lambda x: x[1],
                                reverse=True,
                            )
                            for insurance_name, count in insurance_companies:
                                rows_data.append([insurance_name, count])

                        # Add Grand Total row
                        rows_data.append(["Grand Total", overall_grand_total])

                        # Create dataframe
                        agent_insurance_df = pd.DataFrame(rows_data, columns=columns)

                        # Write Agent Insurance sheet
                        agent_insurance_df.to_excel(
                            writer, sheet_name="Agent Insurance", index=False
                        )

                        # Store agent names for later formatting (we'll add comments after writer closes)
                        # Store in a way that's accessible during formatting
                        agent_insurance_agent_names = sorted_agent_names.copy()

                # Create Priority Appointment Pending sheet
                if processed_df is not None:
                    # Find Office Name and Appointment Date columns
                    office_name_col = None
                    appointment_date_col = None

                    for col in processed_df.columns:
                        if "office" in col.lower() and "name" in col.lower():
                            office_name_col = col
                        if "appointment" in col.lower() and "date" in col.lower():
                            appointment_date_col = col

                    if office_name_col and appointment_date_col:
                        # Get all unique appointment dates (reuse the logic from Priority Status sheet)
                        appointment_dates_dict = (
                            {}
                        )  # key: YYYY-MM-DD, value: MM/DD/YYYY

                        for idx, row in processed_df.iterrows():
                            appt_date = row.get(appointment_date_col)
                            if pd.notna(appt_date):
                                # Convert to date object
                                date_obj = None
                                if hasattr(appt_date, "date"):
                                    date_obj = appt_date.date()
                                elif hasattr(appt_date, "strftime"):
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        date_obj = appt_date
                                else:
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        continue

                                if date_obj:
                                    date_key = date_obj.strftime("%Y-%m-%d")
                                    date_display = date_obj.strftime("%m/%d/%Y")
                                    appointment_dates_dict[date_key] = date_display

                        # Sort dates by key (YYYY-MM-DD) and create lists
                        sorted_date_keys = sorted(appointment_dates_dict.keys())
                        appointment_dates = sorted_date_keys  # For matching
                        appointment_dates_display = [
                            appointment_dates_dict[key] for key in sorted_date_keys
                        ]  # For display

                        # Initialize data structure: office_name -> date -> count
                        office_date_data = {}

                        # Get all unique office names
                        office_names = set()
                        for idx, row in processed_df.iterrows():
                            office_name = row.get(office_name_col)
                            if pd.notna(office_name) and str(office_name).strip():
                                office_name_str = str(office_name).strip()
                                office_names.add(office_name_str)

                        # Initialize counts for each office and date
                        for office_name in office_names:
                            office_date_data[office_name] = {}
                            for date_key in appointment_dates:
                                office_date_data[office_name][date_key] = 0

                        # Count rows by office name and appointment date
                        for idx, row in processed_df.iterrows():
                            office_name = row.get(office_name_col)
                            appt_date = row.get(appointment_date_col)

                            if pd.notna(office_name) and pd.notna(appt_date):
                                office_name_str = str(office_name).strip()

                                # Convert appointment date to YYYY-MM-DD format
                                date_obj = None
                                if hasattr(appt_date, "date"):
                                    date_obj = appt_date.date()
                                elif hasattr(appt_date, "strftime"):
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        try:
                                            date_obj = pd.to_datetime(
                                                str(appt_date)
                                            ).date()
                                        except:
                                            continue
                                else:
                                    try:
                                        date_obj = pd.to_datetime(appt_date).date()
                                    except:
                                        continue

                                if date_obj and office_name_str in office_names:
                                    date_str = date_obj.strftime("%Y-%m-%d")

                                    if date_str in appointment_dates:
                                        office_date_data[office_name_str][date_str] = (
                                            office_date_data[office_name_str].get(
                                                date_str, 0
                                            )
                                            + 1
                                        )

                        # Calculate totals
                        # Totals for each office
                        office_totals = {}
                        for office_name in office_names:
                            office_totals[office_name] = sum(
                                office_date_data[office_name].values()
                            )

                        # Totals for each date column
                        date_totals = {}
                        for date_key in appointment_dates:
                            date_totals[date_key] = sum(
                                office_date_data[office_name][date_key]
                                for office_name in office_names
                            )

                        # Overall grand total
                        overall_grand_total = sum(office_totals.values())

                        # Build the dataframe
                        # Columns: Row Labels, [date columns], Grand Total
                        columns = (
                            ["Row Labels"] + appointment_dates_display + ["Grand Total"]
                        )
                        rows_data = []

                        # Sort office names for consistent ordering
                        sorted_office_names = sorted(office_names)

                        # Add rows for each office
                        for office_name in sorted_office_names:
                            office_total = office_totals[office_name]

                            row_data = [office_name]
                            for date_key in appointment_dates:
                                row_data.append(office_date_data[office_name][date_key])
                            row_data.append(office_total)
                            rows_data.append(row_data)

                        # Add Grand Total row
                        grand_total_row = ["Grand Total"]
                        for date_key in appointment_dates:
                            grand_total_row.append(date_totals[date_key])
                        grand_total_row.append(overall_grand_total)
                        rows_data.append(grand_total_row)

                        # Create dataframe
                        priority_appointment_pending_df = pd.DataFrame(
                            rows_data, columns=columns
                        )

                        # Write Priority Appointment Pending sheet
                        priority_appointment_pending_df.to_excel(
                            writer,
                            sheet_name="Priority Appointment Pending",
                            index=False,
                        )

            # Apply formatting to make certain text bold
            from openpyxl import load_workbook
            from openpyxl.styles import Font

            wb = load_workbook(temp_path)

            # Format Agent Count Summary sheet
            if "Agent Count Summary" in wb.sheetnames:
                ws = wb["Agent Count Summary"]
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1),
                    start=2,
                ):
                    cell_value = row[0].value
                    if cell_value and isinstance(cell_value, str):
                        cell_str = cell_value.strip()
                        # Make only Grand Total bold
                        if cell_str == "Grand Total":
                            # Make entire row bold
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).font = Font(
                                    bold=True
                                )

            # Format Priority Status sheet
            if "Priority Status" in wb.sheetnames:
                ws = wb["Priority Status"]
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1),
                    start=2,
                ):
                    cell_value = row[0].value
                    if cell_value and isinstance(cell_value, str):
                        cell_str = cell_value.strip()
                        # Make only Grand Total bold
                        if cell_str == "Grand Total":
                            # Make entire row bold
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).font = Font(
                                    bold=True
                                )

            # Format Priority Remark sheet
            if "Priority Remark" in wb.sheetnames:
                ws = wb["Priority Remark"]
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1),
                    start=2,
                ):
                    cell_value = row[0].value
                    if cell_value and isinstance(cell_value, str):
                        cell_str = cell_value.strip()
                        # Make First Priority, Second Priority, Third Priority, and Grand Total bold
                        if cell_str in [
                            "First Priority",
                            "Second Priority",
                            "Third Priority",
                            "Grand Total",
                        ]:
                            row[0].font = Font(bold=True)
                            # Make entire row bold
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).font = Font(
                                    bold=True
                                )

            # Format Today Allocation sheet
            if "Today Allocation" in wb.sheetnames:
                ws = wb["Today Allocation"]
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1),
                    start=2,
                ):
                    cell_value = row[0].value
                    if cell_value and isinstance(cell_value, str):
                        cell_str = cell_value.strip()
                        # Make only Grand Total bold
                        if cell_str == "Grand Total":
                            # Make entire row bold
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).font = Font(
                                    bold=True
                                )

            # Format NTBP Allocation sheet
            if "NTBP Allocation" in wb.sheetnames:
                ws = wb["NTBP Allocation"]
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1),
                    start=2,
                ):
                    cell_value = row[0].value
                    if cell_value and isinstance(cell_value, str):
                        cell_str = cell_value.strip()
                        # Make only Grand Total bold
                        if cell_str == "Grand Total":
                            # Make entire row bold
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).font = Font(
                                    bold=True
                                )

            # Format NTC Allocation sheet
            if "NTC Allocation" in wb.sheetnames:
                ws = wb["NTC Allocation"]
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1),
                    start=2,
                ):
                    cell_value = row[0].value
                    if cell_value and isinstance(cell_value, str):
                        cell_str = cell_value.strip()
                        # Make only NTC and Grand Total bold
                        if cell_str in ["Grand Total", "NTC"]:
                            # Make entire row bold
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).font = Font(
                                    bold=True
                                )

            # Format Agent Insurance sheet
            if "Agent Insurance" in wb.sheetnames:
                ws = wb["Agent Insurance"]

                # Check for agent rows using stored agent names (most reliable method)
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1),
                    start=2,
                ):
                    cell_value = row[0].value
                    if cell_value and isinstance(cell_value, str):
                        cell_str = cell_value.strip()
                        if cell_str == "Grand Total":
                            # Make entire row bold
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).font = Font(
                                    bold=True
                                )
                        else:
                            # Only bold rows that are in the stored agent names list
                            # This ensures only agent names are bold, not insurance names
                            if (
                                agent_insurance_agent_names
                                and cell_str in agent_insurance_agent_names
                            ):
                                # Make agent name and count bold
                                row[0].font = Font(bold=True)
                                if ws.max_column >= 2:
                                    ws.cell(row=row_idx, column=2).font = Font(
                                        bold=True
                                    )

            # Format NTC Insurance Name and counts sheet
            if "NTC Insurance Name and counts" in wb.sheetnames:
                ws = wb["NTC Insurance Name and counts"]
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1),
                    start=2,
                ):
                    cell_value = row[0].value
                    if cell_value and isinstance(cell_value, str):
                        cell_str = cell_value.strip()
                        if cell_str == "Grand Total":
                            # Make entire row bold
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).font = Font(
                                    bold=True
                                )

            # Format Priority Appointment Pending sheet
            if "Priority Appointment Pending" in wb.sheetnames:
                ws = wb["Priority Appointment Pending"]
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1),
                    start=2,
                ):
                    cell_value = row[0].value
                    if cell_value and isinstance(cell_value, str):
                        cell_str = cell_value.strip()
                        if cell_str == "Grand Total":
                            # Make entire row bold
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).font = Font(
                                    bold=True
                                )

            wb.save(temp_path)
            wb.close()

            return send_file(temp_path, as_attachment=True, download_name=filename)

        finally:
            # Clean up temporary file
            os.close(temp_fd)
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


def validate_agent_work_file_columns(file_data):
    """Validate that the uploaded file contains all required columns"""
    # Required columns list
    required_columns = [
        "Office Name",
        "Appointment Date",
        "Patient ID",
        "Patient Name",
        "Dental Primary Ins Carr",
        "Dental Secondary Ins Carr",
        "Received date",
        "Type",
        "Status Code",
        "Comment",
        "Group Number",
        "Category",
        "Agent Name",
        "Work Date",
        "Remark",
        "Priority Status",
        "QC Agent",
        "QC Status",
        "QC Comments",
        "QC Date",
    ]

    # Check each sheet in the file
    if not file_data:
        return False, ["No data found in file"]

    # Collect all column names from all sheets (case-insensitive)
    all_columns_found = set()

    for sheet_name, df in file_data.items():
        if df.empty:
            continue

        # Get actual column names from the DataFrame
        actual_columns = [str(col).strip() for col in df.columns]

        # Add all columns (normalized to lowercase for comparison)
        for col in actual_columns:
            all_columns_found.add(col.strip().lower())

    # Check which required columns are missing
    missing_columns = []

    for required_col in required_columns:
        found = False
        required_col_lower = required_col.strip().lower()

        # Check if column exists (case-insensitive)
        for found_col_lower in all_columns_found:
            if found_col_lower == required_col_lower:
                found = True
                break

        if not found:
            missing_columns.append(required_col)

    # Return True if all columns are found
    if not missing_columns:
        return True, []

    return False, missing_columns


@app.route("/upload_work_file", methods=["POST"])
@agent_required
def upload_work_file():
    """Upload agent work file with data changes"""
    if "file" not in request.files:
        return jsonify({"success": False, "message": "No file provided"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"success": False, "message": "No file selected"}), 400

    notes = request.form.get("notes", "")

    try:
        # Get current agent
        user_id = session.get("user_id")
        if not user_id:
            return jsonify({"success": False, "message": "User not found"}), 400

        # Try to find user by ID first, then by email/google_id
        user = User.query.filter_by(id=user_id, is_active=True).first()
        if not user:
            # If not found by ID, try by email (for Google OAuth users)
            user = User.query.filter_by(email=user_id, is_active=True).first()
        if not user:
            # If still not found, try by google_id
            user = User.query.filter_by(google_id=user_id, is_active=True).first()

        if not user:
            return jsonify({"success": False, "message": "User not found"}), 400

        # Save uploaded file
        filename = secure_filename(file.filename)
        file.save(filename)

        # Load and process Excel file
        try:
            # Use parse_dates=False to prevent automatic date parsing that differs between Windows and Mac
            file_data = pd.read_excel(filename, sheet_name=None, parse_dates=False)

            # Validate required columns
            is_valid, missing_columns = validate_agent_work_file_columns(file_data)

            if not is_valid:
                # Clean up uploaded file
                if os.path.exists(filename):
                    os.remove(filename)

                # Format error message
                if len(missing_columns) == 1:
                    error_message = f"Missing required column: {missing_columns[0]}"
                else:
                    error_message = (
                        f'Missing required columns: {", ".join(missing_columns)}'
                    )

                return (
                    jsonify(
                        {
                            "success": False,
                            "message": error_message,
                            "missing_columns": missing_columns,
                        }
                    ),
                    400,
                )

            # Clear all existing agent work files before saving new one
            existing_files = AgentWorkFile.query.filter_by(agent_id=user.id).all()
            for existing_file in existing_files:
                db.session.delete(existing_file)
            db.session.commit()

            # Save new file to database
            work_file = save_agent_work_file(
                agent_id=user.id, filename=filename, file_data=file_data, notes=notes
            )

            # Clean up uploaded file
            if os.path.exists(filename):
                os.remove(filename)

            return jsonify(
                {
                    "success": True,
                    "message": f"Work file uploaded successfully: {filename} (Previous files cleared)",
                }
            )

        except Exception as e:
            # Clean up uploaded file on error
            if os.path.exists(filename):
                os.remove(filename)
            return (
                jsonify(
                    {
                        "success": False,
                        "message": f"Error processing Excel file: {str(e)}",
                    }
                ),
                500,
            )

    except Exception as e:
        return (
            jsonify(
                {"success": False, "message": f"Error uploading work file: {str(e)}"}
            ),
            500,
        )


@app.route("/upload_status", methods=["POST"])
@agent_required
def upload_status_file():
    """Legacy route - redirect to new work file upload"""
    return redirect(url_for("upload_work_file"))


@app.route("/consolidate_agent_files", methods=["POST"])
@admin_required
def consolidate_agent_files():
    """Consolidate all agent work files into one Excel file"""
    try:
        # Get all agent work files (all files regardless of status - includes previously consolidated and newly uploaded)
        work_files = AgentWorkFile.query.order_by(
            AgentWorkFile.upload_date.desc()
        ).all()

        if not work_files:
            flash("No agent work files found to consolidate", "warning")
            return redirect("/")

        # Create Excel buffer
        excel_buffer = io.BytesIO()

        # Helper function to find remark column
        def find_remark_column(df):
            """Find the remark column (case-insensitive)"""
            for col in df.columns:
                if col.lower() in ["remark", "remarks"]:
                    return col
            return None

        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            # First pass: Collect all unique remark statuses across all agents
            all_remark_statuses = set()
            agent_remarks_data = {}  # Store remarks data per agent

            for work_file in work_files:
                file_data = work_file.get_file_data()
                agent_name = work_file.agent.name
                remarks_dict = (
                    {}
                )  # Dictionary to store all remarks with counts for this agent

                if file_data:
                    if isinstance(file_data, dict):
                        # Multiple sheets - process all sheets except Summary
                        for sheet_name, sheet_data in file_data.items():
                            # Skip Summary sheet
                            if sheet_name.lower() == "summary":
                                continue

                            if isinstance(sheet_data, pd.DataFrame):
                                # Get remark column
                                remark_col = find_remark_column(sheet_data)

                                if remark_col:
                                    # Get all remarks (including NaN)
                                    remark_data = sheet_data[remark_col]

                                    # Count all non-empty remarks by status
                                    non_empty_remarks = remark_data.dropna()
                                    for remark in non_empty_remarks:
                                        # Normalize remark (strip whitespace, handle case)
                                        remark_normalized = str(remark).strip()
                                        if remark_normalized:
                                            remark_lower = remark_normalized.lower()
                                            # Use original case for display, but normalize for counting
                                            if remark_lower in remarks_dict:
                                                remarks_dict[remark_lower] = {
                                                    "count": remarks_dict[remark_lower][
                                                        "count"
                                                    ]
                                                    + 1,
                                                    "display": remarks_dict[
                                                        remark_lower
                                                    ][
                                                        "display"
                                                    ],  # Keep first seen case
                                                }
                                            else:
                                                remarks_dict[remark_lower] = {
                                                    "count": 1,
                                                    "display": remark_normalized,  # Keep original case
                                                }
                    elif isinstance(file_data, pd.DataFrame):
                        # Single DataFrame
                        # Get remark column
                        remark_col = find_remark_column(file_data)

                        if remark_col:
                            # Get all remarks (including NaN)
                            remark_data = file_data[remark_col]

                            # Count all non-empty remarks by status
                            non_empty_remarks = remark_data.dropna()
                            for remark in non_empty_remarks:
                                # Normalize remark (strip whitespace, handle case)
                                remark_normalized = str(remark).strip()
                                if remark_normalized:
                                    remark_lower = remark_normalized.lower()
                                    # Use original case for display, but normalize for counting
                                    if remark_lower in remarks_dict:
                                        remarks_dict[remark_lower] = {
                                            "count": remarks_dict[remark_lower]["count"]
                                            + 1,
                                            "display": remarks_dict[remark_lower][
                                                "display"
                                            ],  # Keep first seen case
                                        }
                                    else:
                                        remarks_dict[remark_lower] = {
                                            "count": 1,
                                            "display": remark_normalized,  # Keep original case
                                        }

                # Store remarks data for this agent
                agent_remarks_data[agent_name] = remarks_dict

                # Collect all unique remark statuses (using display name)
                for remark_info in remarks_dict.values():
                    all_remark_statuses.add(remark_info["display"])

            # Add empty remarks as a status if needed
            all_remark_statuses.add("(Empty/No Remark)")

            # Sort remark statuses alphabetically for consistent column order
            sorted_remark_statuses = sorted(all_remark_statuses)

            # Second pass: Create summary data with all remark statuses as columns
            summary_data = []

            for work_file in work_files:
                file_data = work_file.get_file_data()
                agent_name = work_file.agent.name
                total_assigned_count = 0
                completed_count = 0
                empty_remarks_count = 0

                # Calculate counts from file data
                if file_data:
                    if isinstance(file_data, dict):
                        # Multiple sheets - count rows from all sheets except Summary
                        for sheet_name, sheet_data in file_data.items():
                            # Skip Summary sheet
                            if sheet_name.lower() == "summary":
                                continue

                            if isinstance(sheet_data, pd.DataFrame):
                                # Total assigned count = all rows (excluding header)
                                total_assigned_count += len(sheet_data)

                                # Get remark column
                                remark_col = find_remark_column(sheet_data)

                                if remark_col:
                                    # Get all remarks (including NaN)
                                    remark_data = sheet_data[remark_col]

                                    # Count empty/NaN remarks
                                    empty_remarks_count += remark_data.isna().sum()

                                    # Count completed (non-Workable remarks)
                                    non_empty_remarks = remark_data.dropna()
                                    # Convert to string first to handle mixed types
                                    if len(non_empty_remarks) > 0:
                                        non_empty_remarks_str = (
                                            non_empty_remarks.astype(str).str.lower()
                                        )
                                        completed_count += len(
                                            non_empty_remarks_str[
                                                non_empty_remarks_str != "workable"
                                            ]
                                        )
                    elif isinstance(file_data, pd.DataFrame):
                        # Single DataFrame
                        # Total assigned count = all rows (excluding header)
                        total_assigned_count = len(file_data)

                        # Get remark column
                        remark_col = find_remark_column(file_data)

                        if remark_col:
                            # Get all remarks (including NaN)
                            remark_data = file_data[remark_col]

                            # Count empty/NaN remarks
                            empty_remarks_count = remark_data.isna().sum()

                            # Count completed (non-Workable remarks)
                            non_empty_remarks = remark_data.dropna()
                            # Convert to string first to handle mixed types
                            if len(non_empty_remarks) > 0:
                                non_empty_remarks_str = non_empty_remarks.astype(
                                    str
                                ).str.lower()
                                completed_count = len(
                                    non_empty_remarks_str[
                                        non_empty_remarks_str != "workable"
                                    ]
                                )
                            else:
                                completed_count = 0

                # Create row data for this agent
                row_data = {
                    "Agent": agent_name,
                    "Total Assigned Count": total_assigned_count,
                    "Completed Count": completed_count,
                    "Empty Remarks Count": empty_remarks_count,
                }

                # Get remarks data for this agent
                agent_remarks = agent_remarks_data.get(agent_name, {})

                # Add count for each remark status column
                for remark_status in sorted_remark_statuses:
                    if remark_status == "(Empty/No Remark)":
                        # Use the empty_remarks_count
                        row_data[remark_status] = empty_remarks_count
                    else:
                        # Find matching remark in agent's remarks (case-insensitive)
                        count = 0
                        for remark_lower, remark_info in agent_remarks.items():
                            if remark_info["display"] == remark_status:
                                count = remark_info["count"]
                                break
                        row_data[remark_status] = count

                summary_data.append(row_data)

            # Create summary DataFrame
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

            # Combine all agent data into one sheet
            all_agent_data = []
            for work_file in work_files:
                file_data = work_file.get_file_data()
                if file_data:
                    if isinstance(file_data, dict):
                        # Multiple sheets - combine them (excluding Summary sheets)
                        for sheet_name, sheet_data in file_data.items():
                            # Skip Summary sheet
                            if sheet_name.lower() == "summary":
                                continue

                            if isinstance(sheet_data, pd.DataFrame):
                                sheet_data_copy = sheet_data.copy()
                                all_agent_data.append(sheet_data_copy)
                    elif isinstance(file_data, pd.DataFrame):
                        # Single DataFrame
                        file_data_copy = file_data.copy()
                        all_agent_data.append(file_data_copy)

            # Create combined sheet with all agent data
            if all_agent_data:
                combined_df = pd.concat(all_agent_data, ignore_index=True)

                # Format all date columns to MM/DD/YYYY using robust parser to avoid blanks
                for col in combined_df.columns:
                    if "date" in col.lower():
                        try:
                            # Use existing robust parser for each cell; this handles
                            # Excel serials, mixed string formats, timestamps.
                            parsed_series = combined_df[col].apply(parse_excel_date)
                            # Convert to desired string format, blank if None
                            combined_df[col] = parsed_series.apply(
                                lambda d: d.strftime("%m/%d/%Y") if d else ""
                            )
                            # If entire column became blank but original had non-empty raw values,
                            # fall back to original to avoid losing data unexpectedly.
                            if (
                                parsed_series.notna().sum() == 0
                                and combined_df[col]
                                .astype(str)
                                .str.strip()
                                .ne("")
                                .any()
                            ):
                                combined_df[col] = combined_df[
                                    col
                                ]  # keep blanks (intentional) - no fallback needed
                        except Exception:
                            # If robust parsing fails, keep original values unchanged
                            pass

                combined_df.to_excel(writer, sheet_name="All Agent Data", index=False)
            else:
                # Fallback if no data found
                simple_df = pd.DataFrame(
                    [{"Message": "No data available from any agent"}]
                )
                simple_df.to_excel(writer, sheet_name="All Agent Data", index=False)

        excel_buffer.seek(0)

        # Mark files as consolidated
        for work_file in work_files:
            work_file.status = "consolidated"
        db.session.commit()

        # Return file for download
        filename = (
            f"consolidated_agent_files_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        flash(f"Error consolidating agent files: {str(e)}", "error")
        return redirect("/")


@app.route("/download_agent_work_file/<int:file_id>", methods=["GET"])
@admin_required
def download_agent_work_file(file_id):
    """Download a single agent work file"""
    try:
        # Get the work file
        work_file = AgentWorkFile.query.get_or_404(file_id)

        # Get file data
        file_data = work_file.get_file_data()

        if not file_data:
            flash("File data not found", "error")
            return redirect("/")

        # Create Excel buffer
        excel_buffer = io.BytesIO()

        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            if isinstance(file_data, dict):
                # Multiple sheets
                for sheet_name, sheet_data in file_data.items():
                    if isinstance(sheet_data, pd.DataFrame):
                        # Format date columns to MM/DD/YYYY
                        sheet_data_copy = sheet_data.copy()
                        for col in sheet_data_copy.columns:
                            if "date" in col.lower():
                                try:
                                    sheet_data_copy[col] = pd.to_datetime(
                                        sheet_data_copy[col], errors="coerce"
                                    )
                                    sheet_data_copy[col] = sheet_data_copy[col].apply(
                                        lambda x: (
                                            x.strftime("%m/%d/%Y")
                                            if pd.notna(x)
                                            else ""
                                        )
                                    )
                                except Exception:
                                    pass
                        sheet_data_copy.to_excel(
                            writer, sheet_name=sheet_name, index=False
                        )
            elif isinstance(file_data, pd.DataFrame):
                # Single DataFrame
                file_data_copy = file_data.copy()
                # Format date columns to MM/DD/YYYY
                for col in file_data_copy.columns:
                    if "date" in col.lower():
                        try:
                            file_data_copy[col] = pd.to_datetime(
                                file_data_copy[col], errors="coerce"
                            )
                            file_data_copy[col] = file_data_copy[col].apply(
                                lambda x: x.strftime("%m/%d/%Y") if pd.notna(x) else ""
                            )
                        except Exception:
                            pass
                file_data_copy.to_excel(writer, sheet_name="Sheet1", index=False)
            else:
                # Fallback
                pd.DataFrame([{"Message": "No data available"}]).to_excel(
                    writer, sheet_name="Sheet1", index=False
                )

        excel_buffer.seek(0)

        # Create filename with agent name and original filename
        agent_name = work_file.agent.name.replace(" ", "_")
        original_filename = (
            work_file.filename.rsplit(".", 1)[0]
            if "." in work_file.filename
            else work_file.filename
        )
        download_filename = f"{agent_name}_{original_filename}_{work_file.upload_date.strftime('%Y%m%d')}.xlsx"

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=download_filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        flash(f"Error downloading agent file: {str(e)}", "error")
        return redirect("/")


@app.route("/clear_all_agent_files", methods=["POST"])
@admin_required
def clear_all_agent_files():
    """Clear all agent work files from the database"""
    try:
        # Get all agent work files
        all_files = AgentWorkFile.query.all()
        file_count = len(all_files)

        if file_count > 0:
            for work_file in all_files:
                db.session.delete(work_file)
            db.session.commit()
            flash(f"✅ Successfully cleared {file_count} file(s)", "success")
        else:
            flash("ℹ️ No files found to clear", "info")

        return redirect("/")

    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error clearing files: {str(e)}", "error")
        return redirect("/")


@app.route("/get_appointment_dates")
@login_required
def get_appointment_dates():
    global data_file_data

    if not data_file_data:
        return jsonify({"error": "No data file uploaded"}), 400

    try:
        # Get the first sheet from data file
        data_df = list(data_file_data.values())[0]

        # Find the appointment date column
        appointment_date_col = None
        for col in data_df.columns:
            if "appointment" in col.lower() and "date" in col.lower():
                appointment_date_col = col
                break

        if appointment_date_col is None:
            return jsonify({"error": "Appointment Date column not found"}), 400

        # Parse dates using robust date parsing that works consistently across Windows and Mac
        parsed_dates = data_df[appointment_date_col].apply(parse_excel_date)

        # Get unique valid appointment dates (filter out None/invalid dates)
        valid_dates = parsed_dates.dropna().unique()

        # Convert to string format and count rows for each date
        date_data = []
        for date_obj in valid_dates:
            if date_obj is None:
                continue
            date_str = date_obj.strftime("%Y-%m-%d")

            # Count rows for this specific date
            row_count = len(data_df[parsed_dates == date_obj])

            date_data.append({"date": date_str, "row_count": row_count})

        # Sort by date
        date_data.sort(key=lambda x: x["date"])

        return jsonify(
            {
                "appointment_dates": [item["date"] for item in date_data],
                "appointment_dates_with_counts": date_data,
                "column_name": appointment_date_col,
            }
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/get_receive_dates", methods=["GET"])
@admin_required
def get_receive_dates():
    """Get unique receive dates from data file, optionally filtered by appointment dates"""
    global data_file_data

    if not data_file_data:
        return jsonify({"error": "No data file uploaded"}), 400

    try:
        # Get the first sheet from data file
        data_df = list(data_file_data.values())[0]

        # Find the receive date column
        receive_date_col = None
        for col in data_df.columns:
            if "receive" in col.lower() and "date" in col.lower():
                receive_date_col = col
                break

        if receive_date_col is None:
            return jsonify({"error": "Receive Date column not found"}), 400

        # Get appointment dates from query parameters
        appointment_dates = request.args.getlist("appointment_dates")

        # Filter data based on selected appointment dates if provided
        filtered_df = data_df
        if appointment_dates:
            # Find the appointment date column
            appointment_date_col = None
            for col in data_df.columns:
                if "appointment" in col.lower() and "date" in col.lower():
                    appointment_date_col = col
                    break

            if appointment_date_col:
                # Parse appointment dates using robust date parsing
                parsed_appointment_dates = data_df[appointment_date_col].apply(
                    parse_excel_date
                )

                # Convert selected appointment dates to date objects for comparison
                appointment_dates_formatted = []
                for date_str in appointment_dates:
                    try:
                        from datetime import datetime

                        parsed_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                        appointment_dates_formatted.append(parsed_date)
                    except:
                        pass  # Skip invalid dates

                # Filter rows where parsed appointment date matches any of the selected dates
                mask = parsed_appointment_dates.isin(appointment_dates_formatted)
                filtered_df = data_df[mask]

        # Parse receive dates using robust date parsing that works consistently across Windows and Mac
        parsed_receive_dates = filtered_df[receive_date_col].apply(parse_excel_date)

        # Get unique valid receive dates (filter out None/invalid dates)
        valid_receive_dates = parsed_receive_dates.dropna().unique()

        # Convert to string format and sort
        date_strings = []
        for date_obj in valid_receive_dates:
            if date_obj is None:
                continue
            date_str = date_obj.strftime("%Y-%m-%d")
            date_strings.append(date_str)

        date_strings.sort()

        return jsonify(
            {
                "receive_dates": date_strings,
                "column_name": receive_date_col,
                "filtered_by_appointment_dates": (
                    len(appointment_dates) > 0 if appointment_dates else False
                ),
            }
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/get_agent_allocation", methods=["POST"])
@admin_required
def get_agent_allocation():
    global data_file_data, agent_allocations_data

    if not data_file_data or not agent_allocations_data:
        return jsonify({"error": "No data available"}), 400

    agent_id = request.json.get("agent_id")
    agent_name = request.json.get("agent_name")

    if not agent_id and not agent_name:
        return (
            jsonify({"error": "No agent specified (agent_id or agent_name required)"}),
            400,
        )

    try:
        # Find the agent in allocations data
        agent_info = None

        # First try to find by agent_id if provided (most reliable)
        if agent_id:
            for agent in agent_allocations_data:
                if agent.get("id") == agent_id:
                    agent_info = agent
                    break

        # If not found by ID and name is provided, try by name
        if not agent_info and agent_name:
            matching_agents = [
                agent
                for agent in agent_allocations_data
                if agent.get("name") == agent_name
            ]
            if len(matching_agents) == 1:
                agent_info = matching_agents[0]
            elif len(matching_agents) > 1:
                # Multiple agents with same name - require agent_id
                return (
                    jsonify(
                        {
                            "error": f'Multiple agents found with name "{agent_name}". Please use agent_id instead.',
                            "agents": [
                                {"id": a.get("id"), "name": a.get("name")}
                                for a in matching_agents
                            ],
                        }
                    ),
                    400,
                )

        if not agent_info:
            return jsonify({"error": "Agent not found"}), 404

        # Get the processed data
        processed_df = list(data_file_data.values())[0]

        # Get the specific rows allocated to this agent
        agent_rows = agent_info["allocated"]
        row_indices = agent_info.get("row_indices", [])
        agent_name = agent_info.get("name", "Unknown")

        # Use row_indices as primary source (most reliable) since they track actual allocations
        # row_indices are the source of truth - they contain the actual rows allocated to this agent
        if row_indices and len(row_indices) > 0:
            # Filter to only valid indices within the dataframe
            valid_indices = [idx for idx in row_indices if idx < len(processed_df)]

            if valid_indices:
                # Get rows by indices (this is the source of truth)
                agent_df = processed_df.iloc[valid_indices].copy()

                # Then verify/filter by Agent Name column if it exists (for data integrity check)
                if "Agent Name" in agent_df.columns:
                    # Check if Agent Name matches (case-insensitive, handle whitespace)
                    agent_name_matches = (
                        agent_df["Agent Name"].astype(str).str.strip().str.lower()
                        == agent_name.lower().strip()
                    )
                    empty_agent_name = agent_df["Agent Name"].isna() | (
                        agent_df["Agent Name"].astype(str).str.strip() == ""
                    )

                    # If some rows don't have matching Agent Name, still include them (row_indices are source of truth)
                    # But prioritize rows with matching Agent Name if available
                    matching_rows = agent_df[agent_name_matches].copy()
                    empty_rows = agent_df[empty_agent_name].copy()

                    if not matching_rows.empty:
                        # Use rows with matching Agent Name
                        agent_df = matching_rows.copy()
                    elif not empty_rows.empty:
                        # Use rows with empty Agent Name (they were allocated but name not set)
                        agent_df = empty_rows.copy()
                    # else: use all rows from row_indices even if name doesn't match (data integrity issue)
        else:
            agent_df = pd.DataFrame()

        # If no row_indices, try filtering by Agent Name column
        if agent_df.empty and not row_indices:
            if "Agent Name" in processed_df.columns:
                agent_df = processed_df[
                    processed_df["Agent Name"].astype(str).str.strip().str.lower()
                    == agent_name.lower().strip()
                ].copy()
            else:
                # Fallback: if row_indices not available and no Agent Name column, use first N rows
                if len(processed_df) >= agent_rows:
                    agent_df = processed_df.head(agent_rows).copy()
                else:
                    agent_df = processed_df.copy()

        # Add serial number column
        agent_df_with_sr = agent_df.copy()
        agent_df_with_sr.insert(0, "Sr No", range(1, len(agent_df_with_sr) + 1))

        # Convert dataframe to HTML table
        html_table = agent_df_with_sr.to_html(
            classes="modal-table", table_id="agentDataTable", escape=False, index=False
        )

        # Calculate statistics
        total_rows = len(agent_df)
        first_priority = (
            len(agent_df[agent_df["Priority Status"] == "First Priority"])
            if "Priority Status" in agent_df.columns
            else 0
        )
        second_priority = (
            len(agent_df[agent_df["Priority Status"] == "Second Priority"])
            if "Priority Status" in agent_df.columns
            else 0
        )
        third_priority = (
            len(agent_df[agent_df["Priority Status"] == "Third Priority"])
            if "Priority Status" in agent_df.columns
            else 0
        )

        return jsonify(
            {
                "success": True,
                "agent_id": agent_info.get("id"),
                "agent_name": agent_info.get("name"),
                "html_table": html_table,
                "stats": {
                    "total_rows": total_rows,
                    "capacity": agent_info["capacity"],
                    "first_priority": first_priority,
                    "second_priority": second_priority,
                    "third_priority": third_priority,
                },
            }
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/download_agent_file", methods=["POST"])
@admin_required
def download_agent_file():
    global data_file_data, agent_allocations_data

    if not data_file_data or not agent_allocations_data:
        return jsonify({"error": "No data available for download"}), 400

    agent_id = request.form.get("agent_id")
    agent_name = request.form.get("agent_name")

    if not agent_id and not agent_name:
        return (
            jsonify({"error": "No agent specified (agent_id or agent_name required)"}),
            400,
        )

    # Find the agent
    agent_info = None
    if agent_id:
        for agent in agent_allocations_data:
            if agent.get("id") == agent_id:
                agent_info = agent
                break

    if not agent_info and agent_name:
        matching_agents = [
            agent for agent in agent_allocations_data if agent.get("name") == agent_name
        ]
        if len(matching_agents) == 1:
            agent_info = matching_agents[0]
        elif len(matching_agents) > 1:
            return (
                jsonify(
                    {
                        "error": f'Multiple agents found with name "{agent_name}". Please use agent_id instead.',
                        "agents": [
                            {"id": a.get("id"), "name": a.get("name")}
                            for a in matching_agents
                        ],
                    }
                ),
                400,
            )

    if not agent_info:
        return jsonify({"error": "Agent not found"}), 404

    agent_name = agent_info.get("name", "Unknown")

    # Generate filename with agent name and today's date
    from datetime import datetime

    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"{agent_name}_{today}.xlsx"

    try:
        # Get the processed data
        processed_df = list(data_file_data.values())[0]

        # Get the specific rows allocated to this agent
        agent_rows = agent_info["allocated"]
        row_indices = agent_info.get("row_indices", [])

        # CRITICAL: Filter by Agent Name column first to ensure only this agent's rows are included
        # This is the most reliable way to ensure correct filtering
        if "Agent Name" in processed_df.columns:
            # Filter to only include rows where Agent Name matches this agent exactly
            agent_df = processed_df[processed_df["Agent Name"] == agent_name].copy()

            # If no rows found by name, fall back to row_indices
            if agent_df.empty and row_indices:
                # Verify row_indices match Agent Name column
                valid_indices = []
                for idx in row_indices:
                    if idx < len(processed_df):
                        row_agent_name = processed_df.at[idx, "Agent Name"]
                        # Check if agent name matches (handle NaN/empty values)
                        if (
                            pd.notna(row_agent_name)
                            and str(row_agent_name).strip() == agent_name
                        ):
                            valid_indices.append(idx)

                if valid_indices:
                    agent_df = processed_df.iloc[valid_indices].copy()
        else:
            # If Agent Name column doesn't exist, use row_indices
            if (
                row_indices
                and len(row_indices) > 0
                and len(processed_df) > max(row_indices)
            ):
                agent_df = processed_df.iloc[row_indices].copy()
            # Fallback: if row_indices not available, use first N rows
            if len(processed_df) >= agent_rows:
                agent_df = processed_df.head(agent_rows).copy()
            else:
                agent_df = processed_df.copy()

        # Ensure Agent Name column is set correctly for all rows
        agent_df["Agent Name"] = agent_name
        agent_df["Allocated Rows"] = agent_rows
        agent_df["Agent Capacity"] = agent_info["capacity"]

        # Create a temporary file
        temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx")

        try:
            with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
                # Create a copy of the dataframe to avoid modifying the original
                agent_df_copy = agent_df.copy()

                # Find appointment date and received date columns and format them as MM/DD/YYYY
                for col in agent_df_copy.columns:
                    if ("appointment" in col.lower() and "date" in col.lower()) or (
                        "receive" in col.lower() and "date" in col.lower()
                    ):
                        # Convert to datetime and then format as MM/DD/YYYY (no time)
                        agent_df_copy[col] = pd.to_datetime(
                            agent_df_copy[col], errors="coerce"
                        ).dt.strftime("%m/%d/%Y")

                agent_df_copy.to_excel(
                    writer, sheet_name=f"{agent_name}_Allocation", index=False
                )

                # Add a summary sheet
                summary_data = {
                    "Metric": [
                        "Agent Name",
                        "Total Allocated Rows",
                        "Agent Capacity",
                        "First Priority Rows",
                        "Second Priority Rows",
                        "Third Priority Rows",
                    ],
                    "Value": [
                        agent_name,
                        agent_rows,
                        agent_info["capacity"],
                        (
                            len(
                                agent_df[
                                    agent_df["Priority Status"] == "First Priority"
                                ]
                            )
                            if "Priority Status" in agent_df.columns
                            else 0
                        ),
                        (
                            len(
                                agent_df[
                                    agent_df["Priority Status"] == "Second Priority"
                                ]
                            )
                            if "Priority Status" in agent_df.columns
                            else 0
                        ),
                        (
                            len(
                                agent_df[
                                    agent_df["Priority Status"] == "Third Priority"
                                ]
                            )
                            if "Priority Status" in agent_df.columns
                            else 0
                        ),
                    ],
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name="Summary", index=False)

            return send_file(temp_path, as_attachment=True, download_name=filename)

        finally:
            # Clean up temporary file
            os.close(temp_fd)
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/send_approval_email", methods=["POST"])
@admin_required
def send_approval_email():
    try:
        data = request.get_json()
        agent_id = data.get("agent_id")
        agent_name = data.get("agent_name")

        if (not agent_id and not agent_name) or not agent_allocations_data:
            return jsonify({"success": False, "message": "Agent ID or name required"})

        # Find the agent in the allocation data
        agent_info = None
        if agent_id:
            for agent in agent_allocations_data:
                if agent.get("id") == agent_id:
                    agent_info = agent
                    break
        if not agent_info and agent_name:
            matching_agents = [
                agent
                for agent in agent_allocations_data
                if agent.get("name") == agent_name
            ]
            if len(matching_agents) == 1:
                agent_info = matching_agents[0]
            elif len(matching_agents) > 1:
                return jsonify(
                    {
                        "success": False,
                        "message": f'Multiple agents found with name "{agent_name}". Please use agent_id instead.',
                        "agents": [
                            {"id": a.get("id"), "name": a.get("name")}
                            for a in matching_agents
                        ],
                    }
                )

        if not agent_info:
            return jsonify({"success": False, "message": "Agent not found"})

        # Get agent's email from allocation data
        agent_email = agent_info.get("email")
        if not agent_email:
            return jsonify({"success": False, "message": "Agent email not found"})

        # Get allocation summary
        summary = get_allocation_summary(agent_name, agent_info)

        # Create Excel file with agent's allocated data
        excel_buffer = create_agent_excel_file(agent_name, agent_info)

        # Format insurance companies list
        insurance_list = (
            ", ".join(sorted(summary["insurance_companies"]))
            if summary["insurance_companies"]
            else "None"
        )

        # Format first priority deadline
        deadline_text = ""
        if summary["first_priority_deadline"]:
            deadline_text = summary["first_priority_deadline"].strftime(
                "%Y-%m-%d at %I:%M %p"
            )
        else:
            deadline_text = "N/A (No First Priority work assigned)"

        # Prepare email content
        text_content = f"""
Dear {agent_name},

Your work allocation has been approved and is attached to this email.

📊 ALLOCATION SUMMARY:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Total Allocated: {summary['total_allocated']} rows
• Your Capacity: {agent_info['capacity']} rows
• Allocation Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

📋 PRIORITY BREAKDOWN:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• First Priority: {summary['first_priority_count']} rows
• Second Priority: {summary['second_priority_count']} rows
• Third Priority: {summary['third_priority_count']} rows
• Unknown/Other: {summary['unknown_priority_count']} rows

🏥 INSURANCE COMPANIES ({len(summary['insurance_companies'])} unique):
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{insurance_list}

⏰ FIRST PRIORITY DEADLINE:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{deadline_text}

Please find your allocated data in the attached Excel file.

Best regards,
Allocation Management System
        """

        html_content = f"""
            <h2>Work Allocation Approved</h2>
            <p>Dear <strong>{agent_name}</strong>,</p>
            <p>Your work allocation has been approved and is attached to this email.</p>
            
            <div style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <h3 style="margin-top: 0; color: #333;">📊 Allocation Summary</h3>
                <ul style="list-style: none; padding-left: 0;">
                    <li><strong>Total Allocated:</strong> {summary['total_allocated']} rows</li>
                <li><strong>Your Capacity:</strong> {agent_info['capacity']} rows</li>
                <li><strong>Allocation Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</li>
            </ul>
            </div>
            
            <div style="background-color: #e8f4f8; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <h3 style="margin-top: 0; color: #2c5282;">📋 Priority Breakdown</h3>
                <ul style="list-style: none; padding-left: 0;">
                    <li><strong>First Priority:</strong> {summary['first_priority_count']} rows</li>
                    <li><strong>Second Priority:</strong> {summary['second_priority_count']} rows</li>
                    <li><strong>Third Priority:</strong> {summary['third_priority_count']} rows</li>
                    <li><strong>Unknown/Other:</strong> {summary['unknown_priority_count']} rows</li>
                </ul>
            </div>
            
            <div style="background-color: #fff3cd; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <h3 style="margin-top: 0; color: #856404;">🏥 Insurance Companies ({len(summary['insurance_companies'])} unique)</h3>
                <p style="word-wrap: break-word;">{insurance_list}</p>
            </div>
            
            <div style="background-color: #f8d7da; padding: 15px; border-radius: 5px; margin: 20px 0; border-left: 4px solid #dc3545;">
                <h3 style="margin-top: 0; color: #721c24;">⏰ First Priority Completion Deadline</h3>
                <p style="font-size: 18px; font-weight: bold; color: #721c24; margin: 10px 0;">{deadline_text}</p>
                <p style="font-size: 12px; color: #856404;">Please ensure all First Priority work is completed by this deadline.</p>
            </div>
            
            <p>Please find your allocated data in the attached Excel file.</p>
            
            <p>Best regards,<br>
            Allocation Management System</p>
        """

        # Send email using Resend
        attachment_filename = (
            f'{agent_name}_allocation_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        success, message = send_email_with_resend(
            to_email=agent_email,
            subject=f"Your Work Allocation - {agent_name}",
            html_content=html_content,
            text_content=text_content,
            attachment_data=excel_buffer,
            attachment_filename=attachment_filename,
        )

        if success:
            return jsonify(
                {"success": True, "message": f"Approval email sent to {agent_email}"}
            )
        else:
            return jsonify(
                {"success": False, "message": f"Error sending email: {message}"}
            )

    except Exception as e:
        return jsonify({"success": False, "message": f"Error sending email: {str(e)}"})


@app.route("/approve_all_allocations", methods=["POST"])
@admin_required
def approve_all_allocations():
    """Approve all agent allocations and send emails to all agents"""
    try:
        if not agent_allocations_data:
            return jsonify({"success": False, "message": "No allocation data found"})

        successful_sends = []
        failed_sends = []

        # Loop through all agents and send approval emails
        for agent in agent_allocations_data:
            agent_name = agent.get("name")
            agent_email = agent.get("email")
            allocated = agent.get("allocated", 0)

            # Skip agents with no email or no allocated rows
            if not agent_email:
                failed_sends.append(f"{agent_name}: No email address")
                continue

            if allocated == 0:
                # Skip agents with no allocations (no need to send email)
                continue

            try:
                # Get allocation summary
                summary = get_allocation_summary(agent_name, agent)

                # Create Excel file with agent's allocated data
                excel_buffer = create_agent_excel_file(agent_name, agent)

                # Format insurance companies list
                insurance_list = (
                    ", ".join(sorted(summary["insurance_companies"]))
                    if summary["insurance_companies"]
                    else "None"
                )

                # Format first priority deadline
                deadline_text = ""
                if summary["first_priority_deadline"]:
                    deadline_text = summary["first_priority_deadline"].strftime(
                        "%Y-%m-%d at %I:%M %p"
                    )
                else:
                    deadline_text = "N/A (No First Priority work assigned)"

                # Prepare email content
                text_content = f"""
Dear {agent_name},

Your work allocation has been approved and is attached to this email.

📊 ALLOCATION SUMMARY:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Total Allocated: {summary['total_allocated']} rows
• Your Capacity: {agent['capacity']} rows
• Allocation Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

📋 PRIORITY BREAKDOWN:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• First Priority: {summary['first_priority_count']} rows
• Second Priority: {summary['second_priority_count']} rows
• Third Priority: {summary['third_priority_count']} rows
• Unknown/Other: {summary['unknown_priority_count']} rows

🏥 INSURANCE COMPANIES ({len(summary['insurance_companies'])} unique):
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{insurance_list}

⏰ FIRST PRIORITY DEADLINE:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{deadline_text}

Please find your allocated data in the attached Excel file.

Best regards,
Allocation Management System
                """

                html_content = f"""
                    <h2>Work Allocation Approved</h2>
                    <p>Dear <strong>{agent_name}</strong>,</p>
                    <p>Your work allocation has been approved and is attached to this email.</p>
                    
                    <div style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 20px 0;">
                        <h3 style="margin-top: 0; color: #333;">📊 Allocation Summary</h3>
                        <ul style="list-style: none; padding-left: 0;">
                            <li><strong>Total Allocated:</strong> {summary['total_allocated']} rows</li>
                            <li><strong>Your Capacity:</strong> {agent['capacity']} rows</li>
                            <li><strong>Allocation Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</li>
                        </ul>
                    </div>
                    
                    <div style="background-color: #e8f4f8; padding: 15px; border-radius: 5px; margin: 20px 0;">
                        <h3 style="margin-top: 0; color: #2c5282;">📋 Priority Breakdown</h3>
                        <ul style="list-style: none; padding-left: 0;">
                            <li><strong>First Priority:</strong> {summary['first_priority_count']} rows</li>
                            <li><strong>Second Priority:</strong> {summary['second_priority_count']} rows</li>
                            <li><strong>Third Priority:</strong> {summary['third_priority_count']} rows</li>
                            <li><strong>Unknown/Other:</strong> {summary['unknown_priority_count']} rows</li>
                        </ul>
                    </div>
                    
                    <div style="background-color: #fff3cd; padding: 15px; border-radius: 5px; margin: 20px 0;">
                        <h3 style="margin-top: 0; color: #856404;">🏥 Insurance Companies ({len(summary['insurance_companies'])} unique)</h3>
                        <p style="word-wrap: break-word;">{insurance_list}</p>
                    </div>
                    
                    <div style="background-color: #f8d7da; padding: 15px; border-radius: 5px; margin: 20px 0; border-left: 4px solid #dc3545;">
                        <h3 style="margin-top: 0; color: #721c24;">⏰ First Priority Completion Deadline</h3>
                        <p style="font-size: 18px; font-weight: bold; color: #721c24; margin: 10px 0;">{deadline_text}</p>
                        <p style="font-size: 12px; color: #856404;">Please ensure all First Priority work is completed by this deadline.</p>
                    </div>
                    
                    <p>Please find your allocated data in the attached Excel file.</p>
                    
                    <p>Best regards,<br>
                    Allocation Management System</p>
                """

                # Send email using Resend
                attachment_filename = (
                    f'{agent_name}_allocation_{datetime.now().strftime("%Y%m%d")}.xlsx'
                )
                success, message = send_email_with_resend(
                    to_email=agent_email,
                    subject=f"Your Work Allocation - {agent_name}",
                    html_content=html_content,
                    text_content=text_content,
                    attachment_data=excel_buffer,
                    attachment_filename=attachment_filename,
                )

                if success:
                    successful_sends.append(f"{agent_name} ({agent_email})")
                else:
                    failed_sends.append(f"{agent_name}: {message}")

            except Exception as e:
                failed_sends.append(f"{agent_name}: {str(e)}")

        # Prepare response message
        total_agents = len(agent_allocations_data)
        agents_with_allocation = sum(
            1 for a in agent_allocations_data if a.get("allocated", 0) > 0
        )
        successful_count = len(successful_sends)
        failed_count = len(failed_sends)

        if successful_count > 0:
            message = f"Successfully sent approval emails to {successful_count} agent(s): {', '.join([s.split(' (')[0] for s in successful_sends])}"
            if failed_count > 0:
                message += (
                    f". {failed_count} agent(s) failed: {', '.join(failed_sends)}"
                )
        else:
            message = (
                f"No emails sent. Errors: {', '.join(failed_sends)}"
                if failed_sends
                else "No agents with allocations to approve."
            )

        return jsonify(
            {
                "success": successful_count > 0,
                "message": message,
                "details": {
                    "total_agents": total_agents,
                    "agents_with_allocation": agents_with_allocation,
                    "successful": successful_count,
                    "failed": failed_count,
                    "successful_list": successful_sends,
                    "failed_list": failed_sends,
                },
            }
        )

    except Exception as e:
        return jsonify(
            {"success": False, "message": f"Error approving all allocations: {str(e)}"}
        )


@app.route("/view_shift_times", methods=["GET"])
@admin_required
def view_shift_times():
    """Admin endpoint to view all agents' shift information for verification"""
    global agent_allocations_data, allocation_data

    shift_info = []

    # First try to get from agent_allocations_data (after processing)
    if agent_allocations_data:
        try:
            for agent in agent_allocations_data:
                shift_start = agent.get("shift_start_time", "Not set")
                shift_original = agent.get("shift_time_original", "Not set")
                shift_group = agent.get("shift_group")

                # Format shift group name
                group_name = "Not set"
                if shift_group == 1:
                    group_name = "Day Shift"
                elif shift_group == 2:
                    group_name = "Afternoon Shift"
                elif shift_group == 3:
                    group_name = "Night Shift"

                # Format shift start time for display
                start_time_display = shift_start if shift_start else "Not parsed"
                if shift_start:
                    try:
                        hour, minute = map(int, shift_start.split(":"))
                        if hour < 12:
                            start_time_display = (
                                f"{shift_start} ({hour}:{minute:02d} AM)"
                            )
                        elif hour == 12:
                            start_time_display = f"{shift_start} (12:00 PM)"
                        else:
                            start_time_display = (
                                f"{shift_start} ({hour-12}:{minute:02d} PM)"
                            )
                    except:
                        pass

                shift_info.append(
                    {
                        "agent_id": agent.get("id"),
                        "agent_name": agent.get("name"),
                        "email": agent.get("email", "Not set"),
                        "shift_time_original": shift_original,
                        "shift_start_time_parsed": shift_start,
                        "shift_start_time_display": start_time_display,
                        "shift_group": shift_group,
                        "shift_group_name": group_name,
                        "capacity": agent.get("capacity", 0),
                        "allocated": agent.get("allocated", 0),
                    }
                )

            # Sort by shift start time
            shift_info.sort(
                key=lambda x: (
                    (
                        x["shift_start_time_parsed"]
                        if x["shift_start_time_parsed"]
                        and x["shift_start_time_parsed"] != "Not parsed"
                        else "99:99"
                    ),
                    x["agent_name"],
                )
            )

            return jsonify(
                {
                    "success": True,
                    "total_agents": len(shift_info),
                    "agents": shift_info,
                    "source": "processed",
                }
            )
        except Exception as e:
            return (
                jsonify({"error": f"Error retrieving shift information: {str(e)}"}),
                500,
            )

    # If no processed data, try to extract from raw allocation_data
    if allocation_data:
        try:
            # Get the main sheet
            agent_df = None
            if "main" in allocation_data:
                agent_df = allocation_data["main"]
            elif len(allocation_data) > 0:
                agent_df = list(allocation_data.values())[0]

            if agent_df is not None:
                # Find columns
                agent_name_col = None
                agent_id_col = None
                shift_time_col = None
                shift_group_col = None
                email_col = None
                counts_col = None

                for col in agent_df.columns:
                    col_lower = col.lower()
                    if "agent" in col_lower and "name" in col_lower:
                        agent_name_col = col
                    elif col_lower == "id":
                        agent_id_col = col
                    elif "shift" in col_lower and "time" in col_lower:
                        shift_time_col = col
                    elif "shift" in col_lower and "group" in col_lower:
                        shift_group_col = col
                    elif "email" in col_lower and "id" in col_lower:
                        email_col = col
                    elif col_lower == "tfd":
                        counts_col = col

                if agent_name_col:
                    # Parse shift times from raw data
                    for _, row in agent_df.iterrows():
                        agent_name = (
                            str(row[agent_name_col]).strip()
                            if pd.notna(row[agent_name_col])
                            else "Unknown"
                        )

                        # Get agent ID
                        if agent_id_col and pd.notna(row[agent_id_col]):
                            agent_id = str(row[agent_id_col]).strip()
                        else:
                            agent_id = f"{agent_name}_{row.name}"

                        # Get shift time
                        shift_original = None
                        if shift_time_col and pd.notna(row[shift_time_col]):
                            shift_original = str(row[shift_time_col]).strip()

                        # Get shift group
                        shift_group = None
                        if shift_group_col and pd.notna(row[shift_group_col]):
                            try:
                                shift_group = int(
                                    float(str(row[shift_group_col]).strip())
                                )
                            except:
                                pass

                        # Parse shift start time (using same logic as in process_allocation_files_with_dates)
                        shift_start = None
                        shift_start_display = "Not parsed"

                        if shift_original:
                            try:
                                from datetime import time as dt_time
                                import re

                                if "-" in shift_original:
                                    parts = shift_original.split("-")
                                    if len(parts) >= 2:
                                        start_time_str = parts[0].strip()
                                        end_time_str = parts[1].strip()

                                        has_end_am = "am" in end_time_str.lower()
                                        has_end_pm = "pm" in end_time_str.lower()
                                        has_start_am = "am" in start_time_str.lower()
                                        has_start_pm = "pm" in start_time_str.lower()

                                        start_match = re.search(
                                            r"(\d{1,2})", start_time_str
                                        )
                                        if start_match:
                                            hour = int(start_match.group(1))
                                            minute = 0

                                            if has_start_am:
                                                if hour == 12:
                                                    hour = 0
                                            elif has_start_pm:
                                                if hour != 12:
                                                    hour += 12
                                            else:
                                                end_match = re.search(
                                                    r"(\d{1,2})", end_time_str
                                                )
                                                if end_match:
                                                    end_hour_12 = int(
                                                        end_match.group(1)
                                                    )

                                                    if shift_group == 1:
                                                        pass  # Keep as AM
                                                    elif shift_group == 2:
                                                        if hour >= end_hour_12:
                                                            pass
                                                        else:
                                                            if hour != 12:
                                                                hour += 12
                                                    elif shift_group == 3:
                                                        if has_end_am:
                                                            if hour >= end_hour_12:
                                                                if hour != 12:
                                                                    hour += 12
                                                            else:
                                                                if hour == 12:
                                                                    hour = 0
                                                    else:
                                                        if has_end_am:
                                                            if hour >= end_hour_12:
                                                                if hour != 12:
                                                                    hour += 12
                                                            else:
                                                                if hour == 12:
                                                                    hour = 0
                                                        elif has_end_pm:
                                                            if (
                                                                hour >= end_hour_12
                                                                and hour < 12
                                                            ):
                                                                pass
                                                            elif hour < end_hour_12:
                                                                if hour != 12:
                                                                    hour += 12

                                            minute_match = re.search(
                                                r":(\d{2})", start_time_str
                                            )
                                            if minute_match:
                                                minute = int(minute_match.group(1))

                                            shift_start = dt_time(hour % 24, minute)
                                            shift_start_str = shift_start.strftime(
                                                "%H:%M"
                                            )

                                            # Format display
                                            if hour < 12:
                                                shift_start_display = f"{shift_start_str} ({hour}:{minute:02d} AM)"
                                            elif hour == 12:
                                                shift_start_display = (
                                                    f"{shift_start_str} (12:00 PM)"
                                                )
                                            else:
                                                shift_start_display = f"{shift_start_str} ({(hour-12)}:{minute:02d} PM)"
                            except Exception as e:
                                pass

                        # Format shift group name
                        group_name = "Not set"
                        if shift_group == 1:
                            group_name = "Day Shift"
                        elif shift_group == 2:
                            group_name = "Afternoon Shift"
                        elif shift_group == 3:
                            group_name = "Night Shift"

                        # Get other info
                        agent_email = ""
                        if email_col and pd.notna(row[email_col]):
                            agent_email = str(row[email_col]).strip()

                        capacity = 0
                        if counts_col and pd.notna(row[counts_col]):
                            try:
                                capacity = int(
                                    float(str(row[counts_col]).replace(",", ""))
                                )
                            except:
                                pass

                        shift_info.append(
                            {
                                "agent_id": agent_id,
                                "agent_name": agent_name,
                                "email": agent_email or "Not set",
                                "shift_time_original": shift_original or "Not set",
                                "shift_start_time_parsed": (
                                    shift_start.strftime("%H:%M")
                                    if shift_start
                                    else "Not parsed"
                                ),
                                "shift_start_time_display": shift_start_display,
                                "shift_group": shift_group,
                                "shift_group_name": group_name,
                                "capacity": capacity,
                                "allocated": 0,
                            }
                        )

                # Sort by shift start time
                shift_info.sort(
                    key=lambda x: (
                        (
                            x["shift_start_time_parsed"]
                            if x["shift_start_time_parsed"]
                            and x["shift_start_time_parsed"] != "Not parsed"
                            else "99:99"
                        ),
                        x["agent_name"],
                    )
                )

                return jsonify(
                    {
                        "success": True,
                        "total_agents": len(shift_info),
                        "agents": shift_info,
                        "source": "raw_upload",
                        "message": "Showing shift times from uploaded staff details (file not yet processed)",
                    }
                )
        except Exception as e:
            return (
                jsonify(
                    {
                        "error": f"Error extracting shift information from uploaded file: {str(e)}"
                    }
                ),
                500,
            )

    return (
        jsonify(
            {
                "error": "No allocation data available. Please upload staff details file first."
            }
        ),
        400,
    )


def send_reminder_email(agent_info):
    """Send a reminder email to an agent prompting them to upload their work"""
    try:
        agent_name = agent_info.get("name", "Agent")
        agent_email = agent_info.get("email")
        allocated = agent_info.get("allocated", 0)

        if not agent_email:
            return False, "No email address"

        if allocated == 0:
            return False, "No allocated work to remind about"

        # Get allocation summary
        summary = get_allocation_summary(agent_name, agent_info)

        # Format insurance companies list
        insurance_list = (
            ", ".join(sorted(summary["insurance_companies"]))
            if summary["insurance_companies"]
            else "None"
        )

        # Prepare email content
        text_content = f"""
Dear {agent_name},

This is a friendly reminder to upload your completed work.

📊 YOUR CURRENT ALLOCATION:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Total Allocated: {summary['total_allocated']} rows
• First Priority: {summary['first_priority_count']} rows
• Second Priority: {summary['second_priority_count']} rows
• Third Priority: {summary['third_priority_count']} rows

🏥 INSURANCE COMPANIES:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{insurance_list}

⏰ Please log into the system and upload your completed work.

Best regards,
Allocation Management System
        """

        html_content = f"""
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                <h2 style="color: #333;">📧 Work Upload Reminder</h2>
                <p>Dear <strong>{agent_name}</strong>,</p>
                <p>This is a friendly reminder to upload your completed work.</p>
                
                <div style="background-color: #e7f3ff; padding: 15px; border-radius: 5px; margin: 20px 0;">
                    <h3 style="margin-top: 0; color: #0056b3;">📊 Your Current Allocation</h3>
                    <ul style="margin: 10px 0; padding-left: 20px;">
                        <li>Total Allocated: <strong>{summary['total_allocated']} rows</strong></li>
                        <li>First Priority: <strong>{summary['first_priority_count']} rows</strong></li>
                        <li>Second Priority: <strong>{summary['second_priority_count']} rows</strong></li>
                        <li>Third Priority: <strong>{summary['third_priority_count']} rows</strong></li>
                    </ul>
                </div>
                
                <div style="background-color: #fff3cd; padding: 15px; border-radius: 5px; margin: 20px 0;">
                    <h3 style="margin-top: 0; color: #856404;">🏥 Insurance Companies</h3>
                    <p style="word-wrap: break-word;">{insurance_list}</p>
                </div>
                
                <div style="background-color: #d4edda; padding: 15px; border-radius: 5px; margin: 20px 0; border-left: 4px solid #28a745;">
                    <p style="font-size: 16px; font-weight: bold; color: #155724; margin: 10px 0;">⏰ Please log into the system and upload your completed work.</p>
                </div>
                
                <p>Best regards,<br>
                Allocation Management System</p>
            </div>
        """

        # Send reminder email using Resend
        success, message = send_email_with_resend(
            to_email=agent_email,
            subject=f"Reminder: Please Upload Your Work - {agent_name}",
            html_content=html_content,
            text_content=text_content,
        )

        if success:
            return True, f"Reminder sent to {agent_email}"
        else:
            return False, message

    except Exception as e:
        return False, str(e)


def create_consolidated_data():
    """Create consolidated Excel data from all agent work files"""
    try:
        # Get all agent work files
        work_files = AgentWorkFile.query.order_by(
            AgentWorkFile.upload_date.desc()
        ).all()

        if not work_files:
            return None, "No agent work files found"

        # Create Excel buffer
        excel_buffer = io.BytesIO()

        # Helper function to find remark column
        def find_remark_column(df):
            """Find the remark column (case-insensitive)"""
            for col in df.columns:
                if col.lower() in ["remark", "remarks"]:
                    return col
            return None

        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            # First pass: Collect all unique remark statuses across all agents
            all_remark_statuses = set()
            agent_remarks_data = {}  # Store remarks data per agent

            for work_file in work_files:
                file_data = work_file.get_file_data()
                agent_name = work_file.agent.name
                remarks_dict = (
                    {}
                )  # Dictionary to store all remarks with counts for this agent

                if file_data:
                    if isinstance(file_data, dict):
                        # Multiple sheets - process all sheets except Summary
                        for sheet_name, sheet_data in file_data.items():
                            # Skip Summary sheet
                            if sheet_name.lower() == "summary":
                                continue

                            if isinstance(sheet_data, pd.DataFrame):
                                # Get remark column
                                remark_col = find_remark_column(sheet_data)

                                if remark_col:
                                    # Get all remarks (including NaN)
                                    remark_data = sheet_data[remark_col]

                                    # Count all non-empty remarks by status
                                    non_empty_remarks = remark_data.dropna()
                                    for remark in non_empty_remarks:
                                        # Normalize remark (strip whitespace, handle case)
                                        remark_normalized = str(remark).strip()
                                        if remark_normalized:
                                            remark_lower = remark_normalized.lower()
                                            # Use original case for display, but normalize for counting
                                            if remark_lower in remarks_dict:
                                                remarks_dict[remark_lower] = {
                                                    "count": remarks_dict[remark_lower][
                                                        "count"
                                                    ]
                                                    + 1,
                                                    "display": remarks_dict[
                                                        remark_lower
                                                    ][
                                                        "display"
                                                    ],  # Keep first seen case
                                                }
                                            else:
                                                remarks_dict[remark_lower] = {
                                                    "count": 1,
                                                    "display": remark_normalized,  # Keep original case
                                                }
                    elif isinstance(file_data, pd.DataFrame):
                        # Single DataFrame
                        # Get remark column
                        remark_col = find_remark_column(file_data)

                        if remark_col:
                            # Get all remarks (including NaN)
                            remark_data = file_data[remark_col]

                            # Count all non-empty remarks by status
                            non_empty_remarks = remark_data.dropna()
                            for remark in non_empty_remarks:
                                # Normalize remark (strip whitespace, handle case)
                                remark_normalized = str(remark).strip()
                                if remark_normalized:
                                    remark_lower = remark_normalized.lower()
                                    # Use original case for display, but normalize for counting
                                    if remark_lower in remarks_dict:
                                        remarks_dict[remark_lower] = {
                                            "count": remarks_dict[remark_lower]["count"]
                                            + 1,
                                            "display": remarks_dict[remark_lower][
                                                "display"
                                            ],  # Keep first seen case
                                        }
                                    else:
                                        remarks_dict[remark_lower] = {
                                            "count": 1,
                                            "display": remark_normalized,  # Keep original case
                                        }

                # Store remarks data for this agent
                agent_remarks_data[agent_name] = remarks_dict

                # Collect all unique remark statuses (using display name)
                for remark_info in remarks_dict.values():
                    all_remark_statuses.add(remark_info["display"])

            # Add empty remarks as a status if needed
            all_remark_statuses.add("(Empty/No Remark)")

            # Sort remark statuses alphabetically for consistent column order
            sorted_remark_statuses = sorted(all_remark_statuses)

            # Second pass: Create summary data with all remark statuses as columns
            summary_data = []

            for work_file in work_files:
                file_data = work_file.get_file_data()
                agent_name = work_file.agent.name
                total_assigned_count = 0
                completed_count = 0
                empty_remarks_count = 0

                # Calculate counts from file data
                if file_data:
                    if isinstance(file_data, dict):
                        # Multiple sheets - count rows from all sheets except Summary
                        for sheet_name, sheet_data in file_data.items():
                            # Skip Summary sheet
                            if sheet_name.lower() == "summary":
                                continue

                            if isinstance(sheet_data, pd.DataFrame):
                                # Total assigned count = all rows (excluding header)
                                total_assigned_count += len(sheet_data)

                                # Get remark column
                                remark_col = find_remark_column(sheet_data)

                                if remark_col:
                                    # Get all remarks (including NaN)
                                    remark_data = sheet_data[remark_col]

                                    # Count empty/NaN remarks
                                    empty_remarks_count += remark_data.isna().sum()

                                    # Count completed (non-Workable remarks)
                                    non_empty_remarks = remark_data.dropna()
                                    # Convert to string first to handle mixed types
                                    if len(non_empty_remarks) > 0:
                                        non_empty_remarks_str = (
                                            non_empty_remarks.astype(str).str.lower()
                                        )
                                        completed_count += len(
                                            non_empty_remarks_str[
                                                non_empty_remarks_str != "workable"
                                            ]
                                        )
                    elif isinstance(file_data, pd.DataFrame):
                        # Single DataFrame
                        # Total assigned count = all rows (excluding header)
                        total_assigned_count = len(file_data)

                        # Get remark column
                        remark_col = find_remark_column(file_data)

                        if remark_col:
                            # Get all remarks (including NaN)
                            remark_data = file_data[remark_col]

                            # Count empty/NaN remarks
                            empty_remarks_count = remark_data.isna().sum()

                            # Count completed (non-Workable remarks)
                            non_empty_remarks = remark_data.dropna()
                            # Convert to string first to handle mixed types
                            if len(non_empty_remarks) > 0:
                                non_empty_remarks_str = non_empty_remarks.astype(
                                    str
                                ).str.lower()
                                completed_count = len(
                                    non_empty_remarks_str[
                                        non_empty_remarks_str != "workable"
                                    ]
                                )
                            else:
                                completed_count = 0

                # Create row data for this agent
                row_data = {
                    "Agent": agent_name,
                    "Total Assigned Count": total_assigned_count,
                    "Completed Count": completed_count,
                    "Empty Remarks Count": empty_remarks_count,
                }

                # Get remarks data for this agent
                agent_remarks = agent_remarks_data.get(agent_name, {})

                # Add count for each remark status column
                for remark_status in sorted_remark_statuses:
                    if remark_status == "(Empty/No Remark)":
                        # Use the empty_remarks_count
                        row_data[remark_status] = empty_remarks_count
                    else:
                        # Find matching remark in agent's remarks (case-insensitive)
                        count = 0
                        for remark_lower, remark_info in agent_remarks.items():
                            if remark_info["display"] == remark_status:
                                count = remark_info["count"]
                                break
                        row_data[remark_status] = count

                summary_data.append(row_data)

            # Create summary DataFrame
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

            # Combine all agent data into one sheet
            all_agent_data = []
            for work_file in work_files:
                file_data = work_file.get_file_data()
                if file_data:
                    if isinstance(file_data, dict):
                        # Multiple sheets - combine them (excluding Summary sheets)
                        for sheet_name, sheet_data in file_data.items():
                            # Skip Summary sheet
                            if sheet_name.lower() == "summary":
                                continue

                            if isinstance(sheet_data, pd.DataFrame):
                                sheet_data_copy = sheet_data.copy()
                                all_agent_data.append(sheet_data_copy)
                    elif isinstance(file_data, pd.DataFrame):
                        # Single DataFrame
                        file_data_copy = file_data.copy()
                        all_agent_data.append(file_data_copy)

            # Create combined sheet with all agent data
            if all_agent_data:
                combined_df = pd.concat(all_agent_data, ignore_index=True)

                # Format all date columns to MM/DD/YYYY format
                for col in combined_df.columns:
                    if "date" in col.lower():
                        try:
                            # Convert to datetime if not already
                            combined_df[col] = pd.to_datetime(
                                combined_df[col], errors="coerce"
                            )
                            # Format as MM/DD/YYYY, handling NaT (Not a Time) values
                            combined_df[col] = combined_df[col].apply(
                                lambda x: x.strftime("%m/%d/%Y") if pd.notna(x) else ""
                            )
                        except Exception:
                            # If conversion fails, leave column as is
                            pass

                combined_df.to_excel(writer, sheet_name="All Agent Data", index=False)
            else:
                # Fallback if no data found
                simple_df = pd.DataFrame(
                    [{"Message": "No data available from any agent"}]
                )
                simple_df.to_excel(writer, sheet_name="All Agent Data", index=False)

        excel_buffer.seek(0)
        filename = (
            f"consolidated_agent_files_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return excel_buffer, filename

    except Exception as e:
        return None, f"Error creating consolidated data: {str(e)}"


def cleanup_all_agent_files():
    """Delete all agent work files daily at 7 AM to save server space"""
    try:
        with app.app_context():
            # Get all agent work files
            all_files = AgentWorkFile.query.all()
            file_count = len(all_files)

            # Delete all files
            for work_file in all_files:
                db.session.delete(work_file)

            db.session.commit()

            print(
                f"✅ Daily cleanup completed: Deleted {file_count} agent work file(s) at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            )
            return True, file_count
    except Exception as e:
        print(f"❌ Error during daily cleanup: {str(e)}")
        db.session.rollback()
        return False, str(e)


def daily_consolidate_and_cleanup():
    """Consolidate all agent files, email the consolidated workbook, then cleanup."""
    try:
        with app.app_context():
            excel_buffer, filename_or_message = create_consolidated_data()
            if excel_buffer is not None:
                # Use env var for recipient, fallback to sandbox-allowed email for testing
                to_email = os.environ.get("CONSOLIDATION_EMAIL", "amirmursal@gmail.com")
                subject = f"Daily Consolidated Agent Files - {datetime.now().strftime('%Y-%m-%d')}"
                html_content = f"<p>Please find attached the consolidated agent workbook generated at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.</p>"
                success, message = send_email_with_resend(
                    to_email=to_email,
                    subject=subject,
                    html_content=html_content,
                    text_content="Daily consolidated agent files attachment.",
                    attachment_data=excel_buffer.getvalue(),
                    attachment_filename=filename_or_message,
                )
                if success:
                    print(
                        f"✅ Daily consolidation email sent to {to_email}: {filename_or_message}"
                    )
                else:
                    print(f"❌ Failed to send consolidation email: {message}")
            else:
                print(f"⚠️ Skipping email - {filename_or_message}")

            # Perform cleanup after emailing
            all_files = AgentWorkFile.query.all()
            file_count = len(all_files)
            for work_file in all_files:
                db.session.delete(work_file)
            db.session.commit()
            print(
                f"✅ Daily consolidation + cleanup complete. Deleted {file_count} file(s)."
            )
            return True, file_count
    except Exception as e:
        print(f"❌ Error in daily consolidation + cleanup: {str(e)}")
        db.session.rollback()
        return False, str(e)


def check_and_send_reminders():
    """Check which agents need reminders and send them every 2 hours from shift start time"""
    global agent_allocations_for_reminders

    if not agent_allocations_for_reminders:
        return

    # Get timezone from environment (default: IST for local, UTC for Railway)
    # Shift times are stored in IST, so we need to work in IST timezone
    reminder_timezone_str = os.environ.get(
        "REMINDER_TIMEZONE", "Asia/Kolkata"
    )  # Default to IST
    reminder_timezone = pytz.timezone(reminder_timezone_str)

    # Get current time in the specified timezone
    current_time_utc = datetime.now(pytz.UTC)
    current_time = current_time_utc.astimezone(reminder_timezone)
    current_hour = current_time.hour
    current_minute = current_time.minute

    successful_reminders = []
    failed_reminders = []

    for agent in agent_allocations_for_reminders:
        shift_start_time_str = agent.get("shift_start_time")
        agent_email = agent.get("email")
        allocated = agent.get("allocated", 0)

        # Skip if no shift start time, email, or allocated work
        if not shift_start_time_str or not agent_email or allocated == 0:
            continue

        try:
            # Parse shift start time (format: HH:MM) - shift times are in local timezone
            shift_hour, shift_minute = map(int, shift_start_time_str.split(":"))
            # Create shift start time in the reminder timezone
            shift_start_today = reminder_timezone.localize(
                current_time.replace(
                    hour=shift_hour, minute=shift_minute, second=0, microsecond=0
                )
            )

            # If shift hasn't started yet today, skip
            if shift_start_today > current_time:
                continue

            # Calculate hours since shift started today
            hours_since_start = (
                current_time - shift_start_today
            ).total_seconds() / 3600

            # Calculate which reminder interval we're at (0, 2, 4, 6, 8, etc. hours)
            reminder_interval = 2  # hours
            interval_number = int(hours_since_start // reminder_interval)
            next_interval_time = shift_start_today + timedelta(
                hours=interval_number * reminder_interval
            )

            # Check if current time is within 5 minutes before or after a reminder interval
            tolerance_minutes = 5
            time_diff = abs((current_time - next_interval_time).total_seconds() / 60)

            if time_diff <= tolerance_minutes:
                # We're at a reminder interval. Check if we haven't sent one recently
                last_reminder_key = f"last_reminder_{agent.get('id')}"
                if not hasattr(app, "_reminder_tracker"):
                    app._reminder_tracker = {}

                last_reminder_time = app._reminder_tracker.get(last_reminder_key)

                if last_reminder_time:
                    # Convert last reminder time to timezone-aware for comparison
                    if (
                        isinstance(last_reminder_time, datetime)
                        and last_reminder_time.tzinfo is None
                    ):
                        last_reminder_time = reminder_timezone.localize(
                            last_reminder_time
                        )
                    minutes_since_last = (
                        current_time - last_reminder_time
                    ).total_seconds() / 60
                    if (
                        minutes_since_last < 100
                    ):  # Don't send if sent within last 100 minutes
                        continue

                # Send reminder
                success, message = send_reminder_email(agent)
                if success:
                    successful_reminders.append(f"{agent.get('name')} ({agent_email})")
                    if not hasattr(app, "_reminder_tracker"):
                        app._reminder_tracker = {}
                    # Store as timezone-aware datetime
                    app._reminder_tracker[last_reminder_key] = current_time
                else:
                    failed_reminders.append(f"{agent.get('name')}: {message}")

        except Exception as e:
            failed_reminders.append(f"{agent.get('name')}: {str(e)}")

    # Log reminder results
    if successful_reminders or failed_reminders:
        print(
            f"[Reminder System] Sent {len(successful_reminders)} reminders, {len(failed_reminders)} failed at {current_time.strftime('%Y-%m-%d %H:%M:%S %Z')}"
        )
        if failed_reminders:
            print(
                f"[Reminder System] Failed: {', '.join(failed_reminders[:5])}"
            )  # Show first 5 failures


def get_allocation_summary(agent_name, agent_info):
    """Get detailed allocation summary for an agent"""
    global data_file_data

    summary = {
        "total_allocated": agent_info.get("allocated", 0),
        "capacity": agent_info.get("capacity", 0),
        "first_priority_count": 0,
        "second_priority_count": 0,
        "third_priority_count": 0,
        "unknown_priority_count": 0,
        "insurance_companies": set(),
        "first_priority_deadline": None,
    }

    # Get the agent's allocated rows
    row_indices = agent_info.get("row_indices", [])

    if row_indices and data_file_data:
        # Get the main data
        if isinstance(data_file_data, dict):
            first_sheet_name = list(data_file_data.keys())[0]
            main_df = data_file_data[first_sheet_name]
        else:
            main_df = data_file_data

        if len(main_df) > 0:
            # Get allocated rows
            allocated_df = main_df.iloc[row_indices].copy()

            # Count by priority
            if "Priority Status" in allocated_df.columns:
                summary["first_priority_count"] = len(
                    allocated_df[allocated_df["Priority Status"] == "First Priority"]
                )
                summary["second_priority_count"] = len(
                    allocated_df[allocated_df["Priority Status"] == "Second Priority"]
                )
                summary["third_priority_count"] = len(
                    allocated_df[allocated_df["Priority Status"] == "Third Priority"]
                )
                unknown_mask = (
                    allocated_df["Priority Status"].isin(["", "Unknown", None])
                    | allocated_df["Priority Status"].isna()
                )
                summary["unknown_priority_count"] = len(allocated_df[unknown_mask])

            # Get unique insurance companies
            insurance_col = None
            for col in allocated_df.columns:
                if (
                    "dental" in col.lower()
                    and "primary" in col.lower()
                    and "ins" in col.lower()
                ):
                    insurance_col = col
                    break

            if insurance_col:
                insurance_companies = allocated_df[insurance_col].dropna().unique()
                summary["insurance_companies"] = set(
                    [
                        str(ic).strip()
                        for ic in insurance_companies
                        if str(ic).strip() and str(ic).strip().lower() != "unknown"
                    ]
                )

            # Calculate First Priority deadline (2nd business day end of day)
            if summary["first_priority_count"] > 0:
                from datetime import datetime, time

                today = datetime.now().date()
                second_business_day = get_nth_business_day(today, 2)
                # Set deadline to end of business day (5:00 PM) on 2nd business day
                summary["first_priority_deadline"] = datetime.combine(
                    second_business_day, time(17, 0)
                )

    return summary


def create_agent_excel_file(agent_name, agent_info):
    """Create Excel file with agent's allocated data"""
    try:
        # Get the agent's allocated row indices
        row_indices = agent_info.get("row_indices", [])

        if not row_indices or data_file_data is None:
            # If no specific rows or no data, create empty DataFrame
            allocated_df = pd.DataFrame(
                {"Message": ["No data allocated to this agent"]}
            )
        else:
            # data_file_data is a dictionary, get the first sheet (main data)
            if isinstance(data_file_data, dict):
                # Get the first sheet from the dictionary
                first_sheet_name = list(data_file_data.keys())[0]
                main_df = data_file_data[first_sheet_name]
            else:
                # If it's already a DataFrame
                main_df = data_file_data

            # CRITICAL: Filter by Agent Name column to ensure only this agent's rows are included
            if "Agent Name" in main_df.columns:
                # Filter to only include rows where Agent Name matches this agent exactly
                allocated_df = main_df[main_df["Agent Name"] == agent_name].copy()

                # If no rows found by name, verify row_indices match Agent Name
                if allocated_df.empty:
                    valid_indices = []
                    for idx in row_indices:
                        if idx < len(main_df):
                            row_agent_name = main_df.at[idx, "Agent Name"]
                            if (
                                pd.notna(row_agent_name)
                                and str(row_agent_name).strip() == agent_name
                            ):
                                valid_indices.append(idx)

                    if valid_indices:
                        allocated_df = main_df.iloc[valid_indices].copy()
            else:
                # If Agent Name column doesn't exist, use row_indices directly
                allocated_df = main_df.iloc[row_indices].copy()

        # Create Excel buffer
        excel_buffer = io.BytesIO()

        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            # Write main data
            allocated_df.to_excel(writer, sheet_name="Allocated Data", index=False)

            # Create summary sheet
            summary_data = {
                "Agent Name": [agent_name],
                "Total Allocated": [agent_info["allocated"]],
                "Capacity": [agent_info["capacity"]],
                "Allocation Date": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                "Status": ["Approved"],
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

        excel_buffer.seek(0)
        return excel_buffer

    except Exception as e:
        # Return empty Excel file as fallback
        excel_buffer = io.BytesIO()
        empty_df = pd.DataFrame({"Message": ["No data available"]})
        empty_df.to_excel(excel_buffer, index=False)
        excel_buffer.seek(0)
        return excel_buffer


@app.route("/reset_app", methods=["POST"])
@admin_required
def reset_app():
    global allocation_data, data_file_data, allocation_filename, data_filename, processing_result
    global agent_allocations_data

    try:
        # Do NOT clear agent work files - preserve all agent files (both uploaded and consolidated)

        # Clear all allocations from database
        Allocation.query.delete()

        # Reset all global variables
        allocation_data = None
        data_file_data = None
        allocation_filename = None
        data_filename = None
        processing_result = "🔄 Application reset successfully! All uploaded files and data have been cleared. All agent work files have been preserved."
        agent_allocations_data = None

        # Commit database changes
        db.session.commit()

        return redirect("/")

    except Exception as e:
        db.session.rollback()
        processing_result = f"❌ Error resetting application: {str(e)}"
        return redirect("/")


if __name__ == "__main__":
    import os
    import threading
    import time

    # Initialize database
    init_database()

    # Load insurance name mapping at startup
    load_insurance_name_mapping()

    # Start session cleanup thread
    def cleanup_sessions_periodically():
        while True:
            try:
                with app.app_context():
                    cleanup_expired_sessions()
                time.sleep(3600)  # Clean up every hour
            except Exception as e:
                time.sleep(3600)

    cleanup_thread = threading.Thread(target=cleanup_sessions_periodically, daemon=True)
    cleanup_thread.start()

    # Set up scheduler for reminder emails and daily cleanup
    scheduler = BackgroundScheduler()

    # Reminder emails - every 2 hours
    scheduler.add_job(
        func=lambda: check_and_send_reminders(),
        trigger=IntervalTrigger(hours=2),
        id="reminder_check",
        name="Check and send reminder emails every 2 hours",
        replace_existing=True,
    )

    # Cleanup - every day at 7:00 AM (timezone-aware)
    # Get timezone from environment variable (default: IST for local, UTC for Railway)
    # Railway servers run in UTC, so we need to convert local time to UTC
    # IST is UTC+5:30, so 7 AM IST = 1:30 AM UTC
    cleanup_timezone_str = os.environ.get(
        "CLEANUP_TIMEZONE", "Asia/Kolkata"
    )  # Default to IST
    cleanup_timezone = pytz.timezone(cleanup_timezone_str)

    # Schedule time in local timezone (7 AM = 07:00)
    cleanup_hour = int(os.environ.get("CLEANUP_HOUR", "7"))  # 7 AM
    cleanup_minute = int(os.environ.get("CLEANUP_MINUTE", "0"))

    scheduler.add_job(
        func=lambda: daily_consolidate_and_cleanup(),
        trigger=CronTrigger(
            hour=cleanup_hour, minute=cleanup_minute, timezone=cleanup_timezone
        ),
        id="daily_consolidation_cleanup",
        name=f"Daily consolidation email + cleanup at {cleanup_hour:02d}:{cleanup_minute:02d} {cleanup_timezone_str}",
        replace_existing=True,
    )

    scheduler.start()
    print("✅ Reminder email scheduler started - checking every 2 hours")
    # Calculate UTC equivalent for display
    local_time = cleanup_timezone.localize(
        datetime(2025, 1, 1, cleanup_hour, cleanup_minute)
    )
    utc_time = local_time.astimezone(pytz.UTC)
    print(
        f"✅ Cleanup scheduler started - runs every day at {cleanup_hour:02d}:{cleanup_minute:02d} {cleanup_timezone_str} (UTC: {utc_time.strftime('%H:%M')})"
    )

    port = int(os.environ.get("PORT", 5003))
    # Always enable debug + auto-reload for local dev unless explicitly disabled
    debug = True if os.environ.get("DISABLE_DEBUG") != "1" else False

    try:
        app.run(debug=debug, host="0.0.0.0", port=port, use_reloader=debug)
    finally:
        # Shutdown scheduler when app stops
        if scheduler.running:
            scheduler.shutdown()
